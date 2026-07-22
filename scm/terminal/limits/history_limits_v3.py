import sys
import math
import pandas as pd
import statistics
import concurrent.futures
from datetime import datetime, timezone, timedelta
from google.cloud import monitoring_v3
from kubernetes import client, config
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import time

# ================= CONFIG =================
DIAS_A_ANALIZAR = 90
MAX_TRABAJADORES = 10

METRICA_CPU = "kubernetes.io/container/cpu/core_usage_time"
METRICA_MEM = "kubernetes.io/container/memory/used_bytes"

CPU_STEP = 10     # millicores
MEM_STEP = 50     # MiB

# ================= HELPERS =================
def round_up(v, step):
    return int(math.ceil(v / step) * step)

def cpu_to_m(v):
    if not v or v == "N/A":
        return None
    s = str(v)
    try:
        if s.endswith("m"):
            return float(s[:-1])
        return float(s) * 1000
    except:
        return None

def mem_to_mi(v):
    if not v or v == "N/A":
        return None
    s = str(v).upper()
    try:
        if s.endswith("MI"):
            return float(s[:-2])
        if s.endswith("GI"):
            return float(s[:-2]) * 1024
        if s.endswith("M"):
            return float(s[:-1])
        if s.endswith("G"):
            return float(s[:-1]) * 1024
        return float(s)
    except:
        return None

# ================= K8S =================
def configurar_k8s():
    try:
        config.load_incluster_config()
    except:
        config.load_kube_config()

# ================= HPA =================
def obtener_hpas():
    hpas = {}
    try:
        api = client.AutoscalingV2Api()
        for h in api.list_horizontal_pod_autoscaler_for_all_namespaces().items:
            hpas[(h.metadata.namespace, h.spec.scale_target_ref.name)] = {
                "min": h.spec.min_replicas,
                "max": h.spec.max_replicas,
                "target_cpu": next(
                    (
                        m.resource.target.average_utilization
                        for m in (h.spec.metrics or [])
                        if m.type == "Resource" and m.resource.name == "cpu"
                    ),
                    None
                )
            }
    except:
        pass
    return hpas

# ================= DEPLOYMENTS =================
def obtener_contenedores(hpas):
    conts = []
    api = client.AppsV1Api()
    for d in api.list_deployment_for_all_namespaces().items:
        hpa = hpas.get((d.metadata.namespace, d.metadata.name), {})
        for c in d.spec.template.spec.containers:
            r = c.resources or {}
            conts.append({
                "namespace": d.metadata.namespace,
                "deployment": d.metadata.name,
                "container": c.name,
                "request_cpu": (r.requests or {}).get("cpu"),
                "limit_cpu": (r.limits or {}).get("cpu"),
                "request_mem": (r.requests or {}).get("memory"),
                "limit_mem": (r.limits or {}).get("memory"),
                "hpa_min": hpa.get("min"),
                "hpa_max": hpa.get("max"),
                "hpa_target_cpu": hpa.get("target_cpu")
            })
    return conts

# ================= MONITORING =================
def fetch_metric(project, cluster, ns, cont, metric, mem_type=None):
    cli = monitoring_v3.MetricServiceClient()
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=DIAS_A_ANALIZAR)

    filtro = (
        f'metric.type="{metric}" '
        f'AND resource.labels.cluster_name="{cluster}" '
        f'AND resource.labels.namespace_name="{ns}" '
        f'AND resource.labels.container_name="{cont}"'
    )
    if mem_type:
        filtro += f' AND metric.labels.memory_type="{mem_type}"'

    alineador = (
        monitoring_v3.Aggregation.Aligner.ALIGN_RATE
        if "cpu" in metric
        else monitoring_v3.Aggregation.Aligner.ALIGN_MEAN
    )

    valores = []
    try:
        for s in cli.list_time_series(
            request={
                "name": f"projects/{project}",
                "filter": filtro,
                "interval": {
                    "start_time": {"seconds": int(start.timestamp())},
                    "end_time": {"seconds": int(end.timestamp())},
                },
                "aggregation": {
                    "alignment_period": {"seconds": 60},
                    "per_series_aligner": alineador,
                },
            }
        ):
            for p in s.points:
                valores.append(p.value.double_value)
    except:
        pass
    return valores

def summarize(vals, scale):
    if not vals:
        return {}
    v = [x * scale for x in vals]
    v.sort()
    return {
        "avg": statistics.mean(v),
        "min": min(v),
        "max": max(v),
        "p95": v[int(0.95 * (len(v) - 1))],
        "p99": v[int(0.99 * (len(v) - 1))],
    }

# ================= ANALYSIS =================
def analyze(c, project, cluster):
    cpu = summarize(fetch_metric(project, cluster, c["namespace"], c["container"], METRICA_CPU), 1000)
    mem_total = summarize(fetch_metric(project, cluster, c["namespace"], c["container"], METRICA_MEM), 1/(1024*1024))
    mem_non = summarize(fetch_metric(project, cluster, c["namespace"], c["container"], METRICA_MEM, "non-evictable"), 1/(1024*1024))
    mem_ev = summarize(fetch_metric(project, cluster, c["namespace"], c["container"], METRICA_MEM, "evictable"), 1/(1024*1024))

    req_cpu = cpu_to_m(c["request_cpu"])
    lim_cpu = cpu_to_m(c["limit_cpu"])
    req_mem = mem_to_mi(c["request_mem"])
    lim_mem = mem_to_mi(c["limit_mem"])

    rec_req_cpu = rec_lim_cpu = rec_req_mem = rec_lim_mem = None

    if cpu:
        calc = round_up(max(cpu["avg"] * 1.2, cpu["p95"]), CPU_STEP)
        if req_cpu is not None and calc > req_cpu:
            rec_req_cpu = calc

        calc = round_up(cpu["p99"] * 1.5, CPU_STEP)
        if lim_cpu is not None and calc > lim_cpu:
            rec_lim_cpu = calc

    if mem_total:
        calc = round_up(max(mem_total["avg"] * 1.2, mem_total["p95"]), MEM_STEP)
        if req_mem is not None and calc > req_mem:
            rec_req_mem = calc

        calc = round_up(mem_total["p99"] * 1.4, MEM_STEP)
        if lim_mem is not None and calc > lim_mem:
            rec_lim_mem = calc

    c.update({
        # CPU
        "cpu_avg_m": round(cpu.get("avg", 0), 2),
        "cpu_min_m": round(cpu.get("min", 0), 2),
        "cpu_max_m": round(cpu.get("max", 0), 2),
        "cpu_p95_m": round(cpu.get("p95", 0), 2),
        "cpu_p99_m": round(cpu.get("p99", 0), 2),

        # MEM TOTAL
        "mem_avg_Mi": round(mem_total.get("avg", 0), 2),
        "mem_min_Mi": round(mem_total.get("min", 0), 2),
        "mem_max_Mi": round(mem_total.get("max", 0), 2),
        "mem_p95_Mi": round(mem_total.get("p95", 0), 2),
        "mem_p99_Mi": round(mem_total.get("p99", 0), 2),

        # MEM DESGLOSE
        "mem_non_evict_avg_Mi": round(mem_non.get("avg", 0), 2),
        "mem_evict_avg_Mi": round(mem_ev.get("avg", 0), 2),

        # RECOMENDACIONES
        "cpu_request_sugerido_m": rec_req_cpu,
        "cpu_limit_sugerido_m": rec_lim_cpu,
        "mem_request_sugerido_Mi": rec_req_mem,
        "mem_limit_sugerido_Mi": rec_lim_mem,
    })

    return c

# ================= EXCEL =================
def format_excel(f):
    wb = load_workbook(f)
    ws = wb.active
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    wb.save(f)

# ================= PROGRESS =================
class ProgressTracker:
    def __init__(self, total):
        self.total = total
        self.completed = 0
        self.start_time = time.time()
    
    def update(self, item_name=""):
        self.completed += 1
        elapsed = time.time() - self.start_time
        rate = self.completed / elapsed if elapsed > 0 else 0
        remaining = (self.total - self.completed) / rate if rate > 0 else 0
        
        pct = (self.completed / self.total) * 100
        bar_width = 30
        filled = int((self.completed / self.total) * bar_width)
        bar = "█" * filled + "░" * (bar_width - filled)
        
        eta_min = int(remaining / 60)
        eta_sec = int(remaining % 60)
        
        sys.stderr.write(
            f"\r[{self.completed}/{self.total}] {bar} {pct:5.1f}% | "
            f"ETA: {eta_min}m {eta_sec:02d}s | {item_name[:40]:<40}"
        )
        sys.stderr.flush()
    
    def finish(self):
        elapsed = time.time() - self.start_time
        elapsed_min = int(elapsed / 60)
        elapsed_sec = int(elapsed % 60)
        sys.stderr.write(f"\n✅ Completado en {elapsed_min}m {elapsed_sec:02d}s\n")
        sys.stderr.flush()

# ================= MAIN =================
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python history_limits_v3.py <PROJECT_ID> <CLUSTER> [DIAS]")
        sys.exit(1)

    project = sys.argv[1]
    cluster = sys.argv[2]
    if len(sys.argv) == 4:
        DIAS_A_ANALIZAR = int(sys.argv[3])

    print(f"[INICIO] Analizando clúster: {cluster}")
    print(f"[CONFIG] Período: {DIAS_A_ANALIZAR} días | Workers: {MAX_TRABAJADORES}")
    
    print("[PASO 1/5] Configurando Kubernetes...", file=sys.stderr)
    configurar_k8s()
    
    print("[PASO 2/5] Obteniendo HPAs...", file=sys.stderr)
    hpas = obtener_hpas()
    print(f"  → {len(hpas)} HPAs encontrados", file=sys.stderr)
    
    print("[PASO 3/5] Obteniendo contenedores...", file=sys.stderr)
    conts = obtener_contenedores(hpas)
    print(f"  → {len(conts)} contenedores encontrados", file=sys.stderr)
    
    if not conts:
        print("⚠️ No se encontraron contenedores en el clúster", file=sys.stderr)
        sys.exit(0)
    
    print(f"[PASO 4/5] Analizando métricas ({len(conts)} contenedores)...", file=sys.stderr)
    
    progress = ProgressTracker(len(conts))
    
    def analyze_with_progress(c):
        result = analyze(c, project, cluster)
        progress.update(f"{c['namespace']}/{c['deployment']}/{c['container']}")
        return result
    
    with concurrent.futures.ThreadPoolExecutor(MAX_TRABAJADORES) as pool:
        data = list(pool.map(analyze_with_progress, conts))
    
    progress.finish()
    
    print("[PASO 5/5] Generando reporte Excel...", file=sys.stderr)
    df = pd.DataFrame(data)
    out = f"reporte_recursos_{cluster}.xlsx"
    df.to_excel(out, index=False)
    format_excel(out)

    print(f"✅ Reporte generado: {out}")

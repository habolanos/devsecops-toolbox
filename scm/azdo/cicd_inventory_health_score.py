#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline Health Score Report v1.0.0
Orquestador de salud de pipelines CI/CD con scoring multi-dimensional basado en DORA/SRE.

Genera 1 Excel con 3 pestañas:
  1. CI Inventory — datos consolidados de pipelines CI
  2. CD Inventory — datos consolidados de pipelines CD
  3. Health Score — scoring DORA/SRE + recomendaciones

Uso:
    python cicd_inventory_health_score.py --org Coppel-Retail --project "Compras.RMI"
    python cicd_inventory_health_score.py --org Coppel-Retail --project "Compras.RMI" --offline
    python cicd_inventory_health_score.py --org Coppel-Retail --project "Compras.RMI" --skip-incremental

Autor: Harold Adrian Bolanos Rodriguez
"""

import os
import sys
import re
import time
import json
import glob
import math
import subprocess
import requests
import pandas as pd
import argparse
from datetime import datetime, timezone, timedelta
from base64 import b64encode
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed, Future

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = lambda *a, **k: None

try:
    from rich.progress import Progress, SpinnerColumn, BarColumn, TaskProgressColumn, TextColumn, TimeElapsedColumn
    from rich.console import Console
    from rich.table import Table
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    from utils import get_output_dir, resolve_output_path
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
    from datetime import datetime as _dt
    _FMT_EXT = {"excel": ".xlsx", "csv": ".csv", "json": ".json"}
    def resolve_output_path(output_arg, base_name, default_format="excel"):
        output_dir = get_output_dir("outcome")
        output_dir.mkdir(parents=True, exist_ok=True)
        ext = _FMT_EXT.get(default_format, ".xlsx")
        if not output_arg:
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        if output_arg.lower() in _FMT_EXT:
            ext = _FMT_EXT[output_arg.lower()]
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        p = _Path(output_arg)
        if p.suffix == "":
            p = p.with_suffix(ext)
        return str(p.resolve())

load_dotenv(Path(__file__).parent.parent / ".env")

SCRIPT_NAME = "cicd_inventory_health_score"
DEFAULT_ORG = "Coppel-Retail"
DEFAULT_PROJECT = "Compras.RMI"
API_VERSION = "7.1"
DEFAULT_WORKERS = 30
CACHE_TTL_HOURS = 24

# ==========================================================
# UTILIDADES
# ==========================================================

class TeeWriter:
    def __init__(self, log_path):
        self.terminal = sys.__stdout__
        self.log = open(log_path, "w", encoding="utf-8")
        self.log_path = log_path
        self._paused = False
    def write(self, message):
        self.log.write(message)
        if not self._paused:
            self.terminal.write(message)
    def flush(self):
        self.log.flush()
        if not self._paused:
            self.terminal.flush()
    def close(self):
        self.log.close()
    def pause_terminal(self):
        self._paused = True
    def resume_terminal(self):
        self._paused = False


def setup_logging():
    output_dir = get_output_dir("outcome")
    output_dir.mkdir(parents=True, exist_ok=True)
    log_path = output_dir / f"{SCRIPT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeWriter(log_path)
    sys.stdout = tee
    print(f"📝 Log: {log_path.resolve()}")
    print(f"📅 Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    return tee


def teardown_logging(tee):
    print(f"\n📝 Log guardado: {tee.log_path.resolve()}")
    sys.stdout = tee.terminal
    tee.close()


def _progress_context():
    if RICH_AVAILABLE:
        console = Console(file=sys.__stdout__)
        return Progress(
            SpinnerColumn(), TextColumn("[bold blue]{task.description}"),
            BarColumn(bar_width=40), TaskProgressColumn(),
            TextColumn("({task.completed}/{task.total})"), TimeElapsedColumn(),
            console=console,
        )
    return None


def get_headers(pat: str):
    auth = b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {auth}", "Content-Type": "application/json"}


def az_get(url, headers, params=None, max_retries=5):
    params = params or {}
    params["api-version"] = API_VERSION
    for attempt in range(max_retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=30)
            if r.status_code >= 500:
                wait = 2 ** attempt
                print(f"⚠️  {r.status_code} en {url[:60]}... retry {attempt+1}/{max_retries}")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.exceptions.HTTPError:
            raise
        except requests.exceptions.RequestException as e:
            if attempt == max_retries - 1:
                raise
            wait = 2 ** attempt
            print(f"⚠️  Error en {url[:60]}... retry {attempt+1}/{max_retries}: {e}")
            time.sleep(wait)
    return {}


def normalize_org(org: str) -> str:
    """Extrae nombre de organización desde URL o nombre simple."""
    if org.startswith("http"):
        return org.rstrip("/").split("/")[-1]
    return org


def safe_az_get(url, headers, params=None):
    try:
        return az_get(url, headers, params)
    except Exception as e:
        print(f"   ❌ Error: {e}")
        return {}


def _find_latest_cache(pattern_key: str):
    """Busca archivo cache más reciente por patrón de nombre."""
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    pattern = str(cache_dir / f"{pattern_key}_raw_*.json")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    return Path(files[0])


def _cache_is_fresh(cache_path, ttl_hours=CACHE_TTL_HOURS):
    if not cache_path or not cache_path.exists():
        return False
    mtime = cache_path.stat().st_mtime
    age_hours = (time.time() - mtime) / 3600
    return age_hours < ttl_hours


def _load_cache(cache_path):
    with open(cache_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_cache(data, prefix):
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{prefix}_raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return cache_path


# ==========================================================
# FETCH BASE DATA (CI/CD definitions)
# ==========================================================

def fetch_ci_definitions(headers, org, project):
    url = f"https://dev.azure.com/{org}/{project}/_apis/build/definitions"
    data = safe_az_get(url, headers, {"$top": 5000})
    return data.get("value", []) if isinstance(data, dict) else []


def fetch_cd_definitions(headers, org, project):
    url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/definitions"
    data = safe_az_get(url, headers, {"$top": 5000})
    return data.get("value", []) if isinstance(data, dict) else []


# ==========================================================
# FETCH INCREMENTAL: Last 20 executions per pipeline
# ==========================================================

def fetch_ci_builds(def_id, headers, org, project, top=20):
    url = f"https://dev.azure.com/{org}/{project}/_apis/build/builds"
    data = safe_az_get(url, headers, {"definitions": def_id, "$top": top})
    return data.get("value", []) if isinstance(data, dict) else []


def fetch_cd_releases(def_id, headers, org, project, top=20):
    url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/releases"
    data = safe_az_get(url, headers, {"definitionId": def_id, "$top": top})
    return data.get("value", []) if isinstance(data, dict) else []


# ==========================================================
# TECNOLOGÍA Y CICLO DE VIDA
# ==========================================================

TECH_VERSION_PATTERNS = [
    (r"springboot[\s_-]?3", "Spring Boot", "3.x", "Moderna"),
    (r"springboot[\s_-]?2", "Spring Boot", "2.x", "Mantenimiento"),
    (r"spring[\s_-]?boot", "Spring Boot", "", "Desconocido"),
    (r"angular[\s_-]?(19|2[0-9])", "Angular", "19+", "Moderna"),
    (r"angular[\s_-]?(14|15|16|17|18)", "Angular", "14-18", "Mantenimiento"),
    (r"angular", "Angular", "<14", "EOL"),
    (r"react", "React", "", "Moderna"),
    (r"dotnet[\s_-]?(8|9)", ".NET", "8/9", "Moderna"),
    (r"dotnet[\s_-]?6", ".NET", "6", "Mantenimiento"),
    (r"dotnet[\s_-]?(core|5|7)", ".NET Core", "", "Mantenimiento"),
    (r"\.net[\s_-]?framework", ".NET Framework", "4.x", "EOL"),
    (r"net[\s_-]?framework", ".NET Framework", "4.x", "EOL"),
    (r"php[\s_-]?(8\.[2-9])", "PHP", "8.2+", "Moderna"),
    (r"php[\s_-]?8\.1", "PHP", "8.1", "Mantenimiento"),
    (r"php[\s_-]?(7|8\.0|5)", "PHP", "7/8.0/5", "EOL"),
    (r"php", "PHP", "", "Desconocido"),
    (r"java[\s_-]?(21|17)", "Java", "21/17", "Moderna"),
    (r"java[\s_-]?11", "Java", "11", "Mantenimiento"),
    (r"java[\s_-]?8", "Java", "8", "EOL"),
    (r"java", "Java", "", "Desconocido"),
    (r"kotlin", "Kotlin", "", "Moderna"),
    (r"android", "Android", "", "Moderna"),
    (r"nodejs[\s_-]?(20|22)", "Node.js", "20/22", "Moderna"),
    (r"nodejs[\s_-]?18", "Node.js", "18", "Mantenimiento"),
    (r"nodejs[\s_-]?(16|14)", "Node.js", "16/14", "EOL"),
    (r"node[\s_-]?(20|22)", "Node.js", "20/22", "Moderna"),
    (r"node[\s_-]?18", "Node.js", "18", "Mantenimiento"),
    (r"node", "Node.js", "", "Desconocido"),
    (r"python[\s_-]?(3\.(11|12))", "Python", "3.11/3.12", "Moderna"),
    (r"python[\s_-]?(3\.(9|10))", "Python", "3.9/3.10", "Mantenimiento"),
    (r"python[\s_-]?(3\.[0-8]|2)", "Python", "3.8-/2.x", "EOL"),
    (r"python", "Python", "", "Desconocido"),
    (r"gke|k8s|kubernetes", "Kubernetes", "", "Moderna"),
    (r"docker", "Docker", "", "Moderna"),
    (r"aws", "AWS", "", "Moderna"),
]


def detect_technology_status(name: str):
    name_lower = name.lower()
    for pattern, tech, version, status in TECH_VERSION_PATTERNS:
        if re.search(pattern, name_lower):
            return tech, version, status
    return "Desconocido", "", "Desconocido"


# ==========================================================
# SCORING FUNCTIONS
# ==========================================================

def days_since(dt_str):
    if not dt_str:
        return 9999
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return (datetime.now(timezone.utc) - dt).days
    except Exception:
        return 9999


def calculate_recency_score(last_exec_str):
    days = days_since(last_exec_str)
    if days <= 1: return 20
    if days <= 7: return 17
    if days <= 14: return 13
    if days <= 30: return 9
    if days <= 60: return 5
    if days <= 90: return 2
    return 0


def calculate_reliability_score(builds):
    if not builds:
        return 0, 0, 5  # no data, no mttr, 5 pts neutral
    total = len(builds)
    successes = sum(1 for b in builds if b.get("result") == "succeeded")
    failures = total - successes
    success_rate = successes / total if total > 0 else 0
    
    # Success rate pts
    if success_rate >= 1.0: sr_pts = 15
    elif success_rate >= 0.95: sr_pts = 12
    elif success_rate >= 0.90: sr_pts = 9
    elif success_rate >= 0.80: sr_pts = 6
    elif success_rate >= 0.60: sr_pts = 3
    else: sr_pts = 0
    
    # MTTR calculation
    mttr_minutes = 0
    if failures > 0 and total >= 2:
        recovery_times = []
        for i in range(len(builds) - 1):
            if builds[i].get("result") != "succeeded":
                # Look ahead for next success
                for j in range(i + 1, len(builds)):
                    if builds[j].get("result") == "succeeded":
                        t1 = datetime.fromisoformat(builds[i].get("finishTime", "").replace("Z", "+00:00")) if builds[i].get("finishTime") else None
                        t2 = datetime.fromisoformat(builds[j].get("finishTime", "").replace("Z", "+00:00")) if builds[j].get("finishTime") else None
                        if t1 and t2:
                            recovery_times.append((t2 - t1).total_seconds() / 60)
                        break
        if recovery_times:
            mttr_minutes = sum(recovery_times) / len(recovery_times)
    
    # MTTR pts
    if mttr_minutes == 0:
        if failures == 0:
            mttr_pts = 10
        else:
            mttr_pts = 5
    elif mttr_minutes <= 15: mttr_pts = 10
    elif mttr_minutes <= 60: mttr_pts = 8
    elif mttr_minutes <= 240: mttr_pts = 6
    elif mttr_minutes <= 1440: mttr_pts = 3
    else: mttr_pts = 0
    
    return min(sr_pts + mttr_pts, 25), mttr_minutes, sr_pts


def calculate_usage_score(count_30d):
    if count_30d is None: count_30d = 0
    if count_30d > 200: return 20
    if count_30d >= 50: return 17
    if count_30d >= 20: return 13
    if count_30d >= 5: return 8
    if count_30d >= 1: return 3
    return 0


def calculate_freshness_score(modified_str):
    days = days_since(modified_str)
    if days <= 7: return 15
    if days <= 30: return 12
    if days <= 90: return 9
    if days <= 180: return 5
    if days <= 365: return 2
    return 0


def calculate_tech_debt_score(process_type, has_yaml, repo_name, tech_status, pipeline_name):
    penalties = 0
    if process_type == "designerJson":
        penalties += 8
    if tech_status == "EOL":
        penalties += 7
    elif tech_status == "Mantenimiento":
        penalties += 4
    if not has_yaml and process_type != "designerJson":
        penalties += 5
    if not repo_name:
        penalties += 3
    name_lower = pipeline_name.lower()
    for kw in ["obsoleto", "obsolete", "_old", "legacy-", "deprecated"]:
        if kw in name_lower:
            penalties += 2
            break
    return max(0, 20 - penalties)


def calculate_dora_profile(count_30d, failure_rate, mttr_minutes):
    if count_30d > 200 and failure_rate < 0.05 and mttr_minutes < 60:
        return "Elite"
    if count_30d >= 50 and failure_rate < 0.15 and mttr_minutes < 1440:
        return "High"
    if count_30d >= 10 and failure_rate < 0.30 and mttr_minutes < 10080:
        return "Medium"
    return "Low"


def generate_recommendation(score, tech_debt, recency_score, dora_profile, count_30d):
    if score >= 75 and dora_profile in ("Elite", "High"):
        return "Mantener"
    if 50 <= score < 75 and tech_debt < 10:
        return "Evolucionar"
    if 25 <= score < 50 and count_30d < 5:
        return "Consolidar"
    if score < 25 and recency_score == 0:
        return "Deprecar"
    if score < 10 and count_30d == 0:
        return "Eliminar"
    if score >= 75:
        return "Mantener"
    if score >= 50:
        return "Evolucionar"
    return "Consolidar"


def demand_level_text(count_30d):
    if count_30d > 200: return "Elite"
    if count_30d >= 50: return "Alto"
    if count_30d >= 10: return "Medio"
    if count_30d >= 1: return "Bajo"
    return "Nulo"


def rating_text(score):
    if score >= 90: return "Excelente"
    if score >= 75: return "Bueno"
    if score >= 50: return "Regular"
    if score >= 25: return "Bajo"
    return "Crítico"


# ==========================================================
# ENRICH CI PIPELINE WITH INCREMENTAL DATA
# ==========================================================

def enrich_ci_pipeline(base_data, headers, org, project, skip_incremental=False):
    def_id = base_data.get("id")
    if skip_incremental:
        return {**base_data, "builds20": [], "failureRate": 0.0, "mttrMinutes": 0}
    
    builds = fetch_ci_builds(def_id, headers, org, project, top=20)
    total = len(builds)
    failures = sum(1 for b in builds if b.get("result") != "succeeded")
    failure_rate = failures / total if total > 0 else 0.0
    
    rel_score, mttr_min, _ = calculate_reliability_score(builds)
    
    return {
        **base_data,
        "builds20": builds,
        "failureRate": failure_rate,
        "mttrMinutes": mttr_min,
        "reliabilityScore": rel_score,
    }


def enrich_cd_pipeline(base_data, headers, org, project, skip_incremental=False):
    def_id = base_data.get("id")
    if skip_incremental:
        return {**base_data, "releases20": [], "failureRate": 0.0, "mttrMinutes": 0}
    
    releases = fetch_cd_releases(def_id, headers, org, project, top=20)
    # CD releases don't have simple result field like CI builds; use status approximation
    total = len(releases)
    failures = sum(1 for r in releases if r.get("status") in ("rejected", "abandoned", "canceled"))
    failure_rate = failures / total if total > 0 else 0.0
    
    # Simplified MTTR for CD (less granular API)
    mttr_min = 0
    if failures > 0 and total >= 2:
        recovery_times = []
        for i in range(len(releases) - 1):
            if releases[i].get("status") in ("rejected", "abandoned", "canceled"):
                for j in range(i + 1, len(releases)):
                    if releases[j].get("status") in ("succeeded", "active"):
                        t1_str = releases[i].get("createdOn", "")
                        t2_str = releases[j].get("createdOn", "")
                        try:
                            t1 = datetime.fromisoformat(t1_str.replace("Z", "+00:00")) if t1_str else None
                            t2 = datetime.fromisoformat(t2_str.replace("Z", "+00:00")) if t2_str else None
                            if t1 and t2:
                                recovery_times.append((t2 - t1).total_seconds() / 60)
                        except Exception:
                            pass
                        break
        if recovery_times:
            mttr_min = sum(recovery_times) / len(recovery_times)
    
    # Reliability score for CD (simplified, no detailed build-level results)
    if failure_rate == 0: sr_pts = 15
    elif failure_rate < 0.05: sr_pts = 12
    elif failure_rate < 0.10: sr_pts = 9
    elif failure_rate < 0.20: sr_pts = 6
    elif failure_rate < 0.40: sr_pts = 3
    else: sr_pts = 0
    
    if mttr_min == 0:
        mttr_pts = 10 if failures == 0 else 5
    elif mttr_min <= 15: mttr_pts = 10
    elif mttr_min <= 60: mttr_pts = 8
    elif mttr_min <= 240: mttr_pts = 6
    elif mttr_min <= 1440: mttr_pts = 3
    else: mttr_pts = 0
    
    rel_score = min(sr_pts + mttr_pts, 25)
    
    return {
        **base_data,
        "releases20": releases,
        "failureRate": failure_rate,
        "mttrMinutes": mttr_min,
        "reliabilityScore": rel_score,
    }


# ==========================================================
# BUILD INVENTORY ROWS
# ==========================================================

def build_ci_row(data, headers, org, project, skip_incremental=False):
    enriched = enrich_ci_pipeline(data, headers, org, project, skip_incremental)
    tech, version, tech_status = detect_technology_status(enriched.get("name", ""))
    
    last_exec = enriched.get("lastExecution", "")
    modified = enriched.get("modifiedDate", "")
    count_30d = enriched.get("totalExecutions30d", 0) or 0
    
    recency = calculate_recency_score(last_exec)
    reliability = enriched.get("reliabilityScore", 0)
    usage = calculate_usage_score(count_30d)
    freshness = calculate_freshness_score(modified)
    tech_debt = calculate_tech_debt_score(
        enriched.get("processType", ""),
        bool(enriched.get("yamlFilename", "")),
        enriched.get("repositoryName", ""),
        tech_status,
        enriched.get("name", "")
    )
    
    health = recency + reliability + usage + freshness + tech_debt
    dora = calculate_dora_profile(count_30d, enriched.get("failureRate", 0), enriched.get("mttrMinutes", 0))
    rec = generate_recommendation(health, tech_debt, recency, dora, count_30d)
    
    return {
        "pipeline_name": enriched.get("name", ""),
        "pipeline_type": "CI",
        "pipeline_path": enriched.get("path", ""),
        "technology": tech,
        "technology_version": version,
        "technology_status": tech_status,
        "health_score": health,
        "rating": rating_text(health),
        "recency_score": recency,
        "reliability_score": reliability,
        "usage_score": usage,
        "freshness_score": freshness,
        "tech_debt_score": tech_debt,
        "last_execution": last_exec,
        "last_execution_status": enriched.get("lastExecutionState", ""),
        "last_execution_result": enriched.get("lastExecutionResult", ""),
        "total_executions_30d": count_30d,
        "total_executions_90d": enriched.get("totalExecutions90d", 0) or 0,
        "total_failures_30d": int((enriched.get("failureRate", 0) or 0) * max(count_30d, 1)),
        "mttr_minutes": round(enriched.get("mttrMinutes", 0) or 0, 1),
        "last_modified": modified,
        "created_date": enriched.get("createdDate", ""),
        "days_since_creation": days_since(enriched.get("createdDate", "")),
        "demand_level": demand_level_text(count_30d),
        "dora_profile": dora,
        "recommendation": rec,
        "process_type": enriched.get("processType", ""),
        "repository_name": enriched.get("repositoryName", ""),
        "yaml_filename": enriched.get("yamlFilename", ""),
    }


def build_cd_row(data, headers, org, project, skip_incremental=False):
    enriched = enrich_cd_pipeline(data, headers, org, project, skip_incremental)
    tech, version, tech_status = detect_technology_status(enriched.get("name", ""))
    
    last_exec = enriched.get("lastReleaseDate", "")
    modified = enriched.get("modifiedOn", "")
    # For CD, estimate count_30d from last release recency (less granular API)
    # Use a placeholder since CD API doesn't provide direct counts easily
    days = days_since(last_exec)
    if days <= 1: count_30d = 30
    elif days <= 7: count_30d = 4
    elif days <= 30: count_30d = 1
    else: count_30d = 0
    
    recency = calculate_recency_score(last_exec)
    reliability = enriched.get("reliabilityScore", 0)
    usage = calculate_usage_score(count_30d)
    freshness = calculate_freshness_score(modified)
    tech_debt = calculate_tech_debt_score(
        "releaseDefinition",  # CD classic releases are like designerJson
        False,
        "",  # CD doesn't always have direct repo
        tech_status,
        enriched.get("name", "")
    )
    
    health = recency + reliability + usage + freshness + tech_debt
    dora = calculate_dora_profile(count_30d, enriched.get("failureRate", 0), enriched.get("mttrMinutes", 0))
    rec = generate_recommendation(health, tech_debt, recency, dora, count_30d)
    
    return {
        "pipeline_name": enriched.get("name", ""),
        "pipeline_type": "CD",
        "pipeline_path": enriched.get("path", ""),
        "technology": tech,
        "technology_version": version,
        "technology_status": tech_status,
        "health_score": health,
        "rating": rating_text(health),
        "recency_score": recency,
        "reliability_score": reliability,
        "usage_score": usage,
        "freshness_score": freshness,
        "tech_debt_score": tech_debt,
        "last_execution": last_exec,
        "last_execution_status": enriched.get("lastReleaseStatus", ""),
        "last_execution_result": "",
        "total_executions_30d": count_30d,
        "total_executions_90d": 0,
        "total_failures_30d": int((enriched.get("failureRate", 0) or 0) * max(count_30d, 1)),
        "mttr_minutes": round(enriched.get("mttrMinutes", 0) or 0, 1),
        "last_modified": modified,
        "created_date": enriched.get("createdOn", ""),
        "days_since_creation": days_since(enriched.get("createdOn", "")),
        "demand_level": demand_level_text(count_30d),
        "dora_profile": dora,
        "recommendation": rec,
        "environments_count": enriched.get("environmentsCount", 0),
        "environments": enriched.get("environments", ""),
        "is_obsolete": enriched.get("isObsolete", "No"),
    }


# ==========================================================
# CHART GENERATION — openpyxl native Excel charts
# ==========================================================

from openpyxl.chart import BarChart, PieChart, Reference, ScatterChart, RadarChart, BubbleChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

RATING_COLORS_HEX = {
    "Excelente": "2ecc71",
    "Bueno":     "27ae60",
    "Regular":   "f39c12",
    "Bajo":      "e67e22",
    "Crítico":   "e74c3c",
}

DIMENSION_COLORS_HEX = {
    "Recency":    "3498db",
    "Reliability": "2ecc71",
    "Usage":      "9b59b6",
    "Freshness":  "f39c12",
    "TechDebt":   "e74c3c",
}


def _write_chart_data_stacked_bar(sheet, df_health):
    """Escribe datos para P1 Stacked Bar y retorna rango de datos."""
    top = df_health.nlargest(min(30, len(df_health)), "health_score")
    dimensions = ["recency_score", "reliability_score", "usage_score", "freshness_score", "tech_debt_score"]
    headers = ["Pipeline"] + ["Recency", "Reliability", "Usage", "Freshness", "TechDebt"]

    for c, h in enumerate(headers, 1):
        sheet.cell(row=1, column=c, value=h)

    for r, (_, row) in enumerate(top.iterrows(), 2):
        sheet.cell(row=r, column=1, value=row.get("pipeline_name", ""))
        for c, col in enumerate(dimensions, 2):
            sheet.cell(row=r, column=c, value=float(row.get(col, 0)))

    return len(top)


def _build_stacked_bar_chart(sheet, data_rows):
    """P1: Stacked Bar Horizontal — composición del health score."""
    chart = BarChart()
    chart.type = "bar"          # horizontal
    chart.grouping = "stacked"
    chart.title = "Composición Health Score — Top 30 Pipelines"
    chart.y_axis.title = None
    chart.x_axis.title = "Score (0-100)"
    chart.x_axis.scaling.max = 105
    chart.style = 10
    chart.width = 32
    chart.height = max(12, data_rows * 0.45)

    # Categories = pipeline names (col A)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)

    dim_colors = list(DIMENSION_COLORS_HEX.values())
    dim_names = list(DIMENSION_COLORS_HEX.keys())

    for i in range(5):
        data = Reference(sheet, min_col=i + 2, min_row=1, max_row=data_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        s = chart.series[i]
        s.graphicalProperties.solidFill = dim_colors[i]

    chart.legend.position = "b"
    return chart


def _write_chart_data_donut(sheet, df_health):
    """Escribe datos para P2 Donut y retorna número de filas."""
    rating_order = ["Excelente", "Bueno", "Regular", "Bajo", "Crítico"]
    counts = df_health["rating"].value_counts()

    sheet.cell(row=1, column=1, value="Rating")
    sheet.cell(row=1, column=2, value="Cantidad")

    r = 2
    for rating in rating_order:
        c = counts.get(rating, 0)
        if c > 0:
            sheet.cell(row=r, column=1, value=rating)
            sheet.cell(row=r, column=2, value=c)
            r += 1
    return r - 2


def _build_donut_chart(sheet, data_rows):
    """P2: Pie/Donut — distribución de ratings."""
    chart = PieChart()
    chart.style = 10
    chart.title = "Distribución de Ratings"
    chart.width = 16
    chart.height = 14

    data = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Colorear cada slice según rating
    rating_order = ["Excelente", "Bueno", "Regular", "Bajo", "Crítico"]
    for i in range(data_rows):
        pt = DataPoint(idx=i)
        cell_val = sheet.cell(row=i + 2, column=1).value
        color = RATING_COLORS_HEX.get(cell_val, "888888")
        pt.graphicalProperties.solidFill = color
        chart.series[0].data_points.append(pt)

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showPercent = True
    chart.series[0].dLbls.showCatName = True
    chart.series[0].dLbls.showVal = False

    return chart


def _write_chart_data_dora_bar(sheet, df_health):
    """Escribe datos para P3 Grouped Bar y retorna número de filas."""
    dora_order = ["Elite", "High", "Medium", "Low"]
    ci_counts = df_health[df_health["pipeline_type"] == "CI"]["dora_profile"].value_counts()
    cd_counts = df_health[df_health["pipeline_type"] == "CD"]["dora_profile"].value_counts()

    active = [d for d in dora_order if ci_counts.get(d, 0) + cd_counts.get(d, 0) > 0]

    sheet.cell(row=1, column=1, value="DORA Profile")
    sheet.cell(row=1, column=2, value="CI")
    sheet.cell(row=1, column=3, value="CD")

    for r, d in enumerate(active, 2):
        sheet.cell(row=r, column=1, value=d)
        sheet.cell(row=r, column=2, value=ci_counts.get(d, 0))
        sheet.cell(row=r, column=3, value=cd_counts.get(d, 0))

    return len(active)


def _build_dora_grouped_bar(sheet, data_rows):
    """P3: Grouped Bar — DORA profile CI vs CD."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "DORA Profile — CI vs CD"
    chart.y_axis.title = "Cantidad de pipelines"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data_ci = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    data_cd = Reference(sheet, min_col=3, min_row=1, max_row=data_rows + 1)

    chart.add_data(data_ci, titles_from_data=True)
    chart.add_data(data_cd, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.solidFill = "3498db"  # CI blue
    chart.series[1].graphicalProperties.solidFill = "e67e22"  # CD orange

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = True
    chart.series[1].dLbls = DataLabelList()
    chart.series[1].dLbls.showVal = True

    return chart


def _write_chart_data_scatter(sheet, df_health):
    """Escribe datos para P5 Scatter y retorna número de filas."""
    sheet.cell(row=1, column=1, value="Pipeline")
    sheet.cell(row=1, column=2, value="Ejecuciones 30d")
    sheet.cell(row=1, column=3, value="Health Score")
    sheet.cell(row=1, column=4, value="Rating")

    for r, (_, row) in enumerate(df_health.iterrows(), 2):
        sheet.cell(row=r, column=1, value=row.get("pipeline_name", ""))
        sheet.cell(row=r, column=2, value=float(row.get("total_executions_30d", 0)))
        sheet.cell(row=r, column=3, value=float(row.get("health_score", 0)))
        sheet.cell(row=r, column=4, value=row.get("rating", ""))

    return len(df_health)


def _build_scatter_chart(sheet, data_rows):
    """P5: Scatter — Health Score vs Ejecuciones 30d."""
    chart = ScatterChart()
    chart.title = "Health Score vs Uso — Priorización"
    chart.x_axis.title = "Ejecuciones (últimos 30 días)"
    chart.y_axis.title = "Health Score (0-100)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 105
    chart.style = 10
    chart.width = 22
    chart.height = 15

    # One series per rating category for color coding
    rating_order = ["Excelente", "Bueno", "Regular", "Bajo", "Crítico"]
    rating_colors = ["2ecc71", "27ae60", "f39c12", "e67e22", "e74c3c"]

    for rating, color in zip(rating_order, rating_colors):
        # Collect rows matching this rating
        rows_x = []
        rows_y = []
        for r in range(2, data_rows + 2):
            if sheet.cell(row=r, column=4).value == rating:
                rows_x.append(r)
                rows_y.append(r)

        if not rows_x:
            continue

        # Write per-rating data in auxiliary columns (col 6+)
        col_x = 6 + rating_order.index(rating) * 2
        col_y = col_x + 1
        sheet.cell(row=1, column=col_x, value=f"Exec_{rating}")
        sheet.cell(row=1, column=col_y, value=f"Score_{rating}")

        for i, r in enumerate(rows_x, 2):
            sheet.cell(row=i, column=col_x, value=sheet.cell(row=r, column=2).value)
            sheet.cell(row=i, column=col_y, value=sheet.cell(row=r, column=3).value)

        xvalues = Reference(sheet, min_col=col_x, min_row=2, max_row=len(rows_x) + 1)
        yvalues = Reference(sheet, min_col=col_y, min_row=2, max_row=len(rows_y) + 1)

        series = chart.series  # will use append
        from openpyxl.chart import Series as ChartSeries
        s = ChartSeries(yvalues, xvalues, title=rating)
        s.graphicalProperties.line.noFill = True
        s.graphicalProperties.solidFill = color
        s.marker.symbol = "circle"
        s.marker.size = 7
        s.marker.graphicalProperties.solidFill = color
        s.marker.graphicalProperties.line.solidFill = color
        chart.series.append(s)

    chart.legend.position = "r"
    return chart


# ──────────────────────────────────────────────────────────────
# P6: TREEMAP — Technology count + avg health (combo bar + line)
# ──────────────────────────────────────────────────────────────

def _write_chart_data_treemap(sheet, df_health):
    """Escribe datos para P6 Treemap (combo): tecnología, conteo, salud promedio."""
    if "technology" not in df_health.columns:
        return 0
    tech_stats = df_health.groupby("technology").agg(
        count=("health_score", "size"),
        avg_health=("health_score", "mean"),
    ).sort_values("count", ascending=False)

    sheet.cell(row=1, column=1, value="Tecnología")
    sheet.cell(row=1, column=2, value="Cantidad Pipelines")
    sheet.cell(row=1, column=3, value="Salud Promedio")

    for r, (tech, row) in enumerate(tech_stats.iterrows(), 2):
        sheet.cell(row=r, column=1, value=str(tech) if tech else "Sin detectar")
        sheet.cell(row=r, column=2, value=int(row["count"]))
        sheet.cell(row=r, column=3, value=round(float(row["avg_health"]), 1))

    return len(tech_stats)


def _build_treemap_chart(sheet, data_rows):
    """P6: Combo chart — barras (conteo pipelines) + línea (salud promedio) por tecnología."""
    from openpyxl.chart import LineChart
    from openpyxl.chart.axis import NumericAxis

    # Bar chart for count
    bar = BarChart()
    bar.type = "col"
    bar.title = "Tecnologías: Cantidad vs Salud Promedio"
    bar.y_axis.title = "Cantidad de Pipelines"
    bar.style = 10
    bar.width = 24
    bar.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data_count = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    bar.add_data(data_count, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "3498db"

    # Line chart for avg health on secondary axis
    line = LineChart()
    line.y_axis.title = "Salud Promedio (0-100)"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 105
    line.y_axis.axId = 200

    data_health = Reference(sheet, min_col=3, min_row=1, max_row=data_rows + 1)
    line.add_data(data_health, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties.line.solidFill = "e74c3c"
    line.series[0].graphicalProperties.line.width = 25000
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 7
    line.series[0].marker.graphicalProperties.solidFill = "e74c3c"

    bar.y_axis.crosses = "min"
    bar += line

    return bar


# ──────────────────────────────────────────────────────────────
# P7: PARETO — Technologies causing low scores (bar + cumulative %)
# ──────────────────────────────────────────────────────────────

def _write_chart_data_pareto(sheet, df_health):
    """Escribe datos para P7 Pareto: tecnología, count Bajo+Crítico, % acumulado."""
    if "technology" not in df_health.columns:
        return 0
    low_df = df_health[df_health["rating"].isin(["Bajo", "Crítico"])]
    if low_df.empty:
        return 0

    tech_low = low_df.groupby("technology").size().sort_values(ascending=False)
    total_low = tech_low.sum()

    sheet.cell(row=1, column=1, value="Tecnología")
    sheet.cell(row=1, column=2, value="Pipelines Bajo/Crítico")
    sheet.cell(row=1, column=3, value="% Acumulado")

    cum = 0
    for r, (tech, count) in enumerate(tech_low.items(), 2):
        sheet.cell(row=r, column=1, value=str(tech) if tech else "Sin detectar")
        sheet.cell(row=r, column=2, value=int(count))
        cum += count
        sheet.cell(row=r, column=3, value=round(cum / total_low * 100, 1))

    return len(tech_low)


def _build_pareto_chart(sheet, data_rows):
    """P7: Pareto — barras descendentes + línea % acumulado."""
    from openpyxl.chart import LineChart

    bar = BarChart()
    bar.type = "col"
    bar.title = "Pareto: Tecnologías con más pipelines críticos"
    bar.y_axis.title = "Cantidad Bajo/Crítico"
    bar.style = 10
    bar.width = 22
    bar.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data_count = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    bar.add_data(data_count, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "e74c3c"

    # Cumulative % line on secondary axis
    line = LineChart()
    line.y_axis.title = "% Acumulado"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 105
    line.y_axis.axId = 200

    data_cum = Reference(sheet, min_col=3, min_row=1, max_row=data_rows + 1)
    line.add_data(data_cum, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties.line.solidFill = "f39c12"
    line.series[0].graphicalProperties.line.width = 25000
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 6
    line.series[0].marker.graphicalProperties.solidFill = "f39c12"

    bar.y_axis.crosses = "min"
    bar += line

    return bar


# ──────────────────────────────────────────────────────────────
# P8: TREND — Historical health score from cache files
# ──────────────────────────────────────────────────────────────

def _write_chart_data_trend(sheet, df_health):
    """Escribe datos para P8 Trend: escanea cache previo para serie temporal."""
    from openpyxl.chart import LineChart

    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    if not cache_dir.exists():
        # Single data point: current run only
        sheet.cell(row=1, column=1, value="Fecha")
        sheet.cell(row=1, column=2, value="Health Score Promedio")
        sheet.cell(row=1, column=3, value="Total Pipelines")
        sheet.cell(row=2, column=1, value=datetime.now().strftime("%Y-%m-%d"))
        sheet.cell(row=2, column=2, value=round(float(df_health["health_score"].mean()), 1) if len(df_health) > 0 else 0)
        sheet.cell(row=2, column=3, value=len(df_health))
        return 1

    # Scan all health score cache files for historical data
    pattern = str(cache_dir / "cicd_inventory_health_score_*_raw_*.json")
    cache_files = sorted(glob.glob(pattern))

    data_points = []
    for cf in cache_files:
        try:
            with open(cf, "r", encoding="utf-8") as f:
                cache_data = json.load(f)
            meta = cache_data.get("metadata", {})
            rows = cache_data.get("rows", [])
            if rows:
                gen_at = meta.get("generated_at", "")
                dt_str = gen_at[:10] if gen_at else Path(cf).stem.split("_")[-2]
                avg_score = sum(r.get("health_score", 0) for r in rows) / len(rows) if rows else 0
                data_points.append((dt_str, round(avg_score, 1), len(rows)))
        except Exception:
            continue

    # Also add current run
    current_avg = round(float(df_health["health_score"].mean()), 1) if len(df_health) > 0 else 0
    current_date = datetime.now().strftime("%Y-%m-%d")
    if not data_points or data_points[-1][0] != current_date:
        data_points.append((current_date, current_avg, len(df_health)))

    # Deduplicate by date (keep last per day)
    seen = {}
    for dt, avg, cnt in data_points:
        seen[dt] = (avg, cnt)
    data_points = [(dt, v[0], v[1]) for dt, v in sorted(seen.items())]

    sheet.cell(row=1, column=1, value="Fecha")
    sheet.cell(row=1, column=2, value="Health Score Promedio")
    sheet.cell(row=1, column=3, value="Total Pipelines")

    for r, (dt, avg, cnt) in enumerate(data_points, 2):
        sheet.cell(row=r, column=1, value=dt)
        sheet.cell(row=r, column=2, value=avg)
        sheet.cell(row=r, column=3, value=cnt)

    return len(data_points)


def _build_trend_chart(sheet, data_rows):
    """P8: Line chart — tendencia de Health Score promedio en el tiempo."""
    from openpyxl.chart import LineChart

    chart = LineChart()
    chart.title = "Tendencia Health Score Promedio"
    chart.x_axis.title = "Fecha"
    chart.y_axis.title = "Health Score Promedio"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 105
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data_score = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    chart.add_data(data_score, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = "2ecc71"
    chart.series[0].graphicalProperties.line.width = 30000
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 8
    chart.series[0].marker.graphicalProperties.solidFill = "2ecc71"

    if data_rows == 1:
        chart.title = "Tendencia Health Score (solo 1 punto — ejecutar múltiples veces para ver tendencia)"

    return chart


# ──────────────────────────────────────────────────────────────
# HEATMAP: Technology Status vs Rating (conditional formatting table)
# ──────────────────────────────────────────────────────────────

HEATMAP_COLORS = {
    0:  "f7f7f7",   # empty / white
    1:  "d4efdf",   # 1-2 light green
    3:  "a9dfbf",   # 3-5 green
    6:  "f9e79f",   # 6-10 yellow
    11: "f5b041",   # 11-20 orange
    21: "e74c3c",   # 21+ red
}


def _heatmap_color(value):
    """Retorna color hex según valor para heatmap."""
    if value == 0:
        return "f7f7f7"
    if value <= 2:
        return "d4efdf"
    if value <= 5:
        return "a9dfbf"
    if value <= 10:
        return "f9e79f"
    if value <= 20:
        return "f5b041"
    return "e74c3c"


def _build_heatmap_table(df_health):
    """Construye datos para heatmap: technology_status vs rating. Retorna dict."""
    if "technology_status" not in df_health.columns or "rating" not in df_health.columns:
        return {}

    status_order = ["Moderna", "Mantenimiento", "EOL", "Desconocido"]
    rating_order = ["Excelente", "Bueno", "Regular", "Bajo", "Crítico"]

    cross = {}
    for _, row in df_health.iterrows():
        st = row.get("technology_status", "Desconocido") or "Desconocido"
        rt = row.get("rating", "Desconocido") or "Desconocido"
        key = (st, rt)
        cross[key] = cross.get(key, 0) + 1

    return {"cross": cross, "statuses": status_order, "ratings": rating_order}


def _write_heatmap_to_sheet(sheet, df_health, start_row=101):
    """Escribe tabla heatmap con formato condicional en la hoja Charts."""
    hm = _build_heatmap_table(df_health)
    if not hm:
        return

    cross = hm["cross"]
    statuses = hm["statuses"]
    ratings = hm["ratings"]

    from openpyxl.styles import PatternFill, Alignment, Border, Side

    r = start_row
    sheet.cell(row=r, column=1, value="MAPA DE CALOR: Technology Status vs Rating").font = Font(bold=True, size=12)
    r += 1

    # Header row
    sheet.cell(row=r, column=1, value="Technology Status ↓  Rating →").font = Font(bold=True, size=9)
    for c, rating in enumerate(ratings, 2):
        cell = sheet.cell(row=r, column=c, value=rating)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    sheet.cell(row=r, column=len(ratings) + 2, value="Total").font = Font(bold=True, size=9)
    r += 1

    thin_border = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )

    for status in statuses:
        sheet.cell(row=r, column=1, value=status).font = Font(bold=True, size=9)
        row_total = 0
        for c, rating in enumerate(ratings, 2):
            val = cross.get((status, rating), 0)
            row_total += val
            cell = sheet.cell(row=r, column=c, value=val if val > 0 else "")
            cell.fill = PatternFill(start_color=_heatmap_color(val), end_color=_heatmap_color(val), fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if val > 0:
                cell.font = Font(bold=True, size=10)
        sheet.cell(row=r, column=len(ratings) + 2, value=row_total).font = Font(bold=True, size=9)
        r += 1

    # Legend
    r += 1
    sheet.cell(row=r, column=1, value="Leyenda de colores:").font = Font(italic=True, size=8)
    r += 1
    legend_items = [
        ("0", "f7f7f7"), ("1-2", "d4efdf"), ("3-5", "a9dfbf"),
        ("6-10", "f9e79f"), ("11-20", "f5b041"), ("21+", "e74c3c"),
    ]
    for i, (label, color) in enumerate(legend_items):
        cell = sheet.cell(row=r, column=i + 1, value=label)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.font = Font(size=8)


# ──────────────────────────────────────────────────────────────
# P9: TREEMAP RIESGO TECNOLÓGICO — pipeline_path count + avg health
# ──────────────────────────────────────────────────────────────

def _write_chart_data_risk_treemap(sheet, df_health):
    """Escribe datos para P9: agrupa por pipeline_path, conteo y salud promedio."""
    if "pipeline_path" not in df_health.columns:
        return 0
    # Extract top-level area from path (e.g. "\\WMS\\Legacy" -> "WMS")
    def _extract_area(path):
        if not path or not isinstance(path, str):
            return "Raíz"
        parts = [p for p in path.replace("\\", "/").split("/") if p and p != "\\"]
        return parts[0] if parts else "Raíz"

    df_copy = df_health.copy()
    df_copy["area"] = df_copy["pipeline_path"].apply(_extract_area)

    area_stats = df_copy.groupby("area").agg(
        count=("health_score", "size"),
        avg_health=("health_score", "mean"),
    ).sort_values("count", ascending=False)

    sheet.cell(row=1, column=1, value="Área/Path")
    sheet.cell(row=1, column=2, value="Cantidad Pipelines")
    sheet.cell(row=1, column=3, value="Salud Promedio")

    for r, (area, row) in enumerate(area_stats.iterrows(), 2):
        sheet.cell(row=r, column=1, value=str(area))
        sheet.cell(row=r, column=2, value=int(row["count"]))
        sheet.cell(row=r, column=3, value=round(float(row["avg_health"]), 1))

    return len(area_stats)


def _build_risk_treemap_chart(sheet, data_rows):
    """P9: Combo bar+line — barras conteo por área, línea salud promedio."""
    from openpyxl.chart import LineChart

    bar = BarChart()
    bar.type = "col"
    bar.title = "Riesgo Tecnológico por Área"
    bar.y_axis.title = "Cantidad de Pipelines"
    bar.style = 10
    bar.width = 24
    bar.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data_count = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    bar.add_data(data_count, titles_from_data=True)
    bar.set_categories(cats)

    # Color bars by avg health score
    for i in range(data_rows):
        pt = DataPoint(idx=i)
        avg = sheet.cell(row=i + 2, column=3).value or 0
        if avg >= 80:
            pt.graphicalProperties.solidFill = "2ecc71"
        elif avg >= 60:
            pt.graphicalProperties.solidFill = "27ae60"
        elif avg >= 40:
            pt.graphicalProperties.solidFill = "f39c12"
        elif avg >= 20:
            pt.graphicalProperties.solidFill = "e67e22"
        else:
            pt.graphicalProperties.solidFill = "e74c3c"
        bar.series[0].data_points.append(pt)

    # Line for avg health on secondary axis
    line = LineChart()
    line.y_axis.title = "Salud Promedio (0-100)"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 105
    line.y_axis.axId = 200

    data_health = Reference(sheet, min_col=3, min_row=1, max_row=data_rows + 1)
    line.add_data(data_health, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties.line.solidFill = "3498db"
    line.series[0].graphicalProperties.line.width = 25000
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 7
    line.series[0].marker.graphicalProperties.solidFill = "3498db"

    bar.y_axis.crosses = "min"
    bar += line

    return bar


# ──────────────────────────────────────────────────────────────
# P10: SANKEY (stacked bar) — Technology → Recommendation flow
# ──────────────────────────────────────────────────────────────

def _write_chart_data_sankey(sheet, df_health):
    """Escribe datos para P10 Sankey: tecnología vs recomendación (stacked)."""
    if "technology" not in df_health.columns or "recommendation" not in df_health.columns:
        return 0

    # Simplify recommendations to categories
    def _simplify_rec(rec):
        if not rec or not isinstance(rec, str):
            return "Sin recom."
        rec_l = rec.lower()
        if any(k in rec_l for k in ["deprecar", "eliminar", "retirar", "remover"]):
            return "Deprecar"
        if any(k in rec_l for k in ["consolidar", "fusionar", "unificar"]):
            return "Consolidar"
        if any(k in rec_l for k in ["evolucionar", "migrar", "actualizar", "modernizar"]):
            return "Evolucionar"
        if any(k in rec_l for k in ["mantener", "monitorear", "vigilar"]):
            return "Mantener"
        return "Sin recom."

    df_copy = df_health.copy()
    df_copy["rec_cat"] = df_copy["recommendation"].apply(_simplify_rec)

    rec_categories = ["Deprecar", "Consolidar", "Evolucionar", "Mantener", "Sin recom."]
    tech_order = df_copy["technology"].value_counts().index.tolist()[:15]

    sheet.cell(row=1, column=1, value="Tecnología")
    for c, rec in enumerate(rec_categories, 2):
        sheet.cell(row=1, column=c, value=rec)

    for r, tech in enumerate(tech_order, 2):
        sheet.cell(row=r, column=1, value=str(tech) if tech else "Sin detectar")
        tech_df = df_copy[df_copy["technology"] == tech]
        for c, rec in enumerate(rec_categories, 2):
            count = len(tech_df[tech_df["rec_cat"] == rec])
            sheet.cell(row=r, column=c, value=count)

    return len(tech_order)


def _build_sankey_chart(sheet, data_rows):
    """P10: Stacked bar — flujo de tecnología a recomendación."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Flujo Tecnología → Recomendación"
    chart.y_axis.title = "Cantidad de Pipelines"
    chart.style = 10
    chart.width = 24
    chart.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)

    rec_colors = ["e74c3c", "f39c12", "3498db", "2ecc71", "95a5a6"]
    for i in range(5):
        data = Reference(sheet, min_col=i + 2, min_row=1, max_row=data_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[i].graphicalProperties.solidFill = rec_colors[i]

    chart.legend.position = "b"
    return chart


# ──────────────────────────────────────────────────────────────
# P11: RADAR CHART — DORA maturity per area
# ──────────────────────────────────────────────────────────────

def _write_chart_data_radar(sheet, df_health):
    """Escribe datos para P11 Radar: 5 dimensiones por top áreas."""
    if "pipeline_path" not in df_health.columns:
        return 0

    def _extract_area(path):
        if not path or not isinstance(path, str):
            return "Raíz"
        parts = [p for p in path.replace("\\", "/").split("/") if p and p != "\\"]
        return parts[0] if parts else "Raíz"

    df_copy = df_health.copy()
    df_copy["area"] = df_copy["pipeline_path"].apply(_extract_area)

    # Top 5 areas by count
    top_areas = df_copy["area"].value_counts().head(5).index.tolist()
    if not top_areas:
        return 0

    dimensions = ["Recency", "Reliability", "Usage", "Freshness", "TechDebt"]
    dim_cols = ["recency_score", "reliability_score", "usage_score", "freshness_score", "tech_debt_score"]

    sheet.cell(row=1, column=1, value="Dimensión")
    for c, area in enumerate(top_areas, 2):
        sheet.cell(row=1, column=c, value=area)

    for r, (dim, col) in enumerate(zip(dimensions, dim_cols), 2):
        sheet.cell(row=r, column=1, value=dim)
        for c, area in enumerate(top_areas, 2):
            area_df = df_copy[df_copy["area"] == area]
            avg = area_df[col].mean() if len(area_df) > 0 else 0
            sheet.cell(row=r, column=c, value=round(float(avg), 1))

    return len(dimensions)


def _build_radar_chart(sheet, data_rows):
    """P11: Radar chart — comparación dimensional por área."""
    chart = RadarChart()
    chart.type = "filled"
    chart.title = "Radar DORA — Dimensiones por Área (Top 5)"
    chart.style = 10
    chart.width = 20
    chart.height = 16

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)

    # One series per area (columns 2..N)
    n_areas = 0
    for c in range(2, 20):
        if sheet.cell(row=1, column=c).value:
            n_areas += 1
        else:
            break

    area_colors = ["3498db", "e74c3c", "2ecc71", "f39c12", "9b59b6"]
    for i in range(n_areas):
        data = Reference(sheet, min_col=i + 2, min_row=1, max_row=data_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[i].graphicalProperties.solidFill = area_colors[i % len(area_colors)]
        chart.series[i].graphicalProperties.line.solidFill = area_colors[i % len(area_colors)]

    chart.legend.position = "b"
    return chart


# ──────────────────────────────────────────────────────────────
# P12: BUBBLE CHART — Effort vs Impact (antiquity × usage, size=failures)
# ──────────────────────────────────────────────────────────────

def _write_chart_data_bubble(sheet, df_health):
    """Escribe datos para P12 Bubble: X=days_since_creation, Y=exec_30d, size=failures."""
    required = ["days_since_creation", "total_executions_30d", "total_failures_30d"]
    if not all(c in df_health.columns for c in required):
        return 0

    sheet.cell(row=1, column=1, value="Pipeline")
    sheet.cell(row=1, column=2, value="Antigüedad (días)")
    sheet.cell(row=1, column=3, value="Ejecuciones 30d")
    sheet.cell(row=1, column=4, value="Fallos 30d")
    sheet.cell(row=1, column=5, value="Rating")

    for r, (_, row) in enumerate(df_health.iterrows(), 2):
        sheet.cell(row=r, column=1, value=row.get("pipeline_name", ""))
        days = float(row.get("days_since_creation", 0) or 0)
        execs = float(row.get("total_executions_30d", 0) or 0)
        fails = float(row.get("total_failures_30d", 0) or 0)
        sheet.cell(row=r, column=2, value=days)
        sheet.cell(row=r, column=3, value=execs)
        sheet.cell(row=r, column=4, value=max(fails, 1))  # min bubble size
        sheet.cell(row=r, column=5, value=row.get("rating", ""))

    return len(df_health)


def _build_bubble_chart(sheet, data_rows):
    """P12: Bubble chart — Esfuerzo (antigüedad) vs Impacto (uso), tamaño=fallos."""
    chart = BubbleChart()
    chart.title = "Esfuerzo vs Impacto — Antigüedad × Uso, tamaño = Fallos"
    chart.x_axis.title = "Antigüedad (días desde creación)"
    chart.y_axis.title = "Ejecuciones (últimos 30 días)"
    chart.style = 10
    chart.width = 22
    chart.height = 16

    # One series per rating for color coding
    rating_order = ["Excelente", "Bueno", "Regular", "Bajo", "Crítico"]
    rating_colors = ["2ecc71", "27ae60", "f39c12", "e67e22", "e74c3c"]

    for ri, (rating, color) in enumerate(zip(rating_order, rating_colors)):
        # Collect matching rows
        matching_rows = []
        for r in range(2, data_rows + 2):
            if sheet.cell(row=r, column=5).value == rating:
                matching_rows.append(r)

        if not matching_rows:
            continue

        # Write auxiliary data in columns 7+
        col_x = 7 + ri * 3
        col_y = col_x + 1
        col_s = col_x + 2
        sheet.cell(row=1, column=col_x, value=f"Antig_{rating}")
        sheet.cell(row=1, column=col_y, value=f"Uso_{rating}")
        sheet.cell(row=1, column=col_s, value=f"Fallos_{rating}")

        for i, r in enumerate(matching_rows, 2):
            sheet.cell(row=i, column=col_x, value=sheet.cell(row=r, column=2).value)
            sheet.cell(row=i, column=col_y, value=sheet.cell(row=r, column=3).value)
            sheet.cell(row=i, column=col_s, value=sheet.cell(row=r, column=4).value)

        xvalues = Reference(sheet, min_col=col_x, min_row=1, max_row=len(matching_rows) + 1)
        yvalues = Reference(sheet, min_col=col_y, min_row=1, max_row=len(matching_rows) + 1)
        bvalues = Reference(sheet, min_col=col_s, min_row=1, max_row=len(matching_rows) + 1)

        from openpyxl.chart import Series as ChartSeries
        s = ChartSeries(yvalues, xvalues, title=rating)
        s.graphicalProperties.solidFill = color
        chart.series.append(s)

    chart.legend.position = "r"
    return chart


# ──────────────────────────────────────────────────────────────
# P13: HISTOGRAMA MTTR — bins de tiempo de recuperación
# ──────────────────────────────────────────────────────────────

def _write_chart_data_mttr_hist(sheet, df_health):
    """Escribe datos para P13 Histograma: bins de MTTR en minutos."""
    if "mttr_minutes" not in df_health.columns:
        return 0

    mttr_vals = df_health["mttr_minutes"].dropna()
    mttr_vals = mttr_vals[mttr_vals > 0]
    if len(mttr_vals) == 0:
        return 0

    bins = [
        ("0-30 min",    0,   30),
        ("30-60 min",   30,  60),
        ("1-2 hrs",     60,  120),
        ("2-4 hrs",     120, 240),
        ("4-8 hrs",     240, 480),
        ("8-24 hrs",    480, 1440),
        ("24+ hrs",     1440, 999999),
    ]

    sheet.cell(row=1, column=1, value="Rango MTTR")
    sheet.cell(row=1, column=2, value="Cantidad de Pipelines")

    r = 2
    for label, lo, hi in bins:
        count = int(((mttr_vals >= lo) & (mttr_vals < hi)).sum())
        if count > 0 or hi <= 1440:  # always show standard bins
            sheet.cell(row=r, column=1, value=label)
            sheet.cell(row=r, column=2, value=count)
            r += 1

    return r - 2


def _build_mttr_hist_chart(sheet, data_rows):
    """P13: Bar chart — histograma de MTTR."""
    chart = BarChart()
    chart.type = "col"
    chart.title = "Histograma de Tiempo de Recuperación (MTTR)"
    chart.y_axis.title = "Cantidad de Pipelines"
    chart.x_axis.title = "Rango de MTTR"
    chart.style = 10
    chart.width = 20
    chart.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)
    data = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Color gradient: green (fast) to red (slow)
    mttr_colors = ["2ecc71", "27ae60", "f39c12", "e67e22", "e74c3c", "c0392b", "8e44ad"]
    for i in range(data_rows):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = mttr_colors[i % len(mttr_colors)]
        chart.series[0].data_points.append(pt)

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = True
    chart.legend = None

    return chart


# ──────────────────────────────────────────────────────────────
# P14: RUN CHART — Failure rate over time with control limits
# ──────────────────────────────────────────────────────────────

def _write_chart_data_run(sheet, df_health):
    """Escribe datos para P14 Run Chart: tasa de fallos desde cache histórico."""
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"

    # Try to get historical data from cache
    data_points = []
    if cache_dir.exists():
        pattern = str(cache_dir / "cicd_inventory_health_score_*_raw_*.json")
        cache_files = sorted(glob.glob(pattern))
        for cf in cache_files:
            try:
                with open(cf, "r", encoding="utf-8") as f:
                    cache_data = json.load(f)
                meta = cache_data.get("metadata", {})
                rows = cache_data.get("rows", [])
                if rows:
                    gen_at = meta.get("generated_at", "")
                    dt_str = gen_at[:10] if gen_at else Path(cf).stem.split("_")[-2]
                    total = len(rows)
                    # Calculate failure rate: pipelines with rating Bajo or Crítico
                    failed = sum(1 for r in rows if r.get("rating", "") in ("Bajo", "Crítico"))
                    fail_rate = round(failed / total * 100, 1) if total > 0 else 0
                    avg_health = round(sum(r.get("health_score", 0) for r in rows) / total, 1) if total > 0 else 0
                    data_points.append((dt_str, fail_rate, avg_health, total))
            except Exception:
                continue

    # Add current run
    current_date = datetime.now().strftime("%Y-%m-%d")
    total = len(df_health)
    if total > 0:
        failed = sum(1 for _, r in df_health.iterrows() if r.get("rating", "") in ("Bajo", "Crítico"))
        fail_rate = round(failed / total * 100, 1)
        avg_health = round(float(df_health["health_score"].mean()), 1)
    else:
        fail_rate = 0
        avg_health = 0

    if not data_points or data_points[-1][0] != current_date:
        data_points.append((current_date, fail_rate, avg_health, total))

    # Deduplicate by date
    seen = {}
    for dt, fr, ah, cnt in data_points:
        seen[dt] = (fr, ah, cnt)
    data_points = [(dt, v[0], v[1], v[2]) for dt, v in sorted(seen.items())]

    if len(data_points) < 1:
        return 0

    # Calculate control limits (mean ± 2σ)
    rates = [dp[1] for dp in data_points]
    mean_rate = sum(rates) / len(rates) if rates else 0
    variance = sum((r - mean_rate) ** 2 for r in rates) / len(rates) if rates else 0
    std_rate = variance ** 0.5
    ucl = round(mean_rate + 2 * std_rate, 1)
    lcl = round(max(0, mean_rate - 2 * std_rate), 1)
    mean_rate = round(mean_rate, 1)

    sheet.cell(row=1, column=1, value="Fecha")
    sheet.cell(row=1, column=2, value="Tasa Fallos (%)")
    sheet.cell(row=1, column=3, value="Promedio (%)")
    sheet.cell(row=1, column=4, value="Límite Superior (UCL)")
    sheet.cell(row=1, column=5, value="Límite Inferior (LCL)")

    for r, (dt, fr, ah, cnt) in enumerate(data_points, 2):
        sheet.cell(row=r, column=1, value=dt)
        sheet.cell(row=r, column=2, value=fr)
        sheet.cell(row=r, column=3, value=mean_rate)
        sheet.cell(row=r, column=4, value=ucl)
        sheet.cell(row=r, column=5, value=lcl)

    return len(data_points)


def _build_run_chart(sheet, data_rows):
    """P14: Line chart — run chart de tasa de fallos con límites de control."""
    from openpyxl.chart import LineChart

    chart = LineChart()
    chart.title = "Run Chart — Tasa de Fallos (Bajo+Crítico) con Límites de Control"
    chart.x_axis.title = "Fecha"
    chart.y_axis.title = "Tasa de Fallos (%)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(sheet, min_col=1, min_row=2, max_row=data_rows + 1)

    # Actual failure rate
    data_rate = Reference(sheet, min_col=2, min_row=1, max_row=data_rows + 1)
    chart.add_data(data_rate, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "e74c3c"
    chart.series[0].graphicalProperties.line.width = 30000
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 7
    chart.series[0].marker.graphicalProperties.solidFill = "e74c3c"

    # Mean line
    data_mean = Reference(sheet, min_col=3, min_row=1, max_row=data_rows + 1)
    chart.add_data(data_mean, titles_from_data=True)
    chart.series[1].graphicalProperties.line.solidFill = "3498db"
    chart.series[1].graphicalProperties.line.width = 20000
    chart.series[1].graphicalProperties.line.dashStyle = "dash"

    # UCL
    data_ucl = Reference(sheet, min_col=4, min_row=1, max_row=data_rows + 1)
    chart.add_data(data_ucl, titles_from_data=True)
    chart.series[2].graphicalProperties.line.solidFill = "f39c12"
    chart.series[2].graphicalProperties.line.width = 15000
    chart.series[2].graphicalProperties.line.dashStyle = "dot"

    # LCL
    data_lcl = Reference(sheet, min_col=5, min_row=1, max_row=data_rows + 1)
    chart.add_data(data_lcl, titles_from_data=True)
    chart.series[3].graphicalProperties.line.solidFill = "f39c12"
    chart.series[3].graphicalProperties.line.width = 15000
    chart.series[3].graphicalProperties.line.dashStyle = "dot"

    chart.legend.position = "b"

    if data_rows == 1:
        chart.title = "Run Chart Fallos (1 punto — ejecutar múltiples veces para ver tendencia)"

    return chart


def _add_charts_sheet(writer, df_health):
    """Genera pestaña Charts con 14 gráficos nativos de Excel + 1 tabla heatmap."""
    if df_health.empty:
        print("⚠️  Sin datos para generar Charts")
        return

    workbook = writer.book

    # Create hidden data sheets for chart source data
    chart_specs = [
        ("_data_stacked",  _write_chart_data_stacked_bar,  _build_stacked_bar_chart),
        ("_data_donut",    _write_chart_data_donut,        _build_donut_chart),
        ("_data_dora",     _write_chart_data_dora_bar,     _build_dora_grouped_bar),
        ("_data_scatter",  _write_chart_data_scatter,      _build_scatter_chart),
        ("_data_treemap",  _write_chart_data_treemap,      _build_treemap_chart),
        ("_data_pareto",   _write_chart_data_pareto,       _build_pareto_chart),
        ("_data_trend",    _write_chart_data_trend,        _build_trend_chart),
        ("_data_risk",     _write_chart_data_risk_treemap, _build_risk_treemap_chart),
        ("_data_sankey",   _write_chart_data_sankey,       _build_sankey_chart),
        ("_data_radar",    _write_chart_data_radar,         _build_radar_chart),
        ("_data_bubble",   _write_chart_data_bubble,       _build_bubble_chart),
        ("_data_mttr",     _write_chart_data_mttr_hist,    _build_mttr_hist_chart),
        ("_data_run",      _write_chart_data_run,          _build_run_chart),
    ]

    chart_objects = []
    for sheet_name, write_fn, build_fn in chart_specs:
        ds = workbook.create_sheet(sheet_name)
        ds.sheet_state = "hidden"
        data_rows = write_fn(ds, df_health)
        if data_rows > 0:
            ch = build_fn(ds, data_rows)
            chart_objects.append(ch)
        else:
            chart_objects.append(None)

    # Build heatmap table directly on Charts sheet (not a chart, a colored table)
    heatmap_rows = _build_heatmap_table(df_health)

    # Create visible Charts sheet and place charts
    charts_sheet = workbook.create_sheet("Charts")

    chart_titles = [
        "P1 — Composición Health Score (Top 30)",
        "P2 — Distribución de Ratings",
        "P3 — DORA Profile CI vs CD",
        "P5 — Health Score vs Uso (Priorización)",
        "P6 — Tecnologías: Cantidad vs Salud Promedio",
        "P7 — Pareto: Tecnologías con más pipelines críticos",
        "P8 — Tendencia Health Score (histórico)",
        "P9 — Riesgo Tecnológico por Área",
        "P10 — Flujo Tecnología → Recomendación",
        "P11 — Radar DORA por Área",
        "P12 — Esfuerzo vs Impacto (Burbujas)",
        "P13 — Histograma MTTR",
        "P14 — Run Chart Tasa de Fallos",
    ]

    # Layout: 2 columns x 7 rows of charts (13 positions)
    positions = [
        "A1",   "N1",
        "A26",  "N26",
        "A51",  "N51",
        "A76",  "N76",
        "A101", "N101",
        "A126",
    ]

    pos_idx = 0
    for i, (ch, title) in enumerate(zip(chart_objects, chart_titles)):
        if ch is None:
            continue
        ch.title = title
        if pos_idx < len(positions):
            charts_sheet.add_chart(ch, positions[pos_idx])
        else:
            # Overflow: place in next available row
            row = 1 + pos_idx * 25
            charts_sheet.add_chart(ch, f"A{row}")
        pos_idx += 1

    # Heatmap table after last chart
    heatmap_start = 1 + pos_idx * 25
    _write_heatmap_to_sheet(charts_sheet, df_health, start_row=heatmap_start)

    print(f"   📊 Hoja 4 — Charts:         13 gráficos nativos Excel + 1 tabla heatmap")


# ==========================================================
# EXPORT 3-SHEET EXCEL
# ==========================================================

def export_three_sheet_excel(ci_rows, cd_rows, health_rows, output_dir):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = output_dir / f"{SCRIPT_NAME}_{ts}.xlsx"
    
    df_ci = pd.DataFrame(ci_rows) if ci_rows else pd.DataFrame()
    df_cd = pd.DataFrame(cd_rows) if cd_rows else pd.DataFrame()
    df_health = pd.DataFrame(health_rows) if health_rows else pd.DataFrame()
    
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        if not df_ci.empty:
            df_ci.to_excel(writer, sheet_name="CI Inventory", index=False)
        if not df_cd.empty:
            df_cd.to_excel(writer, sheet_name="CD Inventory", index=False)
        if not df_health.empty:
            df_health.to_excel(writer, sheet_name="Health Score", index=False)
            _add_charts_sheet(writer, df_health)
    
    print(f"📊 Excel generado: {excel_path.resolve()}")
    print(f"   📋 Hoja 1 — CI Inventory:     {len(df_ci)} filas")
    print(f"   📋 Hoja 2 — CD Inventory:     {len(df_cd)} filas")
    print(f"   📊 Hoja 3 — Health Score:     {len(df_health)} filas")
    return excel_path


# ==========================================================
# RESUMEN RICH
# ==========================================================

def print_summary(ci_count, cd_count, cache_ci, cache_cd, api_calls, health_rows, duration):
    ratings = {"Excelente": 0, "Bueno": 0, "Regular": 0, "Bajo": 0, "Crítico": 0}
    for row in health_rows:
        r = row.get("rating", "")
        if r in ratings:
            ratings[r] += 1
    
    if not RICH_AVAILABLE:
        print(f"\n{'='*60}")
        print(f"📊 RESUMEN PIPELINE HEALTH SCORE")
        print(f"   Pipelines CI:        {ci_count}")
        print(f"   Pipelines CD:        {cd_count}")
        print(f"   Cache CI usado:      {'Sí' if cache_ci else 'No'}")
        print(f"   Cache CD usado:      {'Sí' if cache_cd else 'No'}")
        print(f"   Llamadas API:        {api_calls}")
        print(f"   Duración:            {duration:.1f}s")
        print(f"   Distribución:")
        for k, v in ratings.items():
            print(f"      {k}: {v}")
        print(f"{'='*60}")
        return
    
    console = Console(file=sys.__stdout__)
    table = Table(title="📊 Resumen Pipeline Health Score", show_header=True, header_style="bold magenta")
    table.add_column("Métrica", style="cyan")
    table.add_column("Valor", style="green")
    table.add_row("Pipelines CI", str(ci_count))
    table.add_row("Pipelines CD", str(cd_count))
    table.add_row("Cache CI usado", "✅ Sí" if cache_ci else "❌ No")
    table.add_row("Cache CD usado", "✅ Sí" if cache_cd else "❌ No")
    table.add_row("Llamadas API", str(api_calls))
    table.add_row("Duración", f"{duration:.1f}s")
    for k, v in ratings.items():
        table.add_row(f"Rating {k}", str(v))
    console.print(table)


# ==========================================================
# MAIN
# ==========================================================

def _launch_proc(label, cmd, env):
    """Lanza un subprocess Popen y retorna el proceso."""
    return subprocess.Popen(cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)


def _fmt_elapsed(seconds):
    """Formatea segundos como Xm Ys."""
    m, s = divmod(int(seconds), 60)
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def _run_inventory_scripts(args, tee):
    """Lanza cicd_inventory_ci_detailed.py y cicd_inventory_cd_detailed.py en paralelo con spinner dinámico."""
    script_dir = Path(__file__).parent
    python_exe = sys.executable
    
    common_env = os.environ.copy()
    common_env["AZDO_PAT"] = args.pat
    
    ci_cmd = [
        python_exe, str(script_dir / "cicd_inventory_ci_detailed.py"),
        "--pat", args.pat,
        "--org", args.org,
        "--project", args.project,
        "--workers", str(args.workers),
    ]
    cd_cmd = [
        python_exe, str(script_dir / "cicd_inventory_cd_detailed.py"),
        "--pat", args.pat,
        "--org", args.org,
        "--project", args.project,
        "--workers", str(args.workers),
    ]
    
    print("\n🔄 Ejecutando inventory CI y CD en paralelo...")
    print(f"   CI: cicd_inventory_ci_detailed.py --org {args.org} --project {args.project}")
    print(f"   CD: cicd_inventory_cd_detailed.py --org {args.org} --project {args.project}")
    
    ci_proc = _launch_proc("CI-Inventory", ci_cmd, common_env)
    cd_proc = _launch_proc("CD-Inventory", cd_cmd, common_env)
    procs = {"CI-Inventory": ci_proc, "CD-Inventory": cd_proc}
    results = {}
    start = time.time()
    
    # Spinner frames
    spinner_frames = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
    frame_idx = 0
    
    if RICH_AVAILABLE:
        tee.pause_terminal()
        console = Console(file=sys.__stdout__)
        from rich.live import Live
        from rich.text import Text
        
        with Live(console=console, refresh_per_second=4, transient=True) as live:
            while procs:
                elapsed = time.time() - start
                frame = spinner_frames[frame_idx % len(spinner_frames)]
                frame_idx += 1
                
                lines = [f"{frame} [bold blue]Inventory en progreso[/] ({_fmt_elapsed(elapsed)})"]
                for label, proc in list(procs.items()):
                    if proc.poll() is not None:
                        rc = proc.returncode
                        icon = "✅" if rc == 0 else "⚠️"
                        lines.append(f"  {icon} {label}: terminado (código {rc})")
                        results[label] = rc
                        del procs[label]
                    else:
                        lines.append(f"  🔄 {label}: ejecutándose...")
                
                live.update(Text.from_markup("\n".join(lines)))
                time.sleep(0.25)
        
        tee.resume_terminal()
    else:
        # Fallback sin Rich: imprimir líneas con \r
        last_print = 0
        while procs:
            elapsed = time.time() - start
            if elapsed - last_print >= 2:
                status_parts = []
                for label, proc in procs.items():
                    status_parts.append(f"{label}: ejecutándose")
                print(f"   ⏳ {_fmt_elapsed(elapsed)} — {', '.join(status_parts)}")
                last_print = elapsed
            
            for label, proc in list(procs.items()):
                if proc.poll() is not None:
                    rc = proc.returncode
                    icon = "✅" if rc == 0 else "⚠️"
                    print(f"   {icon} {label}: terminado (código {rc})")
                    results[label] = rc
                    del procs[label]
            time.sleep(0.5)
    
    # Capturar stderr de procesos terminados para diagnóstico
    for label, rc in results.items():
        proc = ci_proc if label == "CI-Inventory" else cd_proc
        if proc.stderr:
            err_lines = proc.stderr.read().strip().split("\n")[-3:]
            if err_lines and any(l.strip() for l in err_lines):
                print(f"   [{label}] stderr (últimas líneas):")
                for l in err_lines:
                    if l.strip():
                        print(f"      {l.strip()}")
    
    ci_rc = results.get("CI-Inventory", -1)
    cd_rc = results.get("CD-Inventory", -1)
    print(f"   CI-Inventory: código {ci_rc}")
    print(f"   CD-Inventory: código {cd_rc}")
    print("✅ Inventory completado. Continuando con Health Score...\n")


def main():
    parser = argparse.ArgumentParser(description="Pipeline Health Score Report")
    parser.add_argument("--pat", default=os.getenv("AZDO_PAT"), help="Azure DevOps PAT")
    parser.add_argument("--org", default=DEFAULT_ORG, help="Organización")
    parser.add_argument("--project", default=DEFAULT_PROJECT, help="Proyecto")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="Hilos paralelos")
    parser.add_argument("--output", default=None, help="Directorio de salida")
    parser.add_argument("--force-refresh", action="store_true", help="Ignorar cache, re-consultar todo")
    parser.add_argument("--offline", action="store_true", help="Solo usar cache, fallar si no existe")
    parser.add_argument("--skip-incremental", action="store_true", help="No consultar últimas 20 ejecuciones")
    parser.add_argument("--run-inventory", action="store_true", help="Ejecutar CI y CD inventory en hilos paralelos antes de procesar")
    args = parser.parse_args()
    args.org = normalize_org(args.org)

    if not args.pat and not args.offline:
        print("❌ Se requiere --pat o env AZDO_PAT (o usar --offline)")
        sys.exit(1)

    output_dir = get_output_dir(args.output or "outcome")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    tee = setup_logging()
    start_time = time.time()
    api_calls = 0
    ci_rows = []
    cd_rows = []
    health_rows = []
    cache_ci_used = False
    cache_cd_used = False
    
    try:
        # ============================================
        # PASO 0: EJECUTAR INVENTORY EN SEGUNDO PLANO
        # ============================================
        if args.run_inventory and not args.offline:
            _run_inventory_scripts(args, tee)
        
        # ============================================
        # PASO 1: OBTENER DATOS BASE CI
        # ============================================
        ci_base_data = []
        if not args.force_refresh:
            cache_path = _find_latest_cache("cicd_inventory_ci_detailed")
            if cache_path and _cache_is_fresh(cache_path):
                print(f"📦 CI Cache encontrado: {cache_path.name}")
                data = _load_cache(cache_path)
                ci_base_data = data.get("rows", [])
                cache_ci_used = True
            else:
                print("📭 CI Cache no encontrado o > 24h")
        
        if not ci_base_data:
            if args.offline:
                print("❌ Modo offline pero no hay cache CI disponible")
                sys.exit(1)
            print("🔍 Consultando CI definitions...")
            if RICH_AVAILABLE:
                tee.pause_terminal()
                with Progress(SpinnerColumn(), TextColumn("[bold blue]{task.description}"), console=Console(file=sys.__stdout__)) as progress:
                    task = progress.add_task("Fetching CI definitions...", total=None)
                    headers = get_headers(args.pat)
                    definitions = fetch_ci_definitions(headers, args.org, args.project)
                    api_calls += 1
                    progress.update(task, total=1, completed=1)
                tee.resume_terminal()
            else:
                headers = get_headers(args.pat)
                definitions = fetch_ci_definitions(headers, args.org, args.project)
                api_calls += 1
            
            print(f"📋 {len(definitions)} CI definitions encontradas")
            
            # Simplified base extraction for cache-less mode
            for d in definitions:
                ci_base_data.append({
                    "id": d.get("id"),
                    "name": d.get("name", ""),
                    "path": d.get("path", ""),
                    "url": d.get("url", ""),
                    "createdDate": d.get("createdDate", ""),
                    "modifiedDate": d.get("modifiedDate", ""),
                    "processType": d.get("process", {}).get("type", ""),
                    "yamlFilename": d.get("process", {}).get("yamlFilename", ""),
                    "repositoryName": d.get("repository", {}).get("name", ""),
                    "repositoryUrl": d.get("repository", {}).get("url", ""),
                    "defaultBranch": d.get("repository", {}).get("defaultBranch", ""),
                    "lastPipelineModifier": d.get("authoredBy", {}).get("displayName", ""),
                    "lastExecution": "",
                    "lastExecutionState": "",
                    "lastExecutionResult": "",
                    "totalExecutions30d": 0,
                    "totalExecutions90d": 0,
                })
            
            # Save cache for future
            _save_cache({
                "metadata": {"script": "cicd_inventory_health_score_ci", "org": args.org, "project": args.project, "generated_at": datetime.now(timezone.utc).isoformat(), "count": len(ci_base_data)},
                "rows": ci_base_data,
            }, "cicd_inventory_health_score_ci")
            print("💾 CI cache guardado")
        
        # ============================================
        # PASO 2: OBTENER DATOS BASE CD
        # ============================================
        cd_base_data = []
        if not args.force_refresh:
            cache_path = _find_latest_cache("cicd_inventory_cd_detailed")
            if cache_path and _cache_is_fresh(cache_path):
                print(f"📦 CD Cache encontrado: {cache_path.name}")
                data = _load_cache(cache_path)
                cd_base_data = data.get("rows", [])
                cache_cd_used = True
            else:
                print("📭 CD Cache no encontrado o > 24h")
        
        if not cd_base_data:
            if args.offline:
                print("❌ Modo offline pero no hay cache CD disponible")
                sys.exit(1)
            print("🔍 Consultando CD definitions...")
            if RICH_AVAILABLE:
                tee.pause_terminal()
                with Progress(SpinnerColumn(), TextColumn("[bold blue]{task.description}"), console=Console(file=sys.__stdout__)) as progress:
                    task = progress.add_task("Fetching CD definitions...", total=None)
                    headers = get_headers(args.pat)
                    definitions = fetch_cd_definitions(headers, args.org, args.project)
                    api_calls += 1
                    progress.update(task, total=1, completed=1)
                tee.resume_terminal()
            else:
                headers = get_headers(args.pat)
                definitions = fetch_cd_definitions(headers, args.org, args.project)
                api_calls += 1
            
            print(f"📋 {len(definitions)} CD definitions encontradas")
            
            for d in definitions:
                envs = d.get("environments", [])
                cd_base_data.append({
                    "id": d.get("id"),
                    "name": d.get("name", ""),
                    "path": d.get("path", ""),
                    "url": d.get("url", ""),
                    "createdOn": d.get("createdOn", ""),
                    "modifiedOn": d.get("modifiedOn", ""),
                    "environmentsCount": len(envs),
                    "environments": " / ".join([e.get("name", "") for e in envs]),
                    "lastReleaseDate": "",
                    "lastReleaseStatus": "",
                    "isObsolete": "Sí" if any(kw in d.get("name", "").lower() for kw in ["obsoleto", "obsolete", "_old", "legacy-", "deprecated"]) else "No",
                })
            
            _save_cache({
                "metadata": {"script": "cicd_inventory_health_score_cd", "org": args.org, "project": args.project, "generated_at": datetime.now(timezone.utc).isoformat(), "count": len(cd_base_data)},
                "rows": cd_base_data,
            }, "cicd_inventory_health_score_cd")
            print("💾 CD cache guardado")
        
        # ============================================
        # PASO 3: ENRIQUECER CI CON INCREMENTAL
        # ============================================
        print(f"🔧 Enriqueciendo {len(ci_base_data)} CI pipelines...")
        headers = get_headers(args.pat) if args.pat else {}
        
        if RICH_AVAILABLE and ci_base_data:
            tee.pause_terminal()
            with _progress_context() as progress:
                task = progress.add_task("Enriqueciendo CI pipelines", total=len(ci_base_data))
                with ThreadPoolExecutor(max_workers=args.workers) as executor:
                    futures = {executor.submit(build_ci_row, d, headers, args.org, args.project, args.skip_incremental): d for d in ci_base_data}
                    for future in as_completed(futures):
                        try:
                            row = future.result()
                            if row:
                                ci_rows.append(row)
                                health_rows.append(row)
                                api_calls += 1
                        except Exception as e:
                            print(f"❌ Error CI pipeline: {e}")
                        progress.update(task, advance=1)
            tee.resume_terminal()
        else:
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(build_ci_row, d, headers, args.org, args.project, args.skip_incremental): d for d in ci_base_data}
                for i, future in enumerate(as_completed(futures), 1):
                    try:
                        row = future.result()
                        if row:
                            ci_rows.append(row)
                            health_rows.append(row)
                            api_calls += 1
                    except Exception as e:
                        print(f"❌ Error CI pipeline: {e}")
                    if i % 10 == 0 or i == len(ci_base_data):
                        print(f"  CI Progreso: {i}/{len(ci_base_data)}")
        
        # ============================================
        # PASO 4: ENRIQUECER CD CON INCREMENTAL
        # ============================================
        print(f"🔧 Enriqueciendo {len(cd_base_data)} CD pipelines...")
        
        if RICH_AVAILABLE and cd_base_data:
            tee.pause_terminal()
            with _progress_context() as progress:
                task = progress.add_task("Enriqueciendo CD pipelines", total=len(cd_base_data))
                with ThreadPoolExecutor(max_workers=args.workers) as executor:
                    futures = {executor.submit(build_cd_row, d, headers, args.org, args.project, args.skip_incremental): d for d in cd_base_data}
                    for future in as_completed(futures):
                        try:
                            row = future.result()
                            if row:
                                cd_rows.append(row)
                                health_rows.append(row)
                                api_calls += 1
                        except Exception as e:
                            print(f"❌ Error CD pipeline: {e}")
                        progress.update(task, advance=1)
            tee.resume_terminal()
        else:
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(build_cd_row, d, headers, args.org, args.project, args.skip_incremental): d for d in cd_base_data}
                for i, future in enumerate(as_completed(futures), 1):
                    try:
                        row = future.result()
                        if row:
                            cd_rows.append(row)
                            health_rows.append(row)
                            api_calls += 1
                    except Exception as e:
                        print(f"❌ Error CD pipeline: {e}")
                    if i % 10 == 0 or i == len(cd_base_data):
                        print(f"  CD Progreso: {i}/{len(cd_base_data)}")
        
        # ============================================
        # PASO 5: EXPORTAR
        # ============================================
        if health_rows:
            export_three_sheet_excel(ci_rows, cd_rows, health_rows, output_dir)
        else:
            print("⚠️  No hay datos para exportar")
        
        duration = time.time() - start_time
        print_summary(len(ci_rows), len(cd_rows), cache_ci_used, cache_cd_used, api_calls, health_rows, duration)
        
    finally:
        teardown_logging(tee)


if __name__ == "__main__":
    main()

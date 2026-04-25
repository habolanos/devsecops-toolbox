#!/usr/bin/env python3
"""
Branches Created Reporter v1.0.0
Obtiene ramas creadas desde una fecha específica en Azure DevOps
usando la Pushes API con filtro de fecha directo.

Uso:
    python branches_created.py --org Coppel-Retail --project Compras.RMI
    python branches_created.py --org Coppel-Retail --project Compras.RMI --since 2026-01-01

Autor: Harold Adrian (migrado desde Comercial/scripts/ramasCreadasV2.py)
"""

import os
import sys
import base64
import requests
import argparse
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter
from pathlib import Path
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = lambda *a, **k: None  # noqa: E731

try:
    from rich.progress import Progress, SpinnerColumn, BarColumn, TaskProgressColumn, TextColumn, TimeElapsedColumn
    from rich.console import Console
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

# --- Directorio de salida centralizado (DEVSECOPS_OUTPUT_DIR) ---
try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
# -------------------------------------------------------------------

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

load_dotenv(Path(__file__).parent.parent / ".env")


class TeeWriter:
    """Escribe stdout a consola Y archivo de log simultáneamente."""
    def __init__(self, log_path):
        self.terminal = sys.__stdout__
        self.log = open(log_path, "w", encoding="utf-8")
        self.log_path = log_path
        self._paused = False  # Cuando True, solo escribe al archivo log

    def write(self, message):
        self.log.write(message)
        if not self._paused:
            self.terminal.write(message)

    def flush(self):
        self.log.flush()
        if not self._paused:
            self.terminal.flush()

    def close(self):
        self.log.close()

    def pause_terminal(self):
        """Pausa escritura al terminal (Rich toma control)."""
        self._paused = True

    def resume_terminal(self):
        """Reanuda escritura al terminal."""
        self._paused = False


def setup_logging(script_name):
    """Configura TeeWriter para que stdout vaya a consola + archivo log en outcome."""
    output_dir = get_output_dir("outcome")
    output_dir.mkdir(parents=True, exist_ok=True)
    log_path = output_dir / f"{script_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeWriter(log_path)
    sys.stdout = tee
    print(f"📝 Log: {log_path.resolve()}")
    print(f"📅 Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    return tee


def teardown_logging(tee):
    """Restaura stdout y cierra archivo de log."""
    print(f"\n📝 Log guardado: {tee.log_path.resolve()}")
    sys.stdout = tee.terminal
    tee.close()


class _SimpleProgress:
    """Fallback: imprime progreso como texto simple."""
    def __init__(self, desc, total):
        self.desc = desc
        self.total = total
        self.completed = 0
        self._last_pct = -1

    def advance(self, n=1):
        self.completed += n
        pct = int(self.completed / self.total * 100) if self.total else 100
        if pct != self._last_pct and pct % 10 == 0:
            print(f"  {self.desc}: {self.completed}/{self.total} ({pct}%)")
            self._last_pct = pct

    def finish(self, msg):
        print(f"  ✅ {msg}")


def _progress_context():
    """Retorna contexto Rich Progress que escribe al terminal real (no TeeWriter)."""
    if not RICH_AVAILABLE:
        return None
    console = Console(file=sys.__stdout__)
    return Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
    )


# ─── Configuración por defecto ───────────────────────────────────────
DEFAULT_ORG = "Coppel-Retail"
DEFAULT_PROJECT = "Compras.RMI"
API_VERSION = "7.1"
DEFAULT_SINCE = "2026-01-01"
DEFAULT_WORKERS = 30
PUSHES_PAGE_SIZE = 200


def get_headers(pat: str):
    """Genera headers de autenticación."""
    token_b64 = base64.b64encode(f":{pat}".encode()).decode()
    return {
        "Authorization": f"Basic {token_b64}",
        "Content-Type": "application/json"
    }


def normalize_org(org: str) -> str:
    """Extrae el nombre de la organización desde una URL completa o nombre simple."""
    if org.startswith("http"):
        return org.rstrip("/").split("/")[-1]
    return org


def get_repositories(org: str, project: str, headers: dict):
    """Obtiene lista de repositorios del proyecto."""
    url = f"https://dev.azure.com/{org}/{project}/_apis/git/repositories"
    params = {"api-version": API_VERSION}
    resp = requests.get(url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    return data.get("value", [])


def get_pushes_for_repo(repo_id: str, since_date: str, org: str, project: str, headers: dict):
    """Obtiene pushes de un repo desde la fecha indicada (con paginación)."""
    url = f"https://dev.azure.com/{org}/{project}/_apis/git/repositories/{repo_id}/pushes"
    all_pushes = []
    skip = 0
    
    while True:
        params = {
            "api-version": API_VERSION,
            "searchCriteria.fromDate": since_date,
            "searchCriteria.includeRefUpdates": "true",
            "$top": PUSHES_PAGE_SIZE,
            "$skip": skip,
        }
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            page = data.get("value", [])
            all_pushes.extend(page)
            if len(page) < PUSHES_PAGE_SIZE:
                break
            skip += PUSHES_PAGE_SIZE
        except Exception as e:
            print(f"   ⚠️ Error en pushes de repo {repo_id}: {e}")
            break
    
    return all_pushes


def process_repo(repo: dict, since_date: str, org: str, project: str, headers: dict):
    """Procesa un repositorio y retorna ramas nuevas detectadas."""
    repo_id = repo.get("id")
    repo_name = repo.get("name", "unknown")
    
    pushes = get_pushes_for_repo(repo_id, since_date, org, project, headers)
    
    branches = []
    seen_refs = set()
    
    for push in pushes:
        ref_updates = push.get("refUpdates", [])
        for ref in ref_updates:
            ref_name = ref.get("name", "")
            if ref_name.startswith("refs/heads/"):
                branch_name = ref_name.replace("refs/heads/", "")
                if branch_name not in seen_refs:
                    seen_refs.add(branch_name)
                    branches.append({
                        "repo_name": repo_name,
                        "branch_name": branch_name,
                        "created_date": push.get("date", "")[:10] if push.get("date") else "unknown",
                        "created_by": push.get("pushedBy", {}).get("displayName", "unknown"),
                        "push_id": push.get("pushId")
                    })
    
    return repo_name, branches


def export_to_excel(branches: list, org: str, project: str, since_date: str, output_file: str = None):
    """Exporta ramas a Excel con resumen."""
    if not output_file:
        output_dir = get_output_dir()
        default_name = f"branches_created_{org}_{project}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file = str(output_dir / default_name)
    
    wb = Workbook()
    
    # Hoja 1: Ramas
    ws1 = wb.active
    ws1.title = "Ramas Creadas"
    headers = ["Repositorio", "Rama", "Fecha Creación", "Creado por", "Push ID"]
    ws1.append(headers)
    
    for branch in branches:
        ws1.append([
            branch["repo_name"],
            branch["branch_name"],
            branch["created_date"],
            branch["created_by"],
            branch["push_id"]
        ])
    
    # Hoja 2: Resumen por Repo
    ws2 = wb.create_sheet("Resumen por Repo")
    ws2.append(["Repositorio", "Total Ramas Nuevas"])
    
    repo_counts = Counter(b["repo_name"] for b in branches)
    for repo, count in sorted(repo_counts.items(), key=lambda x: x[1], reverse=True):
        ws2.append([repo, count])
    
    # Hoja 3: Resumen por Creador
    ws3 = wb.create_sheet("Resumen por Creador")
    ws3.append(["Creador", "Total Ramas Creadas"])
    
    creator_counts = Counter(b["created_by"] for b in branches)
    for creator, count in sorted(creator_counts.items(), key=lambda x: x[1], reverse=True):
        ws3.append([creator, count])
    
    # Estilos
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar anchos
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 25
    ws1.column_dimensions["E"].width = 15
    
    wb.save(output_file)
    return output_file


def main():
    parser = argparse.ArgumentParser(
        description="Branches Created Reporter - Ramas creadas desde fecha específica",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python branches_created.py
  python branches_created.py --org Coppel-Retail --project Compras.RMI
  python branches_created.py --org Coppel-Retail --project Compras.RMI --since 2025-01-01
  python branches_created.py --org Coppel-Retail --project Compras.RMI --workers 5
        """
    )
    parser.add_argument("--org", default=DEFAULT_ORG,
                       help=f"Organización Azure DevOps (default: {DEFAULT_ORG})")
    parser.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                       help=f"Proyecto (default: {DEFAULT_PROJECT})")
    parser.add_argument("--since", "-s", default=DEFAULT_SINCE,
                       help=f"Fecha desde (ISO format, default: {DEFAULT_SINCE})")
    parser.add_argument("--workers", "-w", type=int, default=DEFAULT_WORKERS,
                       help=f"Workers concurrentes (default: {DEFAULT_WORKERS})")
    parser.add_argument("--pat", default=None,
                       help="PAT de Azure DevOps (default: env AZURE_PAT)")
    parser.add_argument("--output", "-o", default=None,
                       help="Nombre del archivo Excel de salida")
    parser.add_argument("--repo-limit", type=int, default=None,
                       help="Limitar número de repos a procesar (para pruebas)")
    
    args = parser.parse_args()
    
    # Obtener PAT
    pat = args.pat or os.getenv("AZURE_PAT", "")
    if not pat:
        print("❌ ERROR: Variable AZURE_PAT no definida")
        print("   Ejecuta: export AZURE_PAT='tu_token'")
        sys.exit(1)
    
    headers = get_headers(pat)
    since_dt = datetime.fromisoformat(args.since.replace("Z", "+00:00"))
    
    # Configurar logging a archivo en outcome
    tee = setup_logging("cicd_inventory_branches_created")
    
    org = normalize_org(args.org)

    print(f"🔍 Analizando ramas creadas desde: {args.since}")
    print(f"   Org: {org}")
    print(f"   Project: {args.project}")
    print(f"   API: {API_VERSION} | Workers: {args.workers}")
    print("=" * 60)
    
    repos = get_repositories(org, args.project, headers)
    
    if args.repo_limit:
        repos = repos[:args.repo_limit]
        print(f"*** MODO PRUEBA: limitado a {args.repo_limit} repos ***")
    
    all_branches = []
    total = len(repos)
    print(f"\n📦 Consultando {total} repositorios (concurrente)...")
    
    progress = _progress_context()
    
    if progress:
        # Pausar TeeWriter del terminal para que Rich tome control
        tee.pause_terminal()
        with progress:
            task = progress.add_task("🌿 Ramas creadas", total=total)
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(process_repo, r, args.since, org, args.project, headers): r for r in repos}
                for f in as_completed(futures):
                    try:
                        repo_name, branches = f.result()
                        all_branches.extend(branches)
                    except Exception as e:
                        repo_name = futures[f].get("name", "?")
                        print(f"   ⚠️ Error en repo {repo_name}: {e}")
                    progress.advance(task)
        # Reanudar TeeWriter
        tee.resume_terminal()
    else:
        sp = _SimpleProgress("🌿 Ramas creadas", total)
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {executor.submit(process_repo, r, args.since, org, args.project, headers): r for r in repos}
            for f in as_completed(futures):
                try:
                    repo_name, branches = f.result()
                    all_branches.extend(branches)
                except Exception as e:
                    repo_name = futures[f].get("name", "?")
                    print(f"   ⚠️ Error en repo {repo_name}: {e}")
                sp.advance()
        sp.finish(f"{len(all_branches)} ramas nuevas en {total} repos")
    
    if not all_branches:
        print(f"\nℹ️ No se encontraron ramas creadas desde {args.since}")
        sys.exit(0)
    
    print(f"\n✅ Total ramas nuevas encontradas: {len(all_branches)}")
    
    output_file = export_to_excel(all_branches, org, args.project, args.since, args.output)
    excel_path = Path(output_file).resolve()
    print(f"\n💾 Archivo generado: {excel_path}")
    
    teardown_logging(tee)


if __name__ == "__main__":
    main()

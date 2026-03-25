#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
azdo_release_cd_health.py

Analiza todos los Release Pipelines CD de un proyecto Azure DevOps:

  ✦ Lista de stages por pipeline (nombres ordenados por rank)
  ✦ Consistencia de stages entre todos los pipelines del proyecto
  ✦ Detección automática del stage de producción
  ✦ ¿El stage de producción fue ejecutado? ¿Cuándo por última vez?
  ✦ Score de salud 0-100 basado en:
       Recencia    (0-70 pts) — deploy reciente = puntaje alto; escala lineal 365 días
       Estabilidad (0-30 pts) — 1 intento=30, 2=20, 3=10, 4+=0 (último release)

Rating:
  🟢 Excelente   90-100
  🔵 Bueno       70-89
  🟡 Regular     40-69
  🟠 Bajo         1-39
  🔴 Nunca        0

Consistencia de stages:
  ✅ OK       — Idéntico al patrón mayoritario del proyecto
  🟡 Parcial  — Falta algún stage o tiene extras vs. el patrón
  🔴 Diferente — Sin coincidencia con el patrón mayoritario
  ❓ Único    — Pipeline único (sin referencia de comparación)

Autor: Harold Adrian
"""

import argparse
import base64
import csv
import json
import os
import time
from collections import Counter
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Set, Tuple
from zoneinfo import ZoneInfo

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
    from rich.table import Table
    from rich.text import Text
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

from io import BytesIO

try:
    import matplotlib
    matplotlib.use("Agg")  # backend no-interactivo (sin ventana)
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

__version__ = "1.0.0"
__author__ = "Harold Adrian"

# ═══════════════════════════════════════════════════════════════════════════════
# DEFAULTS
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_ORG_URL  = "https://dev.azure.com/Coppel-Retail"
DEFAULT_PROJECT  = "Compras.RMI"
DEFAULT_TIMEZONE = "America/Mazatlan"
API_VERSION_DEFS = "7.2-preview.4"
API_VERSION_RELS = "7.2-preview.8"
DEFAULT_TOP      = 15   # últimos N releases a analizar por pipeline
DEFAULT_THREADS  = 8

PROD_KEYWORDS: Set[str] = {
    "prod", "prd", "production", "produccion", "productivo",
    "producción", "live", "prd01", "prd1", "productivo",
}

CONS_OK      = "OK"
CONS_PARTIAL = "PARCIAL"
CONS_DIFF    = "DIFERENTE"
CONS_UNIQUE  = "ÚNICO"


# ═══════════════════════════════════════════════════════════════════════════════
# ARGS
# ═══════════════════════════════════════════════════════════════════════════════
def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Analiza Release Pipelines CD: stages, consistencia y score de salud"
    )
    p.add_argument("--org", "-g", default=DEFAULT_ORG_URL,
                   help=f"URL de la organización (default: {DEFAULT_ORG_URL})")
    p.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                   help=f"Nombre del proyecto (default: {DEFAULT_PROJECT})")
    p.add_argument("--pat", required=True,
                   help="Personal Access Token con permiso Release (Read)")
    p.add_argument("--filter", "--repo", "-f", "-r", dest="filter", default=None,
                   help="Filtrar pipelines por nombre/repo (substring, case insensitive)")
    p.add_argument("--output", "-o", choices=["json", "csv", "excel"], default=None,
                   help="Exportar resultados (json / csv / excel)")
    p.add_argument("--timezone", "-tz", default=DEFAULT_TIMEZONE,
                   help=f"Zona horaria para fechas (default: {DEFAULT_TIMEZONE})")
    p.add_argument("--threads", type=int, default=DEFAULT_THREADS,
                   help=f"Hilos paralelos (default: {DEFAULT_THREADS})")
    p.add_argument("--top", type=int, default=DEFAULT_TOP,
                   help=f"Últimos N releases a analizar por pipeline (default: {DEFAULT_TOP})")
    p.add_argument("--sort", choices=["score", "name", "date"], default="score",
                   help="Ordenar tabla por (default: score desc)")
    p.add_argument("--debug", action="store_true",
                   help="Mostrar errores HTTP detallados")
    p.add_argument("--diagram", action="store_true",
                   help="Imprime diagrama ASCII de stages por pipeline en consola")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP
# ═══════════════════════════════════════════════════════════════════════════════
def make_headers(pat: str) -> Dict:
    token = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Accept": "application/json"}


def api_get(url: str, headers: Dict, params: Dict = None, debug: bool = False) -> Optional[Any]:
    try:
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code >= 400:
            _api_label = url.split("/_apis")[0].split("/")[-1] if "/_apis" in url else url
            print(f"  ⚠  HTTP {r.status_code} ({_api_label})")
            if debug:
                print(f"[DEBUG] URL: {url}")
                print(f"[DEBUG] Body: {r.text[:400]}")
        r.raise_for_status()
        return r.json()
    except Exception as e:
        if debug:
            print(f"[DEBUG] {url}: {e}")
        return None


def vsrm_base(org_url: str) -> str:
    """Transforma dev.azure.com/{org} → vsrm.dev.azure.com/{org} para Release APIs."""
    return org_url.replace("dev.azure.com", "vsrm.dev.azure.com")


# ═══════════════════════════════════════════════════════════════════════════════
# API CALLS
# ═══════════════════════════════════════════════════════════════════════════════
def get_release_defs_list(org: str, project: str, headers: Dict, debug: bool) -> List[Dict]:
    base = vsrm_base(org)
    data = api_get(f"{base}/{quote(project, safe='')}/_apis/release/definitions",
                   headers, {"api-version": API_VERSION_DEFS, "$top": 500}, debug)
    return data.get("value", []) if data else []


def get_release_def_detail(
    org: str, project: str, def_id: int, headers: Dict, debug: bool
) -> Optional[Dict]:
    base = vsrm_base(org)
    return api_get(
        f"{base}/{quote(project, safe='')}/_apis/release/definitions/{def_id}",
        headers, {"api-version": API_VERSION_DEFS}, debug
    )


def get_latest_releases(
    org: str, project: str, def_id: int, top: int, headers: Dict, debug: bool
) -> List[Dict]:
    base = vsrm_base(org)
    data = api_get(
        f"{base}/{quote(project, safe='')}/_apis/release/releases",
        headers,
        {
            "definitionId": def_id,
            "$expand":      "environments",
            "$top":         top,
            "api-version":  API_VERSION_RELS,
        },
        debug,
    )
    return data.get("value", []) if data else []


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def extract_stages(definition: Dict) -> List[Dict]:
    """Retorna environments/stages ordenados por rank."""
    envs = definition.get("environments", [])
    return sorted(
        [{"name": e.get("name", ""), "rank": e.get("rank", 0)} for e in envs],
        key=lambda s: s["rank"],
    )


def stage_names(stages: List[Dict]) -> List[str]:
    return [s["name"] for s in stages]


def normalize_name(name: str) -> str:
    return name.lower().strip().replace(" ", "").replace("-", "").replace("_", "")


def normalize_stages_tuple(names: List[str]) -> Tuple[str, ...]:
    return tuple(normalize_name(n) for n in names)


def detect_prod_stage(stages: List[Dict]) -> Optional[str]:
    """
    Detecta el stage de producción:
    1. Primer stage cuyo nombre (normalizado) contiene algún keyword de PROD
    2. Fallback: último stage (mayor rank = etapa más avanzada del pipeline)
    """
    for s in stages:
        norm = normalize_name(s["name"])
        if any(kw in norm for kw in PROD_KEYWORDS):
            return s["name"]
    return stages[-1]["name"] if stages else None


# ═══════════════════════════════════════════════════════════════════════════════
# CONSISTENCY
# ═══════════════════════════════════════════════════════════════════════════════
def compute_majority_pattern(all_stage_lists: List[List[str]]) -> Tuple[str, ...]:
    """Patrón de stages más frecuente entre todos los pipelines."""
    if not all_stage_lists:
        return ()
    c = Counter(normalize_stages_tuple(s) for s in all_stage_lists if s)
    return c.most_common(1)[0][0] if c else ()


def compute_consistency(
    my_norm: Tuple[str, ...],
    majority: Tuple[str, ...],
    total_defs: int,
) -> Tuple[str, str]:
    if total_defs <= 1:
        return CONS_UNIQUE, "Pipeline único en el proyecto"
    if not majority:
        return CONS_UNIQUE, "Sin patrón de referencia"
    if my_norm == majority:
        return CONS_OK, f"Coincide con patrón mayoritario ({len(majority)} stages)"

    maj_set = set(majority)
    my_set  = set(my_norm)
    if not (my_set & maj_set):
        return CONS_DIFF, f"Sin coincidencia con patrón ({', '.join(list(majority)[:3])})"

    parts = []
    missing = maj_set - my_set
    extra   = my_set - maj_set
    if missing:
        parts.append(f"Falta: {', '.join(sorted(missing)[:3])}")
    if extra:
        parts.append(f"Extra: {', '.join(sorted(extra)[:3])}")
    if not parts:
        parts.append("Mismo contenido, orden diferente")

    return CONS_PARTIAL, " | ".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# RELEASE ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
def parse_azdo_date(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        clean = s.rstrip("Z").split(".")[0]
        return datetime.fromisoformat(clean).replace(tzinfo=timezone.utc)
    except Exception:
        return None


def analyze_releases(releases: List[Dict], prod_stage: Optional[str]) -> Dict[str, Any]:
    """
    Retorna:
      last_prod_deploy           — datetime del último deploy exitoso a prod (últimos 365 días)
      last_release_prod_attempts — # deploySteps en prod del ÚLTIMO release
      prod_ever_deployed         — bool
      last_release_id            — id del release más reciente
    """
    result: Dict[str, Any] = {
        "last_prod_deploy":           None,
        "last_release_prod_attempts": None,
        "prod_ever_deployed":         False,
        "last_release_id":            None,
    }

    if not releases or not prod_stage:
        return result

    prod_norm    = normalize_name(prod_stage)
    one_year_ago = datetime.now(timezone.utc) - timedelta(days=365)

    # El primer release en la lista es el más reciente
    latest_release = releases[0]
    result["last_release_id"] = latest_release.get("id")

    # Intentos en prod del ÚLTIMO release
    for env in latest_release.get("environments", []):
        if normalize_name(env.get("name", "")) == prod_norm:
            steps = env.get("deploySteps", [])
            result["last_release_prod_attempts"] = len(steps)
            break

    # Último deploy exitoso a prod (buscando hacia atrás en los N releases)
    succeeded_statuses = {"succeeded", "partiallysucceeded"}
    for rel in releases:
        for env in rel.get("environments", []):
            if normalize_name(env.get("name", "")) != prod_norm:
                continue
            if env.get("status", "").lower() not in succeeded_statuses:
                continue
            # Buscar el último deployStep exitoso
            for step in reversed(env.get("deploySteps", [])):
                if step.get("status", "").lower() in succeeded_statuses:
                    dt = parse_azdo_date(
                        step.get("lastModifiedOn") or step.get("requestedOn")
                    )
                    if dt and dt >= one_year_ago:
                        result["last_prod_deploy"]   = dt
                        result["prod_ever_deployed"] = True
                        return result
                    elif dt:
                        # Encontrado pero fuera de ventana de 1 año
                        result["prod_ever_deployed"] = True
                        return result

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SCORE
# ═══════════════════════════════════════════════════════════════════════════════
def compute_score(
    last_prod_deploy: Optional[datetime],
    prod_attempts: Optional[int],
) -> Dict[str, Any]:
    """
    Score 0-100:
      Recencia   (0-70): max(0, 70 * (1 - days/365))   → más reciente = mayor
      Estabilidad (0-30): 1=30, 2=20, 3=10, 4+=0       → menos intentos = mayor
    Si nunca desplegado: 0.
    """
    if last_prod_deploy is None:
        return {"total": 0, "recency": 0, "stability": 0, "days": None}

    days      = max(0, (datetime.now(timezone.utc) - last_prod_deploy).days)
    recency   = max(0.0, 70.0 * (1.0 - days / 365.0))
    attempts  = max(1, prod_attempts or 1)
    # Cada intento adicional resta ~33% de los 30 puntos de estabilidad
    stability = max(0.0, 30.0 - (attempts - 1) * 10.0)
    total     = min(100, round(recency + stability))

    return {
        "total":     total,
        "recency":   round(recency),
        "stability": round(stability),
        "days":      days,
    }


def get_rating(score: int, ever_deployed: bool) -> Tuple[str, str]:
    if not ever_deployed or score == 0:
        return "🔴", "Nunca"
    if score >= 90:
        return "🟢", "Excelente"
    if score >= 70:
        return "🔵", "Bueno"
    if score >= 40:
        return "🟡", "Regular"
    return "🟠", "Bajo"


# ═══════════════════════════════════════════════════════════════════════════════
# SCORE BAR (Rich)
# ═══════════════════════════════════════════════════════════════════════════════
def score_bar_rich(score: int, width: int = 14) -> str:
    filled = round(score / 100 * width)
    bar    = "█" * filled + "░" * (width - filled)
    if score >= 70:
        color = "green"
    elif score >= 40:
        color = "yellow"
    elif score > 0:
        color = "red"
    else:
        color = "dim"
    return f"[{color}]{bar}[/{color}] [bold]{score:>3}[/bold]"


# ═══════════════════════════════════════════════════════════════════════════════
# RICH CELLS
# ═══════════════════════════════════════════════════════════════════════════════
def cons_cell_rich(label: str) -> str:
    return {
        CONS_OK:      "[bold green]✅ OK[/bold green]",
        CONS_PARTIAL: "[bold yellow]🟡 Parcial[/bold yellow]",
        CONS_DIFF:    "[bold red]🔴 Diferente[/bold red]",
        CONS_UNIQUE:  "[dim]❓ Único[/dim]",
    }.get(label, label)


def attempts_cell_rich(attempts: Optional[int]) -> str:
    if attempts is None:
        return "[dim]—[/dim]"
    if attempts == 0:
        return "[dim]0[/dim]"
    if attempts == 1:
        return "[bold green]1 ✅[/bold green]"
    if attempts == 2:
        return "[yellow]2 🟡[/yellow]"
    return f"[bold red]{attempts} 🔴[/bold red]"


def fmt_date(dt: Optional[datetime], tz_name: str) -> str:
    if not dt:
        return "[dim]Nunca[/dim]"
    return dt.astimezone(ZoneInfo(tz_name)).strftime("%Y-%m-%d %H:%M")


def fmt_stages_rich(names: List[str]) -> str:
    if not names:
        return "[dim]Sin stages[/dim]"
    return " [dim]→[/dim] ".join(names)


# ═══════════════════════════════════════════════════════════════════════════════
# PER-PIPELINE PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════
def process_pipeline(
    summary: Dict,
    org: str, project: str,
    headers: Dict,
    top: int, debug: bool,
) -> Optional[Dict]:
    def_id = summary["id"]
    name   = summary.get("name", f"pipeline-{def_id}")

    detail = get_release_def_detail(org, project, def_id, headers, debug)
    if not detail:
        return None

    stages     = extract_stages(detail)
    s_names    = stage_names(stages)
    s_norm     = normalize_stages_tuple(s_names)
    prod_stage = detect_prod_stage(stages)

    releases = get_latest_releases(org, project, def_id, top, headers, debug)
    rel_info = analyze_releases(releases, prod_stage)

    score_data = compute_score(
        rel_info["last_prod_deploy"],
        rel_info["last_release_prod_attempts"],
    )
    r_emoji, r_label = get_rating(score_data["total"], rel_info["prod_ever_deployed"])

    return {
        "id":               def_id,
        "name":             name,
        "stages":           s_names,
        "stages_norm":      s_norm,
        "prod_stage":       prod_stage,
        "last_prod_dt":     rel_info["last_prod_deploy"],
        "prod_attempts":    rel_info["last_release_prod_attempts"],
        "ever_deployed":    rel_info["prod_ever_deployed"],
        "last_release_id":  rel_info["last_release_id"],
        "score":            score_data["total"],
        "score_recency":    score_data["recency"],
        "score_stability":  score_data["stability"],
        "days_since":       score_data["days"],
        "rating_emoji":     r_emoji,
        "rating_label":     r_label,
        "consistency":      CONS_UNIQUE,   # se completa en el paso global
        "cons_detail":      "",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RICH TABLE & SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
def print_rich_table(console: "Console", rows: List[Dict], tz_name: str):
    tbl = Table(
        title="🚀 Release Pipeline CD — Salud & Consistencia de Stages",
        title_style="bold cyan",
        header_style="bold white",
        border_style="dim",
        box=box.ROUNDED,
        show_lines=False,
    )
    tbl.add_column("#",             style="dim",        width=4,  justify="right")
    tbl.add_column("Def ID",        style="cyan",       width=8,  justify="right")
    tbl.add_column("Pipeline CD",   style="bold white", min_width=26)
    tbl.add_column("Stages",        min_width=28, max_width=52)
    tbl.add_column("Consistencia",  justify="center",   width=14)
    tbl.add_column("Stage PROD",    justify="center",   width=14)
    tbl.add_column("Último PROD",   justify="center",   width=18)
    tbl.add_column("Intentos",      justify="center",   width=9)
    tbl.add_column("Score",         justify="left",     width=24)
    tbl.add_column("Rating",        justify="center",   width=13)

    for idx, row in enumerate(rows, 1):
        prod_cell = (
            f"[cyan]{row['prod_stage']}[/cyan]"
            if row["prod_stage"]
            else "[dim]—[/dim]"
        )
        tbl.add_row(
            str(idx),
            str(row["id"]),
            row["name"],
            fmt_stages_rich(row["stages"]),
            cons_cell_rich(row["consistency"]),
            prod_cell,
            fmt_date(row["last_prod_dt"], tz_name),
            attempts_cell_rich(row["prod_attempts"]),
            score_bar_rich(row["score"]),
            f"{row['rating_emoji']} {row['rating_label']}",
        )

    console.print(tbl)
    console.print()


def print_rich_summary(console: "Console", rows: List[Dict], majority: Tuple, elapsed: float):
    total     = len(rows)
    avg_score = round(sum(r["score"] for r in rows) / total) if total else 0
    excellent = sum(1 for r in rows if r["score"] >= 90)
    good      = sum(1 for r in rows if 70 <= r["score"] < 90)
    regular   = sum(1 for r in rows if 40 <= r["score"] < 70)
    low       = sum(1 for r in rows if 0 < r["score"] < 40)
    never     = sum(1 for r in rows if not r["ever_deployed"])
    cons_ok   = sum(1 for r in rows if r["consistency"] == CONS_OK)
    cons_part = sum(1 for r in rows if r["consistency"] == CONS_PARTIAL)
    cons_diff = sum(1 for r in rows if r["consistency"] == CONS_DIFF)

    majority_str = " → ".join(majority) if majority else "N/A"

    console.print(Panel(
        f"[bold white]📋 Pipelines analizados:[/]  [cyan]{total}[/]  "
        f"[bold]· Score promedio: [cyan]{avg_score}/100[/cyan][/]\n\n"
        f"[bold]📊 Distribución de Score:[/]\n"
        f"  [green]🟢 Excelente (90-100):[/] {excellent}\n"
        f"  [blue]🔵 Bueno     (70-89): [/] {good}\n"
        f"  [yellow]🟡 Regular  (40-69): [/] {regular}\n"
        f"  [red]🟠 Bajo       (1-39): [/] {low}\n"
        f"  [dim]🔴 Nunca       (  0): [/] {never}\n\n"
        f"[bold]🔗 Consistencia de Stages:[/]\n"
        f"  [green]✅ OK:         [/] {cons_ok}\n"
        f"  [yellow]🟡 Parcial:   [/] {cons_part}\n"
        f"  [red]🔴 Diferente:  [/] {cons_diff}\n"
        f"[dim]  Patrón mayoritario: {majority_str}[/]\n\n"
        f"[bold]📐 Fórmula Score:[/]\n"
        f"[dim]  Recencia (0-70): 70 × (1 - días/365)[/]\n"
        f"[dim]  Estabilidad (0-30): 30 - (intentos-1) × 10[/]\n\n"
        f"[dim]⏱️  Tiempo total: {elapsed:.2f}s[/]",
        title="📊 Resumen de Salud — Release Pipelines",
        border_style="blue",
        expand=False,
    ))
    console.print()


# ═══════════════════════════════════════════════════════════════════════════════
# CONSOLE PIPELINE DIAGRAM
# ═══════════════════════════════════════════════════════════════════════════════
def _build_ascii_pipeline(stages: List[str]) -> str:
    """Construye 3 líneas ASCII del diagrama horizontal: ┌───┐────▶┌───┐"""
    if not stages:
        return "  (sin stages)"
    PAD    = 1
    widths = [max(len(name) + PAD * 2, 6) for name in stages]
    ARROW  = "────▶"
    GAP    = " " * len(ARROW)
    tops   = ["┌" + "─" * w + "┐" for w in widths]
    mids   = ["│" + n.center(w) + "│" for n, w in zip(stages, widths)]
    bots   = ["└" + "─" * w + "┘" for w in widths]
    return (
        GAP.join(tops) + "\n" +
        ARROW.join(mids) + "\n" +
        GAP.join(bots)
    )


def print_pipeline_diagrams(console: "Console", rows: List[Dict], tz_name: str):
    """Imprime diagrama ASCII por pipeline con stage, score y estado de PROD."""
    CONS_ICONS = {
        CONS_OK:      "[green]✅ OK[/green]",
        CONS_PARTIAL: "[yellow]🟡 Parcial[/yellow]",
        CONS_DIFF:    "[red]🔴 Diferente[/red]",
        CONS_UNIQUE:  "[dim]❓ Único[/dim]",
    }

    console.print()
    console.print(Panel(
        "[bold cyan]📊 Diagramas de Release Pipelines[/]  "
        "[dim](verde=PROD desplegado · rojo=PROD sin deploy · azul=stage normal)[/]",
        border_style="cyan", expand=False,
    ))

    for row in rows:
        stages    = row["stages"]
        prod      = row.get("prod_stage") or ""
        dt_str    = fmt_date(row["last_prod_dt"], tz_name)
        att       = row["prod_attempts"]
        cons_icon = CONS_ICONS.get(row["consistency"], row["consistency"])

        console.print()
        console.rule(
            f"[bold white]{row['name']}[/]  "
            f"{score_bar_rich(row['score'])}  "
            f"{row['rating_emoji']} [bold]{row['rating_label']}[/]",
            style="dim",
        )

        if not stages:
            console.print("  [dim](sin stages definidos)[/dim]")
            continue

        for line in _build_ascii_pipeline(stages).splitlines():
            console.print(f"  [dim]{line}[/dim]")

        prod_icon = "[bold green]✅[/]" if row["ever_deployed"] else "[bold red]⛔[/]"
        att_txt   = f" · {att} intento(s) en último release" if att is not None else ""
        console.print(
            f"  {prod_icon} [bold]Stage PROD:[/] [cyan]{prod or '—'}[/]  "
            f"[dim]Último deploy: {dt_str}{att_txt}[/]"
        )
        console.print(f"  Consistencia: {cons_icon}  [dim]{row['cons_detail']}[/dim]")

    console.print()


# ═══════════════════════════════════════════════════════════════════════════════
# MATPLOTLIB PIPELINE IMAGE (para Excel)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_pipeline_image(row: Dict, tz_name: str) -> Optional[bytes]:
    """
    Genera un PNG del diagrama de stages del pipeline usando matplotlib.
    Retorna bytes PNG o None si matplotlib no está disponible.
    """
    if not MATPLOTLIB_AVAILABLE:
        return None

    stages = row["stages"]
    prod   = row.get("prod_stage") or ""
    n      = len(stages)
    if n == 0:
        return None

    fig_w = max(n * 2.4 + 0.8, 5.0)
    fig_h = 2.0
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    BOX_W = 1.9
    BOX_H = 0.75
    gap   = (fig_w - n * BOX_W) / (n + 1)
    y_mid = 1.25

    for i, name in enumerate(stages):
        x_left = gap * (i + 1) + BOX_W * i
        x_ctr  = x_left + BOX_W / 2

        if name == prod:
            bg = "#27ae60" if row["ever_deployed"] else "#c0392b"
        else:
            bg = "#2980b9"

        rect = FancyBboxPatch(
            (x_left, y_mid - BOX_H / 2), BOX_W, BOX_H,
            boxstyle="round,pad=0.05",
            facecolor=bg, edgecolor="white", linewidth=1.0, zorder=2,
        )
        ax.add_patch(rect)
        ax.text(x_ctr, y_mid, name, ha="center", va="center",
                fontsize=8.5, color="white", fontweight="bold", zorder=3)

        if i < n - 1:
            x_end   = x_left + BOX_W
            x_start = x_end + gap
            ax.annotate(
                "", xy=(x_start, y_mid), xytext=(x_end, y_mid),
                arrowprops=dict(arrowstyle="-|>", color="#444", lw=1.0),
                zorder=1,
            )

    score = row["score"]
    clr   = "#27ae60" if score >= 70 else ("#f39c12" if score >= 40 else "#c0392b")
    prod_info = fmt_date(row["last_prod_dt"], tz_name) if row["last_prod_dt"] else "PROD: Nunca"

    ax.text(fig_w / 2, 0.55,
            f"Score: {score}/100  {row['rating_emoji']} {row['rating_label']}",
            ha="center", va="center", fontsize=8, color=clr, fontweight="bold")
    ax.text(fig_w / 2, 0.20, prod_info,
            ha="center", va="center", fontsize=7, color="#555")

    fig.suptitle(row["name"], fontsize=9, fontweight="bold", color="#222", y=0.97)
    plt.tight_layout(pad=0.3)

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def export_results(
    rows: List[Dict], fmt: str, script_dir: str, tz_name: str
) -> Optional[str]:
    outcome_dir = os.path.join(script_dir, "outcome")
    os.makedirs(outcome_dir, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    flat = [{
        "id":                            r["id"],
        "pipeline_name":                 r["name"],
        "stages":                        " → ".join(r["stages"]),
        "stage_count":                   len(r["stages"]),
        "prod_stage_detected":           r["prod_stage"] or "",
        "consistency":                   r["consistency"],
        "consistency_detail":            r["cons_detail"],
        "ever_deployed_to_prod":         r["ever_deployed"],
        "last_prod_deploy":              r["last_prod_dt"].isoformat() if r["last_prod_dt"] else "",
        "days_since_last_prod":          r["days_since"] if r["days_since"] is not None else "",
        "last_release_id":               r["last_release_id"] or "",
        "prod_attempts_in_last_release": r["prod_attempts"] if r["prod_attempts"] is not None else "",
        "score_total":                   r["score"],
        "score_recency":                 r["score_recency"],
        "score_stability":               r["score_stability"],
        "rating":                        r["rating_label"],
    } for r in rows]

    if fmt == "json":
        filepath = os.path.join(outcome_dir, f"release_cd_health_{ts}.json")
        payload = {
            "metadata": {
                "tool":         "azdo_release_cd_health",
                "version":      __version__,
                "generated_at": datetime.now(ZoneInfo(tz_name)).isoformat(),
            },
            "total":     len(rows),
            "avg_score": round(sum(r["score"] for r in rows) / len(rows)) if rows else 0,
            "data":      flat,
        }
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return filepath

    elif fmt == "csv":
        if not flat:
            return None
        filepath = os.path.join(outcome_dir, f"release_cd_health_{ts}.csv")
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(flat[0].keys()))
            w.writeheader()
            w.writerows(flat)
        return filepath

    elif fmt == "excel":
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            import tempfile

            filepath = os.path.join(outcome_dir, f"release_cd_health_{ts}.xlsx")
            df = pd.DataFrame(flat)
            df.to_excel(filepath, index=False, engine="openpyxl",
                        sheet_name="Health Report")

            # ── Hoja de diagramas matplotlib ─────────────────────────────
            if MATPLOTLIB_AVAILABLE:
                wb      = load_workbook(filepath)
                ws_diag = wb.create_sheet("Pipeline Diagrams")

                # Instrucciones en celda A1
                ws_diag["B1"] = "Release Pipeline CD — Diagramas de Stages"
                ws_diag["B1"].font = Font(bold=True, size=13, color="1F4E79")
                ws_diag["B2"] = (
                    "Verde = PROD desplegado  |  Rojo = PROD sin deploy  |  "
                    "Azul = stage normal  |  Score: Recencia(70) + Estabilidad(30)"
                )
                ws_diag["B2"].font = Font(italic=True, size=9, color="555555")
                ws_diag.column_dimensions["A"].width = 2
                ws_diag.column_dimensions["B"].width = 6

                row_pos = 4  # fila Excel donde empieza el primer diagrama
                tmp_files: List[str] = []

                for r in rows:
                    img_bytes = generate_pipeline_image(r, tz_name)
                    if not img_bytes:
                        continue

                    # Nombre del pipeline sobre la imagen
                    name_cell = ws_diag.cell(row=row_pos, column=2, value=r["name"])
                    name_cell.font      = Font(bold=True, size=10, color="1F4E79")
                    name_cell.alignment = Alignment(vertical="center")

                    # Guardar imagen en archivo temporal e insertar
                    with tempfile.NamedTemporaryFile(
                        suffix=".png", delete=False
                    ) as tmp:
                        tmp.write(img_bytes)
                        tmp_path = tmp.name
                    tmp_files.append(tmp_path)

                    xl_img        = XLImage(tmp_path)
                    xl_img.anchor = f"B{row_pos + 1}"
                    ws_diag.add_image(xl_img)

                    row_pos += 16  # espacio entre diagramas

                wb.save(filepath)

                # Limpiar temporales
                for tmp_path in tmp_files:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass

            return filepath
        except ImportError:
            print("ERROR: pip install pandas openpyxl matplotlib")
            return None

    return None


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    start_time = time.time()

    if not REQUESTS_AVAILABLE:
        print("ERROR: pip install requests rich pandas openpyxl")
        return

    args    = get_args()
    console = Console() if RICH_AVAILABLE else None
    headers = make_headers(args.pat)

    tz_name = args.timezone
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = DEFAULT_TIMEZONE

    rev_time = datetime.now(ZoneInfo(tz_name)).strftime(f"%Y-%m-%d %H:%M:%S ({tz_name})")

    if RICH_AVAILABLE and console:
        console.print()
        console.print(Panel(
            f"[bold cyan]🚀 Release CD Health Analyzer[/]\n"
            f"[dim]🕐 {rev_time}[/]\n"
            f"[dim]🏢 Org:      {args.org}[/]\n"
            f"[dim]📁 Proyecto: {args.project}[/]\n"
            f"[dim]📐 Score:    Recencia (70pt) + Estabilidad (30pt) = 100pt[/]\n"
            f"[dim]📅 Ventana:  últimos 365 días · últimos {args.top} releases/pipeline[/]",
            border_style="cyan", expand=False,
        ))
        console.print()

    # ── 1. Lista de release definitions ─────────────────────────────────────
    if RICH_AVAILABLE and console:
        with Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console) as prog:
            t = prog.add_task("Obteniendo release definitions...", total=None)
            summaries = get_release_defs_list(args.org, args.project, headers, args.debug)
            prog.update(t, description=f"✅ {len(summaries)} release pipelines encontrados")
    else:
        summaries = get_release_defs_list(args.org, args.project, headers, args.debug)
        print(f"{len(summaries)} release pipelines encontrados")

    if not summaries:
        msg = "❌ Sin release definitions. Verifica org, proyecto, PAT y permisos Release (Read)."
        (console.print(f"[red]{msg}[/]") if console else print(msg))
        return

    if args.filter:
        summaries = [s for s in summaries if args.filter.lower() in s.get("name", "").lower()]
        if console:
            console.print(f"[dim]🔍 Filtrado: {len(summaries)} pipelines que contienen '{args.filter}'[/]")

    if not summaries:
        (console.print("[yellow]⚠️ Sin resultados tras el filtro.[/]") if console else print("Sin resultados."))
        return

    console.print() if console else None

    # ── 2. Procesar cada pipeline en paralelo ────────────────────────────────
    rows: List[Dict] = []

    if RICH_AVAILABLE and console:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as prog:
            task = prog.add_task(
                f"Analizando {len(summaries)} pipelines (stages + releases)...",
                total=len(summaries),
            )
            with ThreadPoolExecutor(max_workers=args.threads) as exe:
                futures = {
                    exe.submit(
                        process_pipeline,
                        s, args.org, args.project, headers, args.top, args.debug,
                    ): s.get("name", s["id"])
                    for s in summaries
                }
                for fut in as_completed(futures):
                    try:
                        result = fut.result()
                        if result:
                            rows.append(result)
                    except Exception:
                        pass
                    prog.advance(task)
    else:
        for s in summaries:
            r = process_pipeline(s, args.org, args.project, headers, args.top, args.debug)
            if r:
                rows.append(r)

    if not rows:
        (console.print("[yellow]⚠️ No se pudieron procesar pipelines.[/]") if console else print("Sin datos."))
        return

    # ── 3. Consistencia global (post-proceso sobre todos los pipelines) ──────
    majority = compute_majority_pattern([r["stages"] for r in rows])
    total_defs = len(rows)
    for row in rows:
        label, detail = compute_consistency(row["stages_norm"], majority, total_defs)
        row["consistency"] = label
        row["cons_detail"] = detail

    # ── 4. Ordenar ────────────────────────────────────────────────────────────
    if args.sort == "score":
        rows.sort(key=lambda r: r["score"], reverse=True)
    elif args.sort == "name":
        rows.sort(key=lambda r: r["name"].lower())
    elif args.sort == "date":
        rows.sort(
            key=lambda r: r["last_prod_dt"] or datetime.min.replace(tzinfo=timezone.utc),
            reverse=True,
        )

    # ── 5. Tabla + resumen ───────────────────────────────────────────────────
    elapsed = time.time() - start_time

    if RICH_AVAILABLE and console:
        print_rich_table(console, rows, tz_name)
        print_rich_summary(console, rows, majority, elapsed)
        if args.diagram:
            print_pipeline_diagrams(console, rows, tz_name)
    else:
        hdr = f"{'#':>4}  {'ID':>7}  {'Pipeline':<35} {'Stages':<32} {'Cons':^10} {'Último PROD':^20} {'Int':^5} {'Score':>6}  Rating"
        print(f"\n{'='*len(hdr)}\n{hdr}\n{'='*len(hdr)}")
        for idx, row in enumerate(rows, 1):
            stg = " → ".join(row["stages"])[:30]
            dt  = row["last_prod_dt"].strftime("%Y-%m-%d") if row["last_prod_dt"] else "Nunca"
            att = str(row["prod_attempts"]) if row["prod_attempts"] is not None else "—"
            print(f"{idx:>4}  {row['id']:>7}  {row['name']:<35} {stg:<32} {row['consistency']:^10} "
                  f"{dt:^20} {att:^5} {row['score']:>6}  {row['rating_emoji']} {row['rating_label']}")
        avg = round(sum(r["score"] for r in rows) / len(rows)) if rows else 0
        print(f"\nTotal: {total_defs} | Score promedio: {avg}/100 | Tiempo: {elapsed:.2f}s\n")

    # ── 6. Exportar ───────────────────────────────────────────────────────────
    if args.output:
        fp = export_results(
            rows, args.output,
            os.path.dirname(os.path.abspath(__file__)),
            tz_name,
        )
        if fp:
            msg = f"📁 Exportado: {fp}"
            (console.print(f"[bold green]{msg}[/]\n") if console else print(msg))


if __name__ == "__main__":
    main()

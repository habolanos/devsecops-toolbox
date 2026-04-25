#!/usr/bin/env python3
"""
GKE Pipelines Inventory v1.0.0
Inventario de Release Definitions (pipelines CD) que contienen "GKE" en su nombre,
con detalle de stages y último estado de ejecución.

Uso:
    python gke_pipelines_inventory.py --org Coppel-Retail --project "Compras.RMI"
    python gke_pipelines_inventory.py --org Coppel-Retail --project "Compras.RMI" --keyword "PROD"

Autor: Harold Adrian (migrado desde Comercial/scripts/gkepipe.py)
"""

import requests
import base64
import sys
import json
import os
import argparse
from datetime import datetime
from pathlib import Path
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = lambda *a, **k: None  # noqa: E731

# --- Directorio de salida centralizado (DEVSECOPS_OUTPUT_DIR) ---
try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
# -------------------------------------------------------------------

load_dotenv(Path(__file__).parent.parent / ".env")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

# ── Configuración por defecto ─────────────────────────────────────────
DEFAULT_ORG = "Coppel-Retail"
DEFAULT_PROJECT = "Compras.RMI"
DEFAULT_KEYWORD = "GKE"
API_VERSION = "7.1"
EXPECTED_STAGES = ["Develop", "QA", "Validador", "Production"]


def get_headers(pat: str):
    """Genera headers de autenticación."""
    token_b64 = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token_b64}", "Content-Type": "application/json"}


def resolve_base(org: str, project: str, endpoint: str):
    """Release endpoints van contra vsrm.dev.azure.com."""
    if endpoint.startswith("release/"):
        return f"https://vsrm.dev.azure.com/{org}/{project}/_apis"
    return f"https://dev.azure.com/{org}/{project}/_apis"


def api_get(endpoint: str, org: str, project: str, headers: dict, params=None):
    """GET a Azure DevOps API."""
    url = f"{resolve_base(org, project, endpoint)}/{endpoint}"
    p = {"api-version": API_VERSION}
    if params:
        p.update(params)
    r = requests.get(url, headers=headers, params=p, timeout=30)
    r.raise_for_status()
    return r.json()


def fetch_definitions(org: str, project: str, headers: dict, keyword: str):
    """Obtiene release definitions que contienen el keyword."""
    data = api_get("release/definitions", org, project, headers)
    defs = data.get("value", [])
    return [d for d in defs if keyword.lower() in d.get("name", "").lower()]


def fetch_last_release(definition_id: int, org: str, project: str, headers: dict):
    """Obtiene el último release de una definition."""
    try:
        data = api_get(
            "release/releases",
            org, project, headers,
            {"definitionId": definition_id, "$top": 1, "$expand": "environments"}
        )
        releases = data.get("value", [])
        return releases[0] if releases else None
    except Exception as e:
        print(f"   ⚠️ Error obteniendo último release: {e}")
        return None


def build_stage_status(last_rel: dict):
    """Construye diccionario de estado por stage."""
    if not last_rel:
        return {s: "Sin ejecuciones" for s in EXPECTED_STAGES}
    
    envs = last_rel.get("environments", [])
    status_by_stage = {}
    
    for env in envs:
        name = env.get("name", "")
        status = env.get("status", "unknown")
        status_by_stage[name] = status
    
    return {s: status_by_stage.get(s, "N/A") for s in EXPECTED_STAGES}


def generate_excel(data: list, org: str, project: str, keyword: str, output_file: str):
    """Genera Excel con 3 hojas: Resumen, Detalle Stages, Matriz."""
    wb = Workbook()
    
    # Hoja 1: Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    headers1 = ["Pipeline", "ID", "Último Release", "Fecha", "Estado General", "URL"]
    ws1.append(headers1)
    
    for item in data:
        last_rel = item.get("last_release")
        ws1.append([
            item["name"],
            item["id"],
            last_rel.get("name") if last_rel else "N/A",
            last_rel.get("createdOn", "")[:10] if last_rel else "N/A",
            item["overall_status"],
            item["url"]
        ])
    
    # Hoja 2: Detalle de Stages
    ws2 = wb.create_sheet("Detalle Stages")
    headers2 = ["Pipeline", "Release"] + EXPECTED_STAGES
    ws2.append(headers2)
    
    for item in data:
        last_rel = item.get("last_release")
        row = [
            item["name"],
            last_rel.get("name") if last_rel else "N/A"
        ]
        for stage in EXPECTED_STAGES:
            row.append(item["stages"].get(stage, "N/A"))
        ws2.append(row)
    
    # Hoja 3: Matriz de Cambios
    ws3 = wb.create_sheet("Matriz")
    headers3 = ["Pipeline", "Variable", "Valor Default", "Valor Stage Prod", "Diferente"]
    ws3.append(headers3)
    
    for item in data:
        variables = item.get("variables", {})
        for var_name, var_data in variables.items():
            default_val = var_data.get("value", "")
            prod_val = var_data.get("scope", {}).get("Production", default_val)
            different = "Sí" if default_val != prod_val else "No"
            ws3.append([item["name"], var_name, default_val, prod_val, different])
    
    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata
    wb.properties.title = "Inventario Pipelines CD-GKE"
    wb.properties.subject = f"{org}/{project}"
    wb.properties.creator = "gke_pipelines_inventory.py"
    
    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="GKE Pipelines Inventory - Inventario de pipelines CD con keyword GKE",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python gke_pipelines_inventory.py
  python gke_pipelines_inventory.py --org Coppel-Retail --project "Compras.RMI"
  python gke_pipelines_inventory.py --org Coppel-Retail --project "Compras.RMI" --keyword "PROD"
        """
    )
    parser.add_argument("--org", default=DEFAULT_ORG,
                       help=f"Organización Azure DevOps (default: {DEFAULT_ORG})")
    parser.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                       help=f"Proyecto (default: {DEFAULT_PROJECT})")
    parser.add_argument("--keyword", "-k", default=DEFAULT_KEYWORD,
                       help=f"Keyword para filtrar pipelines (default: {DEFAULT_KEYWORD})")
    parser.add_argument("--pat", default=None,
                       help="PAT de Azure DevOps (default: env AZURE_PAT)")
    parser.add_argument("--output", "-o", default=None,
                       help="Nombre del archivo Excel de salida")
    
    args = parser.parse_args()
    
    # Obtener PAT
    pat = args.pat or os.getenv("AZURE_PAT", "")
    if not pat:
        print("❌ ERROR: Configura tu PAT via --pat o variable de entorno AZURE_PAT")
        print("   Permisos requeridos: Release → Read")
        sys.exit(1)
    
    headers = get_headers(pat)
    
    print(f"🔍 Buscando pipelines CD con keyword '{args.keyword}'...")
    print(f"   Org: {args.org}")
    print(f"   Project: {args.project}")
    
    definitions = fetch_definitions(args.org, args.project, headers, args.keyword)
    print(f"✅ {len(definitions)} pipelines encontrados")
    
    if not definitions:
        print("No se encontraron pipelines. Saliendo.")
        sys.exit(0)
    
    data = []
    for i, d in enumerate(definitions, 1):
        def_id = d.get("id")
        def_name = d.get("name")
        print(f"  [{i}/{len(definitions)}] Analizando: {def_name}...")
        
        last_rel = fetch_last_release(def_id, args.org, args.project, headers)
        stages = build_stage_status(last_rel)
        
        # Estado general
        overall = "Sin ejecuciones"
        if last_rel:
            envs = last_rel.get("environments", [])
            statuses = [e.get("status") for e in envs]
            if any(s == "rejected" for s in statuses):
                overall = "Rechazado"
            elif any(s == "inProgress" for s in statuses):
                overall = "En progreso"
            elif all(s == "succeeded" for s in statuses):
                overall = "Exitoso"
            else:
                overall = "Pendiente/Parcial"
        
        data.append({
            "id": def_id,
            "name": def_name,
            "url": f"https://dev.azure.com/{args.org}/{args.project}/_release?_a=definitions&definitionId={def_id}",
            "last_release": last_rel,
            "stages": stages,
            "overall_status": overall,
            "variables": d.get("variables", {})
        })
    
    # Generar Excel
    output_file = args.output or f"gke_cd_pipelines_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    generate_excel(data, args.org, args.project, args.keyword, output_file)
    
    print(f"\n✅ Reporte generado: {output_file}")
    print(f"   Total pipelines: {len(data)}")


if __name__ == "__main__":
    main()

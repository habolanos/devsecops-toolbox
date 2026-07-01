#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CD Prod Deploy Tracker v1.0.0
Rastrea el último despliegue exitoso a Producción por pipeline CD.
Lee cache CD existente y consulta APIs de releases para obtener fechas, artefactos y vigencia.

Uso:
    python cicd_inventory_prod_deploy.py --org Coppel-Retail --project "Compras.RMI" --deadline 2026-03-01
    python cicd_inventory_prod_deploy.py --org Coppel-Retail --project "Compras.RMI" --deadline 2026-03-01 --force-refresh
    python cicd_inventory_prod_deploy.py --org Coppel-Retail --project "Compras.RMI" --deadline 2026-03-01 --workers 20

Cache-first: verifica cache propio < 24h para skip APIs. Requiere cache CD previo.
Genera Excel + CSV + JSON cache.

Autor: Harold Adrian Bolanos Rodriguez
"""

import os
import sys
import time
import json
import glob
import requests
import pandas as pd
import argparse
from datetime import datetime, timezone, date
from base64 import b64encode
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = lambda *a, **k: None

try:
    from rich.progress import Progress, SpinnerColumn, BarColumn, TaskProgressColumn, TextColumn, TimeElapsedColumn
    from rich.console import Console
    from rich.table import Table
    RICH_AVAILABLE = True
try:
    from export_manager import ExportManager
    EXPORT_MANAGER_AVAILABLE = True
except ImportError:
    EXPORT_MANAGER_AVAILABLE = False


except ImportError:
    RICH_AVAILABLE = False

try:
    from utils import get_output_dir, resolve_output_path
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
    from datetime import datetime as _dt
    _FMT_EXT = {"excel": ".xlsx", "csv": ".csv", "json": ".json"}
    def resolve_output_path(output_arg, base_name, default_format="excel"):
        output_dir = get_output_dir("outcome")
        output_dir.mkdir(parents=True, exist_ok=True)
        ext = _FMT_EXT.get(default_format, ".xlsx")
        if not output_arg:
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        if output_arg.lower() in _FMT_EXT:
            ext = _FMT_EXT[output_arg.lower()]
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        p = _Path(output_arg)
        if p.suffix == "":
            p = p.with_suffix(ext)
        return str(p.resolve())

load_dotenv(Path(__file__).parent.parent / ".env")

SCRIPT_NAME = "cicd_inventory_prod_deploy"
CD_CACHE_SCRIPT = "cicd_inventory_cd_detailed"
DEFAULT_ORG = "Coppel-Retail"
DEFAULT_PROJECT = "Compras.RMI"
API_VERSION = "7.1"
DEFAULT_WORKERS = 30
CACHE_TTL_HOURS = 24

PROD_KEYWORDS = ["producción", "produccion", "production", "prod", "prd", "produc"]

OBSOLETE_KEYWORDS = ["obsoleto", "obsolete", "_old", "legacy-", "deprecated", "deprecated_"]


# ==========================================================
# UTILIDADES COMUNES
# ==========================================================

class TeeWriter:
    def __init__(self, log_path):
        self.terminal = sys.__stdout__
        self.log = open(log_path, "w", encoding="utf-8")
        self.log_path = log_path
        self._paused = False

    def write(self, message):
        self.log.write(message)
        if not self._paused:
            self.terminal.write(message)

    def flush(self):
        self.log.flush()
        if not self._paused:
            self.terminal.flush()

    def close(self):
        self.log.close()

    def pause_terminal(self):
        self._paused = True

    def resume_terminal(self):
        self._paused = False


def setup_logging():
    output_dir = get_output_dir("outcome")
    output_dir.mkdir(parents=True, exist_ok=True)
    log_path = output_dir / f"{SCRIPT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeWriter(log_path)
    sys.stdout = tee
    print(f"📝 Log: {log_path.resolve()}")
    print(f"📅 Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    return tee


def teardown_logging(tee):
    print(f"\n📝 Log guardado: {tee.log_path.resolve()}")
    sys.stdout = tee.terminal
    tee.close()


def _progress_context():
    if RICH_AVAILABLE:
        console = Console(file=sys.__stdout__)
        return Progress(
            SpinnerColumn(),
            TextColumn("[bold blue]{task.description}"),
            BarColumn(bar_width=40),
            TaskProgressColumn(),
            TextColumn("({task.completed}/{task.total})"),
            TimeElapsedColumn(),
            console=console,
        )
    return None


def get_headers(pat: str):
    auth = b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {auth}", "Content-Type": "application/json"}


def az_get(url, headers, params=None, max_retries=5):
    params = params or {}
    params["api-version"] = API_VERSION
    for attempt in range(max_retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=30)
            if r.status_code >= 500:
                wait = 2 ** attempt
                print(f"⚠️  {r.status_code} en {url[:60]}... retry {attempt+1}/{max_retries} (espera {wait}s)")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except requests.exceptions.HTTPError:
            raise
        except requests.exceptions.RequestException as e:
            if attempt == max_retries - 1:
                raise
            wait = 2 ** attempt
            print(f"⚠️  Error en {url[:60]}... retry {attempt+1}/{max_retries}: {e}")
            time.sleep(wait)
    return {}


def normalize_org(org: str) -> str:
    if org.startswith("http"):
        return org.rstrip("/").split("/")[-1]
    return org


def safe_az_get(url, headers, params=None):
    try:
        return az_get(url, headers, params)
    except Exception as e:
        print(f"   ❌ Error: {e}")
        return {}


# ==========================================================
# CACHE
# ==========================================================

def _find_latest_cache():
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    pattern = str(cache_dir / f"{SCRIPT_NAME}_raw_*.json")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    return Path(files[0])


def _find_cd_cache():
    """Busca el cache del script CD detailed."""
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    pattern = str(cache_dir / f"{CD_CACHE_SCRIPT}_raw_*.json")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    return Path(files[0])


def _cache_is_fresh(cache_path, ttl_hours=CACHE_TTL_HOURS):
    if not cache_path or not cache_path.exists():
        return False
    mtime = cache_path.stat().st_mtime
    age_hours = (time.time() - mtime) / 3600
    return age_hours < ttl_hours


def _load_cache(cache_path):
    with open(cache_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_cache(data, script_name=SCRIPT_NAME):
    output_dir = get_output_dir("outcome")
    cache_dir = output_dir / ".cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{script_name}_raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return cache_path


# ==========================================================
# DETECCIÓN
# ==========================================================

def detect_obsolete(name: str) -> str:
    name_lower = name.lower()
    for kw in OBSOLETE_KEYWORDS:
        if kw in name_lower:
            return "Sí"
    return "No"


def _is_prod_env(env_name: str) -> bool:
    """Detecta si un environment es de producción por keywords."""
    env_lower = env_name.lower()
    return any(kw in env_lower for kw in PROD_KEYWORDS)


def _parse_iso_date(date_str):
    """Parsea una fecha ISO a datetime. Retorna None si es vacío/inválido."""
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        # Azure DevOps usa formato ISO 8601: 2026-04-20T14:00:00Z
        dt_str = date_str.replace("Z", "+00:00")
        return datetime.fromisoformat(dt_str)
    except (ValueError, TypeError):
        return None


# ==========================================================
# FETCH PROD DEPLOY PER PIPELINE
# ==========================================================

def _fetch_prod_deploy(cd_row, headers, org, project, deadline_date):
    """Consulta deployments y releases de un pipeline CD y extrae info del último deploy a producción."""
    def_id = cd_row.get("id") or cd_row.get("cd_pipeline_id")
    name = cd_row.get("name") or cd_row.get("cd_pipeline_name", "")
    path = cd_row.get("path") or cd_row.get("cd_pipeline_path", "")
    is_obsolete = cd_row.get("isObsolete") or detect_obsolete(name)

    result = {
        "cd_pipeline_id": def_id,
        "cd_pipeline_name": name,
        "cd_pipeline_path": path,
        "environments": "",
        "last_release_number": "",
        "last_release_id": "",
        "last_release_date": "",
        "last_release_status": "",
        "prod_env_name": "",
        "last_prod_deploy_date": "",
        "last_prod_deploy_status": "",
        "last_prod_release_number": "",
        "last_prod_release_id": "",
        "commit_sha": "",
        "git_commit_sha": "",
        "build_id": "",
        "build_number": "",
        "refresh_release_number": "",
        "refresh_release_id": "",
        "refresh_release_date": "",
        "refresh_release_prod_status": "",
        "refresh_release_prod_date": "",
        "deadline": str(deadline_date) if deadline_date else "",
        "deadline_status": "",
        "days_since_prod_deploy": "",
        "is_obsolete": is_obsolete,
    }

    # ── PASO A: Obtener definition detail (para environments) ────────
    def_url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/definitions/{def_id}"
    try:
        def_data = az_get(def_url, headers)
    except Exception as e:
        print(f"   ⚠️  [{name}] Error definition detail: {e}")
        def_data = {}

    env_names = []
    if isinstance(def_data, dict) and def_data.get("environments"):
        env_names = [e.get("name", "") for e in def_data["environments"]]
    result["environments"] = " / ".join(env_names) if env_names else (cd_row.get("environments", "") or "")

    # ── PASO B: Último release global ───────────────────────────────
    releases_url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/releases"
    try:
        releases_data = az_get(releases_url, headers, {"definitionId": def_id, "$top": 1})
    except Exception as e:
        print(f"   ⚠️  [{name}] Error releases list: {e}")
        releases_data = {}

    releases = releases_data.get("value", []) if isinstance(releases_data, dict) else []
    if not releases:
        result["deadline_status"] = "Sin releases"
        return result

    last_r = releases[0]
    result["last_release_number"] = last_r.get("name", "")
    result["last_release_id"] = last_r.get("id", "")
    result["last_release_date"] = last_r.get("createdOn", "")
    result["last_release_status"] = last_r.get("status", "")

    # ── PASO C: Buscar último deploy exitoso a prod via Deployments API ──
    deploys_url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/deployments"
    try:
        deploys_data = az_get(deploys_url, headers, {"definitionId": def_id, "$top": 100, "queryOrder": "descending"})
    except Exception as e:
        print(f"   ⚠️  [{name}] Error deployments list: {e}")
        deploys_data = {}

    deployments = deploys_data.get("value", []) if isinstance(deploys_data, dict) else []

    # Si no obtuvimos environments del definition detail, extraer de deployments
    if not env_names and deployments:
        dep_env_names = []
        for dep in deployments:
            env_obj = dep.get("releaseEnvironment", {})
            if isinstance(env_obj, dict):
                en = env_obj.get("name", "")
                if en and en not in dep_env_names:
                    dep_env_names.append(en)
        if dep_env_names:
            result["environments"] = " / ".join(dep_env_names)

    # Buscar el deployment exitoso más reciente a un environment de producción
    best_deploy = None  # (datetime, env_name, status, finished_on_str, release_id, release_name)

    for dep in deployments:
        env_name = ""
        # releaseEnvironment es un dict con {id, name}
        env_obj = dep.get("releaseEnvironment", {})
        if isinstance(env_obj, dict):
            env_name = env_obj.get("name", "") or env_obj.get("environmentName", "")
        elif isinstance(env_obj, str):
            env_name = env_obj

        if not _is_prod_env(env_name):
            continue

        dep_status = dep.get("deploymentStatus", "")
        # Deployments API usa completedOn (no finishedOn)
        completed_on = dep.get("completedOn", "") or dep.get("lastModifiedOn", "")
        dt_finished = _parse_iso_date(completed_on)

        if not dt_finished:
            continue

        # Considerar succeeded y partiallySucceeded como exitosos
        if dep_status not in ("succeeded", "partiallySucceeded"):
            continue

        if best_deploy is None or dt_finished > best_deploy[0]:
            rel_obj = dep.get("release", {})
            rel_id = rel_obj.get("id", "") if isinstance(rel_obj, dict) else ""
            rel_name = rel_obj.get("name", "") if isinstance(rel_obj, dict) else ""
            best_deploy = (dt_finished, env_name, dep_status, completed_on, rel_id, rel_name)

    # ── PASO D: Si Deployments API no encontró prod, intentar desde releases ──
    if not best_deploy:
        try:
            releases_full = az_get(releases_url, headers, {"definitionId": def_id, "$top": 100})
        except Exception:
            releases_full = {}

        rel_list = releases_full.get("value", []) if isinstance(releases_full, dict) else []
        for rel in rel_list:
            envs = rel.get("environments", [])
            if not envs:
                continue
            for env in envs:
                env_name = env.get("name", "")
                if not _is_prod_env(env_name):
                    continue
                # Intentar deploySteps
                for step in env.get("deploySteps", []):
                    step_status = step.get("deploymentStatus", "")
                    finished_on = step.get("finishedOn", "")
                    dt_finished = _parse_iso_date(finished_on)
                    if dt_finished and step_status in ("succeeded", "partiallySucceeded"):
                        if best_deploy is None or dt_finished > best_deploy[0]:
                            best_deploy = (dt_finished, env_name, step_status, finished_on, rel.get("id", ""), rel.get("name", ""))
                # Fallback: environment status
                if not env.get("deploySteps"):
                    env_status = env.get("status", "")
                    modified_on = env.get("modifiedOn", "")
                    dt_modified = _parse_iso_date(modified_on)
                    if dt_modified and env_status in ("succeeded", "partiallySucceeded"):
                        if best_deploy is None or dt_modified > best_deploy[0]:
                            best_deploy = (dt_modified, env_name, env_status, modified_on, rel.get("id", ""), rel.get("name", ""))

    # ── PASO E: Completar datos de prod deploy ──────────────────────
    if best_deploy:
        dt_prod, env_name, deploy_status, finished_on_str, prod_rel_id, prod_rel_name = best_deploy
        result["prod_env_name"] = env_name
        result["last_prod_deploy_date"] = finished_on_str
        result["last_prod_deploy_status"] = deploy_status
        result["last_prod_release_number"] = prod_rel_name
        result["last_prod_release_id"] = prod_rel_id

        # ── PASO F: Obtener artefactos del release con deploy exitoso ──
        if prod_rel_id:
            rel_detail_url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/releases/{prod_rel_id}"
            try:
                rel_detail = az_get(rel_detail_url, headers)
            except Exception as e:
                print(f"   ⚠️  [{name}] Error release detail {prod_rel_id}: {e}")
                rel_detail = {}

            if isinstance(rel_detail, dict):
                artifacts = rel_detail.get("artifacts", [])
                # Primario primero para que commit_sha/build_id vengan del artefacto principal
                sorted_arts = sorted(artifacts, key=lambda a: (not a.get("isPrimary", False)))
                for art in sorted_arts:
                    art_type = art.get("type", "")            # Build, Git, GitHub, ExternalGit…
                    is_build_type = art_type == "Build"
                    is_git_type = art_type in ("Git", "GitHub", "ExternalGit", "TFVC")
                    ref = art.get("definitionReference", {})

                    if is_build_type:
                        # commit_sha = commit que compiló el artefacto CI (sourceVersion)
                        sv = ref.get("sourceVersion", {})
                        if isinstance(sv, dict) and sv.get("id") and not result["commit_sha"]:
                            result["commit_sha"] = sv["id"]
                        # build_id / build_number = identificador del run de CI
                        ver = ref.get("version", {})
                        if isinstance(ver, dict):
                            if ver.get("id") and not result["build_id"]:
                                result["build_id"] = str(ver["id"])
                            if ver.get("name") and not result["build_number"]:
                                result["build_number"] = ver["name"]

                    elif is_git_type:
                        # git_commit_sha = commit del artefacto git directo (no CI)
                        ver = ref.get("version", {})
                        if isinstance(ver, dict) and ver.get("id") and not result["git_commit_sha"]:
                            result["git_commit_sha"] = ver["id"]
                        # Fallback: sourceVersion si existe
                        sv = ref.get("sourceVersion", {})
                        if isinstance(sv, dict) and sv.get("id") and not result["git_commit_sha"]:
                            result["git_commit_sha"] = sv["id"]

        # ── PASO G: Calcular days_since_prod_deploy ──────────────────
        now = datetime.now(timezone.utc)
        if dt_prod.tzinfo is None:
            dt_prod = dt_prod.replace(tzinfo=timezone.utc)
        days_elapsed = (now - dt_prod).days
        result["days_since_prod_deploy"] = days_elapsed

        # ── PASO H: Calcular deadline_status ─────────────────────────
        if deadline_date:
            prod_date = dt_prod.date() if hasattr(dt_prod, 'date') else dt_prod
            if prod_date > deadline_date:
                result["deadline_status"] = "Vigente"
            else:
                result["deadline_status"] = "Actualizar release"
        else:
            result["deadline_status"] = ""

        # ── PASO I: Detectar "refresh release" (mismo build, release más nuevo) ──
        # Un refresh release es un release posterior al último deploy a prod
        # que usa el MISMO build artifact → cambio de config/variables, no de código.
        if result["build_number"] and prod_rel_id:
            try:
                refresh_data = az_get(releases_url, headers, {
                    "definitionId": def_id,
                    "$top": 50,
                    "$expand": "artifacts",
                })
                refresh_rels = refresh_data.get("value", []) if isinstance(refresh_data, dict) else []

                for rel in refresh_rels:
                    rel_id_str = str(rel.get("id", ""))
                    if rel_id_str == str(prod_rel_id):
                        break  # Llegamos al release del último prod deploy
                    # Comparar build number del artefacto
                    for art in rel.get("artifacts", []):
                        ver = art.get("definitionReference", {}).get("version", {})
                        if isinstance(ver, dict) and ver.get("name") == result["build_number"]:
                            result["refresh_release_number"] = rel.get("name", "")
                            result["refresh_release_id"] = rel_id_str
                            result["refresh_release_date"] = rel.get("createdOn", "")
                            # Verificar estado de prod en el refresh release (detail call)
                            try:
                                rr_detail = az_get(
                                    f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/releases/{rel_id_str}",
                                    headers
                                )
                                for env in rr_detail.get("environments", []):
                                    if _is_prod_env(env.get("name", "")):
                                        rr_prod_status = env.get("status", "")
                                        result["refresh_release_prod_status"] = rr_prod_status
                                        # Extraer fecha real del deploy a prod del refresh
                                        if rr_prod_status in ("succeeded", "partiallySucceeded"):
                                            rr_prod_date = ""
                                            for step in env.get("deploySteps", []):
                                                if step.get("deploymentStatus") in ("succeeded", "partiallySucceeded"):
                                                    rr_prod_date = step.get("finishedOn", "")
                                                    break
                                            result["refresh_release_prod_date"] = rr_prod_date or env.get("modifiedOn", "")
                                        break
                            except Exception:
                                pass
                            break  # Tomamos el primero (más reciente)
                    if result["refresh_release_number"]:
                        break
            except Exception as e:
                print(f"   ⚠️  [{name}] Error refresh release search: {e}")

        # ── Recalcular deadline_status/days_since con fecha efectiva del refresh ──
        if result["refresh_release_number"]:
            # Prioridad: fecha real del deploy a prod del refresh > fecha de creación del refresh
            refresh_eff = result["refresh_release_prod_date"] or result["refresh_release_date"]
            dt_refresh = _parse_iso_date(refresh_eff)
            if dt_refresh:
                if dt_refresh.tzinfo is None:
                    dt_refresh = dt_refresh.replace(tzinfo=timezone.utc)
                now_r = datetime.now(timezone.utc)
                result["days_since_prod_deploy"] = (now_r - dt_refresh).days
                if deadline_date:
                    if dt_refresh.date() > deadline_date:
                        result["deadline_status"] = "Vigente"
                    else:
                        result["deadline_status"] = "Actualizar release"
    else:
        # No se encontró deploy exitoso a producción
        has_prod_env = any(
            _is_prod_env(e.strip())
            for e in result["environments"].replace(" / ", "/").split("/")
            if e.strip()
        )
        if not has_prod_env:
            result["deadline_status"] = "Sin env. Producción"
        else:
            result["deadline_status"] = "Sin deploy exitoso a prod"

    # Recalcular is_obsolete con el deadline_status final
    result["is_obsolete"] = detect_obsolete(name) or result.get("deadline_status") == "Actualizar release"

    return result


# ==========================================================
# EXPORT
# ==========================================================

# ==========================================================
# CHART HELPERS
# ==========================================================

# Colores para deadline_status
STATUS_COLORS = {
    "Vigente":               "27AE60",   # verde
    "Actualizar release":    "E67E22",   # naranja
    "Sin env. Producción":   "95A5A6",   # gris
    "Sin releases":          "BDC3C7",   # gris claro
    "Sin deploy exitoso a prod": "E74C3C", # rojo
}

# Colores para bins de antigüedad
BIN_COLORS = ["27AE60", "2ECC71", "F1C40F", "E67E22", "E74C3C", "C0392B"]


def _add_charts_sheet(excel_path, df, deadline_date):
    """Agrega hoja Charts con 2 gráficos nativos Excel."""
    wb = load_workbook(excel_path)

    # ---- Hoja oculta para datos de gráficos ----
    data_sheet = wb.create_sheet("_chart_data")

    # ── Chart 1: Distribución Deadline Status (Donut) ──────────────
    status_counts = df["deadline_status"].value_counts()
    data_sheet["A1"] = "Estado"
    data_sheet["B1"] = "Cantidad"
    for i, (status, count) in enumerate(status_counts.items(), start=2):
        data_sheet[f"A{i}"] = status
        data_sheet[f"B{i}"] = int(count)

    chart1 = PieChart()
    chart1.style = 10
    chart1.title = f"Distribución Vigencia vs Deadline ({deadline_date})"
    chart1.width = 18
    chart1.height = 14

    labels_ref = Reference(data_sheet, min_col=1, min_row=2, max_row=1 + len(status_counts))
    data_ref = Reference(data_sheet, min_col=2, min_row=1, max_row=1 + len(status_counts))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(labels_ref)

    # Colores por estado
    for idx, status in enumerate(status_counts.index):
        pt = DataPoint(idx=idx)
        color = STATUS_COLORS.get(status, "BDC3C7")
        pt.graphicalProperties.solidFill = color
        chart1.series[0].data_points.append(pt)

    chart1.series[0].dLbls = DataLabelList()
    chart1.series[0].dLbls.showPercent = True
    chart1.series[0].dLbls.showCatName = True
    chart1.series[0].dLbls.showVal = True

    # Donut (hole size 50%)
    from openpyxl.chart.series import DataPoint as DP
    chart1.series[0].graphicalProperties.line.noFill = True
    try:
        chart1.series[0].explosion = 0
    except Exception:
        pass

    # ── Chart 2: Antigüedad desde último deploy a Prod (Histograma) ─
    # Bins: 0-30, 31-60, 61-90, 91-180, 181-365, >365 días
    bins_labels = ["0-30", "31-60", "61-90", "91-180", "181-365", ">365"]
    bins_edges = [0, 30, 60, 90, 180, 365, 99999]
    bins_counts = [0] * 6

    days_col = df["days_since_prod_deploy"]
    for val in days_col:
        if val == "" or val is None:
            continue
        try:
            d = int(val)
        except (ValueError, TypeError):
            continue
        for b in range(6):
            if bins_edges[b] < d <= bins_edges[b + 1]:
                bins_counts[b] += 1
                break

    data_sheet["D1"] = "Rango días"
    data_sheet["E1"] = "Pipelines"
    for i, (label, count) in enumerate(zip(bins_labels, bins_counts), start=2):
        data_sheet[f"D{i}"] = label
        data_sheet[f"E{i}"] = count

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Antigüedad del Último Deploy a Producción (días)"
    chart2.y_axis.title = "Cantidad de pipelines"
    chart2.x_axis.title = "Rango de días desde último deploy"
    chart2.width = 18
    chart2.height = 14

    cats_ref = Reference(data_sheet, min_col=4, min_row=2, max_row=7)
    vals_ref = Reference(data_sheet, min_col=5, min_row=1, max_row=7)
    chart2.add_data(vals_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.shape = 4

    # Colores por bin (verde→rojo)
    for idx, color in enumerate(BIN_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        chart2.series[0].data_points.append(pt)

    chart2.series[0].dLbls = DataLabelList()
    chart2.series[0].dLbls.showVal = True

    # ── Crear hoja Charts ────────────────────────────────────────────
    charts_sheet = wb.create_sheet("Charts")
    charts_sheet.add_chart(chart1, "A1")
    charts_sheet.add_chart(chart2, "A32")

    # Ocultar hoja de datos
    data_sheet.sheet_state = "hidden"

    wb.save(excel_path)
    wb.close()


def export_results(rows, output_dir, script_name=SCRIPT_NAME, deadline_date=None):

    """Exporta resultados usando ExportManager centralizado con fallback."""

    if not EXPORT_MANAGER_AVAILABLE:
        # Fallback a exportación manual con pandas
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        df = pd.DataFrame(rows)
        excel_path = output_dir / f"{script_name}_{ts}.xlsx"
        df.to_excel(excel_path, index=False, engine="openpyxl")

        # Agregar charts si hay datos
        if not df.empty and deadline_date:
            try:
                _add_charts_sheet(excel_path, df, deadline_date)
                print(f"📊 Excel: {excel_path.resolve()} (2 gráficos)")
            except Exception as e:
                print(f"⚠️  Error generando gráficos: {e}")
                print(f"📊 Excel: {excel_path.resolve()}")
        else:
            print(f"📊 Excel: {excel_path.resolve()}")

        csv_path = output_dir / f"{script_name}_{ts}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"📄 CSV:  {csv_path.resolve()}")
        return excel_path, csv_path
    
    # Usar ExportManager
    manager = ExportManager(script_name, "1.0.0")
    
    summary = {
        "total_deployments": len(rows),
    }
    
    # Exportar a JSON
    json_path = manager.export_json(rows, summary=summary)
    if json_path:
        print(f"📋 JSON: {json_path}")
    
    # Exportar a CSV
    csv_path = manager.export_csv(rows)
    if csv_path:
        print(f"📄 CSV:  {csv_path}")
    
    # Exportar a Excel
    excel_path = manager.export_excel(rows, sheet_name="Prod Deployments", summary=summary)
    if excel_path:
        print(f"📊 Excel: {excel_path}")
    
    return excel_path, csv_path


# ==========================================================
# RESUMEN RICH
# ==========================================================

def print_summary(total, vigentes, actualizar, sin_prod, sin_releases, api_calls, cache_used, duration_seconds, deadline):
    if not RICH_AVAILABLE:
        print(f"\n{'='*60}")
        print(f"📊 RESUMEN — Prod Deploy Tracker")
        print(f"   Deadline:            {deadline or 'N/A'}")
        print(f"   Total pipelines CD:  {total}")
        print(f"   Vigentes:            {vigentes}")
        print(f"   Actualizar release:  {actualizar}")
        print(f"   Sin env. Producción: {sin_prod}")
        print(f"   Sin releases:        {sin_releases}")
        print(f"   Cache usado:         {'Sí' if cache_used else 'No'}")
        print(f"   Llamadas API:        {api_calls}")
        print(f"   Duración:            {duration_seconds:.1f}s")
        print(f"{'='*60}")
        return

    console = Console(file=sys.__stdout__)
    table = Table(title="📊 Resumen — Prod Deploy Tracker", show_header=True, header_style="bold magenta")
    table.add_column("Métrica", style="cyan")
    table.add_column("Valor", style="green")
    table.add_row("Deadline", str(deadline or "N/A"))
    table.add_row("Total pipelines CD", str(total))
    table.add_row("✅ Vigentes", str(vigentes))
    table.add_row("⚠️  Actualizar release", str(actualizar))
    table.add_row("❌ Sin env. Producción", str(sin_prod))
    table.add_row("📭 Sin releases", str(sin_releases))
    table.add_row("Cache usado", "✅ Sí" if cache_used else "❌ No")
    table.add_row("Llamadas API", str(api_calls))
    table.add_row("Duración", f"{duration_seconds:.1f}s")
    console.print(table)


# ==========================================================
# MAIN
# ==========================================================

def main():
    parser = argparse.ArgumentParser(description="CD Prod Deploy Tracker — Rastrea último deploy exitoso a Producción")
    parser.add_argument("--pat", default=os.getenv("AZDO_PAT"), help="Azure DevOps PAT")
    parser.add_argument("--org", default=DEFAULT_ORG, help="Organización Azure DevOps")
    parser.add_argument("--project", default=DEFAULT_PROJECT, help="Proyecto")
    parser.add_argument("--deadline", required=True, help="Fecha deadline (YYYY-MM-DD). Deploy posterior = Vigente, igual o anterior = Actualizar release")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="Hilos paralelos")
    parser.add_argument("--output", default=None, help="Directorio de salida")
    parser.add_argument("--force-refresh", action="store_true", help="Ignorar cache propio, consultar APIs")
    parser.add_argument("--skip-cache", action="store_true", help="Alias de --force-refresh")
    args = parser.parse_args()
    args.org = normalize_org(args.org)

    if not args.pat:
        print("❌ Se requiere --pat o env AZDO_PAT")
        sys.exit(1)

    # Parse deadline
    try:
        deadline_date = date.fromisoformat(args.deadline)
    except ValueError:
        print(f"❌ Formato de deadline inválido: {args.deadline}. Usar YYYY-MM-DD")
        sys.exit(1)

    output_dir = get_output_dir(args.output or "outcome")
    output_dir.mkdir(parents=True, exist_ok=True)

    tee = setup_logging()
    start_time = time.time()
    api_calls = 0
    cache_used = False
    rows = []

    try:
        # ============================================
        # PASO 1: Verificar cache propio
        # ============================================
        if not args.force_refresh and not args.skip_cache:
            cache_path = _find_latest_cache()
            if cache_path and _cache_is_fresh(cache_path):
                print(f"📦 Cache encontrado: {cache_path.name} (fresh)")
                # Verificar que el deadline coincide
                data = _load_cache(cache_path)
                cached_deadline = data.get("metadata", {}).get("deadline", "")
                if cached_deadline == str(deadline_date):
                    print("⏭️  Mismo deadline — generando outputs desde cache...")
                    rows = data.get("rows", [])
                    cache_used = True
                else:
                    print(f"⚠️  Deadline diferente (cache={cached_deadline}, solicitado={deadline_date}). Re-consultando...")
            else:
                print("📭 Cache propio no encontrado o > 24h. Consultando APIs...")
        else:
            print("🔄 Force refresh — consultando APIs...")

        if not rows:
            # ============================================
            # PASO 2: Cargar datos base desde cache CD
            # ============================================
            cd_cache_path = _find_cd_cache()
            if not cd_cache_path:
                print("❌ No se encontró cache CD. Ejecutar herramienta 15 (CD Inventory) primero.")
                print("   Comando: python cicd_inventory_cd_detailed.py --org {org} --project {project}")
                sys.exit(1)

            print(f"📦 Cache CD encontrado: {cd_cache_path.name}")
            cd_data = _load_cache(cd_cache_path)
            cd_rows = cd_data.get("rows", [])

            if not cd_rows:
                print("❌ Cache CD vacío. Ejecutar herramienta 15 (CD Inventory) primero.")
                sys.exit(1)

            print(f"📋 {len(cd_rows)} pipelines CD cargados desde cache")

            # ============================================
            # PASO 3: Consultar releases por pipeline (paralelo)
            # ============================================
            headers = get_headers(args.pat)
            total = len(cd_rows)
            processed = 0

            print(f"🔍 Consultando releases con environments + artifacts ({total} pipelines)...")

            if RICH_AVAILABLE:
                tee.pause_terminal()
                with _progress_context() as progress:
                    task = progress.add_task("Rastreando deploys a Prod", total=total)
                    with ThreadPoolExecutor(max_workers=args.workers) as executor:
                        futures = {
                            executor.submit(_fetch_prod_deploy, cd_row, headers, args.org, args.project, deadline_date): cd_row
                            for cd_row in cd_rows
                        }
                        for future in as_completed(futures):
                            try:
                                result = future.result()
                                if result:
                                    rows.append(result)
                                    api_calls += 3  # definition + releases + deployments (+1 optional release detail)
                            except Exception as e:
                                cd_name = futures[future].get("name", futures[future].get("cd_pipeline_name", "?"))
                                print(f"❌ Error en pipeline {cd_name}: {e}")
                            processed += 1
                            progress.update(task, advance=1)
                tee.resume_terminal()
            else:
                with ThreadPoolExecutor(max_workers=args.workers) as executor:
                    futures = {
                        executor.submit(_fetch_prod_deploy, cd_row, headers, args.org, args.project, deadline_date): cd_row
                        for cd_row in cd_rows
                    }
                    for i, future in enumerate(as_completed(futures), 1):
                        try:
                            result = future.result()
                            if result:
                                rows.append(result)
                                api_calls += 3  # definition + releases + deployments (+1 optional release detail)
                        except Exception as e:
                            cd_name = futures[future].get("name", futures[future].get("cd_pipeline_name", "?"))
                            print(f"❌ Error en pipeline {cd_name}: {e}")
                        if i % 10 == 0 or i == total:
                            print(f"  Progreso: {i}/{total} ({int(i/total*100)}%)")

            # ============================================
            # PASO 4: Guardar cache propio
            # ============================================
            cache_data = {
                "metadata": {
                    "script": SCRIPT_NAME,
                    "org": args.org,
                    "project": args.project,
                    "deadline": str(deadline_date),
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                    "count": len(rows),
                },
                "rows": rows,
            }
            cache_path = _save_cache(cache_data)
            print(f"💾 Cache guardado: {cache_path.name}")

        # ============================================
        # PASO 5: Exportar resultados
        # ============================================
        if rows:
            export_results(rows, output_dir, deadline_date=deadline_date)
        else:
            print("⚠️  No hay datos para exportar")

        # ============================================
        # PASO 6: Resumen
        # ============================================
        vigentes = sum(1 for r in rows if r.get("deadline_status") == "Vigente")
        actualizar = sum(1 for r in rows if r.get("deadline_status") == "Actualizar release")
        sin_prod = sum(1 for r in rows if r.get("deadline_status") == "Sin env. Producción")
        sin_releases = sum(1 for r in rows if r.get("deadline_status") == "Sin releases")

        duration = time.time() - start_time
        print_summary(
            total=len(rows),
            vigentes=vigentes,
            actualizar=actualizar,
            sin_prod=sin_prod,
            sin_releases=sin_releases,
            api_calls=api_calls,
            cache_used=cache_used,
            duration_seconds=duration,
            deadline=deadline_date,
        )

    finally:
        teardown_logging(tee)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
azdo_repo_properties_branch_diff.py

Compara la configuración de un componente (carpeta) entre dos ramas de un
repositorio de propiedades en Azure DevOps.  Detecta diferencias que puedan
impactar la calidad y estabilidad de un despliegue a producción.

Flujo:
  1. Selecciona el repositorio de propiedades  (--repo  o prompt interactivo)
  2. Selecciona el componente / carpeta        (--component o prompt)
  3. Define rama origen y destino              (--source / --target o prompt)
  4. Recupera la lista de archivos de ambas ramas vía AzDO REST API
  5. Calcula diff de conjuntos y de contenido por archivo
  6. Clasifica severidad por archivo
  7. Muestra tabla-resumen + detalle de diff (panel Rich con sintaxis diff)
  8. Exporta resultados  (--output json / csv / excel)

Severidad por tipo de cambio:
  🚨 CRITICAL  → Archivo eliminado en source  (config que existe en destino se perderá)
  🔴 HIGH      → Contenido modificado con cambios funcionales  (valores reales cambiados)
  🟡 MEDIUM    → Archivo nuevo en source  (nueva config que se introducirá en destino)
  🔵 LOW       → Sólo diferencias de formato / comentarios  (impacto funcional nulo)
  ⚪ NONE      → Sin diferencias entre las ramas

Casos de uso:
  - Validar que la config de una release coincide con master antes del despliegue
  - Auditar diferencias entre develop y QA antes de una integración
  - Generar evidencia de configuración para gates de aprobación en CD pipelines
  - Detectar configuraciones huérfanas o inconsistencias de entorno

Códigos de salida:
  0 → Sin diferencias o sólo MEDIUM/LOW/NONE
  1 → Diferencias HIGH detectadas
  2 → Diferencias CRITICAL detectadas

Autor: Harold Adrian
"""

import argparse
import base64
import csv
import difflib
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

# --- Directorio de salida centralizado (DEVSECOPS_OUTPUT_DIR) ---
try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
# -------------------------------------------------------------------

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn
    from rich.syntax import Syntax
    from rich.table import Table
    from rich.text import Text
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

__version__ = "1.0.0"
__author__  = "Harold Adrian"

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_ORG_URL  = "https://dev.azure.com/Coppel-Retail"
DEFAULT_PROJECT  = "Compras.RMI"
DEFAULT_TIMEZONE = "America/Mazatlan"
API_VERSION      = "7.1"

SEV_CRITICAL = "CRITICAL"
SEV_HIGH     = "HIGH"
SEV_MEDIUM   = "MEDIUM"
SEV_LOW      = "LOW"
SEV_NONE     = "NONE"

SEV_ORDER: Dict[str, int] = {
    SEV_NONE: 0, SEV_LOW: 1, SEV_MEDIUM: 2, SEV_HIGH: 3, SEV_CRITICAL: 4
}

SEV_EMOJI = {
    SEV_CRITICAL: "🚨",
    SEV_HIGH:     "🔴",
    SEV_MEDIUM:   "🟡",
    SEV_LOW:      "🔵",
    SEV_NONE:     "⚪",
}

SEV_COLOR = {
    SEV_CRITICAL: "bold red",
    SEV_HIGH:     "red",
    SEV_MEDIUM:   "yellow",
    SEV_LOW:      "blue",
    SEV_NONE:     "dim",
}

SEV_BORDER = {
    SEV_CRITICAL: "red",
    SEV_HIGH:     "red",
    SEV_MEDIUM:   "yellow",
    SEV_LOW:      "blue",
    SEV_NONE:     "dim",
}

CHANGE_ADD    = "add"
CHANGE_DELETE = "delete"
CHANGE_EDIT   = "edit"
CHANGE_RENAME = "rename"
CHANGE_NONE   = "none"

CHANGE_LABEL: Dict[str, Tuple[str, str]] = {
    CHANGE_ADD:    ("green",    "➕ ADD"),
    CHANGE_DELETE: ("bold red", "🗑  DEL"),
    CHANGE_EDIT:   ("yellow",   "✏  EDIT"),
    CHANGE_RENAME: ("cyan",     "♻  REN"),
    CHANGE_NONE:   ("dim",      "⚪ SAME"),
}

# Prefijos de líneas informativas de kubectl / devtools (no son errores reales)
_COMMENT_PREFIXES = ("#", "!", "//", "/*", " *")


# ═══════════════════════════════════════════════════════════════════════════════
# ARGS
# ═══════════════════════════════════════════════════════════════════════════════
def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Compara configuración de un componente entre dos ramas de un repo de propiedades en AzDO",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python azdo_repo_properties_branch_diff.py --pat TOKEN --repo retail-properties \\
         --component ps-om-com-customerorder --source release/release-1.6.0 --target master

  python azdo_repo_properties_branch_diff.py --pat TOKEN --repo retail-properties \\
         --component ps-om-com-customerorder --source develop --target master --output excel

  python azdo_repo_properties_branch_diff.py --pat TOKEN   # modo interactivo completo
        """,
    )
    p.add_argument("--org",        "-g",  default=DEFAULT_ORG_URL,
                   help=f"URL de la organización (default: {DEFAULT_ORG_URL})")
    p.add_argument("--project",    "-p",  default=DEFAULT_PROJECT,
                   help=f"Nombre del proyecto (default: {DEFAULT_PROJECT})")
    p.add_argument("--pat",               required=True,
                   help="Personal Access Token con permisos: Code (Read)")
    p.add_argument("--repo",       "-r",  default=None,
                   help="Nombre del repositorio de propiedades (substring, case insensitive)")
    p.add_argument("--component",  "-c",  default=None,
                   help="Nombre del componente (carpeta dentro del repo, ej: ps-om-com-customerorder)")
    p.add_argument("--source",     "-s",  default=None,
                   help="Rama origen — la que se quiere desplegar (ej: release/release-1.6.0)")
    p.add_argument("--target",     "-t",  default=None,
                   help="Rama destino — el entorno receptor (ej: master)")
    p.add_argument("--context",           type=int, default=3,
                   help="Líneas de contexto en el diff (default: 3)")
    p.add_argument("--severity",          default=None,
                   choices=["CRITICAL", "HIGH", "MEDIUM", "LOW", "NONE"],
                   help="Mostrar solo archivos con severidad >= especificada")
    p.add_argument("--only-diff",         action="store_true",
                   help="Mostrar solo archivos con diferencias (omitir NONE)")
    p.add_argument("--no-content",        action="store_true",
                   help="Omitir detalle de diff de contenido (solo tabla resumen)")
    p.add_argument("--output",     "-o",  choices=["json", "csv", "excel"], default=None,
                   help="Exportar resultados")
    p.add_argument("--timezone",   "-tz", default=DEFAULT_TIMEZONE,
                   help=f"Zona horaria para fechas (default: {DEFAULT_TIMEZONE})")
    p.add_argument("--debug",             action="store_true",
                   help="Mostrar errores HTTP detallados")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP
# ═══════════════════════════════════════════════════════════════════════════════
def make_headers(pat: str) -> Dict:
    token = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Accept": "application/json"}


def api_get(
    url: str, headers: Dict, params: Dict = None, debug: bool = False,
    raw_content: bool = False,
) -> Optional[Any]:
    try:
        h = dict(headers)
        if raw_content:
            h["Accept"] = "application/octet-stream"
        r = requests.get(url, headers=h, params=params, timeout=30)
        if r.status_code == 404:
            return None
        if r.status_code >= 400:
            if debug:
                print(f"[DEBUG] HTTP {r.status_code} → {url}")
                print(f"[DEBUG] {r.text[:300]}")
            r.raise_for_status()
        if raw_content:
            return r.text
        return r.json()
    except Exception as exc:
        if debug:
            print(f"[DEBUG] {exc} → {url}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# AzDO GIT REST API
# ═══════════════════════════════════════════════════════════════════════════════
def get_repositories(
    org: str, project: str, headers: Dict, debug: bool = False
) -> List[Dict]:
    url  = f"{org}/{quote(project, safe='')}/_apis/git/repositories"
    data = api_get(url, headers, {"api-version": API_VERSION}, debug)
    return data.get("value", []) if data else []


def get_branches(
    org: str, project: str, repo_id: str, headers: Dict, debug: bool = False
) -> List[str]:
    """Retorna lista de nombres de ramas sin el prefijo refs/heads/."""
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/refs"
    params = {"api-version": API_VERSION, "filter": "heads/", "$top": 1000}
    data   = api_get(url, headers, params, debug)
    refs   = data.get("value", []) if data else []
    return [r["name"].replace("refs/heads/", "") for r in refs]


def get_items_in_path(
    org: str, project: str, repo_id: str, scope_path: str, branch: str,
    headers: Dict, debug: bool = False,
) -> List[Dict]:
    """Lista todos los archivos (blob) bajo scope_path en la rama indicada."""
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/items"
    params = {
        "api-version":                   API_VERSION,
        "scopePath":                     scope_path,
        "recursionLevel":                "Full",
        "versionDescriptor.versionType": "Branch",
        "versionDescriptor.version":     branch,
        "includeContentMetadata":        "true",
    }
    data = api_get(url, headers, params, debug)
    if not data:
        return []
    return [i for i in data.get("value", []) if i.get("gitObjectType") == "blob"]


def get_file_content(
    org: str, project: str, repo_id: str, file_path: str, branch: str,
    headers: Dict, debug: bool = False,
) -> Optional[str]:
    """Retorna el contenido de texto de un archivo en una rama específica."""
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/items"
    params = {
        "api-version":                   API_VERSION,
        "path":                          file_path,
        "versionDescriptor.versionType": "Branch",
        "versionDescriptor.version":     branch,
        "$format":                       "octetStream",
    }
    return api_get(url, headers, params, debug, raw_content=True)


def get_branch_diffs(
    org: str, project: str, repo_id: str,
    source_branch: str, target_branch: str,
    headers: Dict, debug: bool = False,
) -> List[Dict]:
    """
    Retorna cambios entre target (base/old) y source (new).
    changeType: add / delete / edit / rename
      add    = archivo presente en source pero no en target (nueva config)
      delete = archivo presente en target pero no en source (config que se eliminará)
      edit   = mismo archivo, contenido distinto
    """
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/diffs/commits"
    params = {
        "api-version":       API_VERSION,
        "baseVersionType":   "branch",
        "baseVersion":       target_branch,   # old / destino
        "targetVersionType": "branch",
        "targetVersion":     source_branch,   # new / origen
        "$top":              5000,
    }
    data = api_get(url, headers, params, debug)
    return data.get("changes", []) if data else []


# ═══════════════════════════════════════════════════════════════════════════════
# DIFF ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
def classify_diff_content(diff_lines: List[str]) -> str:
    """
    Determina si el diff contiene cambios funcionales (HIGH) o solo de
    formato / comentarios (LOW).
    """
    for line in diff_lines:
        if not (line.startswith("+") or line.startswith("-")):
            continue
        if line.startswith("+++") or line.startswith("---"):
            continue
        stripped = line[1:].strip()
        if not stripped:
            continue
        # Solo comentarios o whitespace → LOW
        if any(stripped.startswith(p) for p in _COMMENT_PREFIXES):
            continue
        return SEV_HIGH
    return SEV_LOW


def build_unified_diff(
    source_content: Optional[str],
    target_content: Optional[str],
    filename: str,
    source_branch: str,
    target_branch: str,
    context: int = 3,
) -> Tuple[List[str], int, int]:
    """
    Genera unified diff entre target_content (old) y source_content (new).
    Returns: (diff_lines, lines_added, lines_removed)
    """
    old = (target_content or "").splitlines(keepends=True)
    new = (source_content or "").splitlines(keepends=True)
    diff = list(difflib.unified_diff(
        old, new,
        fromfile=f"{target_branch}/{filename}",
        tofile=f"{source_branch}/{filename}",
        lineterm="",
        n=context,
    ))
    added   = sum(1 for l in diff if l.startswith("+") and not l.startswith("+++"))
    removed = sum(1 for l in diff if l.startswith("-") and not l.startswith("---"))
    return diff, added, removed


def build_side_by_side_rows(
    old_content: str, new_content: str
) -> List[Tuple[str, str, str]]:
    """
    Alinea las líneas de ambas versiones con SequenceMatcher.
    Retorna lista de (tag, left_line, right_line):
      'equal'   → misma línea en ambas ramas
      'delete'  → solo en target/old (izquierda)
      'insert'  → solo en source/new (derecha)
      'replace' → línea modificada (izquierda=old, derecha=new)
    """
    old_lines = (old_content or "").splitlines()
    new_lines = (new_content or "").splitlines()
    matcher   = difflib.SequenceMatcher(None, old_lines, new_lines, autojunk=False)
    rows: List[Tuple[str, str, str]] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for ol, nl in zip(old_lines[i1:i2], new_lines[j1:j2]):
                rows.append(("equal", ol, nl))
        elif tag == "replace":
            old_chunk = old_lines[i1:i2]
            new_chunk = new_lines[j1:j2]
            n = max(len(old_chunk), len(new_chunk))
            for k in range(n):
                ol = old_chunk[k] if k < len(old_chunk) else ""
                nl = new_chunk[k] if k < len(new_chunk) else ""
                rows.append(("replace", ol, nl))
        elif tag == "delete":
            for ol in old_lines[i1:i2]:
                rows.append(("delete", ol, ""))
        elif tag == "insert":
            for nl in new_lines[j1:j2]:
                rows.append(("insert", "", nl))
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════
class FileDiff:
    """Resultado de la comparación de un archivo entre dos ramas."""

    __slots__ = (
        "path", "filename", "change_type", "severity",
        "diff_lines", "lines_added", "lines_removed",
        "source_content", "target_content",
    )

    def __init__(
        self,
        path: str,
        change_type: str,
        severity: str,
        diff_lines: List[str],
        lines_added: int,
        lines_removed: int,
        source_content: Optional[str],
        target_content: Optional[str],
    ):
        self.path           = path
        self.filename       = path.rsplit("/", 1)[-1]
        self.change_type    = change_type
        self.severity       = severity
        self.diff_lines     = diff_lines
        self.lines_added    = lines_added
        self.lines_removed  = lines_removed
        self.source_content = source_content
        self.target_content = target_content

    def to_dict(self) -> Dict:
        return {
            "path":          self.path,
            "filename":      self.filename,
            "change_type":   self.change_type,
            "severity":      self.severity,
            "lines_added":   self.lines_added,
            "lines_removed": self.lines_removed,
            "diff":          "\n".join(self.diff_lines),
        }


# ═══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS CORE
# ═══════════════════════════════════════════════════════════════════════════════
class _NullCtx:
    """Contexto nulo para reemplazar Progress cuando Rich no está disponible."""
    def __enter__(self):  return self
    def __exit__(self, *_): pass
    def add_task(self, *_, **__): return None
    def advance(self, _):  pass
    def update(self, *_, **__): pass


def _normalize_change_type(raw: str) -> str:
    raw = raw.lower()
    if "delete" in raw: return CHANGE_DELETE
    if "rename" in raw: return CHANGE_RENAME
    if "add"    in raw: return CHANGE_ADD
    return CHANGE_EDIT


def _initial_severity(change_type: str) -> str:
    if change_type == CHANGE_DELETE: return SEV_CRITICAL
    if change_type == CHANGE_ADD:    return SEV_MEDIUM
    return SEV_HIGH   # edit / rename


def analyze_component(
    org: str, project: str, repo_id: str,
    component_path: str, source_branch: str, target_branch: str,
    headers: Dict, context: int, debug: bool,
    console: Optional[Any] = None,
) -> List["FileDiff"]:
    """
    Punto de entrada del análisis.
    Intenta primero con la API diffs/commits (más eficiente);
    si no retorna resultados, compara manualmente item por item.
    """
    if console and RICH_AVAILABLE:
        console.print(
            f"[dim]  📡 Obteniendo diffs "
            f"[yellow]{source_branch}[/] ← [cyan]{target_branch}[/]...[/]"
        )

    all_changes = get_branch_diffs(
        org, project, repo_id, source_branch, target_branch, headers, debug
    )

    # Filtrar cambios dentro del componente
    prefix = component_path.strip("/") + "/"
    component_changes = [
        c for c in all_changes
        if c.get("item", {}).get("path", "").lstrip("/").startswith(prefix)
    ]

    if not component_changes:
        if console and RICH_AVAILABLE:
            console.print(
                "[dim]  ℹ️  diffs/commits sin resultados para el componente, "
                "comparando items directamente...[/]"
            )
        return _analyze_by_items(
            org, project, repo_id, component_path,
            source_branch, target_branch, headers, context, debug, console,
        )

    if console and RICH_AVAILABLE:
        console.print(
            f"[dim]  🔍 {len(component_changes)} cambio(s) en [{prefix}][/]"
        )

    return _process_changes(
        component_changes, org, project, repo_id,
        source_branch, target_branch, headers, context, debug, console,
    )


def _process_changes(
    changes: List[Dict],
    org: str, project: str, repo_id: str,
    source_branch: str, target_branch: str,
    headers: Dict, context: int, debug: bool,
    console: Optional[Any],
) -> List["FileDiff"]:
    results: List[FileDiff] = []
    ctx = (
        Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                 console=console)
        if RICH_AVAILABLE and console else _NullCtx()
    )
    with ctx as progress:
        task = progress.add_task("Analizando archivos...", total=len(changes)) \
               if RICH_AVAILABLE and console else None

        for change in changes:
            file_path   = change.get("item", {}).get("path", "")
            change_type = _normalize_change_type(change.get("changeType", "edit"))
            severity    = _initial_severity(change_type)

            sc = None if change_type == CHANGE_DELETE else \
                 get_file_content(org, project, repo_id, file_path, source_branch, headers, debug)
            tc = None if change_type == CHANGE_ADD else \
                 get_file_content(org, project, repo_id, file_path, target_branch, headers, debug)

            diff_lines, added, removed = build_unified_diff(
                sc, tc, file_path.rsplit("/", 1)[-1], source_branch, target_branch, context
            )

            if change_type == CHANGE_EDIT and diff_lines:
                severity = classify_diff_content(diff_lines)

            results.append(FileDiff(
                path=file_path, change_type=change_type, severity=severity,
                diff_lines=diff_lines, lines_added=added, lines_removed=removed,
                source_content=sc, target_content=tc,
            ))
            if task is not None:
                progress.advance(task)

    return sorted(results, key=lambda x: SEV_ORDER.get(x.severity, 0), reverse=True)


def _analyze_by_items(
    org: str, project: str, repo_id: str,
    component_path: str, source_branch: str, target_branch: str,
    headers: Dict, context: int, debug: bool,
    console: Optional[Any],
) -> List["FileDiff"]:
    """Fallback: compara la lista completa de items de ambas ramas."""
    src_items = get_items_in_path(
        org, project, repo_id, component_path, source_branch, headers, debug
    )
    tgt_items = get_items_in_path(
        org, project, repo_id, component_path, target_branch, headers, debug
    )

    src_map = {i["path"]: i for i in src_items}
    tgt_map = {i["path"]: i for i in tgt_items}
    all_paths = sorted(set(src_map) | set(tgt_map))

    if console and RICH_AVAILABLE:
        console.print(
            f"[dim]  📁 {len(src_map)} archivos en [{source_branch}]"
            f" | {len(tgt_map)} archivos en [{target_branch}][/]"
        )

    results: List[FileDiff] = []
    ctx = (
        Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                 console=console)
        if RICH_AVAILABLE and console else _NullCtx()
    )
    with ctx as progress:
        task = progress.add_task("Comparando archivos...", total=len(all_paths)) \
               if RICH_AVAILABLE and console else None

        for path in all_paths:
            in_src = path in src_map
            in_tgt = path in tgt_map
            fname  = path.rsplit("/", 1)[-1]

            if in_src and not in_tgt:
                change_type = CHANGE_ADD
                severity    = SEV_MEDIUM
                sc = get_file_content(org, project, repo_id, path, source_branch, headers, debug)
                tc = None
            elif not in_src and in_tgt:
                change_type = CHANGE_DELETE
                severity    = SEV_CRITICAL
                sc = None
                tc = get_file_content(org, project, repo_id, path, target_branch, headers, debug)
            else:
                sc = get_file_content(org, project, repo_id, path, source_branch, headers, debug)
                tc = get_file_content(org, project, repo_id, path, target_branch, headers, debug)
                diff_lines_tmp, _, _ = build_unified_diff(
                    sc, tc, fname, source_branch, target_branch, context
                )
                if not diff_lines_tmp:
                    change_type = CHANGE_NONE
                    severity    = SEV_NONE
                else:
                    change_type = CHANGE_EDIT
                    severity    = classify_diff_content(diff_lines_tmp)

            diff_lines, added, removed = build_unified_diff(
                sc, tc, fname, source_branch, target_branch, context
            )
            results.append(FileDiff(
                path=path, change_type=change_type, severity=severity,
                diff_lines=diff_lines, lines_added=added, lines_removed=removed,
                source_content=sc, target_content=tc,
            ))
            if task is not None:
                progress.advance(task)

    return sorted(results, key=lambda x: SEV_ORDER.get(x.severity, 0), reverse=True)


# ═══════════════════════════════════════════════════════════════════════════════
# DISPLAY
# ═══════════════════════════════════════════════════════════════════════════════
def _impact_desc(r: FileDiff) -> str:
    if r.change_type == CHANGE_DELETE:
        return "Archivo eliminado en origen — se perderá en destino tras merge"
    if r.change_type == CHANGE_ADD:
        return "Archivo nuevo en origen — se introducirá en destino"
    if r.change_type == CHANGE_RENAME:
        return "Archivo renombrado — validar referencias en la aplicación"
    if r.severity == SEV_LOW:
        return "Solo cambios de formato/comentarios — impacto funcional nulo"
    if r.severity == SEV_HIGH:
        return f"Valores de configuración cambiados (+{r.lines_added}/-{r.lines_removed} líneas)"
    return "Sin diferencias"


def print_summary_table(
    results: List[FileDiff],
    source_branch: str,
    target_branch: str,
    component: str,
    console: Any,
):
    counts = {s: sum(1 for r in results if r.severity == s)
              for s in (SEV_CRITICAL, SEV_HIGH, SEV_MEDIUM, SEV_LOW, SEV_NONE)}
    max_sev = max(results, key=lambda x: SEV_ORDER.get(x.severity, 0)).severity \
              if results else SEV_NONE
    border  = SEV_BORDER.get(max_sev, "dim")

    console.print(Panel(
        f"[bold]Componente:[/]     [cyan]{component}[/]\n"
        f"[bold]Rama origen:[/]    [green]{source_branch}[/]\n"
        f"[bold]Rama destino:[/]   [yellow]{target_branch}[/]\n"
        f"[bold]Archivos:[/]       {len(results)}\n"
        f"[bold]Severidad máx:[/]  "
        f"{SEV_EMOJI.get(max_sev, '')} "
        f"[{SEV_COLOR.get(max_sev, 'white')}]{max_sev}[/]",
        title="📊 Resumen de Diferencias de Configuración",
        border_style=border,
    ))

    table = Table(
        title=f"🔍 [{component}]  —  {source_branch} vs {target_branch}",
        title_style="bold magenta",
        header_style="bold cyan",
        border_style="dim",
        box=box.ROUNDED,
    )
    table.add_column("Archivo",   style="white",  min_width=28)
    table.add_column("Cambio",    justify="center", width=10)
    table.add_column("Severidad", justify="center", width=14)
    table.add_column("+Líneas",   justify="right",  width=8)
    table.add_column("-Líneas",   justify="right",  width=8)
    table.add_column("Impacto",   justify="left",   max_width=50)

    for r in results:
        ch_style, ch_label = CHANGE_LABEL.get(r.change_type, ("dim", r.change_type))
        sev_color = SEV_COLOR.get(r.severity, "white")
        sev_label = f"{SEV_EMOJI.get(r.severity, '')} {r.severity}"
        added_s   = f"[green]+{r.lines_added}[/]"   if r.lines_added   else "[dim]0[/]"
        removed_s = f"[red]-{r.lines_removed}[/]"   if r.lines_removed else "[dim]0[/]"
        table.add_row(
            r.filename,
            f"[{ch_style}]{ch_label}[/{ch_style}]",
            f"[{sev_color}]{sev_label}[/{sev_color}]",
            added_s, removed_s,
            _impact_desc(r),
        )

    console.print(table)

    # Footer estadísticas
    parts = []
    if counts[SEV_CRITICAL]: parts.append(f"[bold red]🚨 CRITICAL:{counts[SEV_CRITICAL]}[/]")
    if counts[SEV_HIGH]:     parts.append(f"[red]🔴 HIGH:{counts[SEV_HIGH]}[/]")
    if counts[SEV_MEDIUM]:   parts.append(f"[yellow]🟡 MEDIUM:{counts[SEV_MEDIUM]}[/]")
    if counts[SEV_LOW]:      parts.append(f"[blue]🔵 LOW:{counts[SEV_LOW]}[/]")
    if counts[SEV_NONE]:     parts.append(f"[dim]⚪ NONE:{counts[SEV_NONE]}[/]")
    if parts:
        console.print("  " + "  │  ".join(parts))

    # Leyenda de severidad
    console.print()
    console.print(
        "[dim]Leyenda: 🚨 config eliminada → CRITICAL | 🔴 valores cambiados → HIGH | "
        "🟡 config nueva → MEDIUM | 🔵 solo formato → LOW | ⚪ sin cambios → NONE[/]"
    )


def _print_side_by_side(
    r: FileDiff,
    context: int,
    border: str,
    title: str,
    source_branch: str,
    target_branch: str,
    console: Any,
):
    """Muestra diff lado-a-lado en una tabla Rich de dos columnas."""
    rows = build_side_by_side_rows(r.target_content or "", r.source_content or "")

    changed = {i for i, (tag, _, _) in enumerate(rows) if tag != "equal"}
    visible: set = set()
    for ci in changed:
        for off in range(-context, context + 1):
            idx = ci + off
            if 0 <= idx < len(rows):
                visible.add(idx)

    t = Table(
        show_header=True,
        header_style="bold white",
        box=box.SIMPLE_HEAD,
        border_style=border,
        expand=True,
        show_lines=True,
        padding=(0, 1),
    )
    t.add_column(f"◀  {target_branch}  (actual)",    ratio=1, no_wrap=False, overflow="fold")
    t.add_column(f"▶  {source_branch}  (entrante)", ratio=1, no_wrap=False, overflow="fold")

    prev = -1
    for i, (tag, left, right) in enumerate(rows):
        if i not in visible:
            continue
        if prev >= 0 and i > prev + 1:
            t.add_row("[dim]  ···[/]", "[dim]  ···[/]")
        if tag == "equal":
            t.add_row(f"[dim]{left}[/]", f"[dim]{right}[/]")
        elif tag == "delete":
            t.add_row(f"[bold red]- {left}[/]", "")
        elif tag == "insert":
            t.add_row("", f"[bold green]+ {right}[/]")
        else:
            t.add_row(f"[red]~ {left}[/]", f"[green]~ {right}[/]")
        prev = i

    console.print(Panel(t, title=title, border_style=border, expand=True))


def print_file_diffs(results: List[FileDiff], console: Any,
                    context: int = 3,
                    source_branch: str = "source",
                    target_branch: str = "target"):
    """Muestra diff lado-a-lado (o contenido completo) por cada archivo con diferencias."""
    diff_files = [r for r in results if r.change_type != CHANGE_NONE and
                  (r.diff_lines or r.source_content or r.target_content)]
    if not diff_files:
        return

    console.print()
    console.print("[bold cyan]══════════════════════ DETALLE POR ARCHIVO ══════════════════════[/]")

    for r in diff_files:
        sev_color       = SEV_COLOR.get(r.severity, "white")
        ch_style, ch_label = CHANGE_LABEL.get(r.change_type, ("dim", r.change_type))
        title = (
            f"{SEV_EMOJI.get(r.severity, '')} "
            f"[{sev_color}]{r.severity}[/{sev_color}]  │  "
            f"[{ch_style}]{ch_label}[/{ch_style}]  │  "
            f"[bold white]{r.path}[/]"
        )
        border = SEV_BORDER.get(r.severity, "dim")

        if r.change_type == CHANGE_ADD and r.source_content:
            syntax = Syntax(r.source_content, "yaml", theme="ansi_dark",
                            line_numbers=True, word_wrap=True)
            console.print(Panel(syntax,
                                title=f"[green]➕ NUEVO ARCHIVO[/]  {r.path}",
                                border_style="green", expand=False))
        elif r.change_type == CHANGE_DELETE and r.target_content:
            syntax = Syntax(r.target_content, "yaml", theme="ansi_dark",
                            line_numbers=True, word_wrap=True)
            console.print(Panel(syntax,
                                title=f"[red]🗑  ARCHIVO ELIMINADO[/]  {r.path}",
                                border_style="red", expand=False))
        elif r.source_content is not None or r.target_content is not None:
            _print_side_by_side(r, context, border, title, source_branch, target_branch, console)
        elif r.diff_lines:
            diff_text = "\n".join(r.diff_lines)
            syntax    = Syntax(diff_text, "diff", theme="ansi_dark",
                               line_numbers=False, word_wrap=True)
            console.print(Panel(syntax, title=title, border_style=border, expand=False))
        console.print()


def print_plain(
    results: List[FileDiff], source_branch: str, target_branch: str, component: str
):
    print(f"\n{'='*70}")
    print(f"COMPARACIÓN: {component}")
    print(f"Origen : {source_branch}")
    print(f"Destino: {target_branch}")
    print(f"{'='*70}")
    print(f"{'Archivo':<40} {'Cambio':<8} {'Severidad':<10} {'+':>5} {'-':>5}")
    print("-" * 70)
    for r in results:
        print(f"{r.filename:<40} {r.change_type:<8} {r.severity:<10}"
              f" {r.lines_added:>5} {r.lines_removed:>5}")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def export_results(
    results: List[FileDiff], fmt: str,
    org: str, project: str, repo_name: str,
    component: str, source_branch: str, target_branch: str,
    tz_name: str, console: Optional[Any],
):
    output_dir = get_output_dir("outcome")
    tz         = ZoneInfo(tz_name)
    ts         = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    safe_cmp   = component.replace("/", "_").replace("\\", "_")
    base       = f"properties_diff_{safe_cmp}_{ts}"

    summary = {s: sum(1 for r in results if r.severity == s)
               for s in (SEV_CRITICAL, SEV_HIGH, SEV_MEDIUM, SEV_LOW, SEV_NONE)}
    metadata = {
        "tool":          "azdo_repo_properties_branch_diff",
        "version":       __version__,
        "organization":  org,
        "project":       project,
        "repository":    repo_name,
        "component":     component,
        "source_branch": source_branch,
        "target_branch": target_branch,
        "generated_at":  datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S %Z"),
        "total_files":   len(results),
        "summary":       summary,
    }

    if fmt == "json":
        filepath = Path(output_dir) / f"{base}.json"
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({"metadata": metadata, "diffs": [r.to_dict() for r in results]},
                      f, indent=2, ensure_ascii=False)

    elif fmt == "csv":
        filepath = Path(output_dir) / f"{base}.csv"
        fields   = ["path", "filename", "change_type", "severity",
                    "lines_added", "lines_removed", "diff"]
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for r in results:
                w.writerow(r.to_dict())

    elif fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            msg = "openpyxl no instalado. pip install openpyxl"
            (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
            return

        filepath = Path(output_dir) / f"{base}.xlsx"
        wb       = openpyxl.Workbook()

        SEV_FILL = {
            SEV_CRITICAL: PatternFill("solid", fgColor="FF4444"),
            SEV_HIGH:     PatternFill("solid", fgColor="FF9999"),
            SEV_MEDIUM:   PatternFill("solid", fgColor="FFDD57"),
            SEV_LOW:      PatternFill("solid", fgColor="90CAF9"),
            SEV_NONE:     PatternFill("solid", fgColor="E0E0E0"),
        }
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(bold=True, color="FFFFFF")

        # ── Hoja 1: Resumen ────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"
        cols = ["Archivo", "Cambio", "Severidad", "+Líneas", "-Líneas", "Impacto", "Path"]
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(results, 2):
            row_vals = [
                r.filename, r.change_type.upper(),
                f"{SEV_EMOJI.get(r.severity,'')} {r.severity}",
                r.lines_added, r.lines_removed, _impact_desc(r), r.path,
            ]
            fill = SEV_FILL.get(r.severity, PatternFill())
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
        for ci in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 22

        # ── Hoja 2: Diff detallado ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Diff Detallado")
        for ci, h in enumerate(["Path", "Cambio", "Diff"], 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill, c.font = hdr_fill, hdr_font
        for ri, r in enumerate(results, 2):
            ws2.cell(row=ri, column=1, value=r.path)
            ws2.cell(row=ri, column=2, value=r.change_type.upper())
            ws2.cell(row=ri, column=3, value="\n".join(r.diff_lines) if r.diff_lines else "")
            ws2.cell(row=ri, column=3).alignment = Alignment(wrap_text=True)
        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["C"].width = 80

        # ── Hoja 3: Metadata ───────────────────────────────────────────────────
        ws3 = wb.create_sheet("Metadata")
        for ki, (key, val) in enumerate(metadata.items(), 1):
            ws3.cell(row=ki, column=1, value=key).font = Font(bold=True)
            ws3.cell(row=ki, column=2, value=str(val))

        wb.save(filepath)

    msg = f"📁 Exportado: {filepath}"
    (console.print(f"[bold green]{msg}[/]") if RICH_AVAILABLE and console else print(msg))


# ═══════════════════════════════════════════════════════════════════════════════
# PROMPTS INTERACTIVOS
# ═══════════════════════════════════════════════════════════════════════════════
def _prompt(text: str, default: str = "", console: Optional[Any] = None) -> str:
    if RICH_AVAILABLE and console:
        console.print(
            f"[bold]{text}[/] [dim](default: {default})[/]: " if default
            else f"[bold]{text}:[/] ",
            end="",
        )
    else:
        print(f"{text} [{default}]: " if default else f"{text}: ", end="")
    val = input().strip()
    return val or default


def prompt_select_repo(repos: List[Dict], console: Optional[Any]) -> Optional[Dict]:
    if RICH_AVAILABLE and console:
        t = Table(title="Repositorios disponibles", header_style="bold cyan", border_style="dim")
        t.add_column("#", justify="right", width=4)
        t.add_column("Repositorio", style="white")
        for i, r in enumerate(repos, 1):
            t.add_row(str(i), r["name"])
        console.print(t)
    else:
        print("\nRepositorios disponibles:")
        for i, r in enumerate(repos, 1):
            print(f"  {i:3}. {r['name']}")

    choice = _prompt("Seleccione # o nombre del repositorio de propiedades", console=console)
    if not choice:
        return None
    if choice.isdigit():
        idx = int(choice) - 1
        return repos[idx] if 0 <= idx < len(repos) else None
    exact = [r for r in repos if choice.lower() == r["name"].lower()]
    if exact:
        return exact[0]
    hits = [r for r in repos if choice.lower() in r["name"].lower()]
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1 and RICH_AVAILABLE and console:
        console.print(f"[yellow]Múltiples coincidencias con '{choice}': "
                      f"{[h['name'] for h in hits]}[/]")
    return None


def prompt_select_component(
    org: str, project: str, repo_id: str, source_branch: str,
    headers: Dict, debug: bool, console: Optional[Any],
) -> str:
    root_items = get_items_in_path(
        org, project, repo_id, "/", source_branch, headers, debug
    )
    folders = sorted({
        i["path"].lstrip("/").split("/")[0]
        for i in root_items
        if i.get("gitObjectType") in ("tree", "blob")
           and "/" in i["path"].lstrip("/")
    })

    if folders:
        if RICH_AVAILABLE and console:
            t = Table(title="Componentes disponibles", header_style="bold cyan", border_style="dim")
            t.add_column("#",          justify="right", width=4)
            t.add_column("Componente", style="cyan")
            for i, f in enumerate(folders, 1):
                t.add_row(str(i), f)
            console.print(t)
        else:
            print("\nComponentes disponibles:")
            for i, f in enumerate(folders, 1):
                print(f"  {i:3}. {f}")

    choice = _prompt("Nombre del componente (carpeta)", console=console)
    if choice.isdigit() and folders:
        idx = int(choice) - 1
        return folders[idx] if 0 <= idx < len(folders) else choice
    return choice


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main() -> int:
    start  = time.time()
    args   = get_args()
    console = Console() if RICH_AVAILABLE else None

    if not REQUESTS_AVAILABLE:
        print("ERROR: requests no instalado. pip install requests")
        return 1

    headers = make_headers(args.pat)

    if RICH_AVAILABLE and console:
        console.print(Panel(
            f"[bold cyan]Properties Branch Diff Analyzer[/]  "
            f"[dim]v{__version__} | {__author__}[/]\n"
            "[dim]Validación de configuración entre ramas para garantía "
            "de calidad en despliegues productivos[/]",
            border_style="cyan",
        ))

    # ── 1. Repositorio ──────────────────────────────────────────────────────────
    ctx = (Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console)
           if RICH_AVAILABLE and console else _NullCtx())
    with ctx as prog:
        t = prog.add_task("Obteniendo repositorios...", total=None) \
            if RICH_AVAILABLE and console else None
        repos = get_repositories(args.org, args.project, headers, args.debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

    if not repos:
        msg = "❌ No se pudieron obtener repositorios. Verifica org/project/PAT."
        (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
        return 1

    if args.repo:
        exact = [r for r in repos if args.repo.lower() == r["name"].lower()]
        if exact:
            selected_repo = exact[0]
        else:
            hits = [r for r in repos if args.repo.lower() in r["name"].lower()]
            if not hits:
                msg = f"❌ No se encontró repositorio con '{args.repo}'."
                (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
                return 1
            selected_repo = hits[0] if len(hits) == 1 else prompt_select_repo(hits, console)
    else:
        selected_repo = prompt_select_repo(repos, console)

    if not selected_repo:
        print("❌ No se seleccionó repositorio.")
        return 1

    repo_id, repo_name = selected_repo["id"], selected_repo["name"]
    if RICH_AVAILABLE and console:
        console.print(f"[green]✅ Repositorio:[/] [bold]{repo_name}[/]")

    # ── 2. Ramas ────────────────────────────────────────────────────────────────
    ctx = (Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console)
           if RICH_AVAILABLE and console else _NullCtx())
    with ctx as prog:
        t = prog.add_task("Obteniendo ramas...", total=None) \
            if RICH_AVAILABLE and console else None
        branch_names = get_branches(args.org, args.project, repo_id, headers, args.debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

    if not branch_names:
        msg = "❌ No se pudieron obtener ramas del repositorio."
        (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
        return 1

    if RICH_AVAILABLE and console:
        preview = branch_names[:8]
        suffix  = f" [dim]+{len(branch_names)-8} más[/]" if len(branch_names) > 8 else ""
        console.print(
            f"[dim]🌿 Ramas: {', '.join(f'[cyan]{b}[/]' for b in preview)}{suffix}[/]"
        )

    source_branch = args.source or _prompt(
        "Rama ORIGEN (lo que se desplegará)", "develop", console
    )
    target_branch = args.target or _prompt(
        "Rama DESTINO (entorno receptor)", "master", console
    )

    for br, label in [(source_branch, "origen"), (target_branch, "destino")]:
        if br not in branch_names:
            msg = f"⚠️  Rama {label} '{br}' no encontrada en el repositorio."
            (console.print(f"[yellow]{msg}[/]") if RICH_AVAILABLE and console else print(msg))

    if RICH_AVAILABLE and console:
        console.print(
            f"[green]✅ Ramas:[/] "
            f"[bold green]{source_branch}[/] → [bold yellow]{target_branch}[/]"
        )

    # ── 3. Componente ───────────────────────────────────────────────────────────
    if args.component:
        component = args.component.strip("/")
    else:
        if RICH_AVAILABLE and console:
            console.print("[dim]  📂 Listando estructura del repositorio...[/]")
        component = prompt_select_component(
            args.org, args.project, repo_id, source_branch,
            headers, args.debug, console,
        ).strip("/")

    if not component:
        msg = "❌ El componente no puede estar vacío."
        (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
        return 1

    component_path = f"/{component}"
    if RICH_AVAILABLE and console:
        console.print(f"[green]✅ Componente:[/] [bold]{component}[/]")

    # ── 4. Análisis ─────────────────────────────────────────────────────────────
    if RICH_AVAILABLE and console:
        console.print()
        console.print(
            f"[bold cyan]🔍 Analizando [{component}] — "
            f"{source_branch} vs {target_branch}...[/]"
        )
        console.print()

    results = analyze_component(
        args.org, args.project, repo_id, component_path,
        source_branch, target_branch,
        headers, args.context, args.debug, console,
    )

    # Filtros opcionales
    if args.severity:
        min_order = SEV_ORDER.get(args.severity, 0)
        results   = [r for r in results if SEV_ORDER.get(r.severity, 0) >= min_order]

    if args.only_diff:
        results = [r for r in results if r.change_type != CHANGE_NONE]

    # ── 5. Resultados ───────────────────────────────────────────────────────────
    if not results:
        msg = (f"⚪ Sin diferencias en [{component}] "
               f"entre [{source_branch}] y [{target_branch}]")
        if RICH_AVAILABLE and console:
            console.print(Panel(
                f"[green]{msg}[/]\n[dim]La configuración es idéntica en ambas ramas.[/]",
                border_style="green",
            ))
        else:
            print(msg)
        return 0

    if RICH_AVAILABLE and console:
        print_summary_table(results, source_branch, target_branch, component, console)
        if not args.no_content:
            print_file_diffs(results, console, args.context, source_branch, target_branch)
    else:
        print_plain(results, source_branch, target_branch, component)

    # ── 6. Exportar ─────────────────────────────────────────────────────────────
    if args.output:
        export_results(
            results, args.output,
            args.org, args.project, repo_name,
            component, source_branch, target_branch,
            args.timezone, console,
        )

    # ── 7. Footer ────────────────────────────────────────────────────────────────
    elapsed = time.time() - start
    if RICH_AVAILABLE and console:
        console.print(f"\n[dim]⏱  Tiempo de ejecución: {elapsed:.1f}s[/]")

    # Exit code según severidad máxima (útil como quality gate en CD)
    max_order = max((SEV_ORDER.get(r.severity, 0) for r in results), default=0)
    if max_order >= SEV_ORDER[SEV_CRITICAL]:
        return 2
    if max_order >= SEV_ORDER[SEV_HIGH]:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())

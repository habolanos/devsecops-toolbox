#!/usr/bin/env python3
"""
Pending Approvals Reporter v1.0.0
Lista releases con aprobaciones pendientes en Azure DevOps,
mostrando también el estado del stage "Validador".

Uso:
    python pending_approvals.py --org Soluciones-Corporativas --project Juridico
    python pending_approvals.py --org Coppel-Retail --project Compras.RMI

Autor: Harold Adrian (migrado desde Comercial/scripts/pending.py)
"""

import requests
import os
import sys
import argparse
from datetime import datetime
from pathlib import Path
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = lambda *a, **k: None  # noqa: E731

# --- Directorio de salida centralizado (DEVSECOPS_OUTPUT_DIR) ---
try:
    from utils import get_output_dir, resolve_output_path
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
    from datetime import datetime as _dt
    _FMT_EXT = {"excel": ".xlsx", "csv": ".csv", "json": ".json"}
    def resolve_output_path(output_arg, base_name, default_format="excel"):
        output_dir = get_output_dir("outcome")
        output_dir.mkdir(parents=True, exist_ok=True)
        ext = _FMT_EXT.get(default_format, ".xlsx")
        if not output_arg:
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        if output_arg.lower() in _FMT_EXT:
            ext = _FMT_EXT[output_arg.lower()]
            return str(output_dir / f"{base_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        p = _Path(output_arg)
        if p.suffix == "":
            p = p.with_suffix(ext)
        return str(p.resolve())
# -------------------------------------------------------------------

load_dotenv(Path(__file__).parent.parent / ".env")


class TeeWriter:
    """Escribe stdout a consola Y archivo de log simultáneamente."""
    def __init__(self, log_path):
        self.terminal = sys.__stdout__
        self.log = open(log_path, "w", encoding="utf-8")
        self.log_path = log_path

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        self.log.close()


def setup_logging(script_name):
    """Configura TeeWriter para que stdout vaya a consola + archivo log en outcome."""
    output_dir = get_output_dir("outcome")
    output_dir.mkdir(parents=True, exist_ok=True)
    log_path = output_dir / f"{script_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeWriter(log_path)
    sys.stdout = tee
    print(f"📝 Log: {log_path.resolve()}")
    print(f"📅 Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    return tee


def teardown_logging(tee):
    """Restaura stdout y cierra archivo de log."""
    print(f"\n📝 Log guardado: {tee.log_path.resolve()}")
    sys.stdout = tee.terminal
    tee.close()


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl no instalado. Ejecuta: pip install openpyxl")
    exit(1)

# ── Configuración por defecto ─────────────────────────────────────────
DEFAULT_ORG = "Soluciones-Corporativas"
DEFAULT_PROJECT = "Juridico"
API_VERSION = "7.1"


def get_headers(pat: str):
    """Genera headers de autenticación."""
    return {"Authorization": f"Basic :{pat}"}


def normalize_org(org: str) -> str:
    """Extrae el nombre de la organización desde una URL completa o nombre simple."""
    if org.startswith("http"):
        return org.rstrip("/").split("/")[-1]
    return org


def get_pending_approvals(org: str, project: str, pat: str):
    """Obtiene aprobaciones pendientes del proyecto."""
    base_url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release"
    url = f"{base_url}/approvals?statusFilter=pending&api-version={API_VERSION}"
    
    all_approvals = []
    next_url = url
    
    while next_url:
        resp = requests.get(next_url, auth=("", pat), timeout=30)
        resp.raise_for_status()
        data = resp.json()
        all_approvals.extend(data.get("value", []))
        next_url = data.get("nextLink")
    
    return all_approvals


def get_release_detail(release_id: int, org: str, project: str, pat: str):
    """Obtiene detalle completo de un release."""
    url = f"https://vsrm.dev.azure.com/{org}/{project}/_apis/release/releases/{release_id}?api-version={API_VERSION}"
    try:
        resp = requests.get(url, auth=("", pat), timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"   ⚠️ Error obteniendo release {release_id}: {e}")
        return None


def get_validador_status(release_id: int, org: str, project: str, pat: str):
    """Busca el stage 'Validador' y retorna su estado."""
    release = get_release_detail(release_id, org, project, pat)
    if not release:
        return "Error"
    
    for env in release.get("environments", []):
        if env.get("name", "").lower() == "validador":
            return env.get("status", "notStarted")
    return "No existe"


def export_to_excel(approvals: list, org: str, project: str, pat: str, filename: str = None):
    """Exporta aprobaciones a Excel."""
    if not filename or filename.lower() in ("excel", "csv", "json"):
        filename = resolve_output_path(filename, f"pending_approvals_{org}_{project}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Approvals"
    
    # Headers
    headers = [
        "Release Name", "Release ID", "Pipeline", "Stage", "Approver",
        "Created On", "Pending Since", "Validador Status", "Approval ID"
    ]
    ws.append(headers)
    
    # Colores para Validador
    VALIDADOR_COLORS = {
        "succeeded": "C6EFCE",      # Verde
        "inProgress": "FFEB9C",     # Amarillo
        "notStarted": "B8CCE4",    # Azul claro
        "rejected": "FFC7CE",       # Rojo
        "No existe": "D9D9D9",      # Gris
        "Error": "FF0000",          # Rojo fuerte
    }
    
    # Datos
    for approval in approvals:
        release_ref = approval.get("release", {})
        release_id = release_ref.get("id")
        release_name = release_ref.get("name", "N/A")
        
        stage = approval.get("releaseEnvironment", {}).get("name", "N/A")
        pipeline = approval.get("releaseDefinition", {}).get("name", "N/A")
        approver = approval.get("approver", {}).get("displayName", "N/A")
        created_on = release_ref.get("createdOn", "")[:10] if release_ref.get("createdOn") else "N/A"
        pending_since = approval.get("createdOn", "")[:10] if approval.get("createdOn") else "N/A"
        approval_id = approval.get("id", "N/A")
        
        # Obtener estado del Validador
        validador_status = get_validador_status(release_id, org, project, pat)
        
        row = [
            release_name, release_id, pipeline, stage, approver,
            created_on, pending_since, validador_status, approval_id
        ]
        ws.append(row)
    
    # Estilos
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0078D4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Colorear columna de Validador (columna H = 8)
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            status = cell.value
            if status in VALIDADOR_COLORS:
                cell.fill = PatternFill("solid", fgColor=VALIDADOR_COLORS[status])
                if status == "rejected":
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    # Ajustar anchos
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 40
    
    wb.save(filename)
    return filename


def main():
    parser = argparse.ArgumentParser(
        description="Pending Approvals Reporter - Releases con aprobaciones pendientes",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python pending_approvals.py
  python pending_approvals.py --org Soluciones-Corporativas --project Juridico
  python pending_approvals.py --org Coppel-Retail --project Compras.RMI --output aprobaciones.xlsx
        """
    )
    parser.add_argument("--org", default=DEFAULT_ORG,
                       help=f"Organización Azure DevOps (default: {DEFAULT_ORG})")
    parser.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                       help=f"Proyecto (default: {DEFAULT_PROJECT})")
    parser.add_argument("--pat", default=None,
                       help="PAT de Azure DevOps (default: env AZURE_PAT)")
    parser.add_argument("--output", "-o", default=None,
                       help="Nombre del archivo Excel de salida")
    
    args = parser.parse_args()
    
    # Obtener PAT
    pat = args.pat or os.getenv("AZURE_PAT", "")
    if not pat:
        print("❌ ERROR: Configura tu PAT via --pat o variable de entorno AZURE_PAT")
        exit(1)
    
    org = normalize_org(args.org)
    
    # Configurar logging a archivo en outcome
    tee = setup_logging("cicd_inventory_pending_approvals")

    print(f"🔍 Consultando aprobaciones pendientes...")
    print(f"   Org: {org}")
    print(f"   Project: {args.project}")
    print("=" * 60)
    
    approvals = get_pending_approvals(org, args.project, pat)
    
    if not approvals:
        print("\n✅ No hay aprobaciones pendientes.")
        return
    
    print(f"\n📝 {len(approvals)} aprobación(es) pendiente(s) encontrada(s)")
    
    filename = export_to_excel(approvals, org, args.project, pat, args.output)
    excel_path = Path(filename).resolve()
    print(f"\n✅ Archivo generado: {excel_path}")
    
    teardown_logging(tee)


if __name__ == "__main__":
    main()

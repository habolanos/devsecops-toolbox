#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
azdo_pipeline_drift.py

Compara el estado ACTUAL de cada Release Pipeline CD contra el snapshot
almacenado en el último release ejecutado, detectando "drift" en:

  ✦ B — Stage diff       → stages añadidos / eliminados en la definición
  ✦ C — Variable drift   → variables añadidas / eliminadas (solo keys, no valores)
  ✦ D — Approval drift   → cambios en gates de aprobación (crítico para compliance)
  ✦ F — Task diff        → tasks añadidas, eliminadas o con versión cambiada por stage

Severidad de drift:
  🚨 CRITICAL  → Approval gates cambiaron
  🔴 HIGH      → Stages o tasks añadidas/eliminadas
  🟡 MEDIUM    → Versión de task actualizada
  🔵 LOW       → Solo variables añadidas/eliminadas
  ⚪ NONE      → Sin drift detectado

Autor: Harold Adrian
"""

import argparse
import base64
import csv
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
    from rich.table import Table
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

__version__ = "1.0.0"
__author__  = "Harold Adrian"

# ═══════════════════════════════════════════════════════════════════════════════
# DEFAULTS / CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_ORG_URL  = "https://dev.azure.com/Coppel-Retail"
DEFAULT_PROJECT  = "Compras.RMI"
DEFAULT_TIMEZONE = "America/Mazatlan"
API_VERSION_DEFS = "7.2-preview.4"
API_VERSION_RELS = "7.2-preview.8"
DEFAULT_THREADS  = 8

SEV_NONE     = "NONE"
SEV_LOW      = "LOW"
SEV_MEDIUM   = "MEDIUM"
SEV_HIGH     = "HIGH"
SEV_CRITICAL = "CRITICAL"

SEV_ORDER = {SEV_NONE: 0, SEV_LOW: 1, SEV_MEDIUM: 2, SEV_HIGH: 3, SEV_CRITICAL: 4}

SEV_EMOJI = {
    SEV_NONE:     "⚪",
    SEV_LOW:      "🔵",
    SEV_MEDIUM:   "🟡",
    SEV_HIGH:     "🔴",
    SEV_CRITICAL: "🚨",
}

SEV_BORDER = {
    SEV_NONE:     "dim",
    SEV_LOW:      "blue",
    SEV_MEDIUM:   "yellow",
    SEV_HIGH:     "red",
    SEV_CRITICAL: "red",
}


# ═══════════════════════════════════════════════════════════════════════════════
# ARGS
# ═══════════════════════════════════════════════════════════════════════════════
def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Detecta drift entre Release Pipeline CD actual y último release ejecutado"
    )
    p.add_argument("--org", "-g", default=DEFAULT_ORG_URL,
                   help=f"URL de la organización (default: {DEFAULT_ORG_URL})")
    p.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                   help=f"Nombre del proyecto (default: {DEFAULT_PROJECT})")
    p.add_argument("--pat", required=True,
                   help="Personal Access Token con permiso Release (Read)")
    p.add_argument("--filter", "--repo", "-f", "-r", dest="filter", default=None,
                   help="Filtrar pipelines por nombre/repo (substring, case insensitive)")
    p.add_argument("--severity", "-s",
                   choices=["NONE", "LOW", "MEDIUM", "HIGH", "CRITICAL"], default=None,
                   help="Mostrar solo pipelines con severidad >= especificada")
    p.add_argument("--output", "-o", choices=["json", "csv", "excel"], default=None,
                   help="Exportar resultados (json / csv / excel)")
    p.add_argument("--timezone", "-tz", default=DEFAULT_TIMEZONE,
                   help=f"Zona horaria para fechas (default: {DEFAULT_TIMEZONE})")
    p.add_argument("--threads", type=int, default=DEFAULT_THREADS,
                   help=f"Hilos paralelos (default: {DEFAULT_THREADS})")
    p.add_argument("--sort", choices=["severity", "name", "gap"], default="severity",
                   help="Ordenar tabla por (default: severity desc)")
    p.add_argument("--debug", action="store_true",
                   help="Mostrar errores HTTP detallados")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP
# ═══════════════════════════════════════════════════════════════════════════════
def make_headers(pat: str) -> Dict:
    token = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Accept": "application/json"}


def api_get(
    url: str, headers: Dict, params: Dict = None, debug: bool = False
) -> Optional[Any]:
    try:
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code >= 400:
            _api_label = url.split("/_apis")[0].split("/")[-1] if "/_apis" in url else url
            print(f"  ⚠  HTTP {r.status_code} ({_api_label})")
            if debug:
                print(f"[DEBUG] URL: {url}")
                print(f"[DEBUG] Body: {r.text[:400]}")
        r.raise_for_status()
        return r.json()
    except Exception as e:
        if debug:
            print(f"[DEBUG] {url}: {e}")
        return None


def vsrm_base(org_url: str) -> str:
    return org_url.replace("dev.azure.com", "vsrm.dev.azure.com")


def fmt_date(dt: Optional[datetime], tz_name: str) -> str:
    if not dt:
        return "—"
    try:
        return dt.astimezone(ZoneInfo(tz_name)).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return dt.strftime("%Y-%m-%d %H:%M")


def parse_iso(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _dt_default(obj: Any) -> str:
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError(f"Not serializable: {type(obj)}")


# ═══════════════════════════════════════════════════════════════════════════════
# API CALLS
# ═══════════════════════════════════════════════════════════════════════════════
def get_release_definitions(
    org: str, project: str, headers: Dict, debug: bool
) -> List[Dict]:
    vsrm = vsrm_base(org)
    url  = f"{vsrm}/{quote(project, safe='')}/_apis/release/definitions"
    data = api_get(url, headers, {"api-version": API_VERSION_DEFS, "$top": 500}, debug)
    return (data or {}).get("value", [])


def get_release_definition_detail(
    org: str, project: str, def_id: int, headers: Dict, debug: bool
) -> Optional[Dict]:
    vsrm = vsrm_base(org)
    url  = f"{vsrm}/{quote(project, safe='')}/_apis/release/definitions/{def_id}"
    return api_get(url, headers,
                   {"api-version": API_VERSION_DEFS, "$expand": "environments"}, debug)


def get_last_release_full(
    org: str, project: str, def_id: int, headers: Dict, debug: bool
) -> Optional[Dict]:
    """
    Obtiene el último release con snapshot completo de environments.
    El snapshot incluye deployPhases.workflowTasks (task comparison F).
    """
    vsrm = vsrm_base(org)

    list_data = api_get(
        f"{vsrm}/{quote(project, safe='')}/_apis/release/releases", headers,
        {
            "api-version": API_VERSION_RELS,
            "definitionId": def_id,
            "$top": 1,
            "$orderby": "createdOn desc",
        },
        debug,
    )
    releases = (list_data or {}).get("value", [])
    if not releases:
        return None

    rel_id = releases[0]["id"]
    return api_get(
        f"{vsrm}/{quote(project, safe='')}/_apis/release/releases/{rel_id}",
        headers,
        {"api-version": API_VERSION_RELS},
        debug,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# COMPARISON HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _extract_tasks(env: Dict) -> List[Dict]:
    """
    Extrae workflowTasks de un environment, probando múltiples rutas:
      1. env.deployPhases[].workflowTasks             (definición actual)
      2. env.releaseDefinitionEnvironment.deployPhases (snapshot del release)
    """
    tasks: List[Dict] = []
    for phase in env.get("deployPhases", []):
        tasks.extend(phase.get("workflowTasks", []))
    if not tasks:
        inner = env.get("releaseDefinitionEnvironment", {})
        for phase in inner.get("deployPhases", []):
            tasks.extend(phase.get("workflowTasks", []))
    return tasks


def compare_stages(current_def: Dict, snapshot: Dict) -> Dict:
    cur   = {e["name"] for e in current_def.get("environments", [])}
    snap  = {e["name"] for e in snapshot.get("environments", [])}
    return {
        "added":   sorted(cur  - snap),
        "removed": sorted(snap - cur),
        "common":  sorted(cur  & snap),
    }


def compare_var_keys(cur_vars: Dict, snap_vars: Dict) -> Dict:
    cur  = set(cur_vars.keys())
    snap = set(snap_vars.keys())
    return {"added": sorted(cur - snap), "removed": sorted(snap - cur)}


def compare_approvals(cur_env: Dict, snap_env: Dict) -> Tuple[bool, str]:
    """
    Compara preDeployApprovals entre definición actual y snapshot del release.
    Retorna (changed: bool, detalle: str).
    """
    def _sig(env: Dict) -> Dict:
        pre      = env.get("preDeployApprovals", {})
        opts     = pre.get("options", {})
        approvals = pre.get("approvals", [])
        manual   = [a for a in approvals if not a.get("isAutomated", True)]
        return {
            "count":         len(manual),
            "req_count":     opts.get("requiredApproverCount", 0),
            "creator_can":   opts.get("releaseCreatorCanBeApprover", False),
            "approver_ids":  sorted(
                str(a.get("approver", {}).get("id", "")) for a in manual
            ),
        }

    cur_sig  = _sig(cur_env)
    snap_sig = _sig(snap_env)

    if cur_sig == snap_sig:
        return False, "Sin cambios"

    details: List[str] = []
    if cur_sig["count"] != snap_sig["count"]:
        details.append(f"Aprobadores: {snap_sig['count']} → {cur_sig['count']}")
    if cur_sig["req_count"] != snap_sig["req_count"]:
        details.append(
            f"Mínimo requerido: {snap_sig['req_count']} → {cur_sig['req_count']}"
        )
    if cur_sig["creator_can"] != snap_sig["creator_can"]:
        details.append(
            f"Creator puede aprobar: {snap_sig['creator_can']} → {cur_sig['creator_can']}"
        )
    if cur_sig["approver_ids"] != snap_sig["approver_ids"]:
        details.append("Lista de aprobadores cambió")

    return True, " · ".join(details) if details else "Cambios en aprobaciones"


def compare_tasks(cur_env: Dict, snap_env: Dict) -> Dict:
    """
    Compara workflowTasks entre la definición actual y el snapshot del release.
    Index por taskId (o name como fallback).
    """
    cur_tasks  = _extract_tasks(cur_env)
    snap_tasks = _extract_tasks(snap_env)
    available  = bool(cur_tasks or snap_tasks)

    def _idx(tasks: List[Dict]) -> Dict:
        return {t.get("taskId") or t.get("name", ""): t for t in tasks}

    cur_idx  = _idx(cur_tasks)
    snap_idx = _idx(snap_tasks)
    cur_ids  = set(cur_idx)
    snap_ids = set(snap_idx)

    added   = [cur_idx[k]["name"] for k in sorted(cur_ids  - snap_ids)]
    removed = [snap_idx[k]["name"] for k in sorted(snap_ids - cur_ids)]

    version_changed: List[str] = []
    for k in sorted(cur_ids & snap_ids):
        cv = cur_idx[k].get("version", "")
        sv = snap_idx[k].get("version", "")
        if cv != sv:
            version_changed.append(f"{cur_idx[k]['name']}: {sv} → {cv}")

    return {
        "added":           added,
        "removed":         removed,
        "version_changed": version_changed,
        "available":       available,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SEVERITY
# ═══════════════════════════════════════════════════════════════════════════════
def compute_severity(result: Dict) -> str:
    if not result.get("has_drift"):
        return SEV_NONE

    for sd in result.get("stage_diffs", {}).values():
        if sd.get("approvals_changed"):
            return SEV_CRITICAL

    sd_top = result.get("stages_diff", {})
    if sd_top.get("added") or sd_top.get("removed"):
        return SEV_HIGH
    for sd in result.get("stage_diffs", {}).values():
        if sd.get("tasks_added") or sd.get("tasks_removed"):
            return SEV_HIGH

    for sd in result.get("stage_diffs", {}).values():
        if sd.get("tasks_version_changed"):
            return SEV_MEDIUM

    return SEV_LOW


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ANALYSIS PER PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════
def analyze_pipeline_drift(
    summary: Dict,
    org: str, project: str, headers: Dict, debug: bool,
) -> Dict:
    def_id   = summary["id"]
    def_name = summary.get("name", str(def_id))

    result: Dict = {
        "id":                def_id,
        "name":              def_name,
        "current_revision":  None,
        "snapshot_revision": None,
        "revision_gap":      None,
        "last_release_id":   None,
        "last_release_name": None,
        "last_release_date": None,
        "stages_diff":       {"added": [], "removed": [], "common": []},
        "pipeline_vars_diff": {"added": [], "removed": []},
        "stage_diffs":       {},
        "has_drift":         False,
        "severity":          SEV_NONE,
        "error":             None,
    }

    current_def = get_release_definition_detail(org, project, def_id, headers, debug)
    if not current_def:
        result["error"] = "No se pudo obtener la definición"
        return result

    result["current_revision"] = current_def.get("revision")

    last_release = get_last_release_full(org, project, def_id, headers, debug)
    if not last_release:
        result["error"] = "Sin releases ejecutados"
        return result

    snap_rev = last_release.get("releaseDefinitionRevision")
    result["snapshot_revision"]  = snap_rev
    result["last_release_id"]    = last_release.get("id")
    result["last_release_name"]  = last_release.get("name", "")
    result["last_release_date"]  = parse_iso(last_release.get("createdOn"))
    result["revision_gap"]       = (result["current_revision"] or 0) - (snap_rev or 0)

    result["stages_diff"] = compare_stages(current_def, last_release)

    result["pipeline_vars_diff"] = compare_var_keys(
        current_def.get("variables", {}),
        last_release.get("variables", {}),
    )

    cur_env_idx  = {e["name"]: e for e in current_def.get("environments", [])}
    snap_env_idx = {e["name"]: e for e in last_release.get("environments", [])}

    stage_diffs: Dict = {}
    for stage_name in result["stages_diff"]["common"]:
        cur_env  = cur_env_idx.get(stage_name, {})
        snap_env = snap_env_idx.get(stage_name, {})

        vars_diff             = compare_var_keys(
            cur_env.get("variables", {}), snap_env.get("variables", {})
        )
        approvals_changed, approval_detail = compare_approvals(cur_env, snap_env)
        tasks_diff            = compare_tasks(cur_env, snap_env)

        stage_diffs[stage_name] = {
            "vars_added":            vars_diff["added"],
            "vars_removed":          vars_diff["removed"],
            "approvals_changed":     approvals_changed,
            "approval_detail":       approval_detail,
            "tasks_added":           tasks_diff["added"],
            "tasks_removed":         tasks_diff["removed"],
            "tasks_version_changed": tasks_diff["version_changed"],
            "tasks_available":       tasks_diff["available"],
        }

    result["stage_diffs"] = stage_diffs

    pd = result["pipeline_vars_diff"]
    has_drift = bool(
        result["stages_diff"]["added"] or result["stages_diff"]["removed"]
        or pd["added"] or pd["removed"]
        or (result["revision_gap"] and result["revision_gap"] > 0)
    )
    if not has_drift:
        for sd in stage_diffs.values():
            if (sd["vars_added"] or sd["vars_removed"]
                    or sd["approvals_changed"]
                    or sd["tasks_added"] or sd["tasks_removed"]
                    or sd["tasks_version_changed"]):
                has_drift = True
                break

    result["has_drift"] = has_drift
    result["severity"]  = compute_severity(result)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# RICH OUTPUT — SUMMARY TABLE
# ═══════════════════════════════════════════════════════════════════════════════
def print_summary_table(console: "Console", results: List[Dict], tz_name: str):
    t = Table(
        title="🔍 Pipeline Drift Report",
        box=box.ROUNDED, border_style="cyan",
        show_header=True, header_style="bold cyan",
    )
    t.add_column("#",             justify="right",  style="dim",   width=4)
    t.add_column("Pipeline",      justify="left",   style="white", min_width=28)
    t.add_column("Rev\nGap",      justify="center", width=6)
    t.add_column("Stages Δ",      justify="center", width=9)
    t.add_column("Vars Δ",        justify="center", width=8)
    t.add_column("Approvals Δ",   justify="center", width=12)
    t.add_column("Tasks Δ",       justify="center", width=10)
    t.add_column("Último Release", justify="center", width=17)
    t.add_column("Severity",      justify="center", width=12)

    for idx, r in enumerate(results, 1):
        emoji = SEV_EMOJI.get(r["severity"], "")
        gap   = str(r["revision_gap"]) if r["revision_gap"] is not None else "—"

        s_add = len(r["stages_diff"]["added"])
        s_rem = len(r["stages_diff"]["removed"])
        stages_d = (f"[green]+{s_add}[/][red]-{s_rem}[/]"
                    if s_add or s_rem else "[dim]—[/]")

        pv    = r["pipeline_vars_diff"]
        va    = len(pv.get("added", []))   + sum(len(sd.get("vars_added",   [])) for sd in r["stage_diffs"].values())
        vr    = len(pv.get("removed", [])) + sum(len(sd.get("vars_removed", [])) for sd in r["stage_diffs"].values())
        vars_d = (f"[green]+{va}[/][red]-{vr}[/]" if va or vr else "[dim]—[/]")

        app_ch = sum(1 for sd in r["stage_diffs"].values() if sd.get("approvals_changed"))
        app_d  = f"[bold red]{app_ch} stage(s)[/]" if app_ch else "[dim]—[/]"

        ta = sum(len(sd.get("tasks_added",   [])) for sd in r["stage_diffs"].values())
        tr = sum(len(sd.get("tasks_removed", [])) for sd in r["stage_diffs"].values())
        tv = sum(len(sd.get("tasks_version_changed", [])) for sd in r["stage_diffs"].values())
        tasks_d = (f"[green]+{ta}[/][red]-{tr}[/][yellow]~{tv}[/]"
                   if ta or tr or tv else "[dim]—[/]")

        gap_style = f"[yellow]{gap}[/]" if gap not in ("0", "—") else f"[dim]{gap}[/]"
        sev_style = {
            SEV_CRITICAL: f"[bold reverse red]{emoji} {r['severity']}[/]",
            SEV_HIGH:     f"[bold red]{emoji} {r['severity']}[/]",
            SEV_MEDIUM:   f"[yellow]{emoji} {r['severity']}[/]",
            SEV_LOW:      f"[cyan]{emoji} {r['severity']}[/]",
            SEV_NONE:     f"[dim]{emoji} {r['severity']}[/]",
        }.get(r["severity"], r["severity"])

        t.add_row(
            str(idx),
            r["name"],
            gap_style,
            stages_d, vars_d, app_d, tasks_d,
            f"[dim]{fmt_date(r['last_release_date'], tz_name)}[/]",
            sev_style,
        )

    console.print(t)


# ═══════════════════════════════════════════════════════════════════════════════
# RICH OUTPUT — PER-PIPELINE DETAIL
# ═══════════════════════════════════════════════════════════════════════════════
def print_pipeline_detail(console: "Console", r: Dict, tz_name: str):
    if r["severity"] == SEV_NONE:
        return

    lines: List[str] = [
        f"{SEV_EMOJI.get(r['severity'], '')} [bold]{r['name']}[/]  "
        f"[dim]Rev actual:[/] [white]{r['current_revision']}[/]  "
        f"[dim]Rev snapshot:[/] [white]{r['snapshot_revision']}[/]  "
        f"[dim]Gap:[/] [yellow]{r['revision_gap']}[/]  "
        f"[dim]Último:[/] [cyan]{r['last_release_name']}[/] "
        f"[dim]({fmt_date(r['last_release_date'], tz_name)})[/]",
        "",
    ]

    sd = r["stages_diff"]
    if sd["added"] or sd["removed"]:
        lines.append("[bold underline]STAGES[/]")
        for s in sd["added"]:
            lines.append(f"  [green]➕  {s}[/]  [dim]añadido en def actual — sin deploy[/]")
        for s in sd["removed"]:
            lines.append(f"  [red]🗑️   {s}[/]  [dim]eliminado de la def — existía en release[/]")
        lines.append("")

    pv = r["pipeline_vars_diff"]
    if pv["added"] or pv["removed"]:
        lines.append("[bold underline]VARIABLES PIPELINE[/]")
        for v in pv["added"]:
            lines.append(f"  [green]➕  {v}[/]")
        for v in pv["removed"]:
            lines.append(f"  [red]🗑️   {v}[/]")
        lines.append("")

    has_stage_block = False
    for stage_name, sdd in r["stage_diffs"].items():
        stage_lines: List[str] = []

        if sdd.get("approvals_changed"):
            stage_lines.append(
                f"    [bold red]🚨 APPROVALS:[/] {sdd['approval_detail']}"
            )
        for v in sdd.get("vars_added", []):
            stage_lines.append(f"    [green]➕  VAR: {v}[/]")
        for v in sdd.get("vars_removed", []):
            stage_lines.append(f"    [red]🗑️   VAR: {v}[/]")
        for t in sdd.get("tasks_added", []):
            stage_lines.append(f"    [green]➕  TASK: {t}[/]")
        for t in sdd.get("tasks_removed", []):
            stage_lines.append(f"    [red]🗑️   TASK: {t}[/]")
        for t in sdd.get("tasks_version_changed", []):
            stage_lines.append(f"    [yellow]🔄  TASK VERSION: {t}[/]")
        if not sdd.get("tasks_available") and (sdd.get("tasks_added") is not None):
            stage_lines.append(
                "    [dim]ℹ️  Tasks: snapshot no disponible (release sin workflowTasks)[/]"
            )

        if stage_lines:
            if not has_stage_block:
                lines.append("[bold underline]STAGES — DETALLE[/]")
                has_stage_block = True
            lines.append(f"  [bold cyan]┌ {stage_name}[/]")
            lines.extend(stage_lines)
            lines.append("")

    border = SEV_BORDER.get(r["severity"], "dim")
    console.print(Panel("\n".join(lines), border_style=border, expand=False))


# ═══════════════════════════════════════════════════════════════════════════════
# RICH OUTPUT — FULL REPORT
# ═══════════════════════════════════════════════════════════════════════════════
def print_drift_report(
    console: "Console", results: List[Dict], tz_name: str, elapsed: float
):
    print_summary_table(console, results, tz_name)
    console.print()

    drifted = [r for r in results if r["has_drift"] and not r.get("error")]
    if drifted:
        console.print(Panel(
            "[bold cyan]🔍 Detalle de Drift por Pipeline[/]",
            border_style="cyan", expand=False,
        ))
        console.print()
        for r in drifted:
            print_pipeline_detail(console, r, tz_name)

    total    = len(results)
    critical = sum(1 for r in results if r["severity"] == SEV_CRITICAL)
    high     = sum(1 for r in results if r["severity"] == SEV_HIGH)
    medium   = sum(1 for r in results if r["severity"] == SEV_MEDIUM)
    low      = sum(1 for r in results if r["severity"] == SEV_LOW)
    none_    = sum(1 for r in results if r["severity"] == SEV_NONE)
    errors   = sum(1 for r in results if r.get("error"))

    console.print(Panel(
        f"[bold white]📊 Pipelines analizados:[/] [cyan]{total}[/]\n\n"
        f"[bold]Distribución de severidad:[/]\n"
        f"  [bold red]🚨 CRITICAL: [/] {critical}  [dim](approvals cambiaron)[/]\n"
        f"  [red]🔴 HIGH:     [/] {high}  [dim](stages o tasks añadidas/eliminadas)[/]\n"
        f"  [yellow]🟡 MEDIUM:   [/] {medium}  [dim](versión de task actualizada)[/]\n"
        f"  [cyan]🔵 LOW:      [/] {low}  [dim](solo variables)[/]\n"
        f"  [dim]⚪ NONE:     [/] {none_}  [dim](sin drift)[/]\n"
        f"  [dim]⚠️  Errores:  [/] {errors}\n\n"
        f"[bold]Columnas tabla:[/]\n"
        f"[dim]  Rev Gap → revisiones del pipeline sin desplegar\n"
        f"  Stages Δ → +añadidos / -eliminados desde último release\n"
        f"  Vars Δ   → +añadidas / -eliminadas (keys, no valores)\n"
        f"  Approvals Δ → stages con gates de aprobación cambiados\n"
        f"  Tasks Δ → +añadidas / -eliminadas / ~versión cambiada[/]\n\n"
        f"[dim]⏱️  Tiempo total: {elapsed:.2f}s[/]",
        title="📋 Resumen — Pipeline Drift Report",
        border_style="blue",
        expand=False,
    ))


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def _flatten(r: Dict) -> Dict:
    pd = r.get("pipeline_vars_diff", {})
    ta = sum(len(sd.get("tasks_added",           [])) for sd in r.get("stage_diffs", {}).values())
    tr = sum(len(sd.get("tasks_removed",         [])) for sd in r.get("stage_diffs", {}).values())
    tv = sum(len(sd.get("tasks_version_changed", [])) for sd in r.get("stage_diffs", {}).values())
    va = len(pd.get("added",   [])) + sum(len(sd.get("vars_added",   [])) for sd in r.get("stage_diffs", {}).values())
    vr = len(pd.get("removed", [])) + sum(len(sd.get("vars_removed", [])) for sd in r.get("stage_diffs", {}).values())
    app_stages = [n for n, sd in r.get("stage_diffs", {}).items() if sd.get("approvals_changed")]

    return {
        "id":                         r["id"],
        "pipeline_name":              r["name"],
        "current_revision":           r.get("current_revision", ""),
        "snapshot_revision":          r.get("snapshot_revision", ""),
        "revision_gap":               r.get("revision_gap", ""),
        "last_release_id":            r.get("last_release_id", ""),
        "last_release_name":          r.get("last_release_name", ""),
        "last_release_date":          r["last_release_date"].isoformat() if r.get("last_release_date") else "",
        "severity":                   r["severity"],
        "has_drift":                  r["has_drift"],
        "stages_added":               ", ".join(r["stages_diff"]["added"]),
        "stages_removed":             ", ".join(r["stages_diff"]["removed"]),
        "total_vars_added":           va,
        "total_vars_removed":         vr,
        "approval_stages_changed":    ", ".join(app_stages),
        "total_tasks_added":          ta,
        "total_tasks_removed":        tr,
        "total_tasks_version_changed": tv,
        "error":                      r.get("error", ""),
    }


def export_results(
    results: List[Dict], fmt: str, script_dir: str, tz_name: str
) -> Optional[str]:
    outcome_dir = os.path.join(script_dir, "outcome")
    os.makedirs(outcome_dir, exist_ok=True)
    ts   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    flat = [_flatten(r) for r in results]

    if fmt == "json":
        filepath = os.path.join(outcome_dir, f"pipeline_drift_{ts}.json")
        payload  = {
            "metadata": {
                "tool":         "azdo_pipeline_drift",
                "version":      __version__,
                "generated_at": datetime.now(ZoneInfo(tz_name)).isoformat(),
            },
            "total": len(results),
            "summary": flat,
        }
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, default=_dt_default, ensure_ascii=False)
        return filepath

    elif fmt == "csv":
        if not flat:
            return None
        filepath = os.path.join(outcome_dir, f"pipeline_drift_{ts}.csv")
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(flat[0].keys()))
            w.writeheader()
            w.writerows(flat)
        return filepath

    elif fmt == "excel":
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            filepath = os.path.join(outcome_dir, f"pipeline_drift_{ts}.xlsx")
            df = pd.DataFrame(flat)
            df.to_excel(filepath, index=False, engine="openpyxl", sheet_name="Drift Report")

            wb = load_workbook(filepath)
            ws = wb["Drift Report"]

            sev_fills = {
                SEV_CRITICAL: PatternFill("solid", fgColor="FF4444"),
                SEV_HIGH:     PatternFill("solid", fgColor="FF8C42"),
                SEV_MEDIUM:   PatternFill("solid", fgColor="FFD166"),
                SEV_LOW:      PatternFill("solid", fgColor="87CEEB"),
                SEV_NONE:     PatternFill("solid", fgColor="90EE90"),
            }

            headers = [cell.value for cell in ws[1]]
            sev_col    = (headers.index("severity")  + 1) if "severity"  in headers else None
            drift_col  = (headers.index("has_drift") + 1) if "has_drift" in headers else None

            for row_idx in range(2, len(flat) + 2):
                if sev_col:
                    cell = ws.cell(row=row_idx, column=sev_col)
                    fill = sev_fills.get(str(cell.value))
                    if fill:
                        cell.fill = fill
                        cell.font = Font(bold=True)
                if drift_col:
                    cell = ws.cell(row=row_idx, column=drift_col)
                    cell.alignment = Alignment(horizontal="center")

            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22

            wb.save(filepath)
            return filepath

        except ImportError:
            print("ERROR: pip install pandas openpyxl")
            return None

    return None


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    start_time = time.time()

    if not REQUESTS_AVAILABLE:
        print("ERROR: pip install requests rich pandas openpyxl")
        return

    args    = get_args()
    console = Console() if RICH_AVAILABLE else None
    headers = make_headers(args.pat)

    tz_name = args.timezone
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = DEFAULT_TIMEZONE

    rev_time = datetime.now(ZoneInfo(tz_name)).strftime(f"%Y-%m-%d %H:%M:%S ({tz_name})")

    if RICH_AVAILABLE and console:
        console.print()
        console.print(Panel(
            f"[bold cyan]🔍 Pipeline Drift Analyzer[/]\n"
            f"[dim]🕐 {rev_time}[/]\n"
            f"[dim]🏢 Org:      {args.org}[/]\n"
            f"[dim]📁 Proyecto: {args.project}[/]\n"
            f"[dim]📐 Dimensiones: B-Stages · C-Variables · D-Approvals · F-Tasks[/]\n"
            f"[dim]⚡ Severidad:   NONE < LOW < MEDIUM < HIGH < CRITICAL[/]",
            border_style="cyan", expand=False,
        ))
        console.print()

    if RICH_AVAILABLE and console:
        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      console=console) as prog:
            t = prog.add_task("Obteniendo release definitions...", total=None)
            summaries = get_release_definitions(args.org, args.project, headers, args.debug)
            prog.update(t, description=f"✅ {len(summaries)} release pipelines encontrados")
    else:
        summaries = get_release_definitions(args.org, args.project, headers, args.debug)
        print(f"{len(summaries)} release pipelines encontrados")

    if not summaries:
        msg = "❌ Sin release definitions. Verifica org, proyecto, PAT y permisos Release (Read)."
        (console.print(f"[red]{msg}[/]") if console else print(msg))
        return

    if args.filter:
        summaries = [s for s in summaries
                     if args.filter.lower() in s.get("name", "").lower()]
        if console:
            console.print(
                f"[dim]🔍 Filtrado: {len(summaries)} pipelines con '{args.filter}'[/]"
            )

    if not summaries:
        (console.print("[yellow]⚠️ Sin resultados tras el filtro.[/]") if console
         else print("Sin resultados."))
        return

    console.print() if console else None

    results: List[Dict] = []

    if RICH_AVAILABLE and console:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(),
            console=console,
        ) as prog:
            task = prog.add_task(
                f"Analizando drift en {len(summaries)} pipelines...",
                total=len(summaries),
            )
            with ThreadPoolExecutor(max_workers=args.threads) as exe:
                futures = {
                    exe.submit(
                        analyze_pipeline_drift,
                        s, args.org, args.project, headers, args.debug,
                    ): s
                    for s in summaries
                }
                for fut in as_completed(futures):
                    try:
                        results.append(fut.result())
                    except Exception:
                        pass
                    prog.advance(task)
    else:
        for s in summaries:
            results.append(
                analyze_pipeline_drift(s, args.org, args.project, headers, args.debug)
            )

    if not results:
        (console.print("[yellow]⚠️ No se pudieron analizar pipelines.[/]") if console
         else print("Sin datos."))
        return

    if args.severity:
        min_order = SEV_ORDER.get(args.severity, 0)
        results   = [r for r in results if SEV_ORDER.get(r["severity"], 0) >= min_order]
        if console:
            console.print(
                f"[dim]🔍 Severidad >= {args.severity}: {len(results)} pipelines[/]"
            )

    if args.sort == "severity":
        results.sort(key=lambda r: SEV_ORDER.get(r["severity"], 0), reverse=True)
    elif args.sort == "name":
        results.sort(key=lambda r: r["name"].lower())
    elif args.sort == "gap":
        results.sort(key=lambda r: (r.get("revision_gap") or 0), reverse=True)

    elapsed = time.time() - start_time

    if RICH_AVAILABLE and console:
        print_drift_report(console, results, tz_name, elapsed)
    else:
        hdr = f"{'Sev':<10} {'Pipeline':<40} {'Gap':>4}  {'S+':>3} {'S-':>3}  {'App':>4}  {'T+':>3} {'T-':>3} {'T~':>3}"
        print(f"\n{'='*len(hdr)}\n{hdr}\n{'='*len(hdr)}")
        for r in results:
            s_add = len(r["stages_diff"]["added"])
            s_rem = len(r["stages_diff"]["removed"])
            app   = sum(1 for sd in r["stage_diffs"].values() if sd.get("approvals_changed"))
            ta    = sum(len(sd.get("tasks_added",   [])) for sd in r["stage_diffs"].values())
            tr    = sum(len(sd.get("tasks_removed", [])) for sd in r["stage_diffs"].values())
            tv    = sum(len(sd.get("tasks_version_changed", [])) for sd in r["stage_diffs"].values())
            gap   = str(r.get("revision_gap", "—"))
            print(f"{r['severity']:<10} {r['name']:<40} {gap:>4}  {s_add:>3} {s_rem:>3}  {app:>4}  {ta:>3} {tr:>3} {tv:>3}")
        print(f"\nTotal: {len(results)} | Tiempo: {elapsed:.2f}s")

    if args.output:
        fp = export_results(
            results, args.output,
            os.path.dirname(os.path.abspath(__file__)),
            tz_name,
        )
        if fp:
            msg = f"📁 Exportado: {fp}"
            (console.print(f"[bold green]{msg}[/]\n") if console else print(msg))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
azdo_repo_branch_diff.py

Analiza el impacto de los cambios entre dos ramas de cualquier repositorio
en Azure DevOps.  Genera un informe ejecutivo de riesgo orientado a garantizar
la calidad y trazabilidad de un despliegue a producción.

Secciones del informe:
  1. Resumen Ejecutivo  — repo / ramas / commits adelante-atrás / score de impacto
  2. Distribución de Riesgo — archivos por categoría y nivel de riesgo
  3. Archivos Modificados   — tabla por archivo con riesgo, categoría y cambio
  4. Commits incluidos      — SHA / autor / fecha / mensaje
  5. Estadísticas por Autor — commits por colaborador
  6. Recomendaciones        — acciones sugeridas según los hallazgos

Clasificación de riesgo por archivo:
  🚨 CRITICAL → CI/CD pipelines, Dockerfiles, manifests K8s, archivos de seguridad
  🔴 HIGH     → Build files (pom, gradle, package.json), migraciones BD, configuración
  🟡 MEDIUM   → Código fuente (.java, .py, .js, .ts ...)
  🔵 LOW      → Tests, documentación, assets estáticos
  ⚪ NONE     → Archivos no clasificados

Score de impacto (0–100):
  Ponderado por riesgo de cada archivo + bonus por volumen de commits.
  Útil como quality gate en CD pipelines (exit 0 / 1 / 2).

Códigos de salida:
  0 → Impacto bajo / sin cambios críticos o altos
  1 → Archivos HIGH detectados
  2 → Archivos CRITICAL detectados

Autor: Harold Adrian
"""

import argparse
import base64
import csv
import json
import math
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

# --- Directorio de salida centralizado ---
try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn
    from rich.rule import Rule
    from rich.table import Table
    from rich.text import Text
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

__version__ = "1.0.0"
__author__  = "Harold Adrian"

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_ORG_URL  = "https://dev.azure.com/Coppel-Retail"
DEFAULT_PROJECT  = "Compras.RMI"
DEFAULT_TIMEZONE = "America/Mazatlan"
API_VERSION      = "7.1"
MAX_COMMITS      = 500
MAX_FILES        = 5000

RISK_CRITICAL = "CRITICAL"
RISK_HIGH     = "HIGH"
RISK_MEDIUM   = "MEDIUM"
RISK_LOW      = "LOW"
RISK_NONE     = "NONE"

RISK_ORDER: Dict[str, int] = {
    RISK_NONE: 0, RISK_LOW: 1, RISK_MEDIUM: 2, RISK_HIGH: 3, RISK_CRITICAL: 4
}
RISK_EMOJI = {
    RISK_CRITICAL: "🚨", RISK_HIGH: "🔴", RISK_MEDIUM: "🟡",
    RISK_LOW: "🔵", RISK_NONE: "⚪",
}
RISK_COLOR = {
    RISK_CRITICAL: "bold red", RISK_HIGH: "red",
    RISK_MEDIUM: "yellow", RISK_LOW: "blue", RISK_NONE: "dim",
}
RISK_BORDER = {
    RISK_CRITICAL: "red", RISK_HIGH: "red",
    RISK_MEDIUM: "yellow", RISK_LOW: "blue", RISK_NONE: "dim",
}

# Puntos base por nivel de riesgo para el cálculo del score de impacto
RISK_POINTS = {RISK_CRITICAL: 25, RISK_HIGH: 10, RISK_MEDIUM: 3, RISK_LOW: 0.5, RISK_NONE: 0}

CHANGE_ADD    = "add"
CHANGE_DELETE = "delete"
CHANGE_EDIT   = "edit"
CHANGE_RENAME = "rename"

CHANGE_LABEL: Dict[str, Tuple[str, str]] = {
    CHANGE_ADD:    ("green",    "➕ ADD"),
    CHANGE_DELETE: ("bold red", "🗑  DEL"),
    CHANGE_EDIT:   ("yellow",   "✏  EDIT"),
    CHANGE_RENAME: ("cyan",     "♻  REN"),
}

# Categorías con emojis de display
CATEGORY_META: Dict[str, Dict] = {
    "cicd":     {"label": "CI/CD Pipelines",  "emoji": "⚙️ ", "color": "bold red"},
    "security": {"label": "Seguridad",         "emoji": "🔐", "color": "bold red"},
    "infra":    {"label": "Infraestructura",   "emoji": "🏗️ ", "color": "red"},
    "build":    {"label": "Build/Dependencias","emoji": "📦", "color": "red"},
    "database": {"label": "Base de Datos",     "emoji": "🗄️ ", "color": "yellow"},
    "config":   {"label": "Configuración",     "emoji": "⚙️ ", "color": "yellow"},
    "code":     {"label": "Código Fuente",     "emoji": "💻", "color": "cyan"},
    "test":     {"label": "Pruebas",           "emoji": "🧪", "color": "blue"},
    "docs":     {"label": "Documentación",     "emoji": "📄", "color": "dim"},
    "assets":   {"label": "Assets Estáticos",  "emoji": "🖼️ ", "color": "dim"},
    "other":    {"label": "Otros",             "emoji": "📁", "color": "dim"},
}

# ═══════════════════════════════════════════════════════════════════════════════
# MOTOR DE CLASIFICACIÓN DE RIESGO
# ═══════════════════════════════════════════════════════════════════════════════
# Reglas evaluadas en orden (primera coincidencia gana)
# Formato: (regex_sobre_path_en_minúsculas, nivel_riesgo, categoría)
_RISK_RULES: List[Tuple[str, str, str]] = [
    # CRITICAL — CI/CD
    (r"dockerfile|docker-compose", RISK_CRITICAL, "cicd"),
    (r"\.github/|\.gitlab-ci|azure-pipelines\.yml|jenkinsfile|\.circleci|sonar-project", RISK_CRITICAL, "cicd"),
    # CRITICAL — Seguridad
    (r"secret|credential|password|\.key$|\.pem$|\.p12$|\.jks$|\.pfx$|keystore", RISK_CRITICAL, "security"),
    # CRITICAL — Infraestructura K8s/Helm
    (r"k8s/|kubernetes/|helm/|manifests?/", RISK_CRITICAL, "infra"),
    # HIGH — Build / Dependencias
    (r"pom\.xml$|build\.gradle$|settings\.gradle$|build\.sbt$|\.gradle$", RISK_HIGH, "build"),
    (r"package\.json$|package-lock\.json$|yarn\.lock$|requirements\.txt$|pipfile$|go\.mod$|go\.sum$", RISK_HIGH, "build"),
    # HIGH — Base de Datos
    (r"flyway|liquibase|migration|\.sql$|schema\.|db/migrate", RISK_HIGH, "database"),
    # HIGH — Configuración de aplicación
    (r"application.*\.(yml|yaml|properties)$|bootstrap\.(yml|yaml)$", RISK_HIGH, "config"),
    (r"logback.*\.xml$|log4j.*\.xml$|\.env$|\.env\.", RISK_HIGH, "config"),
    # LOW — Tests (evaluado ANTES que código fuente para que TestXxx.java sea LOW)
    (r"test|spec|mock|stub|fixture|__tests__|test_|_test\.|_spec\.", RISK_LOW, "test"),
    # MEDIUM — Código fuente JVM
    (r"\.(java|kt|scala|groovy|clj)$", RISK_MEDIUM, "code"),
    # MEDIUM — Código Python / Ruby / PHP / Go / Rust / C
    (r"\.(py|rb|php|go|rs|c|cpp|h|hpp|m|swift)$", RISK_MEDIUM, "code"),
    # MEDIUM — Código frontend
    (r"\.(js|ts|jsx|tsx|vue|svelte|mjs|cjs)$", RISK_MEDIUM, "code"),
    # MEDIUM — Código .NET
    (r"\.(cs|vb|fs|csproj|vbproj|fsproj|sln)$", RISK_MEDIUM, "code"),
    # MEDIUM — XML/JSON/YAML genérico (no clasificado antes)
    (r"\.(xml|json|yml|yaml)$", RISK_MEDIUM, "config"),
    # LOW — Documentación
    (r"\.(md|txt|rst|adoc|html|htm|pdf)$|readme|changelog|license|contributing", RISK_LOW, "docs"),
    # LOW — Assets estáticos
    (r"\.(png|jpg|jpeg|gif|svg|ico|webp|woff|woff2|ttf|eot)$", RISK_LOW, "assets"),
]


def classify_file(path: str) -> Tuple[str, str]:
    """
    Clasifica un archivo por nivel de riesgo y categoría basándose en su path.
    Returns: (risk_level, category)
    """
    lower = path.lower()
    for pattern, risk, category in _RISK_RULES:
        if re.search(pattern, lower):
            return risk, category
    return RISK_NONE, "other"


def _normalize_change_type(raw: str) -> str:
    raw = raw.lower()
    if "delete" in raw: return CHANGE_DELETE
    if "rename" in raw: return CHANGE_RENAME
    if "add"    in raw: return CHANGE_ADD
    return CHANGE_EDIT


# ═══════════════════════════════════════════════════════════════════════════════
# MODELOS DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════
class FileChange:
    __slots__ = ("path", "filename", "ext", "change_type", "risk", "category")

    def __init__(self, path: str, change_type: str):
        self.path        = path
        self.filename    = path.rsplit("/", 1)[-1]
        self.ext         = ("." + self.filename.rsplit(".", 1)[-1]).lower() \
                           if "." in self.filename else ""
        self.change_type = change_type
        self.risk, self.category = classify_file(path)

    def to_dict(self) -> Dict:
        return {
            "path":        self.path,
            "filename":    self.filename,
            "ext":         self.ext,
            "change_type": self.change_type,
            "risk":        self.risk,
            "category":    self.category,
        }


class CommitInfo:
    __slots__ = ("sha", "short_sha", "author", "date_iso", "message")

    def __init__(self, sha: str, author: str, date_iso: str, message: str):
        self.sha      = sha
        self.short_sha = sha[:8]
        self.author   = author
        self.date_iso = date_iso
        self.message  = message.strip().splitlines()[0][:80] if message else ""

    def to_dict(self) -> Dict:
        return {
            "sha":     self.sha,
            "author":  self.author,
            "date":    self.date_iso,
            "message": self.message,
        }


class BranchDiffReport:
    def __init__(
        self, repo_name: str, source_branch: str, target_branch: str,
        ahead_count: int, behind_count: int, common_commit: str,
        files: List[FileChange], commits: List[CommitInfo],
        generated_at: str,
    ):
        self.repo_name      = repo_name
        self.source_branch  = source_branch
        self.target_branch  = target_branch
        self.ahead_count    = ahead_count
        self.behind_count   = behind_count
        self.common_commit  = common_commit
        self.files          = files
        self.commits        = commits
        self.generated_at   = generated_at
        self.impact_score   = self._calc_impact()
        self.max_risk       = self._max_risk()

    def _calc_impact(self) -> int:
        """Score 0–100 ponderado por riesgo × volumen."""
        raw = sum(RISK_POINTS.get(f.risk, 0) for f in self.files)
        if self.ahead_count > 0:
            raw += math.log10(self.ahead_count + 1) * 5
        return min(100, int(raw))

    def _max_risk(self) -> str:
        if not self.files:
            return RISK_NONE
        return max(self.files, key=lambda f: RISK_ORDER.get(f.risk, 0)).risk

    def category_stats(self) -> Dict[str, Dict]:
        """Agrupa archivos por categoría."""
        stats: Dict[str, Dict] = {}
        for f in self.files:
            c = f.category
            if c not in stats:
                stats[c] = {"count": 0, "max_risk": RISK_NONE,
                             "add": 0, "edit": 0, "delete": 0}
            stats[c]["count"] += 1
            if RISK_ORDER.get(f.risk, 0) > RISK_ORDER.get(stats[c]["max_risk"], 0):
                stats[c]["max_risk"] = f.risk
            stats[c][f.change_type] = stats[c].get(f.change_type, 0) + 1
        return dict(sorted(stats.items(),
                    key=lambda x: RISK_ORDER.get(x[1]["max_risk"], 0), reverse=True))

    def author_stats(self) -> List[Dict]:
        """Agrupa commits por autor."""
        authors: Dict[str, int] = {}
        for c in self.commits:
            authors[c.author] = authors.get(c.author, 0) + 1
        return sorted(
            [{"author": a, "commits": n} for a, n in authors.items()],
            key=lambda x: x["commits"], reverse=True,
        )

    def to_dict(self) -> Dict:
        return {
            "repo":          self.repo_name,
            "source_branch": self.source_branch,
            "target_branch": self.target_branch,
            "ahead_count":   self.ahead_count,
            "behind_count":  self.behind_count,
            "common_commit": self.common_commit,
            "impact_score":  self.impact_score,
            "max_risk":      self.max_risk,
            "generated_at":  self.generated_at,
            "files":         [f.to_dict() for f in self.files],
            "commits":       [c.to_dict() for c in self.commits],
        }


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP / AzDO REST API
# ═══════════════════════════════════════════════════════════════════════════════
def make_headers(pat: str) -> Dict:
    token = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Accept": "application/json"}


def api_get(
    url: str, headers: Dict, params: Dict = None, debug: bool = False
) -> Optional[Any]:
    try:
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code == 404:
            return None
        if r.status_code >= 400:
            if debug:
                print(f"[DEBUG] HTTP {r.status_code} → {url}\n[DEBUG] {r.text[:300]}")
            r.raise_for_status()
        return r.json()
    except Exception as exc:
        if debug:
            print(f"[DEBUG] {exc} → {url}")
        return None


def get_repositories(
    org: str, project: str, headers: Dict, debug: bool = False
) -> List[Dict]:
    url  = f"{org}/{quote(project, safe='')}/_apis/git/repositories"
    data = api_get(url, headers, {"api-version": API_VERSION}, debug)
    return data.get("value", []) if data else []


def get_branches(
    org: str, project: str, repo_id: str, headers: Dict, debug: bool = False
) -> List[str]:
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/refs"
    params = {"api-version": API_VERSION, "filter": "heads/", "$top": 1000}
    data   = api_get(url, headers, params, debug)
    return [r["name"].replace("refs/heads/", "") for r in (data or {}).get("value", [])]


def get_branch_diff_raw(
    org: str, project: str, repo_id: str,
    source: str, target: str,
    headers: Dict, debug: bool = False,
) -> Dict:
    """
    Llama a diffs/commits y retorna el dict completo.
    base    = target (old)  |  targetVersion = source (new)
    """
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/diffs/commits"
    params = {
        "api-version":       API_VERSION,
        "baseVersionType":   "branch",
        "baseVersion":       target,
        "targetVersionType": "branch",
        "targetVersion":     source,
        "$top":              MAX_FILES,
    }
    data = api_get(url, headers, params, debug)
    return data or {}


def get_commits_ahead(
    org: str, project: str, repo_id: str,
    source: str, target: str,
    headers: Dict, debug: bool = False,
) -> List[Dict]:
    """Retorna commits presentes en source que NO están en target."""
    url    = f"{org}/{quote(project, safe='')}/_apis/git/repositories/{repo_id}/commits"
    params = {
        "api-version":                              API_VERSION,
        "searchCriteria.itemVersion.version":       source,
        "searchCriteria.itemVersion.versionType":   "branch",
        "searchCriteria.compareVersion.version":    target,
        "searchCriteria.compareVersion.versionType":"branch",
        "$top":                                     MAX_COMMITS,
    }
    data = api_get(url, headers, params, debug)
    return (data or {}).get("value", [])


# ═══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS CORE
# ═══════════════════════════════════════════════════════════════════════════════
class _NullCtx:
    def __enter__(self): return self
    def __exit__(self, *_): pass
    def add_task(self, *_, **__): return None
    def advance(self, _): pass
    def update(self, *_, **__): pass


def analyze_branches(
    org: str, project: str, repo_id: str, repo_name: str,
    source: str, target: str,
    headers: Dict, debug: bool,
    console: Optional[Any], tz_name: str,
) -> BranchDiffReport:
    ctx = (Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                    console=console)
           if RICH_AVAILABLE and console else _NullCtx())

    with ctx as prog:
        # ── 1. File changes ──────────────────────────────────────────────────
        t = prog.add_task("Obteniendo cambios de archivos...", total=None) \
            if RICH_AVAILABLE and console else None
        raw_diff = get_branch_diff_raw(org, project, repo_id, source, target, headers, debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

        raw_changes  = raw_diff.get("changes", [])
        ahead_count  = raw_diff.get("aheadCount",  0)
        behind_count = raw_diff.get("behindCount", 0)
        common_commit = raw_diff.get("commonCommit", {}).get("commitId", "") \
                        if isinstance(raw_diff.get("commonCommit"), dict) \
                        else str(raw_diff.get("commonCommit", ""))

        files: List[FileChange] = []
        for ch in raw_changes:
            item = ch.get("item", {})
            path = item.get("path", "")
            if not path or item.get("gitObjectType") == "tree":
                continue
            change_type = _normalize_change_type(ch.get("changeType", "edit"))
            files.append(FileChange(path, change_type))

        # ── 2. Commits ───────────────────────────────────────────────────────
        t = prog.add_task("Obteniendo commits...", total=None) \
            if RICH_AVAILABLE and console else None
        raw_commits = get_commits_ahead(org, project, repo_id, source, target, headers, debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

        # Si ahead_count no vino del diff, usar len de commits
        if not ahead_count and raw_commits:
            ahead_count = len(raw_commits)

        commits: List[CommitInfo] = []
        for rc in raw_commits:
            author  = rc.get("author", {}).get("name", "") \
                      or rc.get("committer", {}).get("name", "")
            date_v  = rc.get("author", {}).get("date", "") \
                      or rc.get("committer", {}).get("date", "")
            message = rc.get("comment", rc.get("message", ""))
            commits.append(CommitInfo(
                sha=rc.get("commitId", ""),
                author=author,
                date_iso=date_v,
                message=message,
            ))

    generated_at = datetime.now(ZoneInfo(tz_name)).strftime("%Y-%m-%d %H:%M:%S %Z")
    return BranchDiffReport(
        repo_name=repo_name,
        source_branch=source,
        target_branch=target,
        ahead_count=ahead_count,
        behind_count=behind_count,
        common_commit=common_commit,
        files=files,
        commits=commits,
        generated_at=generated_at,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# DISPLAY
# ═══════════════════════════════════════════════════════════════════════════════
def _score_bar(score: int, width: int = 20) -> str:
    """Barra visual de score 0-100."""
    filled = int(score / 100 * width)
    color  = "red" if score >= 70 else "yellow" if score >= 40 else "green"
    bar    = f"[{color}]{'█' * filled}[/{color}]{'░' * (width - filled)}"
    return bar


def _risk_badge(risk: str) -> str:
    color = RISK_COLOR.get(risk, "white")
    return f"[{color}]{RISK_EMOJI.get(risk,'')} {risk}[/{color}]"


def print_executive_summary(report: BranchDiffReport, console: Any):
    max_risk     = report.max_risk
    border       = RISK_BORDER.get(max_risk, "dim")
    score_bar    = _score_bar(report.impact_score)

    # Conteos por tipo de cambio
    adds    = sum(1 for f in report.files if f.change_type == CHANGE_ADD)
    edits   = sum(1 for f in report.files if f.change_type == CHANGE_EDIT)
    deletes = sum(1 for f in report.files if f.change_type == CHANGE_DELETE)

    # Conteos por riesgo
    crits   = sum(1 for f in report.files if f.risk == RISK_CRITICAL)
    highs   = sum(1 for f in report.files if f.risk == RISK_HIGH)
    meds    = sum(1 for f in report.files if f.risk == RISK_MEDIUM)
    lows    = sum(1 for f in report.files if f.risk == RISK_LOW)

    behind_warn = f"\n  [yellow]⚠️  {report.behind_count} commit(s) en destino no presentes en origen[/]" \
                  if report.behind_count > 0 else ""

    body = (
        f"  [bold]Repositorio:[/]    [cyan]{report.repo_name}[/]\n"
        f"  [bold]Rama origen:[/]    [green]{report.source_branch}[/]\n"
        f"  [bold]Rama destino:[/]   [yellow]{report.target_branch}[/]\n"
        f"  [bold]Analizado:[/]      {report.generated_at}\n"
        f"\n"
        f"  [bold]Commits adelante:[/]  [green]{report.ahead_count}[/]"
        f"   [bold]Atrás:[/]  [red]{report.behind_count}[/]{behind_warn}\n"
        f"\n"
        f"  [bold]Archivos cambiados:[/] {len(report.files)}"
        f"  ([green]+{adds}[/] [yellow]~{edits}[/] [red]-{deletes}[/])\n"
        f"  [bold]Por riesgo:[/]     "
        f"[bold red]🚨{crits}[/]  [red]🔴{highs}[/]  [yellow]🟡{meds}[/]  [blue]🔵{lows}[/]\n"
        f"\n"
        f"  [bold]Score de impacto:[/]  {score_bar} [bold]{report.impact_score}/100[/]\n"
        f"  [bold]Nivel de riesgo:[/]   {_risk_badge(max_risk)}"
    )
    console.print(Panel(body, title="📊 Resumen Ejecutivo", border_style=border))


def print_risk_distribution(report: BranchDiffReport, console: Any):
    stats = report.category_stats()
    if not stats:
        return
    total = len(report.files) or 1

    table = Table(
        title="📂 Distribución de Cambios por Categoría y Riesgo",
        title_style="bold magenta", header_style="bold cyan",
        border_style="dim", box=box.ROUNDED,
    )
    table.add_column("Categoría",   style="white", min_width=20)
    table.add_column("Archivos",    justify="right", width=9)
    table.add_column("Riesgo máx", justify="center", width=14)
    table.add_column("Dist.",       justify="left", min_width=22)
    table.add_column("ADD",         justify="right", width=5)
    table.add_column("EDIT",        justify="right", width=5)
    table.add_column("DEL",         justify="right", width=5)

    for cat, data in stats.items():
        meta  = CATEGORY_META.get(cat, CATEGORY_META["other"])
        count = data["count"]
        pct   = count / total
        bar_w = int(pct * 20)
        bar   = f"{'█' * bar_w}{'░' * (20 - bar_w)}  {pct:.0%}"
        risk_badge = _risk_badge(data["max_risk"])
        cat_label  = f"{meta['emoji']} {meta['label']}"
        table.add_row(
            f"[{meta['color']}]{cat_label}[/{meta['color']}]",
            str(count),
            risk_badge,
            f"[dim]{bar}[/]",
            f"[green]{data.get(CHANGE_ADD, 0)}[/]",
            f"[yellow]{data.get(CHANGE_EDIT, 0)}[/]",
            f"[red]{data.get(CHANGE_DELETE, 0)}[/]",
        )
    console.print(table)


def print_files_table(report: BranchDiffReport, console: Any, top_n: int = 60):
    files = sorted(report.files, key=lambda f: RISK_ORDER.get(f.risk, 0), reverse=True)
    total = len(files)
    shown = files[:top_n]

    table = Table(
        title=f"🗂  Archivos Modificados ({total} total"
              f"{f', mostrando top {top_n}' if total > top_n else ''})",
        title_style="bold magenta", header_style="bold cyan",
        border_style="dim", box=box.ROUNDED, show_lines=False,
    )
    table.add_column("Archivo",   style="white", min_width=30, max_width=55)
    table.add_column("Categoría", justify="left", width=18)
    table.add_column("Cambio",    justify="center", width=10)
    table.add_column("Riesgo",    justify="center", width=14)

    for f in shown:
        ch_style, ch_label = CHANGE_LABEL.get(f.change_type, ("dim", f.change_type))
        risk_badge = _risk_badge(f.risk)
        meta       = CATEGORY_META.get(f.category, CATEGORY_META["other"])
        cat_label  = f"[{meta['color']}]{meta['emoji']} {meta['label']}[/{meta['color']}]"
        table.add_row(
            f.filename,
            cat_label,
            f"[{ch_style}]{ch_label}[/{ch_style}]",
            risk_badge,
        )

    console.print(table)
    if total > top_n:
        console.print(
            f"[dim]  ... y {total - top_n} archivos más. "
            f"Usa --top-files 0 para ver todos, o --output excel para exportar.[/]"
        )


def print_commits_table(report: BranchDiffReport, console: Any, top_n: int = 25):
    commits = report.commits
    total   = len(commits)
    shown   = commits[:top_n]

    table = Table(
        title=f"📋 Commits Incluidos ({total} total"
              f"{f', mostrando últimos {top_n}' if total > top_n else ''})",
        title_style="bold magenta", header_style="bold cyan",
        border_style="dim", box=box.SIMPLE,
    )
    table.add_column("SHA",    width=10, style="cyan")
    table.add_column("Autor",  width=22, style="white")
    table.add_column("Fecha",  width=12, style="dim")
    table.add_column("Mensaje", max_width=65, style="white")

    for c in shown:
        date_str = c.date_iso[:10] if c.date_iso else "—"
        table.add_row(c.short_sha, c.author, date_str, c.message)

    console.print(table)


def print_author_stats(report: BranchDiffReport, console: Any):
    authors = report.author_stats()
    if not authors:
        return
    table = Table(
        title="👥 Estadísticas por Autor",
        title_style="bold magenta", header_style="bold cyan",
        border_style="dim", box=box.SIMPLE,
    )
    table.add_column("#",       justify="right", width=4)
    table.add_column("Autor",   style="white", min_width=25)
    table.add_column("Commits", justify="right", width=8)
    table.add_column("Participación", justify="left", min_width=22)

    total_c = sum(a["commits"] for a in authors)
    for i, a in enumerate(authors[:15], 1):
        pct = a["commits"] / max(total_c, 1)
        bar = f"{'█' * int(pct * 20)}  {pct:.0%}"
        table.add_row(str(i), a["author"], str(a["commits"]), f"[dim]{bar}[/]")

    console.print(table)


def _build_recommendations(report: BranchDiffReport) -> List[str]:
    recs: List[str] = []
    cats = {f.category for f in report.files}
    risks = {f.risk for f in report.files}

    if RISK_CRITICAL in risks:
        recs.append("🚨 [bold red]Revisión obligatoria[/] con arquitecto / equipo de infraestructura "
                    "antes de promover a producción.")
    if "security" in cats:
        recs.append("🔐 [bold red]Auditoría de seguridad[/] — se detectaron cambios en archivos "
                    "relacionados con credenciales o certificados.")
    if "cicd" in cats:
        recs.append("⚙️  Validar los [yellow]pipelines CI/CD[/] en staging antes de activar el "
                    "despliegue productivo.")
    if "infra" in cats:
        recs.append("🏗️  Los cambios de [yellow]infraestructura (K8s/Helm)[/] requieren validación "
                    "con el equipo de plataforma.")
    if "database" in cats:
        recs.append("🗄️  Planificar [yellow]ventana de mantenimiento[/] para las migraciones de "
                    "base de datos detectadas.")
    if "build" in cats:
        recs.append("📦 Cambios en dependencias de build — ejecutar [cyan]análisis de "
                    "vulnerabilidades[/] (OWASP Dependency-Check / Trivy).")
    if report.behind_count > 0:
        recs.append(f"⚠️  La rama origen está [yellow]{report.behind_count} commit(s) atrás[/] de "
                    "la rama destino. Considerar merge o rebase antes del despliegue.")
    if len(report.files) > 50:
        recs.append("📏 Volumen alto de cambios (> 50 archivos). Evaluar [cyan]despliegue "
                    "incremental[/] o [cyan]feature flags[/] para reducir el riesgo.")
    if report.ahead_count > 30:
        recs.append("📦 Muchos commits pendientes — verificar que todos hayan pasado por "
                    "[cyan]revisión de código (PR)[/] antes de promover.")
    if not recs:
        recs.append("✅ Sin hallazgos críticos. Proceder con las revisiones estándar de "
                    "código y pruebas de regresión.")
    return recs


def print_recommendations(report: BranchDiffReport, console: Any):
    recs   = _build_recommendations(report)
    border = RISK_BORDER.get(report.max_risk, "dim")
    body   = "\n".join(f"  • {r}" for r in recs)
    console.print(Panel(body, title="💡 Recomendaciones", border_style=border))


def print_plain_report(report: BranchDiffReport):
    print(f"\n{'='*70}")
    print(f"BRANCH DIFF REPORT  |  {report.repo_name}")
    print(f"Origen : {report.source_branch}")
    print(f"Destino: {report.target_branch}")
    print(f"Commits adelante: {report.ahead_count}  |  Atrás: {report.behind_count}")
    print(f"Archivos cambiados: {len(report.files)}")
    print(f"Score: {report.impact_score}/100  |  Riesgo: {report.max_risk}")
    print(f"{'='*70}")
    print(f"{'Archivo':<45} {'Cambio':<8} {'Riesgo':<10} {'Categoría'}")
    print("-"*70)
    for f in sorted(report.files, key=lambda x: RISK_ORDER.get(x.risk, 0), reverse=True):
        print(f"{f.filename:<45} {f.change_type:<8} {f.risk:<10} {f.category}")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def export_results(
    report: BranchDiffReport, fmt: str,
    org: str, project: str, tz_name: str, console: Optional[Any],
):
    out_dir  = Path(get_output_dir("outcome"))
    ts       = datetime.now(ZoneInfo(tz_name)).strftime("%Y%m%d_%H%M%S")
    safe_src = report.source_branch.replace("/", "_")
    safe_tgt = report.target_branch.replace("/", "_")
    base     = f"branch_diff_{report.repo_name}_{safe_src}_vs_{safe_tgt}_{ts}"

    if fmt == "json":
        fp = out_dir / f"{base}.json"
        with open(fp, "w", encoding="utf-8") as fh:
            json.dump(report.to_dict(), fh, indent=2, ensure_ascii=False)

    elif fmt == "csv":
        fp = out_dir / f"{base}.csv"
        fields = ["path", "filename", "ext", "change_type", "risk", "category"]
        with open(fp, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=fields)
            w.writeheader()
            for f in report.files:
                w.writerow(f.to_dict())

    elif fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            msg = "openpyxl no instalado. pip install openpyxl"
            (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
            return

        fp  = out_dir / f"{base}.xlsx"
        wb  = openpyxl.Workbook()
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(bold=True, color="FFFFFF")

        RISK_FILL = {
            RISK_CRITICAL: PatternFill("solid", fgColor="FF4444"),
            RISK_HIGH:     PatternFill("solid", fgColor="FF9999"),
            RISK_MEDIUM:   PatternFill("solid", fgColor="FFDD57"),
            RISK_LOW:      PatternFill("solid", fgColor="90CAF9"),
            RISK_NONE:     PatternFill("solid", fgColor="E8E8E8"),
        }

        # ── Hoja 1: Archivos ───────────────────────────────────────────────
        ws = wb.active
        ws.title = "Archivos"
        for ci, col in enumerate(["Archivo", "Categoría", "Cambio", "Riesgo", "Path"], 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(horizontal="center")
        files_sorted = sorted(report.files,
                               key=lambda f: RISK_ORDER.get(f.risk, 0), reverse=True)
        for ri, f in enumerate(files_sorted, 2):
            meta = CATEGORY_META.get(f.category, CATEGORY_META["other"])
            vals = [f.filename, meta["label"], f.change_type.upper(),
                    f"{RISK_EMOJI.get(f.risk,'')} {f.risk}", f.path]
            fill = RISK_FILL.get(f.risk, PatternFill())
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill
        for ci in range(1, 6):
            ws.column_dimensions[get_column_letter(ci)].width = [25, 22, 10, 14, 55][ci - 1]

        # ── Hoja 2: Commits ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Commits")
        for ci, col in enumerate(["SHA", "Autor", "Fecha", "Mensaje"], 1):
            c = ws2.cell(row=1, column=ci, value=col)
            c.fill, c.font = hdr_fill, hdr_font
        for ri, cm in enumerate(report.commits, 2):
            ws2.cell(row=ri, column=1, value=cm.short_sha)
            ws2.cell(row=ri, column=2, value=cm.author)
            ws2.cell(row=ri, column=3, value=cm.date_iso[:10] if cm.date_iso else "")
            ws2.cell(row=ri, column=4, value=cm.message)
        for ci, w in enumerate([12, 28, 12, 80], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

        # ── Hoja 3: Resumen ────────────────────────────────────────────────
        ws3 = wb.create_sheet("Resumen")
        meta_rows = [
            ("Repositorio",      report.repo_name),
            ("Rama origen",      report.source_branch),
            ("Rama destino",     report.target_branch),
            ("Commits adelante", report.ahead_count),
            ("Commits atrás",    report.behind_count),
            ("Archivos totales", len(report.files)),
            ("Score impacto",    f"{report.impact_score}/100"),
            ("Nivel riesgo",     report.max_risk),
            ("Generado",         report.generated_at),
            ("",                 ""),
            ("Archivos CRITICAL",sum(1 for f in report.files if f.risk == RISK_CRITICAL)),
            ("Archivos HIGH",    sum(1 for f in report.files if f.risk == RISK_HIGH)),
            ("Archivos MEDIUM",  sum(1 for f in report.files if f.risk == RISK_MEDIUM)),
            ("Archivos LOW",     sum(1 for f in report.files if f.risk == RISK_LOW)),
            ("",                 ""),
            ("Recomendaciones",  ""),
        ]
        for ki, (key, val) in enumerate(meta_rows, 1):
            ws3.cell(row=ki, column=1, value=key).font = Font(bold=True)
            ws3.cell(row=ki, column=2, value=str(val) if val != "" else "")
        recs = _build_recommendations(report)
        for i, rec in enumerate(recs):
            clean_rec = re.sub(r"\[/?[^\]]+\]", "", rec)
            ws3.cell(row=len(meta_rows) + i, column=2, value=f"• {clean_rec}")
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 80

        wb.save(fp)

    msg = f"📁 Exportado: {fp}"
    (console.print(f"[bold green]{msg}[/]") if RICH_AVAILABLE and console else print(msg))


# ═══════════════════════════════════════════════════════════════════════════════
# PROMPTS INTERACTIVOS
# ═══════════════════════════════════════════════════════════════════════════════
def _prompt(text: str, default: str = "", console: Optional[Any] = None) -> str:
    if RICH_AVAILABLE and console:
        suffix = f" [dim](default: {default})[/]: " if default else ": "
        console.print(f"[bold]{text}[/]{suffix}", end="")
    else:
        print(f"{text} [{default}]: " if default else f"{text}: ", end="")
    return input().strip() or default


def prompt_select_repo(repos: List[Dict], console: Optional[Any]) -> Optional[Dict]:
    if RICH_AVAILABLE and console:
        t = Table(title="Repositorios disponibles", header_style="bold cyan", border_style="dim")
        t.add_column("#", justify="right", width=4)
        t.add_column("Repositorio", style="white", min_width=30)
        for i, r in enumerate(repos, 1):
            t.add_row(str(i), r["name"])
        console.print(t)
    else:
        for i, r in enumerate(repos, 1):
            print(f"  {i:3}. {r['name']}")

    choice = _prompt("Seleccione # o nombre del repositorio", console=console)
    if not choice:
        return None
    if choice.isdigit():
        idx = int(choice) - 1
        return repos[idx] if 0 <= idx < len(repos) else None
    hits = [r for r in repos if choice.lower() in r["name"].lower()]
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1 and RICH_AVAILABLE and console:
        console.print(f"[yellow]Múltiples coincidencias: {[h['name'] for h in hits]}[/]")
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# ARGS
# ═══════════════════════════════════════════════════════════════════════════════
def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Analiza el impacto de cambios entre dos ramas de un repositorio en AzDO",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python azdo_repo_branch_diff.py --pat TOKEN --repo mi-servicio --source release/1.6.0 --target master
  python azdo_repo_branch_diff.py --pat TOKEN --repo mi-servicio --source develop --target master --output excel
  python azdo_repo_branch_diff.py --pat TOKEN   # modo interactivo completo
        """,
    )
    p.add_argument("--org",       "-g",  default=DEFAULT_ORG_URL,
                   help=f"URL organización (default: {DEFAULT_ORG_URL})")
    p.add_argument("--project",   "-p",  default=DEFAULT_PROJECT,
                   help=f"Nombre del proyecto (default: {DEFAULT_PROJECT})")
    p.add_argument("--pat",              required=True,
                   help="Personal Access Token con permisos: Code (Read)")
    p.add_argument("--repo",      "-r",  default=None,
                   help="Nombre del repositorio (substring, case insensitive)")
    p.add_argument("--source",    "-s",  default=None,
                   help="Rama origen — la que se quiere desplegar")
    p.add_argument("--target",    "-t",  default=None,
                   help="Rama destino — el entorno receptor (ej: master)")
    p.add_argument("--top-files",        type=int, default=60,
                   help="Máx. archivos en tabla de detalle (0=todos, default: 60)")
    p.add_argument("--top-commits",      type=int, default=25,
                   help="Máx. commits en tabla (default: 25)")
    p.add_argument("--no-commits",       action="store_true",
                   help="Omitir tabla de commits del informe")
    p.add_argument("--no-authors",       action="store_true",
                   help="Omitir estadísticas por autor")
    p.add_argument("--severity",         default=None,
                   choices=["CRITICAL", "HIGH", "MEDIUM", "LOW", "NONE"],
                   help="Mostrar solo archivos con riesgo >= especificado")
    p.add_argument("--output",    "-o",  choices=["json", "csv", "excel"], default=None,
                   help="Exportar resultados")
    p.add_argument("--timezone",  "-tz", default=DEFAULT_TIMEZONE)
    p.add_argument("--debug",            action="store_true")
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main() -> int:
    start   = time.time()
    args    = get_args()
    console = Console() if RICH_AVAILABLE else None

    if not REQUESTS_AVAILABLE:
        print("ERROR: requests no instalado. pip install requests")
        return 1

    headers = make_headers(args.pat)

    if RICH_AVAILABLE and console:
        console.print(Panel(
            f"[bold cyan]Repository Branch Diff Analyzer[/]  "
            f"[dim]v{__version__} | {__author__}[/]\n"
            "[dim]Informe ejecutivo de impacto de cambios para garantía de calidad "
            "en despliegues productivos[/]",
            border_style="cyan",
        ))

    # ── 1. Repositorio ──────────────────────────────────────────────────────────
    ctx = (Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console)
           if RICH_AVAILABLE and console else _NullCtx())
    with ctx as prog:
        t = prog.add_task("Obteniendo repositorios...", total=None) \
            if RICH_AVAILABLE and console else None
        repos = get_repositories(args.org, args.project, headers, args.debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

    if not repos:
        msg = "❌ No se pudieron obtener repositorios. Verifica org/project/PAT."
        (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
        return 1

    if args.repo:
        hits = [r for r in repos if args.repo.lower() in r["name"].lower()]
        if not hits:
            msg = f"❌ No se encontró repositorio con '{args.repo}'."
            (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
            return 1
        selected = hits[0] if len(hits) == 1 else prompt_select_repo(hits, console)
    else:
        selected = prompt_select_repo(repos, console)

    if not selected:
        print("❌ No se seleccionó repositorio.")
        return 1

    repo_id, repo_name = selected["id"], selected["name"]
    if RICH_AVAILABLE and console:
        console.print(f"[green]✅ Repositorio:[/] [bold]{repo_name}[/]")

    # ── 2. Ramas ────────────────────────────────────────────────────────────────
    ctx = (Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console)
           if RICH_AVAILABLE and console else _NullCtx())
    with ctx as prog:
        t = prog.add_task("Obteniendo ramas...", total=None) \
            if RICH_AVAILABLE and console else None
        branch_names = get_branches(args.org, args.project, repo_id, headers, args.debug)
        if t and RICH_AVAILABLE and console:
            prog.update(t, completed=True)

    if not branch_names:
        msg = "❌ No se pudieron obtener ramas."
        (console.print(f"[red]{msg}[/]") if RICH_AVAILABLE and console else print(msg))
        return 1

    if RICH_AVAILABLE and console:
        preview = branch_names[:8]
        suffix  = f" [dim]+{len(branch_names)-8} más[/]" if len(branch_names) > 8 else ""
        console.print(
            f"[dim]🌿 Ramas: {', '.join(f'[cyan]{b}[/]' for b in preview)}{suffix}[/]"
        )

    source = args.source or _prompt("Rama ORIGEN (la que se desplegará)", "develop", console)
    target = args.target or _prompt("Rama DESTINO (entorno receptor)",   "master",  console)

    for br, label in [(source, "origen"), (target, "destino")]:
        if br not in branch_names:
            msg = f"⚠️  Rama {label} '{br}' no encontrada — se intentará de todas formas."
            (console.print(f"[yellow]{msg}[/]") if RICH_AVAILABLE and console else print(msg))

    if RICH_AVAILABLE and console:
        console.print(
            f"[green]✅ Comparación:[/] [bold green]{source}[/] → [bold yellow]{target}[/]"
        )

    # ── 3. Análisis ─────────────────────────────────────────────────────────────
    if RICH_AVAILABLE and console:
        console.print()

    report = analyze_branches(
        args.org, args.project, repo_id, repo_name,
        source, target, headers, args.debug, console, args.timezone,
    )

    # Filtro de severidad opcional
    if args.severity:
        min_order  = RISK_ORDER.get(args.severity, 0)
        report.files = [f for f in report.files
                        if RISK_ORDER.get(f.risk, 0) >= min_order]

    top_files = len(report.files) if args.top_files == 0 else args.top_files

    # ── 4. Informe ──────────────────────────────────────────────────────────────
    if not report.files and not report.commits:
        if RICH_AVAILABLE and console:
            console.print(Panel(
                f"[green]⚪ Sin diferencias detectadas entre "
                f"[bold]{source}[/] y [bold]{target}[/][/]",
                border_style="green",
            ))
        return 0

    if RICH_AVAILABLE and console:
        console.print()
        print_executive_summary(report, console)
        console.print()
        if report.files:
            print_risk_distribution(report, console)
            console.print()
            print_files_table(report, console, top_files)
            console.print()
        if not args.no_commits and report.commits:
            print_commits_table(report, console, args.top_commits)
            console.print()
        if not args.no_authors and report.commits:
            print_author_stats(report, console)
            console.print()
        print_recommendations(report, console)
    else:
        print_plain_report(report)

    # ── 5. Export ────────────────────────────────────────────────────────────────
    if args.output:
        export_results(report, args.output, args.org, args.project, args.timezone, console)

    # ── 6. Footer ────────────────────────────────────────────────────────────────
    elapsed = time.time() - start
    if RICH_AVAILABLE and console:
        console.print(f"\n[dim]⏱  Tiempo de ejecución: {elapsed:.1f}s[/]")

    max_order = RISK_ORDER.get(report.max_risk, 0)
    if max_order >= RISK_ORDER[RISK_CRITICAL]:
        return 2
    if max_order >= RISK_ORDER[RISK_HIGH]:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())

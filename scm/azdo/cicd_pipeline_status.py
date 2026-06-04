#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cicd_pipeline_status.py

Reporte consolidado de estado de pipelines CI + CD en Azure DevOps.

Columnas:
  # | Tipo | Nombre | Estado | Deprecado | Última Actualización | Último Run | Días Inactivo

Lógica de "Deprecado":
  CI : queueStatus == 'disabled'  OR  sin ejecuciones en --inactive-days días (default: 365)
  CD : sin releases en --inactive-days días (default: 365)  OR  nunca ejecutado

Paginación:
  Ambas APIs (build definitions / release definitions) devuelven máximo ~1000 registros
  por página vía x-ms-continuationtoken. api_get_paginated() acumula todas las páginas.

Resumen:
  Total | Activos | Deshabilitados | Deprecados | Sin actividad N días

Autor: Harold Adrian
"""

import argparse
import base64
import concurrent.futures
import csv
import glob
import json
import os
import threading
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import quote
from typing import Any, Dict, List, Optional, Tuple
try:
    from zoneinfo import ZoneInfo
except ImportError:  # Python < 3.9
    try:
        from backports.zoneinfo import ZoneInfo  # type: ignore[no-redef]
    except ImportError:
        ZoneInfo = None  # type: ignore[assignment,misc]

try:
    from utils import get_output_dir
except ImportError:
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
    from rich.table import Table
    from rich.text import Text
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    import types as _types
    requests = _types.SimpleNamespace(get=None, HTTPError=Exception)  # type: ignore[assignment]
    REQUESTS_AVAILABLE = False

__version__     = "1.1.1"
__author__      = "Harold Adrian"
SCRIPT_NAME     = "cicd_pipeline_status"
CACHE_TTL_HOURS = 24

# ═══════════════════════════════════════════════════════════════════════════════
# DEFAULTS
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_ORG_URL       = "https://dev.azure.com/Coppel-Retail"
DEFAULT_PROJECT       = "Compras.RMI"
DEFAULT_TIMEZONE      = "America/Mazatlan"
DEFAULT_INACTIVE_DAYS = 365
API_VERSION           = "7.1"

QUEUE_STATUS_LABEL = {
    "enabled":  "✅ Activo",
    "paused":   "⏸️  Pausado",
    "disabled": "🔴 Deshabilitado",
}

DEPRECADO_SI  = "⚠️  Sí"
DEPRECADO_NO  = "No"

BUCKETS       = ["0-30d", "31-60d", "61-90d", "91-180d", ">180d", "Nunca"]

BUCKET_COLORS = {
    "0-30d":   "2ecc71",
    "31-60d":  "f1c40f",
    "61-90d":  "e67e22",
    "91-180d": "e74c3c",
    ">180d":   "922b21",
    "Nunca":   "7f8c8d",
}

# Reglas de clasificación por grupo — se evalúan en orden (primera coincidencia gana)
_GRUPO_RULES: List[Tuple[str, List[str]]] = [
    ("WMS", ["wm", "wms", "ayr", "ims", "rdm"]),
    ("OMS", ["oms"]),
    ("CSC", ["csc"]),
    ("TMS", ["tms", "cmanager", "torrecontrol"]),
]
GRUPO_SIN_COINCIDENCIA = "Sin Coincidencia"

GRUPO_COLORS_RICH = {
    "WMS": "cyan",
    "OMS": "green",
    "CSC": "yellow",
    "TMS": "magenta",
    GRUPO_SIN_COINCIDENCIA: "dim",
}

GRUPO_COLORS_EXCEL = {
    "WMS": "1abc9c",
    "OMS": "2ecc71",
    "CSC": "f1c40f",
    "TMS": "9b59b6",
    GRUPO_SIN_COINCIDENCIA: "bdc3c7",
}


def classify_pipeline_group(path: str) -> str:
    """Clasifica un pipeline en WMS/OMS/CSC/TMS según palabras clave en su ruta.

    Evalúa _GRUPO_RULES en orden; devuelve la primera coincidencia.
    La comparación es case-insensitive sobre el path completo.
    """
    path_lower = (path or "").lower()
    for group, keywords in _GRUPO_RULES:
        if any(kw in path_lower for kw in keywords):
            return group
    return GRUPO_SIN_COINCIDENCIA


def _bucket(dias_str: str) -> str:
    if dias_str == "Nunca":
        return "Nunca"
    d = int(dias_str)
    if d <= 30:  return "0-30d"
    if d <= 60:  return "31-60d"
    if d <= 90:  return "61-90d"
    if d <= 180: return "91-180d"
    return ">180d"


# ═══════════════════════════════════════════════════════════════════════════════
# ARGS
# ═══════════════════════════════════════════════════════════════════════════════
def get_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Reporte consolidado de estado de pipelines CI + CD"
    )
    parser.add_argument("--org",  "-g", default=DEFAULT_ORG_URL,
                        help=f"URL de la organización (default: {DEFAULT_ORG_URL})")
    parser.add_argument("--project", "-p", default=DEFAULT_PROJECT,
                        help=f"Nombre del proyecto (default: {DEFAULT_PROJECT})")
    parser.add_argument("--pat", required=True,
                        help="Personal Access Token: Build (Read) + Release (Read)")
    parser.add_argument("--workers", "-w", type=int, default=10,
                        help="Hilos paralelos para consultas CD (default: 10)")
    parser.add_argument("--inactive-days", type=int, default=DEFAULT_INACTIVE_DAYS,
                        help=f"Días sin actividad para marcar como deprecado (default: {DEFAULT_INACTIVE_DAYS})")
    parser.add_argument("--type", choices=["ci", "cd", "all"], default="all",
                        help="Tipo de pipeline a mostrar (default: all)")
    parser.add_argument("--only-deprecated", action="store_true",
                        help="Mostrar solo pipelines deprecados")
    parser.add_argument("--output", "-o", choices=["json", "csv", "excel"], default=None,
                        help="Exportar resultados (json / csv / excel)")
    parser.add_argument("--timezone", "-tz", default=DEFAULT_TIMEZONE,
                        help=f"Zona horaria (default: {DEFAULT_TIMEZONE})")
    parser.add_argument("--force-refresh", action="store_true",
                        help="Ignorar cache y consultar APIs")
    parser.add_argument("--use-cache-only", action="store_true",
                        help="Solo usar cache; falla si no existe o es > 24h")
    parser.add_argument("--debug", action="store_true",
                        help="Mostrar errores HTTP detallados")
    return parser.parse_args()


# ═══════════════════════════════════════════════════════════════════════════════
# HTTP
# ═══════════════════════════════════════════════════════════════════════════════
def make_headers(pat: str) -> Dict:
    token = base64.b64encode(f":{pat}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}


def api_get(url: str, headers: Dict, params: Dict = None, debug: bool = False) -> Optional[Any]:
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        if resp.status_code >= 400:
            if debug:
                print(f"[DEBUG] HTTP {resp.status_code} → {url}")
                print(f"[DEBUG] {resp.text[:300]}")
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        if debug:
            print(f"[DEBUG] Exception: {e}")
        return None


def api_get_paginated(url: str, headers: Dict, params: Dict = None,
                     debug: bool = False) -> List[Dict]:
    """GET paginado: acumula todas las páginas vía x-ms-continuationtoken.

    Azure DevOps limita a ~1000 registros por llamada. Cuando hay más,
    la respuesta incluye el header 'x-ms-continuationtoken'. Este helper
    itera hasta que no haya más token, devolviendo la lista completa.
    """
    all_values: List[Dict] = []
    continuation_token: Optional[str] = None
    page = 0

    while True:
        page += 1
        p = dict(params or {})
        if continuation_token:
            p["continuationToken"] = continuation_token
        try:
            resp = requests.get(url, headers=headers, params=p, timeout=30)
            if resp.status_code >= 400:
                if debug:
                    print(f"[DEBUG] HTTP {resp.status_code} pág.{page} → {url}")
                    print(f"[DEBUG] {resp.text[:300]}")
                resp.raise_for_status()
            data = resp.json()
            batch = data.get("value", [])
            all_values.extend(batch)
            if debug:
                print(f"[DEBUG] Pág.{page}: {len(batch)} registros (total acum. {len(all_values)})")
            continuation_token = resp.headers.get("x-ms-continuationtoken")
            if not continuation_token:
                break
        except Exception as e:
            if debug:
                print(f"[DEBUG] Exception en pág.{page}: {e}")
            break

    return all_values


# ═══════════════════════════════════════════════════════════════════════════════
# DATE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def parse_date(value: str) -> Optional[datetime]:
    if not value:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(value[:26].rstrip("Z"), fmt.rstrip("Z")).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _safe_tz(name: str):
    """Devuelve ZoneInfo(name) o timezone.utc si tzdata no está disponible."""
    try:
        return ZoneInfo(name)
    except Exception:
        return timezone.utc


def days_since(dt: Optional[datetime]) -> Optional[int]:
    if dt is None:
        return None
    return (datetime.now(timezone.utc) - dt).days


def fmt_date(value: str, tz_name: str) -> str:
    dt = parse_date(value)
    if dt is None:
        return "—"
    try:
        return dt.astimezone(ZoneInfo(tz_name)).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return dt.strftime("%Y-%m-%d %H:%M")


# ═══════════════════════════════════════════════════════════════════════════════
# CI — FETCH
# ═══════════════════════════════════════════════════════════════════════════════
def get_ci_definitions(org: str, project: str, headers: Dict, debug: bool = False) -> List[Dict]:
    """Obtiene TODAS las build definitions con info del último build.

    Estrategia en dos pasos (necesaria porque includeLatestBuilds=true
    desactiva la paginación vía x-ms-continuationtoken en Azure DevOps):

    Paso 1 — Paginación completa sin includeLatestBuilds:
        Obtiene todos los IDs/metadata usando continuationToken.

    Paso 2 — Batch-fetch con includeLatestBuilds=true (lotes de 200):
        Enriquece cada lote con latestCompletedBuild / latestBuild.
    """
    org_name = org.rstrip("/").split("/")[-1]
    url = f"https://dev.azure.com/{org_name}/{quote(project, safe='')}/_apis/build/definitions"

    # ── Paso 1: obtener TODOS los IDs con paginación ────────────────────────
    # Sin includeLatestBuilds el API soporta $top hasta 5000+ en una sola página.
    # includeLatestBuilds=true tiene un límite interno de ~1000 por página.
    all_defs_raw = api_get_paginated(
        url, headers, {"api-version": API_VERSION, "$top": 5000}, debug
    )

    if not all_defs_raw:
        return []

    if debug:
        print(f"[DEBUG CI] Total definitions encontradas: {len(all_defs_raw)}")

    # ── Paso 2: enriquecer en lotes de 200 con includeLatestBuilds=true ──────
    result: List[Dict] = []
    batch_size = 200
    for i in range(0, len(all_defs_raw), batch_size):
        batch     = all_defs_raw[i : i + batch_size]
        batch_ids = ",".join(str(d["id"]) for d in batch)
        data = api_get(
            url, headers,
            {"api-version": API_VERSION, "definitionIds": batch_ids, "includeLatestBuilds": "true"},
            debug,
        )
        if data:
            result.extend(data.get("value", []))
        elif debug:
            print(f"[DEBUG CI] Lote {i // batch_size + 1}: sin datos")

    if debug:
        print(f"[DEBUG CI] Definitions enriquecidas: {len(result)}")

    return result


def build_ci_row(defn: Dict, inactive_days: int, tz_name: str) -> Dict:
    queue_status   = defn.get("queueStatus", "enabled")
    modified_raw   = defn.get("modifiedDate", "")
    latest_build   = defn.get("latestCompletedBuild") or defn.get("latestBuild") or {}
    last_run_raw   = latest_build.get("finishTime", "")
    last_run_dt    = parse_date(last_run_raw)
    days_inactive  = days_since(last_run_dt)

    # Deprecado: deshabilitado explícitamente OR sin ejecución en inactive_days
    deprecated = (
        queue_status == "disabled"
        or (days_inactive is not None and days_inactive > inactive_days)
        or (days_inactive is None)  # nunca ejecutado
    )

    path = defn.get("path", "")
    return {
        "tipo":           "CI",
        "id":             defn.get("id", ""),
        "nombre":         defn.get("name", ""),
        "path":           path,
        "grupo":          classify_pipeline_group(path),
        "estado":         QUEUE_STATUS_LABEL.get(queue_status, queue_status),
        "queue_status":   queue_status,
        "deprecado":      DEPRECADO_SI if deprecated else DEPRECADO_NO,
        "ultima_act":     fmt_date(modified_raw, tz_name),
        "ultima_act_raw": modified_raw,
        "ultimo_run":     fmt_date(last_run_raw, tz_name),
        "ultimo_run_raw": last_run_raw,
        "dias_inactivo":  str(days_inactive) if days_inactive is not None else "Nunca",
        "url":            defn.get("url", ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CD — FETCH
# ═══════════════════════════════════════════════════════════════════════════════
def get_cd_definitions(org: str, project: str, headers: Dict, debug: bool = False) -> List[Dict]:
    """Obtiene TODAS las release definitions con paginación via x-ms-continuationtoken."""
    org_name = org.rstrip("/").split("/")[-1]
    url = f"https://vsrm.dev.azure.com/{org_name}/{quote(project, safe='')}/_apis/release/definitions"
    params = {"api-version": API_VERSION, "$top": 1000}
    return api_get_paginated(url, headers, params, debug)


def get_latest_release(defn_id: int, org: str, project: str, headers: Dict, debug: bool) -> Optional[Dict]:
    org_name = org.rstrip("/").split("/")[-1]
    url = (f"https://vsrm.dev.azure.com/{org_name}/{quote(project, safe='')}/_apis/release/releases")
    params = {
        "api-version":   API_VERSION,
        "definitionId":  defn_id,
        "$top":          1,
        "$expand":       "none",
    }
    data = api_get(url, headers, params, debug)
    releases = data.get("value", []) if data else []
    return releases[0] if releases else None


def _cd_worker(defn: Dict, org: str, project: str, headers: Dict,
               inactive_days: int, tz_name: str, debug: bool) -> Dict:
    latest    = get_latest_release(defn["id"], org, project, headers, debug)
    modified  = defn.get("modifiedOn", "")
    last_raw  = latest.get("createdOn", "") if latest else ""
    last_dt   = parse_date(last_raw)
    days_inac = days_since(last_dt)

    deprecated = (days_inac is None) or (days_inac > inactive_days)

    # Estado inferido por actividad
    if days_inac is None:
        estado = "⬜ Sin ejecución"
    elif days_inac <= 30:
        estado = "✅ Activo"
    elif days_inac <= inactive_days:
        estado = "⚠️  Inactivo"
    else:
        estado = "🔴 Sin uso"

    path = defn.get("path", "")
    return {
        "tipo":           "CD",
        "id":             defn.get("id", ""),
        "nombre":         defn.get("name", ""),
        "path":           path,
        "grupo":          classify_pipeline_group(path),
        "estado":         estado,
        "queue_status":   "",
        "deprecado":      DEPRECADO_SI if deprecated else DEPRECADO_NO,
        "ultima_act":     fmt_date(modified, tz_name),
        "ultima_act_raw": modified,
        "ultimo_run":     fmt_date(last_raw, tz_name),
        "ultimo_run_raw": last_raw,
        "dias_inactivo":  str(days_inac) if days_inac is not None else "Nunca",
        "url":            defn.get("url", ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT — RICH
# ═══════════════════════════════════════════════════════════════════════════════
def _row_style(row: Dict) -> str:
    if row["deprecado"] == DEPRECADO_SI:
        return "dim"
    return ""


def print_rich_table(console: "Console", rows: List[Dict]):
    table = Table(
        title="📊  Pipeline Status Report — CI + CD",
        title_style="bold cyan",
        header_style="bold white",
        border_style="dim",
        box=box.ROUNDED,
        show_lines=False,
    )
    table.add_column("#",           style="dim",          width=4,  justify="right")
    table.add_column("Tipo",        justify="center",     width=5)
    table.add_column("Nombre",      style="bold white",   min_width=35)
    table.add_column("Grupo",       justify="center",     min_width=16)
    table.add_column("Estado",      min_width=18)
    table.add_column("Deprecado",   justify="center",     min_width=10)
    table.add_column("Última Act.", style="dim",          min_width=17)
    table.add_column("Último Run",  style="dim",          min_width=17)
    table.add_column("Días Inact.", justify="right",      min_width=11)

    for idx, row in enumerate(rows, 1):
        tipo_label  = "[blue]CI[/]" if row["tipo"] == "CI" else "[magenta]CD[/]"
        dep_label   = "[bold red]⚠️  Sí[/]" if row["deprecado"] == DEPRECADO_SI else "[green]No[/]"
        grupo       = row.get("grupo", GRUPO_SIN_COINCIDENCIA)
        grupo_color = GRUPO_COLORS_RICH.get(grupo, "dim")
        grupo_label = f"[{grupo_color}]{grupo}[/]"
        dias        = row["dias_inactivo"]
        dias_label  = (
            f"[bold red]{dias}[/]" if dias != "Nunca" and int(dias) > 90
            else f"[yellow]{dias}[/]" if dias != "Nunca" and int(dias) > 30
            else f"[dim]{dias}[/]"
        ) if dias != "Nunca" else "[bold red]Nunca[/]"

        table.add_row(
            str(idx),
            tipo_label,
            row["nombre"],
            grupo_label,
            row["estado"],
            dep_label,
            row["ultima_act"],
            row["ultimo_run"],
            dias_label,
        )

    console.print(table)
    console.print()


def print_rich_summary(console: "Console", rows: List[Dict], elapsed: float, inactive_days: int):
    ci_rows   = [r for r in rows if r["tipo"] == "CI"]
    cd_rows   = [r for r in rows if r["tipo"] == "CD"]
    total     = len(rows)
    dep_total = sum(1 for r in rows if r["deprecado"] == DEPRECADO_SI)
    disabled  = sum(1 for r in ci_rows if r["queue_status"] == "disabled")
    paused    = sum(1 for r in ci_rows if r["queue_status"] == "paused")
    sin_act   = sum(1 for r in rows if r["dias_inactivo"] == "Nunca"
                    or (r["dias_inactivo"] != "Nunca" and int(r["dias_inactivo"]) > inactive_days))

    console.print(Panel(
        f"[bold white]📊 Total pipelines:[/]        [cyan]{total}[/]  "
        f"([blue]CI: {len(ci_rows)}[/] | [magenta]CD: {len(cd_rows)}[/])\n"
        f"[bold green]✅ Activos:[/]                [green]{total - dep_total}[/]\n"
        f"[bold red]🔴 Deshabilitados (CI):[/]    [red]{disabled}[/]\n"
        f"[bold yellow]⏸️  Pausados (CI):[/]          [yellow]{paused}[/]\n"
        f"[bold red]⚠️  Deprecados:[/]             [red]{dep_total}[/]  "
        f"([blue]CI: {sum(1 for r in ci_rows if r['deprecado']==DEPRECADO_SI)}[/] | "
        f"[magenta]CD: {sum(1 for r in cd_rows if r['deprecado']==DEPRECADO_SI)}[/])\n"
        f"[dim]📅 Sin actividad >{inactive_days}d:    {sin_act}[/]\n"
        f"[dim]⏱️  Tiempo: {elapsed:.2f}s[/]",
        title="📋 Resumen",
        border_style="blue",
        expand=False,
    ))
    console.print()


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT — FALLBACK
# ═══════════════════════════════════════════════════════════════════════════════
def print_plain_table(rows: List[Dict], elapsed: float, inactive_days: int):
    hdr = f"{'#':>4}  {'T':^3}  {'Nombre':<38} {'Grupo':<16} {'Estado':<20} {'Dep':^5} {'Última Act.':<17} {'Último Run':<17} Días"
    sep = "=" * len(hdr)
    print(f"\n{sep}\n{hdr}\n{sep}")
    for idx, row in enumerate(rows, 1):
        dep   = "SI" if row["deprecado"] == DEPRECADO_SI else "No"
        grupo = row.get("grupo", GRUPO_SIN_COINCIDENCIA)[:16]
        print(f"{idx:>4}  {row['tipo']:^3}  {row['nombre']:<38} {grupo:<16} {row['estado'][:20]:<20} "
              f"{dep:^5} {row['ultima_act']:<17} {row['ultimo_run']:<17} {row['dias_inactivo']}")
    ci_rows = [r for r in rows if r["tipo"] == "CI"]
    cd_rows = [r for r in rows if r["tipo"] == "CD"]
    dep     = sum(1 for r in rows if r["deprecado"] == DEPRECADO_SI)
    print(f"\nTotal: {len(rows)} (CI:{len(ci_rows)} CD:{len(cd_rows)}) | "
          f"Deprecados: {dep} | ⏱️ {elapsed:.2f}s\n")


# ═══════════════════════════════════════════════════════════════════════════════
# CACHE
# ═══════════════════════════════════════════════════════════════════════════════
def _find_latest_cache() -> Optional[Path]:
    output_dir = get_output_dir("outcome")
    cache_dir  = Path(str(output_dir)) / ".cache"
    files      = sorted(glob.glob(str(cache_dir / f"{SCRIPT_NAME}_raw_*.json")), reverse=True)
    return Path(files[0]) if files else None


def _cache_is_fresh(cache_path: Optional[Path]) -> bool:
    if not cache_path or not cache_path.exists():
        return False
    age_hours = (time.time() - cache_path.stat().st_mtime) / 3600
    return age_hours < CACHE_TTL_HOURS


def _load_cache(cache_path: Path) -> Dict:
    with open(cache_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_cache(rows: List[Dict], org: str, project: str, inactive_days: int) -> Path:
    output_dir = Path(str(get_output_dir("outcome")))
    cache_dir  = output_dir / ".cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    cache_path = cache_dir / f"{SCRIPT_NAME}_raw_{ts}.json"
    payload    = {
        "metadata": {
            "script":        SCRIPT_NAME,
            "version":       __version__,
            "org":           org,
            "project":       project,
            "inactive_days": inactive_days,
            "generated_at":  datetime.now(timezone.utc).isoformat(),
        },
        "rows": rows,
    }
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return cache_path


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL CHARTS
# ═══════════════════════════════════════════════════════════════════════════════
def _add_excel_charts(filepath: str, rows: List[Dict]) -> None:
    """Agrega pestanas 'Resumen' y 'Charts' con 4 graficos nativos al workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, PieChart, RadarChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    wb = load_workbook(filepath)

    # ── Hoja Resumen: tablas de datos para los graficos ─────────────────────────
    ws = wb.create_sheet("Resumen")

    ci_rows = [r for r in rows if r["tipo"] == "CI"]
    cd_rows = [r for r in rows if r["tipo"] == "CD"]

    # Tabla 1: Estado CI (col A-B, filas 1-5)
    ws["A1"] = "Estado CI";  ws["B1"] = "Cantidad"
    ws["A1"].font = Font(bold=True)
    ci_estados = {"Activo": 0, "Pausado": 0, "Deshabilitado": 0}
    for r in ci_rows:
        qs = r.get("queue_status", "enabled")
        lbl = {"enabled": "Activo", "paused": "Pausado", "disabled": "Deshabilitado"}.get(qs, "Activo")
        ci_estados[lbl] = ci_estados.get(lbl, 0) + 1
    for i, (lbl, cnt) in enumerate(ci_estados.items(), 2):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=cnt)

    # Tabla 2: Estado CD (col A-B, filas 7-12)
    ws["A7"] = "Estado CD";  ws["B7"] = "Cantidad"
    ws["A7"].font = Font(bold=True)
    cd_estados = {"Activo (<30d)": 0, "Inactivo (31-90d)": 0, "Sin uso (>90d)": 0, "Sin ejecucion": 0}
    for r in cd_rows:
        est = r.get("estado", "")
        if "✅" in est:   cd_estados["Activo (<30d)"] += 1
        elif "⚠" in est: cd_estados["Inactivo (31-90d)"] += 1
        elif "🔴" in est: cd_estados["Sin uso (>90d)"] += 1
        else:              cd_estados["Sin ejecucion"] += 1
    for i, (lbl, cnt) in enumerate(cd_estados.items(), 8):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=cnt)

    # Tabla 3: Distribucion por bucket CI vs CD (col A-C, filas 14-21)
    ws["A14"] = "Bucket"; ws["B14"] = "CI"; ws["C14"] = "CD"
    ws["A14"].font = Font(bold=True)
    ci_bkt = {b: 0 for b in BUCKETS}
    cd_bkt = {b: 0 for b in BUCKETS}
    for r in ci_rows:
        ci_bkt[_bucket(r["dias_inactivo"])] += 1
    for r in cd_rows:
        cd_bkt[_bucket(r["dias_inactivo"])] += 1
    for i, b in enumerate(BUCKETS, 15):
        ws.cell(row=i, column=1, value=b)
        ws.cell(row=i, column=2, value=ci_bkt[b])
        ws.cell(row=i, column=3, value=cd_bkt[b])

    # Tabla 4: Resumen ejecutivo (col A-B, filas 23-29)
    ws["A23"] = "Categoria"; ws["B23"] = "Cantidad"
    ws["A23"].font = Font(bold=True)
    dep_ci = sum(1 for r in ci_rows if r["deprecado"] == DEPRECADO_SI)
    dep_cd = sum(1 for r in cd_rows if r["deprecado"] == DEPRECADO_SI)
    resumen_data = [
        ("Total CI",        len(ci_rows)),
        ("Total CD",        len(cd_rows)),
        ("Deprecados CI",   dep_ci),
        ("Deprecados CD",   dep_cd),
        ("Activos CI",      len(ci_rows) - dep_ci),
        ("Activos CD",      len(cd_rows) - dep_cd),
    ]
    for i, (lbl, cnt) in enumerate(resumen_data, 24):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=cnt)

    # Tabla 5: Distribución por Grupo (col E-G, filas 1-7)
    grupos_order = [g for g, _ in _GRUPO_RULES] + [GRUPO_SIN_COINCIDENCIA]
    ws["E1"] = "Grupo";  ws["F1"] = "CI";  ws["G1"] = "CD"
    ws["E1"].font = Font(bold=True)
    grupo_ci = {g: sum(1 for r in ci_rows if r.get("grupo") == g) for g in grupos_order}
    grupo_cd = {g: sum(1 for r in cd_rows if r.get("grupo") == g) for g in grupos_order}
    for i, g in enumerate(grupos_order, 2):
        ws.cell(row=i, column=5, value=g)
        ws.cell(row=i, column=6, value=grupo_ci[g])
        ws.cell(row=i, column=7, value=grupo_cd[g])
        fill_color = GRUPO_COLORS_EXCEL.get(g, "bdc3c7")
        for col in (5, 6, 7):
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=fill_color)

    # ── Hoja Consolidado — Estilo ejecutivo sobrio ────────────────────────────
    # Paleta ejecutiva
    _NAV1  = "1a2e4a"   # marino oscuro — título / total
    _NAV2  = "243d5c"   # marino medio — headers columna
    _WHT   = "FFFFFF"
    _CHR   = "2c3e50"   # charcoal — texto datos
    # Tintes muy suaves para cada grupo (fondo de fila)
    _TINTS = {
        "WMS": "eaf7f4", "OMS": "eafaf1", "CSC": "fefce8",
        "TMS": "f4ecfb", GRUPO_SIN_COINCIDENCIA: "f7f8f8",
    }

    def _brd_accent(color_hex):
        _s = Side(style="thin",   color="d5d8dc")
        return Border(left=Side(style="medium", color=color_hex),
                      right=_s, top=_s, bottom=_s)

    def _brd_plain():
        _s = Side(style="thin", color="d5d8dc")
        return Border(left=_s, right=_s, top=_s, bottom=_s)

    wcon = wb.create_sheet("Consolidado")
    _NUM_COLS = 6
    _CON_HDRS   = ["Grupo", "Total Pipeline CI", "Total Pipeline CD",
                   "Total Deprecados", "Total Propuestos para Deprecar", "Total Pipelines"]
    _CON_WIDTHS = [24, 20, 20, 22, 30, 18]

    # Fila 1: Título fusionado
    wcon.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NUM_COLS)
    _tc = wcon.cell(row=1, column=1, value="Resumen Ejecutivo de Pipelines por Grupo")
    _tc.font      = Font(name="Calibri", bold=True, size=13, color=_WHT)
    _tc.fill      = PatternFill("solid", fgColor=_NAV1)
    _tc.alignment = Alignment(horizontal="center", vertical="center")
    wcon.row_dimensions[1].height = 30

    # Fila 2: Cabeceras de columna
    for c_i, (hdr, w) in enumerate(zip(_CON_HDRS, _CON_WIDTHS), 1):
        cell = wcon.cell(row=2, column=c_i, value=hdr)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHT)
        cell.fill      = PatternFill("solid", fgColor=_NAV2)
        cell.border    = _brd_plain()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wcon.column_dimensions[get_column_letter(c_i)].width = w
    wcon.row_dimensions[2].height = 36

    # Filas de datos (comienzan en fila 3)
    _grupos_con = [g for g, _ in _GRUPO_RULES] + [GRUPO_SIN_COINCIDENCIA]
    for r_i, grupo in enumerate(_grupos_con, 3):
        g_ci   = [r for r in ci_rows if r.get("grupo") == grupo]
        g_cd   = [r for r in cd_rows if r.get("grupo") == grupo]
        g_dep  = sum(1 for r in g_ci + g_cd if r["deprecado"] == DEPRECADO_SI)
        g_prop = (
            sum(1 for r in g_ci if r.get("queue_status") == "enabled"
                and r["dias_inactivo"] == "Nunca")
            + sum(1 for r in g_cd if r["dias_inactivo"] == "Nunca")
        )
        g_total  = len(g_ci) + len(g_cd)
        row_vals = [grupo, len(g_ci), len(g_cd), g_dep, g_prop, g_total]
        tint     = _TINTS.get(grupo, "f7f8f8")
        accent   = GRUPO_COLORS_EXCEL.get(grupo, "95a5a6")
        for c_i, val in enumerate(row_vals, 1):
            cell = wcon.cell(row=r_i, column=c_i, value=val)
            cell.fill   = PatternFill("solid", fgColor=tint)
            cell.border = _brd_accent(accent)
            if c_i == 1:
                cell.font      = Font(name="Calibri", bold=True, size=10, color=_CHR)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.font          = Font(name="Calibri", size=10, color=_CHR)
                cell.alignment     = Alignment(horizontal="center", vertical="center")
                cell.number_format = "#,##0"
        wcon.row_dimensions[r_i].height = 22

    # Fila TOTAL
    r_tot        = len(_grupos_con) + 3
    tot_ci_all   = len(ci_rows)
    tot_cd_all   = len(cd_rows)
    tot_dep_all  = sum(1 for r in rows if r["deprecado"] == DEPRECADO_SI)
    tot_prop_all = (
        sum(1 for r in ci_rows if r.get("queue_status") == "enabled"
            and r["dias_inactivo"] == "Nunca")
        + sum(1 for r in cd_rows if r["dias_inactivo"] == "Nunca")
    )
    tot_total_all = len(rows)
    tot_vals = ["TOTAL", tot_ci_all, tot_cd_all, tot_dep_all, tot_prop_all, tot_total_all]
    for c_i, val in enumerate(tot_vals, 1):
        cell = wcon.cell(row=r_tot, column=c_i, value=val)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHT)
        cell.fill      = PatternFill("solid", fgColor=_NAV1)
        cell.border    = _brd_plain()
        if c_i == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            cell.alignment     = Alignment(horizontal="center", vertical="center")
            cell.number_format = "#,##0"
    wcon.row_dimensions[r_tot].height = 24

    # Mover "Consolidado" a la segunda posición (después de "Datos")
    wb.move_sheet("Consolidado", offset=-(len(wb.sheetnames) - 2))

    # ── Hoja Charts ───────────────────────────────────────────────────
    wc = wb.create_sheet("Charts")

    # Grafico 1: Donut Estado CI
    c1 = PieChart()
    c1.title  = "Estado Pipelines CI"
    c1.style  = 10
    c1.width  = 15
    c1.height = 13
    c1.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    c1.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ci_colors = ["2ecc71", "f1c40f", "e74c3c"]
    for j, color in enumerate(ci_colors):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = color
        c1.series[0].data_points.append(pt)
    c1.series[0].dLbls = DataLabelList()
    c1.series[0].dLbls.showPercent  = True
    c1.series[0].dLbls.showCatName  = True
    c1.series[0].dLbls.showVal      = False
    wc.add_chart(c1, "A1")

    # Grafico 2: Donut Estado CD
    c2 = PieChart()
    c2.title  = "Estado Pipelines CD"
    c2.style  = 10
    c2.width  = 15
    c2.height = 13
    c2.add_data(Reference(ws, min_col=2, min_row=7, max_row=11), titles_from_data=True)
    c2.set_categories(Reference(ws, min_col=1, min_row=8, max_row=11))
    cd_colors = ["2ecc71", "f39c12", "e74c3c", "7f8c8d"]
    for j, color in enumerate(cd_colors):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = color
        c2.series[0].data_points.append(pt)
    c2.series[0].dLbls = DataLabelList()
    c2.series[0].dLbls.showPercent  = True
    c2.series[0].dLbls.showCatName  = True
    c2.series[0].dLbls.showVal      = False
    wc.add_chart(c2, "I1")

    # Grafico 3: Barras agrupadas CI vs CD por bucket
    c3 = BarChart()
    c3.type      = "col"
    c3.grouping  = "clustered"
    c3.title     = "Distribucion de Inactividad — CI vs CD"
    c3.y_axis.title = "Pipelines"
    c3.x_axis.title = "Dias inactivo"
    c3.width     = 24
    c3.height    = 13
    cats3 = Reference(ws, min_col=1, min_row=15, max_row=20)
    data3_ci = Reference(ws, min_col=2, min_row=14, max_row=20)
    data3_cd = Reference(ws, min_col=3, min_row=14, max_row=20)
    c3.add_data(data3_ci, titles_from_data=True)
    c3.add_data(data3_cd, titles_from_data=True)
    c3.set_categories(cats3)
    c3.series[0].graphicalProperties.solidFill = "3498db"
    c3.series[1].graphicalProperties.solidFill = "8e44ad"
    c3.series[0].dLbls = DataLabelList(); c3.series[0].dLbls.showVal = True
    c3.series[1].dLbls = DataLabelList(); c3.series[1].dLbls.showVal = True
    wc.add_chart(c3, "A16")

    # Grafico 4: Barras Resumen ejecutivo
    c4 = BarChart()
    c4.type      = "bar"
    c4.grouping  = "clustered"
    c4.title     = "Resumen Ejecutivo CI + CD"
    c4.x_axis.title = "Cantidad"
    c4.width     = 15
    c4.height    = 13
    cats4  = Reference(ws, min_col=1, min_row=24, max_row=29)
    data4  = Reference(ws, min_col=2, min_row=23, max_row=29)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    exec_colors = ["3498db", "8e44ad", "e74c3c", "c0392b", "2ecc71", "27ae60"]
    for j, color in enumerate(exec_colors):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = color
        c4.series[0].data_points.append(pt)
    c4.series[0].dLbls = DataLabelList(); c4.series[0].dLbls.showVal = True
    wc.add_chart(c4, "I16")

    # Grafico 5: Radar — Distribucion por Grupo CI vs CD
    c5 = RadarChart()
    c5.type   = "filled"
    c5.title  = "Pipelines por Grupo — CI vs CD"
    c5.style  = 10
    c5.width  = 18
    c5.height = 14
    n_grupos  = len(grupos_order)
    cats5    = Reference(ws, min_col=5, min_row=2, max_row=1 + n_grupos)
    data5_ci = Reference(ws, min_col=6, min_row=1, max_row=1 + n_grupos)
    data5_cd = Reference(ws, min_col=7, min_row=1, max_row=1 + n_grupos)
    c5.add_data(data5_ci, titles_from_data=True)
    c5.add_data(data5_cd, titles_from_data=True)
    c5.set_categories(cats5)
    c5.series[0].graphicalProperties.solidFill = "3498db"
    c5.series[1].graphicalProperties.solidFill = "8e44ad"
    wc.add_chart(c5, "A32")

    wb.save(filepath)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
EXPORT_FIELDS = ["tipo", "id", "nombre", "path", "grupo", "estado", "deprecado",
                 "ultima_act", "ultimo_run", "dias_inactivo", "url"]


def export_results(rows: List[Dict], output_format: str, tz_name: str) -> Optional[str]:
    outcome_dir = str(get_output_dir("outcome"))
    os.makedirs(outcome_dir, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    flat = [{f: r.get(f, "") for f in EXPORT_FIELDS} for r in rows]

    if output_format == "json":
        filepath = os.path.join(outcome_dir, f"pipeline_status_{ts}.json")
        ci_rows  = [r for r in rows if r["tipo"] == "CI"]
        cd_rows  = [r for r in rows if r["tipo"] == "CD"]
        payload  = {
            "metadata": {
                "tool": "cicd_pipeline_status", "version": __version__,
                "generated_at": datetime.now(_safe_tz(tz_name)).isoformat(),
            },
            "summary": {
                "total": len(rows),
                "ci":    len(ci_rows),
                "cd":    len(cd_rows),
                "deprecated": sum(1 for r in rows if r["deprecado"] == DEPRECADO_SI),
            },
            "data": flat,
        }
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return filepath

    elif output_format == "csv":
        filepath = os.path.join(outcome_dir, f"pipeline_status_{ts}.csv")
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=EXPORT_FIELDS)
            writer.writeheader()
            writer.writerows(flat)
        return filepath

    elif output_format == "excel":
        try:
            import pandas as pd
            filepath = os.path.join(outcome_dir, f"pipeline_status_{ts}.xlsx")
            pd.DataFrame(flat).to_excel(filepath, index=False,
                                        sheet_name="Datos", engine="openpyxl")
            _add_excel_charts(filepath, rows)
            return filepath
        except ImportError:
            print("ERROR: Instala pandas y openpyxl para exportar a Excel.")
            return None

    return None


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    start_time = time.time()

    if not REQUESTS_AVAILABLE:
        print("ERROR: pip install requests rich pandas openpyxl")
        return

    args    = get_args()
    console = Console() if RICH_AVAILABLE else None
    headers = make_headers(args.pat)

    tz_name = args.timezone
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = DEFAULT_TIMEZONE

    revision_time = datetime.now(_safe_tz(tz_name)).strftime(f"%Y-%m-%d %H:%M:%S ({tz_name})")

    # ── Cache check ──────────────────────────────────────────────────────────
    cache_used = False
    cache_path = _find_latest_cache()

    if args.use_cache_only:
        if cache_path and _cache_is_fresh(cache_path):
            cache_used = True
        else:
            msg = "❌ Cache no disponible para --use-cache-only (no existe o > 24h)."
            (console.print(f"[red]{msg}[/]") if console else print(msg))
            return
    elif not args.force_refresh and cache_path and _cache_is_fresh(cache_path):
        cache_used = True

    cache_label = (
        f"[green]💾 Cache[/] ({cache_path.name})" if cache_used
        else "[yellow]🔄 APIs[/]" if not args.force_refresh
        else "[yellow]🔄 Force refresh[/]"
    ) if RICH_AVAILABLE else ""

    if RICH_AVAILABLE and console:
        console.print()
        console.print(Panel(
            f"[bold cyan]📊  Pipeline Status Report[/]\n"
            f"[dim]🕐 {revision_time}[/]\n"
            f"[dim]🏢 Org:      {args.org}[/]\n"
            f"[dim]📁 Proyecto: {args.project}[/]\n"
            f"[dim]📅 Inactivo: >{args.inactive_days} días → deprecado[/]\n"
            f"[dim]📦 Fuente:   {cache_label}[/]",
            border_style="cyan",
            expand=False,
        ))
        console.print()
    else:
        src = "cache" if cache_used else "APIs"
        print(f"Pipeline Status Report | {revision_time} | fuente: {src}")

    rows: List[Dict] = []

    # ── Cargar desde cache si aplica ─────────────────────────────────────────
    if cache_used and cache_path:
        data  = _load_cache(cache_path)
        rows  = data.get("rows", [])
        msg   = f"📦 Cache cargado: {cache_path.name} ({len(rows)} pipelines)"
        (console.print(f"[dim]{msg}[/]\n") if console else print(msg))

    # ── CI ───────────────────────────────────────────────────────────────────
    if not cache_used and args.type in ("ci", "all"):
        if RICH_AVAILABLE and console:
            with Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console) as p:
                t = p.add_task("Obteniendo CI pipelines...", total=None)
                ci_defs = get_ci_definitions(args.org, args.project, headers, args.debug)
                p.update(t, description=f"✅ {len(ci_defs)} CI pipelines obtenidos (con último build incluido)")
        else:
            ci_defs = get_ci_definitions(args.org, args.project, headers, args.debug)
            print(f"{len(ci_defs)} CI pipelines")

        for d in ci_defs:
            rows.append(build_ci_row(d, args.inactive_days, tz_name))

        if console:
            console.print()

    # ── CD ───────────────────────────────────────────────────────────────────────
    if not cache_used and args.type in ("cd", "all"):
        if RICH_AVAILABLE and console:
            with Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console) as p:
                t = p.add_task("Obteniendo CD pipelines...", total=None)
                cd_defs = get_cd_definitions(args.org, args.project, headers, args.debug)
                p.update(t, description=f"✅ {len(cd_defs)} CD pipelines — consultando último release...")
        else:
            cd_defs = get_cd_definitions(args.org, args.project, headers, args.debug)
            print(f"{len(cd_defs)} CD pipelines — consultando último release...")

        if console:
            console.print()

        workers = max(1, min(args.workers, len(cd_defs))) if cd_defs else 1
        cd_lock = threading.Lock()

        if RICH_AVAILABLE and console:
            with Progress(
                SpinnerColumn(), TextColumn("{task.description}"),
                BarColumn(), TaskProgressColumn(), console=console,
            ) as p:
                t = p.add_task(f"Último release CD ({workers} hilos)...", total=len(cd_defs))
                with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
                    futures = {
                        pool.submit(
                            _cd_worker, d, args.org, args.project, headers,
                            args.inactive_days, tz_name, args.debug
                        ): d for d in cd_defs
                    }
                    for future in concurrent.futures.as_completed(futures):
                        result = future.result()
                        if result:
                            with cd_lock:
                                rows.append(result)
                        p.advance(t)
                p.update(t, description=f"✅ {len(cd_defs)} CD procesados")
            console.print()
        else:
            counter  = {"n": 0}
            cnt_lock = threading.Lock()
            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {
                    pool.submit(
                        _cd_worker, d, args.org, args.project, headers,
                        args.inactive_days, tz_name, args.debug
                    ): d for d in cd_defs
                }
                for future in concurrent.futures.as_completed(futures):
                    result = future.result()
                    if result:
                        with cnt_lock:
                            rows.append(result)
                            counter["n"] += 1
                            print(f"  [{counter['n']}/{len(cd_defs)}] CD procesados...", end="\r")
            print()

    # ── Guardar cache tras consulta API ─────────────────────────────────
    if not cache_used and rows:
        saved = _save_cache(rows, args.org, args.project, args.inactive_days)
        msg   = f"💾 Cache guardado: {saved.name}"
        (console.print(f"[dim]{msg}[/]") if console else print(msg))

    elapsed = time.time() - start_time

    if not rows:
        msg = "❌ No se encontraron pipelines."
        (console.print(f"[red]{msg}[/]") if console else print(msg))
        return

    # ── Filtro --only-deprecated ──────────────────────────────────────────────
    if args.only_deprecated:
        rows = [r for r in rows if r["deprecado"] == DEPRECADO_SI]
        if console:
            console.print(f"[dim]🔍 Filtrado: {len(rows)} pipelines deprecados[/]\n")

    # ── Ordenar: tipo asc → deprecado desc → nombre asc ──────────────────────
    rows.sort(key=lambda r: (
        r["tipo"],
        0 if r["deprecado"] == DEPRECADO_SI else 1,
        r["nombre"].lower()
    ))

    # ── Mostrar ───────────────────────────────────────────────────────────────
    if RICH_AVAILABLE and console:
        print_rich_table(console, rows)
        print_rich_summary(console, rows, elapsed, args.inactive_days)
    else:
        print_plain_table(rows, elapsed, args.inactive_days)

    # ── Exportar ──────────────────────────────────────────────────────────────
    if args.output:
        filepath = export_results(rows, args.output, tz_name)
        if filepath:
            msg = f"📁 Exportado: {filepath}"
            (console.print(f"[bold green]{msg}[/]\n") if console else print(msg))


if __name__ == "__main__":
    main()

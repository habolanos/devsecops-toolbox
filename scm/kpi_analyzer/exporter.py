#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KPI Analyzer Pro Exporter
Exporta datos a múltiples formatos: JSON, CSV, HTML, Excel
"""

import sys
import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import logging

try:
    from rich.console import Console
    from rich.progress import Progress
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

console = Console() if RICH_AVAILABLE else None

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p


class ExporterPro:
    """Exporta datos a múltiples formatos profesionales"""
    
    def __init__(self, data: Dict[str, Any], output_dir: Optional[str] = None):
        self.data = data
        self.output_dir = Path(output_dir) if output_dir else get_output_dir("outcome/kpi_analyzer")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def to_json(self, filepath: Optional[str] = None) -> bool:
        """Exporta a JSON"""
        try:
            if filepath is None:
                filepath = self.output_dir / f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            else:
                filepath = Path(filepath)
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"✅ Exportado a JSON: {filepath}")
            if RICH_AVAILABLE and console:
                console.print(f"[green]✅ Exportado a JSON: {filepath}[/green]")
            return True
        except Exception as e:
            logger.error(f"Error exportando JSON: {e}")
            if RICH_AVAILABLE and console:
                console.print(f"[red]❌ Error exportando JSON: {e}[/red]")
            return False
    
    def to_csv(self, filepath: Optional[str] = None) -> bool:
        """Exporta a CSV"""
        try:
            if filepath is None:
                filepath = self.output_dir / f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            else:
                filepath = Path(filepath)
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Aplanar datos para CSV
            rows = self._flatten_data(self.data)
            
            if not rows:
                logger.warning("No hay datos para exportar a CSV")
                return False
            
            # Obtener todas las claves
            fieldnames = set()
            for row in rows:
                fieldnames.update(row.keys())
            fieldnames = sorted(list(fieldnames))
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            
            logger.info(f"✅ Exportado a CSV: {filepath}")
            if RICH_AVAILABLE and console:
                console.print(f"[green]✅ Exportado a CSV: {filepath}[/green]")
            return True
        except Exception as e:
            logger.error(f"Error exportando CSV: {e}")
            if RICH_AVAILABLE and console:
                console.print(f"[red]❌ Error exportando CSV: {e}[/red]")
            return False
    
    def to_html(self, filepath: Optional[str] = None) -> bool:
        """Exporta a HTML profesional"""
        try:
            if filepath is None:
                filepath = self.output_dir / f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            else:
                filepath = Path(filepath)
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            html_content = self._generate_html()
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"✅ Exportado a HTML: {filepath}")
            if RICH_AVAILABLE and console:
                console.print(f"[green]✅ Exportado a HTML: {filepath}[/green]")
            return True
        except Exception as e:
            logger.error(f"Error exportando HTML: {e}")
            if RICH_AVAILABLE and console:
                console.print(f"[red]❌ Error exportando HTML: {e}[/red]")
            return False
    
    def to_excel(self, filepath: Optional[str] = None) -> bool:
        """Exporta a Excel con formatos profesionales"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl no está instalado. Instalando...")
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        
        try:
            if filepath is None:
                filepath = self.output_dir / f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            else:
                filepath = Path(filepath)
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "KPI Report"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplanar datos
            rows = self._flatten_data(self.data)
            
            if not rows:
                logger.warning("No hay datos para exportar a Excel")
                return False
            
            # Obtener todas las claves
            fieldnames = set()
            for row in rows:
                fieldnames.update(row.keys())
            fieldnames = sorted(list(fieldnames))
            
            # Escribir encabezados
            for col_idx, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = fieldname
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Escribir datos
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, fieldname in enumerate(fieldnames, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = row_data.get(fieldname, "")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar ancho de columnas
            for col_idx, fieldname in enumerate(fieldnames, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(fieldname) + 2)
            
            # Guardar
            wb.save(filepath)
            
            logger.info(f"✅ Exportado a Excel: {filepath}")
            if RICH_AVAILABLE and console:
                console.print(f"[green]✅ Exportado a Excel: {filepath}[/green]")
            return True
        except Exception as e:
            logger.error(f"Error exportando Excel: {e}")
            if RICH_AVAILABLE and console:
                console.print(f"[red]❌ Error exportando Excel: {e}[/red]")
            return False
    
    def export_all(self) -> Dict[str, bool]:
        """Exporta a todos los formatos"""
        results = {
            "json": self.to_json(),
            "csv": self.to_csv(),
            "html": self.to_html(),
            "excel": self.to_excel()
        }
        
        if RICH_AVAILABLE and console:
            console.print("\n[bold]📊 Resumen de Exportación:[/bold]")
            for fmt, success in results.items():
                status = "[green]✅[/green]" if success else "[red]❌[/red]"
                console.print(f"  {status} {fmt.upper()}")
        
        return results
    
    def _flatten_data(self, data: Dict, parent_key: str = '', sep: str = '_') -> List[Dict]:
        """Aplana datos anidados para CSV/Excel"""
        items = []
        
        def flatten(obj, prefix=''):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    new_key = f"{prefix}{sep}{k}" if prefix else k
                    if isinstance(v, (dict, list)):
                        flatten(v, new_key)
                    else:
                        items.append({new_key: v})
            elif isinstance(obj, list):
                for idx, item in enumerate(obj):
                    new_key = f"{prefix}[{idx}]"
                    if isinstance(item, (dict, list)):
                        flatten(item, new_key)
                    else:
                        items.append({new_key: item})
        
        flatten(data)
        
        # Consolidar items en un diccionario único
        if items:
            result = {}
            for item in items:
                result.update(item)
            return [result]
        
        return []
    
    def _generate_html(self) -> str:
        """Genera HTML profesional"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        html = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>KPI Report - DevSecOps Toolbox</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        
        .header p {{
            font-size: 1.1em;
            opacity: 0.9;
        }}
        
        .content {{
            padding: 40px;
        }}
        
        .section {{
            margin-bottom: 40px;
        }}
        
        .section h2 {{
            color: #667eea;
            border-bottom: 3px solid #667eea;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        
        th {{
            background: #f5f5f5;
            color: #333;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            border-bottom: 2px solid #667eea;
        }}
        
        td {{
            padding: 12px;
            border-bottom: 1px solid #eee;
        }}
        
        tr:hover {{
            background: #f9f9f9;
        }}
        
        .footer {{
            background: #f5f5f5;
            padding: 20px;
            text-align: center;
            color: #666;
            font-size: 0.9em;
        }}
        
        .metric {{
            display: inline-block;
            background: #f5f5f5;
            padding: 15px 25px;
            margin: 10px;
            border-radius: 5px;
            border-left: 4px solid #667eea;
        }}
        
        .metric-value {{
            font-size: 1.5em;
            font-weight: bold;
            color: #667eea;
        }}
        
        .metric-label {{
            font-size: 0.9em;
            color: #666;
            margin-top: 5px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 KPI Report</h1>
            <p>DevSecOps Toolbox - Reporte Profesional</p>
        </div>
        
        <div class="content">
            <div class="section">
                <h2>📈 Resumen de Datos</h2>
                <pre>{json.dumps(self.data, indent=2, ensure_ascii=False)}</pre>
            </div>
        </div>
        
        <div class="footer">
            <p>Generado: {timestamp}</p>
            <p>DevSecOps Toolbox v1.9.6</p>
        </div>
    </div>
</body>
</html>"""
        
        return html


def main():
    """Función principal"""
    import argparse
    
    parser = argparse.ArgumentParser(description="KPI Analyzer Pro Exporter")
    parser.add_argument("--format", choices=["json", "csv", "html", "excel", "all"], default="all", help="Formato de exportación")
    parser.add_argument("--output", help="Directorio de salida")
    
    args = parser.parse_args()
    
    # Datos de ejemplo
    sample_data = {
        "timestamp": datetime.now().isoformat(),
        "metrics": {
            "health_score": 85.5,
            "deployment_frequency": 2.5,
            "lead_time": 4.5,
            "mttr": 2.0,
            "change_failure_rate": 12.5
        }
    }
    
    exporter = ExporterPro(sample_data, output_dir=args.output)
    
    if args.format == "all":
        exporter.export_all()
    elif args.format == "json":
        exporter.to_json()
    elif args.format == "csv":
        exporter.to_csv()
    elif args.format == "html":
        exporter.to_html()
    elif args.format == "excel":
        exporter.to_excel()


if __name__ == "__main__":
    main()

"""
Generadores de Reportes para Service Accounts
Genera reportes en múltiples formatos (JSON, CSV, Excel, HTML)
"""

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


class BaseReportGenerator:
    """Clase base para generadores de reportes."""
    
    def __init__(self, output_dir: str = "outcome", debug: bool = False):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.debug = debug
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def _get_output_path(self, extension: str) -> Path:
        """Genera ruta de salida."""
        filename = f"sa_report_{self.timestamp}.{extension}"
        return self.output_dir / filename
    
    def _log(self, message: str):
        """Registra mensaje de debug."""
        if self.debug:
            print(f"[{self.__class__.__name__}] {message}")


class JSONReportGenerator(BaseReportGenerator):
    """Genera reportes en formato JSON."""
    
    def generate(self, data: Dict) -> str:
        """Genera reporte JSON."""
        output_path = self._get_output_path("json")
        
        report = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "format": "json",
                "version": "1.0.0"
            },
            "data": data
        }
        
        try:
            with open(output_path, 'w') as f:
                json.dump(report, f, indent=2, default=str)
            
            self._log(f"✅ Reporte JSON generado: {output_path}")
            return str(output_path)
        except Exception as e:
            self._log(f"❌ Error generando JSON: {e}")
            return None


class CSVReportGenerator(BaseReportGenerator):
    """Genera reportes en formato CSV."""
    
    def generate(self, data: Dict) -> str:
        """Genera reporte CSV."""
        output_path = self._get_output_path("csv")
        
        try:
            rows = self._flatten_data(data)
            
            if not rows:
                self._log("⚠️  No hay datos para generar CSV")
                return None
            
            # Obtener headers del primer row
            headers = rows[0].keys()
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)
            
            self._log(f"✅ Reporte CSV generado: {output_path}")
            return str(output_path)
        except Exception as e:
            self._log(f"❌ Error generando CSV: {e}")
            return None
    
    def _flatten_data(self, data: Dict) -> List[Dict]:
        """Aplana estructura de datos para CSV."""
        rows = []
        
        for project_data in data.get('by_project', {}).values():
            service_accounts = project_data.get('service_accounts', [])
            
            for sa in service_accounts:
                roles_analysis = sa.get('roles_analysis', {})
                bindings = roles_analysis.get('iam_bindings', [])
                
                if not bindings:
                    # Service account sin roles
                    rows.append({
                        'proyecto': project_data.get('project_id', ''),
                        'service_account': sa.get('email', ''),
                        'rol': 'N/A',
                        'titulo_rol': 'N/A',
                        'permisos': 0,
                        'otorgado_en': sa.get('created_at', ''),
                        'duracion_dias': 'N/A',
                        'fecha_expiracion': 'N/A',
                        'dias_restantes': 'N/A',
                        'nivel_riesgo': 'N/A'
                    })
                else:
                    # Service account con roles
                    for binding in bindings:
                        rows.append({
                            'proyecto': project_data.get('project_id', ''),
                            'service_account': sa.get('email', ''),
                            'rol': binding.get('role', ''),
                            'titulo_rol': binding.get('role_title', ''),
                            'permisos': binding.get('permission_count', 0),
                            'otorgado_en': binding.get('granted_at', ''),
                            'duracion_dias': binding.get('requested_duration_days', 'Permanente'),
                            'fecha_expiracion': binding.get('expiration_date', 'N/A'),
                            'dias_restantes': binding.get('days_remaining', 'N/A'),
                            'nivel_riesgo': binding.get('risk_level', 'N/A')
                        })
        
        return rows


class ExcelReportGenerator(BaseReportGenerator):
    """Genera reportes en formato Excel."""
    
    def generate(self, data: Dict) -> str:
        """Genera reporte Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self._log("⚠️  openpyxl no instalado. Instala con: pip install openpyxl")
            return None
        
        output_path = self._get_output_path("xlsx")
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remover hoja por defecto
            
            # Tab 1: Resumen Ejecutivo
            self._create_summary_tab(wb, data)
            
            # Tab 2: Roles por Service Account
            self._create_roles_tab(wb, data)
            
            # Tab 3: Roles Expirando Pronto
            self._create_expiring_soon_tab(wb, data)
            
            # Tab 4: Análisis de Riesgos
            self._create_risks_tab(wb, data)
            
            wb.save(output_path)
            self._log(f"✅ Reporte Excel generado: {output_path}")
            return str(output_path)
        except Exception as e:
            self._log(f"❌ Error generando Excel: {e}")
            return None
    
    def _create_summary_tab(self, wb, data: Dict):
        """Crea tab de resumen ejecutivo."""
        ws = wb.create_sheet("Resumen Ejecutivo")
        
        # Calcular métricas
        total_sa = sum(len(proj.get('service_accounts', []))
                      for proj in data.get('by_project', {}).values())
        total_roles = sum(len(sa.get('roles_analysis', {}).get('iam_bindings', []))
                         for proj in data.get('by_project', {}).values()
                         for sa in proj.get('service_accounts', []))
        
        # Escribir datos
        ws['A1'] = "RESUMEN DE SERVICE ACCOUNTS"
        ws['A3'] = "Métrica"
        ws['B3'] = "Valor"
        
        metrics = [
            ("Total Proyectos", len(data.get('by_project', {}))),
            ("Total Service Accounts", total_sa),
            ("Total Roles", total_roles),
        ]
        
        row = 4
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
    
    def _create_roles_tab(self, wb, data: Dict):
        """Crea tab de roles por service account."""
        ws = wb.create_sheet("Roles por SA")
        
        # Headers
        headers = ['Proyecto', 'Service Account', 'Rol', 'Duración', 'Expiración', 'Días Rest.', 'Riesgo']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        row = 2
        for project_data in data.get('by_project', {}).values():
            for sa in project_data.get('service_accounts', []):
                for binding in sa.get('roles_analysis', {}).get('iam_bindings', []):
                    ws.cell(row=row, column=1, value=project_data.get('project_id', ''))
                    ws.cell(row=row, column=2, value=sa.get('email', ''))
                    ws.cell(row=row, column=3, value=binding.get('role', ''))
                    ws.cell(row=row, column=4, value=binding.get('requested_duration_days', 'Permanente'))
                    ws.cell(row=row, column=5, value=binding.get('expiration_date', 'N/A'))
                    ws.cell(row=row, column=6, value=binding.get('days_remaining', 'N/A'))
                    ws.cell(row=row, column=7, value=binding.get('risk_level', 'N/A'))
                    row += 1
    
    def _create_expiring_soon_tab(self, wb, data: Dict):
        """Crea tab de roles expirando pronto."""
        ws = wb.create_sheet("Expirando Pronto")
        
        # Headers
        headers = ['Proyecto', 'Service Account', 'Rol', 'Expira En', 'Acción']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        row = 2
        for project_data in data.get('by_project', {}).values():
            for sa in project_data.get('service_accounts', []):
                for binding in sa.get('roles_analysis', {}).get('iam_bindings', []):
                    days_remaining = binding.get('days_remaining')
                    if days_remaining and 0 <= days_remaining < 30:
                        ws.cell(row=row, column=1, value=project_data.get('project_id', ''))
                        ws.cell(row=row, column=2, value=sa.get('email', ''))
                        ws.cell(row=row, column=3, value=binding.get('role', ''))
                        ws.cell(row=row, column=4, value=f"{days_remaining} días")
                        ws.cell(row=row, column=5, value="Renovar" if days_remaining > 7 else "URGENTE")
                        row += 1
    
    def _create_risks_tab(self, wb, data: Dict):
        """Crea tab de análisis de riesgos."""
        ws = wb.create_sheet("Riesgos")
        
        # Headers
        headers = ['Proyecto', 'Service Account', 'Rol', 'Riesgo', 'Factores', 'Recomendación']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        row = 2
        for project_data in data.get('by_project', {}).values():
            for sa in project_data.get('service_accounts', []):
                for binding in sa.get('roles_analysis', {}).get('iam_bindings', []):
                    ws.cell(row=row, column=1, value=project_data.get('project_id', ''))
                    ws.cell(row=row, column=2, value=sa.get('email', ''))
                    ws.cell(row=row, column=3, value=binding.get('role', ''))
                    ws.cell(row=row, column=4, value=binding.get('risk_level', ''))
                    ws.cell(row=row, column=5, value='; '.join(binding.get('risk_factors', [])))
                    
                    risk_level = binding.get('risk_level', '')
                    if risk_level in ['HIGH', 'CRITICAL']:
                        ws.cell(row=row, column=6, value="Revisar urgentemente")
                    else:
                        ws.cell(row=row, column=6, value="Monitorear")
                    
                    row += 1


class HTMLReportGenerator(BaseReportGenerator):
    """Genera reportes en formato HTML."""
    
    def generate(self, data: Dict) -> str:
        """Genera reporte HTML."""
        output_path = self._get_output_path("html")
        
        try:
            html_content = self._generate_html(data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self._log(f"✅ Reporte HTML generado: {output_path}")
            return str(output_path)
        except Exception as e:
            self._log(f"❌ Error generando HTML: {e}")
            return None
    
    def _generate_html(self, data: Dict) -> str:
        """Genera contenido HTML."""
        html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Reporte de Service Accounts</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 5px; }
        h1 { color: #333; border-bottom: 3px solid #007bff; padding-bottom: 10px; }
        h2 { color: #555; margin-top: 30px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th { background-color: #007bff; color: white; padding: 12px; text-align: left; }
        td { padding: 10px; border-bottom: 1px solid #ddd; }
        tr:hover { background-color: #f9f9f9; }
        .risk-critical { background-color: #ffcccc; }
        .risk-high { background-color: #ffe6e6; }
        .risk-medium { background-color: #fff3cd; }
        .risk-low { background-color: #d4edda; }
        .metric { display: inline-block; margin: 10px 20px; }
        .metric-value { font-size: 24px; font-weight: bold; color: #007bff; }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Reporte de Service Accounts Multi-Proyecto</h1>
        <p>Generado: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """</p>
        
        <h2>Resumen Ejecutivo</h2>
        <div>
            <div class="metric">
                <div>Total Proyectos</div>
                <div class="metric-value">""" + str(len(data.get('by_project', {}))) + """</div>
            </div>
            <div class="metric">
                <div>Total Service Accounts</div>
                <div class="metric-value">""" + str(sum(len(proj.get('service_accounts', []))
                                                        for proj in data.get('by_project', {}).values())) + """</div>
            </div>
        </div>
        
        <h2>Roles por Service Account</h2>
        <table>
            <tr>
                <th>Proyecto</th>
                <th>Service Account</th>
                <th>Rol</th>
                <th>Duración</th>
                <th>Expiración</th>
                <th>Días Restantes</th>
                <th>Riesgo</th>
            </tr>
"""
        
        for project_data in data.get('by_project', {}).values():
            for sa in project_data.get('service_accounts', []):
                for binding in sa.get('roles_analysis', {}).get('iam_bindings', []):
                    risk_class = f"risk-{binding.get('risk_level', 'low').lower()}"
                    html += f"""
            <tr class="{risk_class}">
                <td>{project_data.get('project_id', '')}</td>
                <td>{sa.get('email', '')}</td>
                <td>{binding.get('role', '')}</td>
                <td>{binding.get('requested_duration_days', 'Permanente')}</td>
                <td>{binding.get('expiration_date', 'N/A')}</td>
                <td>{binding.get('days_remaining', 'N/A')}</td>
                <td>{binding.get('risk_level', 'N/A')}</td>
            </tr>
"""
        
        html += """
        </table>
    </div>
</body>
</html>
"""
        return html

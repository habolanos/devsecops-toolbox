#!/usr/bin/env python3
# =============================================================================
# Script: generar-inventario-csv-combinar-a-excel.py
# Versión con gráficos consolidados por entorno y proyecto
# =============================================================================

import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.chart import BarChart, PieChart, RadarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import GradientFillProperties, GradientStop, PathShadeProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.text import RichText
from openpyxl.chart.title import Title
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# Detecta automáticamente la carpeta donde está el script
SCRIPT_DIR = Path(__file__).parent.resolve()
OUTCOME_DIR = SCRIPT_DIR / "outcome"

BASE_NAME = "Inventario_Completo_GKE_CloudSQL"
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = SCRIPT_DIR / f"{BASE_NAME}_{timestamp}.xlsx"

print("Script ubicado en :", SCRIPT_DIR)
print("Buscando CSVs en   :", OUTCOME_DIR)
print(f"Generando Excel    : {OUTPUT_EXCEL.name}")
print("=" * 90)

# Definición exacta de columnas por hoja
TIPOS = {
    "clusters":    ["PROYECTO", "NAME", "LOCATION", "VERSION", "CURRENT_VERSION", "STATUS", "MACHINE_TYPE"],
    "deployments": ["PROYECTO", "NAMESPACE", "CLUSTER", "DEPLOYMENT", "IMAGES"],
    "services":    ["PROYECTO", "NAMESPACE", "CLUSTER", "NAME", "TYPE", "CLUSTER-IP", "EXTERNAL-IP", "PORTS"],
    "cloudsql":        ["PROYECTO", "NAME", "DATABASE_VERSION", "REGION", "TIER", "STATE", "PUBLIC_IP", "PRIVATE_IP", "AUTO_RESIZE", "BACKUP_ENABLED"],
    "clouddatabases":  ["PROYECTO", "INSTANCE", "DATABASE", "CHARSET", "COLLATION"],
    "ingress":         ["PROYECTO", "NAMESPACE", "CLUSTER", "NAME", "HOSTS", "ADDRESS", "PORTS"],
    "cloudrun":        ["PROYECTO", "NAME", "REGION", "URL", "LAST_DEPLOYED", "IMAGE"],
    "pubsub":          ["PROYECTO", "NAME", "LABELS"]
}

# --- Helpers para extraer entorno y nombre base del proyecto ---
ENV_PATTERN = re.compile(r"-(dev|qa|stag|prod)-", re.IGNORECASE)

def extract_env(proyecto: str) -> str:
    m = ENV_PATTERN.search(proyecto)
    return m.group(1).upper() if m else "OTRO"

def extract_base(proyecto: str) -> str:
    m = ENV_PATTERN.search(proyecto)
    if m:
        return proyecto[:m.start()]
    return proyecto

# --- Colores para gráficos ---
ENV_COLORS = {"DEV": "4472C4", "QA": "ED7D31", "STAG": "70AD47", "PROD": "FF0000", "OTRO": "A5A5A5"}

# =============================================================================
# 1. Leer todos los CSVs y generar hojas de datos
# =============================================================================
archivos_encontrados = False
data_frames = {}

with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
    for tipo, columnas_esperadas in TIPOS.items():
        all_dfs = []
        print(f"\nProcesando → {tipo.upper()}")

        for csv_file in sorted(OUTCOME_DIR.rglob(f"*{tipo}.csv")):
            try:
                folder_name = csv_file.parent.name
                project_name = folder_name.replace("inventario-", "").split("-20")[0]

                if tipo in ("deployments", "clusters", "cloudrun", "pubsub"):
                    df = pd.read_csv(csv_file, sep=';', dtype=str, on_bad_lines='skip')
                else:
                    df = pd.read_csv(csv_file, sep=None, engine='python', dtype=str, on_bad_lines='skip')

                df.insert(0, "PROYECTO", project_name)
                df = df.reindex(columns=columnas_esperadas, fill_value="")
                df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

                all_dfs.append(df)
                print(f"  ✓ {csv_file.name} → {project_name} ({len(df)} filas)")
                archivos_encontrados = True

            except Exception as e:
                print(f"  ✗ Error leyendo {csv_file.name}: {e}")

        if all_dfs:
            combined_df = pd.concat(all_dfs, ignore_index=True)

            # Agregar columnas de entorno y nombre base
            combined_df["ENTORNO"] = combined_df["PROYECTO"].apply(extract_env)
            combined_df["BASE"] = combined_df["PROYECTO"].apply(extract_base)

            data_frames[tipo] = combined_df
            sheet_name = tipo.upper()[:31]
            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  → Hoja '{sheet_name}' creada correctamente ({len(combined_df)} filas)")
        else:
            print(f"  → No se encontraron archivos para '{tipo}'")

    # =========================================================================
    # 2. Generar hojas de gráficos (solo si hay datos)
    # =========================================================================
    if not data_frames:
        print("\n⚠️ No hay datos para generar gráficos.")
    else:
        wb = writer.book
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        TOTAL_FONT = Font(bold=True, size=11)

        def style_header(ws, row, max_col):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")

        def style_total_row(ws, row, max_col):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

        # =====================================================================
        # 2a. RESUMEN — Totales por categoría y entorno
        # =====================================================================
        print("\nGenerando → RESUMEN")
        ws_resumen = wb.create_sheet("RESUMEN")

        envs_resumen = ["DEV", "QA", "STAG"]
        tipos_list = list(TIPOS.keys())

        # Tabla: filas=tipos recurso, columnas=entornos + TOTAL
        ws_resumen["A1"] = "Recurso"
        for j, env in enumerate(envs_resumen, 2):
            ws_resumen.cell(row=1, column=j, value=env)
        ws_resumen.cell(row=1, column=len(envs_resumen) + 2, value="TOTAL")
        style_header(ws_resumen, 1, len(envs_resumen) + 2)

        for i, tipo in enumerate(tipos_list, 2):
            ws_resumen.cell(row=i, column=1, value=tipo.upper())
            row_total = 0
            for j, env in enumerate(envs_resumen, 2):
                df = data_frames.get(tipo, pd.DataFrame())
                cnt = len(df[df["ENTORNO"] == env]) if not df.empty else 0
                ws_resumen.cell(row=i, column=j, value=cnt)
                row_total += cnt
            ws_resumen.cell(row=i, column=len(envs_resumen) + 2, value=row_total)

        total_row = len(tipos_list) + 2
        ws_resumen.cell(row=total_row, column=1, value="TOTAL")
        style_total_row(ws_resumen, total_row, len(envs_resumen) + 2)
        for j, env in enumerate(envs_resumen, 2):
            env_total = sum(
                len(data_frames.get(tipo, pd.DataFrame())[data_frames.get(tipo, pd.DataFrame())["ENTORNO"] == env])
                if not data_frames.get(tipo, pd.DataFrame()).empty else 0
                for tipo in tipos_list
            )
            ws_resumen.cell(row=total_row, column=j, value=env_total)
        grand_total = sum(len(data_frames.get(t, pd.DataFrame())) for t in tipos_list)
        ws_resumen.cell(row=total_row, column=len(envs_resumen) + 2, value=grand_total)

        ws_resumen.column_dimensions["A"].width = 20
        for j in range(2, len(envs_resumen) + 3):
            ws_resumen.column_dimensions[get_column_letter(j)].width = 12

        # Gráfico de barras agrupadas por entorno
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.style = 26
        chart.legend.position = 'r'
        chart.title = "Inventario Consolidado por Entorno"
        chart.y_axis.title = "Cantidad"
        chart.x_axis.title = "Recurso"
        data_ref = Reference(ws_resumen, min_col=2, max_col=len(envs_resumen) + 1,
                             min_row=1, max_row=len(tipos_list) + 1)
        cats_ref = Reference(ws_resumen, min_col=1, min_row=2, max_row=len(tipos_list) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 18
        chart.height = 12
        ws_resumen.add_chart(chart, "A" + str(total_row + 3))

        # =====================================================================
        # 2b. POR ENTORNO — Clusters/Deployments/Services/CloudSQL por dev/qa/stag
        # =====================================================================
        print("Generando → POR ENTORNO")
        ws_entorno = wb.create_sheet("POR ENTORNO")

        envs_ordered = ["DEV", "QA", "STAG", "PROD", "OTRO"]
        envs_present = []
        for e in envs_ordered:
            for tipo in TIPOS:
                df = data_frames.get(tipo, pd.DataFrame())
                if not df.empty and e in df["ENTORNO"].values:
                    if e not in envs_present:
                        envs_present.append(e)

        # Tabla
        ws_entorno["A1"] = "Entorno"
        for j, tipo in enumerate(TIPOS, 2):
            ws_entorno.cell(row=1, column=j, value=tipo.upper())
        ws_entorno.cell(row=1, column=len(TIPOS) + 2, value="TOTAL")
        style_header(ws_entorno, 1, len(TIPOS) + 2)

        for i, env in enumerate(envs_present, 2):
            ws_entorno.cell(row=i, column=1, value=env)
            row_total = 0
            for j, tipo in enumerate(TIPOS, 2):
                df = data_frames.get(tipo, pd.DataFrame())
                cnt = len(df[df["ENTORNO"] == env]) if not df.empty else 0
                ws_entorno.cell(row=i, column=j, value=cnt)
                row_total += cnt
            ws_entorno.cell(row=i, column=len(TIPOS) + 2, value=row_total)

        total_r = len(envs_present) + 2
        ws_entorno.cell(row=total_r, column=1, value="TOTAL")
        style_total_row(ws_entorno, total_r, len(TIPOS) + 2)
        for j, tipo in enumerate(TIPOS, 2):
            df = data_frames.get(tipo, pd.DataFrame())
            ws_entorno.cell(row=total_r, column=j, value=len(df) if not df.empty else 0)
        ws_entorno.cell(row=total_r, column=len(TIPOS) + 2,
                        value=sum(len(data_frames.get(t, pd.DataFrame())) for t in TIPOS))

        ws_entorno.column_dimensions["A"].width = 12
        for j in range(2, len(TIPOS) + 3):
            ws_entorno.column_dimensions[get_column_letter(j)].width = 16

        # Gráfico de barras agrupadas
        chart_env = BarChart()
        chart_env.type = "col"
        chart_env.grouping = "clustered"
        chart_env.style = 26
        chart_env.legend.position = 'r'
        chart_env.title = "Recursos por Entorno"
        chart_env.y_axis.title = "Cantidad"
        chart_env.x_axis.title = "Entorno"
        data_ref = Reference(ws_entorno, min_col=2, max_col=len(TIPOS) + 1,
                             min_row=1, max_row=len(envs_present) + 1)
        cats_ref = Reference(ws_entorno, min_col=1, min_row=2, max_row=len(envs_present) + 1)
        chart_env.add_data(data_ref, titles_from_data=True)
        chart_env.set_categories(cats_ref)
        chart_env.width = 22
        chart_env.height = 14
        ws_entorno.add_chart(chart_env, "A" + str(total_r + 3))

        # =====================================================================
        # 2c. POR PROYECTO — Consolidado por nombre base con entornos
        # =====================================================================
        print("Generando → POR PROYECTO")
        ws_proyecto = wb.create_sheet("POR PROYECTO")

        all_bases = set()
        for tipo in TIPOS:
            df = data_frames.get(tipo, pd.DataFrame())
            if not df.empty:
                all_bases.update(df["BASE"].unique())
        bases_sorted = sorted(all_bases)
        envs_proj = ["DEV", "QA", "STAG"]
        tipos_list = list(TIPOS.keys())

        # Tabla resumen: Proyecto | CLUSTERS_DEV | CLUSTERS_QA | CLUSTERS_STAG | ... | TOTAL
        ws_proyecto["A1"] = "Proyecto"
        col = 2
        tipo_col_start = {}
        for tipo in tipos_list:
            tipo_col_start[tipo] = col
            for env in envs_proj:
                ws_proyecto.cell(row=1, column=col, value=f"{tipo.upper()}_{env}")
                col += 1
            ws_proyecto.cell(row=1, column=col, value=f"{tipo.upper()}_TOTAL")
            col += 1
        total_col = col
        ws_proyecto.cell(row=1, column=total_col, value="TOTAL")
        style_header(ws_proyecto, 1, total_col)

        for i, base in enumerate(bases_sorted, 2):
            ws_proyecto.cell(row=i, column=1, value=base)
            grand_total = 0
            col = 2
            for tipo in tipos_list:
                df = data_frames.get(tipo, pd.DataFrame())
                tipo_total = 0
                for env in envs_proj:
                    cnt = len(df[(df["BASE"] == base) & (df["ENTORNO"] == env)]) if not df.empty else 0
                    ws_proyecto.cell(row=i, column=col, value=cnt)
                    tipo_total += cnt
                    grand_total += cnt
                    col += 1
                ws_proyecto.cell(row=i, column=col, value=tipo_total)
                col += 1
            ws_proyecto.cell(row=i, column=total_col, value=grand_total)

        total_r = len(bases_sorted) + 2
        ws_proyecto.cell(row=total_r, column=1, value="TOTAL")
        style_total_row(ws_proyecto, total_r, total_col)
        col = 2
        for tipo in tipos_list:
            tipo_grand = 0
            for env in envs_proj:
                env_sum = 0
                for tipo2 in tipos_list:
                    df = data_frames.get(tipo2, pd.DataFrame())
                    if tipo2 == tipo:
                        env_sum += len(df[df["ENTORNO"] == env]) if not df.empty else 0
                ws_proyecto.cell(row=total_r, column=col, value=env_sum)
                tipo_grand += env_sum
                col += 1
            ws_proyecto.cell(row=total_r, column=col, value=tipo_grand)
            col += 1
        ws_proyecto.cell(row=total_r, column=total_col,
                         value=sum(len(data_frames.get(t, pd.DataFrame())) for t in TIPOS))

        ws_proyecto.column_dimensions["A"].width = 35
        for j in range(2, total_col + 1):
            ws_proyecto.column_dimensions[get_column_letter(j)].width = 14

        # Tabla para gráfico de líneas: Proyecto | Entorno | CLUSTERS | DEPLOYMENTS | SERVICES | CLOUDSQL
        chart_start = total_r + 3
        ws_proyecto.cell(row=chart_start, column=1, value="Proyecto")
        ws_proyecto.cell(row=chart_start, column=2, value="Entorno")
        for j, tipo in enumerate(tipos_list, 3):
            ws_proyecto.cell(row=chart_start, column=j, value=tipo.upper())
        style_header(ws_proyecto, chart_start, len(tipos_list) + 2)

        row = chart_start + 1
        for base in bases_sorted:
            for env in envs_proj:
                ws_proyecto.cell(row=row, column=1, value=base)
                ws_proyecto.cell(row=row, column=2, value=env)
                for j, tipo in enumerate(tipos_list, 3):
                    df = data_frames.get(tipo, pd.DataFrame())
                    cnt = len(df[(df["BASE"] == base) & (df["ENTORNO"] == env)]) if not df.empty else 0
                    ws_proyecto.cell(row=row, column=j, value=cnt)
                row += 1
        chart_data_end = row - 1

        # Gráfico de líneas: series=CLUSTERS/DEPLOYMENTS/SERVICES/CLOUDSQL, eje X=Proyecto-Entorno
        chart_line = LineChart()
        chart_line.style = 26
        chart_line.legend.position = 'r'
        chart_line.title = "Proyectos por Entorno y Tipo de Recurso"
        chart_line.y_axis.title = "Cantidad"
        chart_line.x_axis.title = "Proyecto - Entorno"
        data_ref = Reference(ws_proyecto, min_col=3, max_col=len(tipos_list) + 2,
                             min_row=chart_start, max_row=chart_data_end)
        cats_ref = Reference(ws_proyecto, min_col=1, min_row=chart_start + 1,
                             max_row=chart_data_end)
        chart_line.add_data(data_ref, titles_from_data=True)
        chart_line.set_categories(cats_ref)
        for s in chart_line.series:
            s.graphicalProperties.line.width = 22000
        chart_line.width = 30
        chart_line.height = 16
        ws_proyecto.add_chart(chart_line, "A" + str(chart_data_end + 2))

        # =====================================================================
        # 2d. MACHINE TYPE — Distribución de tipos de máquina (clusters)
        # =====================================================================
        if "clusters" in data_frames and not data_frames["clusters"].empty:
            print("Generando → MACHINE TYPE")
            ws_mt = wb.create_sheet("MACHINE TYPE")

            df_cl = data_frames["clusters"]
            # Expandir machine types separados por |
            mt_rows = []
            for _, row in df_cl.iterrows():
                mts = str(row.get("MACHINE_TYPE", "")).split("|")
                for mt in mts:
                    mt = mt.strip()
                    if mt and mt != "":
                        mt_rows.append({"MACHINE_TYPE": mt, "ENTORNO": row.get("ENTORNO", "")})

            if mt_rows:
                df_mt = pd.DataFrame(mt_rows)
                mt_counts = df_mt["MACHINE_TYPE"].value_counts().reset_index()
                mt_counts.columns = ["MACHINE_TYPE", "CANTIDAD"]

                envs_mt = ["DEV", "QA", "STAG"]
                ws_mt["A1"] = "Machine Type"
                for j, env in enumerate(envs_mt, 2):
                    ws_mt.cell(row=1, column=j, value=env)
                ws_mt.cell(row=1, column=len(envs_mt) + 2, value="TOTAL")
                style_header(ws_mt, 1, len(envs_mt) + 2)
                for i, (_, r) in enumerate(mt_counts.iterrows(), 2):
                    mt_name = r["MACHINE_TYPE"]
                    ws_mt.cell(row=i, column=1, value=mt_name)
                    row_total = 0
                    for j, env in enumerate(envs_mt, 2):
                        cnt = len(df_mt[(df_mt["MACHINE_TYPE"] == mt_name) & (df_mt["ENTORNO"] == env)])
                        ws_mt.cell(row=i, column=j, value=cnt)
                        row_total += cnt
                    ws_mt.cell(row=i, column=len(envs_mt) + 2, value=row_total)
                total_r = len(mt_counts) + 2
                ws_mt.cell(row=total_r, column=1, value="TOTAL")
                style_total_row(ws_mt, total_r, len(envs_mt) + 2)
                for j, env in enumerate(envs_mt, 2):
                    ws_mt.cell(row=total_r, column=j, value=len(df_mt[df_mt["ENTORNO"] == env]))
                ws_mt.cell(row=total_r, column=len(envs_mt) + 2, value=int(mt_counts["CANTIDAD"].sum()))
                ws_mt.column_dimensions["A"].width = 25
                for j in range(2, len(envs_mt) + 3):
                    ws_mt.column_dimensions[get_column_letter(j)].width = 12
                # Stacked bar chart
                chart_mt = BarChart()
                chart_mt.type = "col"
                chart_mt.grouping = "stacked"
                chart_mt.style = 26
                chart_mt.legend.position = 'r'
                chart_mt.title = "Machine Types por Entorno"
                chart_mt.y_axis.title = "Cantidad"
                chart_mt.x_axis.title = "Machine Type"
                data_ref = Reference(ws_mt, min_col=2, max_col=len(envs_mt) + 1, min_row=1, max_row=len(mt_counts) + 1)
                cats_ref = Reference(ws_mt, min_col=1, min_row=2, max_row=len(mt_counts) + 1)
                chart_mt.add_data(data_ref, titles_from_data=True)
                chart_mt.set_categories(cats_ref)
                chart_mt.width = 22
                chart_mt.height = 14
                ws_mt.add_chart(chart_mt, "A" + str(total_r + 3))

        # =====================================================================
        # 2e. SERVICE TYPE — Distribución de tipos de Service (ClusterIP, etc.)
        # =====================================================================
        if "services" in data_frames and not data_frames["services"].empty:
            print("Generando → SERVICE TYPE")
            ws_st = wb.create_sheet("SERVICE TYPE")

            df_sv = data_frames["services"]
            st_counts = df_sv["TYPE"].value_counts().reset_index()
            st_counts.columns = ["SERVICE_TYPE", "CANTIDAD"]
            envs_st = ["DEV", "QA", "STAG"]
            bases_st = sorted(df_sv["BASE"].unique()) if "BASE" in df_sv.columns else []

            # Tabla: Service Type | DEV | QA | STAG | TOTAL
            ws_st["A1"] = "Service Type"
            for j, env in enumerate(envs_st, 2):
                ws_st.cell(row=1, column=j, value=env)
            ws_st.cell(row=1, column=len(envs_st) + 2, value="TOTAL")
            style_header(ws_st, 1, len(envs_st) + 2)
            for i, (_, r) in enumerate(st_counts.iterrows(), 2):
                st_name = r["SERVICE_TYPE"]
                ws_st.cell(row=i, column=1, value=st_name)
                row_total = 0
                for j, env in enumerate(envs_st, 2):
                    cnt = len(df_sv[(df_sv["TYPE"] == st_name) & (df_sv["ENTORNO"] == env)])
                    ws_st.cell(row=i, column=j, value=cnt)
                    row_total += cnt
                ws_st.cell(row=i, column=len(envs_st) + 2, value=row_total)
            total_r = len(st_counts) + 2
            ws_st.cell(row=total_r, column=1, value="TOTAL")
            style_total_row(ws_st, total_r, len(envs_st) + 2)
            for j, env in enumerate(envs_st, 2):
                ws_st.cell(row=total_r, column=j, value=len(df_sv[df_sv["ENTORNO"] == env]))
            ws_st.cell(row=total_r, column=len(envs_st) + 2, value=int(st_counts["CANTIDAD"].sum()))
            ws_st.column_dimensions["A"].width = 20
            for j in range(2, len(envs_st) + 3):
                ws_st.column_dimensions[get_column_letter(j)].width = 12

            # Stacked bar chart por entorno
            chart_st = BarChart()
            chart_st.type = "col"
            chart_st.grouping = "stacked"
            chart_st.style = 26
            chart_st.legend.position = 'r'
            chart_st.title = "Service Types por Entorno"
            chart_st.y_axis.title = "Cantidad"
            chart_st.x_axis.title = "Service Type"
            data_ref = Reference(ws_st, min_col=2, max_col=len(envs_st) + 1,
                                 min_row=1, max_row=len(st_counts) + 1)
            cats_ref = Reference(ws_st, min_col=1, min_row=2, max_row=len(st_counts) + 1)
            chart_st.add_data(data_ref, titles_from_data=True)
            chart_st.set_categories(cats_ref)
            chart_st.width = 22
            chart_st.height = 14
            ws_st.add_chart(chart_st, "A" + str(total_r + 3))

            # Tabla por proyecto: Service Type | Proyecto | DEV | QA | STAG | TOTAL
            if bases_st:
                proj_start_row = total_r + 20
                ws_st.cell(row=proj_start_row, column=1, value="Service Type")
                ws_st.cell(row=proj_start_row, column=2, value="Proyecto")
                for j, env in enumerate(envs_st, 3):
                    ws_st.cell(row=proj_start_row, column=j, value=env)
                ws_st.cell(row=proj_start_row, column=len(envs_st) + 3, value="TOTAL")
                style_header(ws_st, proj_start_row, len(envs_st) + 3)
                row = proj_start_row + 1
                for st_name in st_counts["SERVICE_TYPE"]:
                    for base in bases_st:
                        sub = df_sv[(df_sv["TYPE"] == st_name) & (df_sv["BASE"] == base)]
                        if sub.empty:
                            continue
                        ws_st.cell(row=row, column=1, value=st_name)
                        ws_st.cell(row=row, column=2, value=base)
                        row_total = 0
                        for j, env in enumerate(envs_st, 3):
                            cnt = len(sub[sub["ENTORNO"] == env])
                            ws_st.cell(row=row, column=j, value=cnt)
                            row_total += cnt
                        ws_st.cell(row=row, column=len(envs_st) + 3, value=row_total)
                        row += 1
                ws_st.column_dimensions["B"].width = 30

        # =====================================================================
        # 2f. RADAR PROYECTO - Ejes=CLUSTERS/DEPLOYMENTS/SERVICES/CLOUDSQL/CLOUDSQL_DATABASES/INGRESS/CLOUDRUN/PUBSUB
        #     Series=DEV/QA/STAG (lineas, no relleno)
        # =====================================================================
        if data_frames:
            print("Generando -> RADAR PROYECTO")
            ws_radar = wb.create_sheet("RADAR PROYECTO")

            envs_for_radar = ["DEV", "QA", "STAG"]
            tipos_list = list(TIPOS.keys())

            base_env_counts = {}
            for tipo in tipos_list:
                df = data_frames.get(tipo, pd.DataFrame())
                if df.empty:
                    continue
                for base in df["BASE"].unique():
                    sub = df[df["BASE"] == base]
                    envs_found = set(sub["ENTORNO"].unique()) & set(envs_for_radar)
                    base_env_counts.setdefault(base, set()).update(envs_found)

            bases_with_data = sorted(b for b, ev in base_env_counts.items() if ev)

            if bases_with_data:
                current_row = 1

                for base in bases_with_data:
                    ws_radar.cell(row=current_row, column=1, value=base)
                    ws_radar.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="1F4E79")

                    # Tabla transpuesta: filas=tipos recurso (ejes), columnas=entornos (series)
                    header_row = current_row + 1
                    ws_radar.cell(row=header_row, column=1, value="Recurso")
                    for j, env in enumerate(envs_for_radar, 2):
                        ws_radar.cell(row=header_row, column=j, value=env)
                    style_header(ws_radar, header_row, len(envs_for_radar) + 1)

                    for i, tipo in enumerate(tipos_list):
                        data_row = header_row + 1 + i
                        ws_radar.cell(row=data_row, column=1, value=tipo.upper())
                        for j, env in enumerate(envs_for_radar, 2):
                            df = data_frames.get(tipo, pd.DataFrame())
                            if df.empty:
                                cnt = 0
                            else:
                                cnt = len(df[(df["BASE"] == base) & (df["ENTORNO"] == env)])
                            ws_radar.cell(row=data_row, column=j, value=cnt)

                    last_data_row = header_row + len(tipos_list)

                    # Radar chart: lineas (no relleno) para ver todas las series
                    chart_radar = RadarChart()
                    chart_radar.title = f"{base} - Perfil por Entorno"
                    # Titulo en esquina superior izquierda
                    chart_radar.title.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode='edge', yMode='edge', x=0.0, y=0.0, w=0.5, h=0.1
                        )
                    )
                    # Leyenda en esquina inferior derecha
                    chart_radar.legend.position = 'tr'
                    chart_radar.legend.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode='edge', yMode='edge', x=0.7, y=0.85, w=0.28, h=0.12
                        )
                    )
                    chart_radar.legend.overlay = True
                    # Chart area: degradado radial oscuro (centro mas claro, bordes oscuros)
                    chart_radar.graphical_properties = GraphicalProperties(
                        gradFill=GradientFillProperties(
                            path=PathShadeProperties(path='circle'),
                            gsLst=(
                                GradientStop(pos=0, srgbClr='555555'),
                                GradientStop(pos=100000, srgbClr='2d2d2d')
                            )
                        )
                    )
                    # Plot area: transparente (usar graphicalProperties camelCase)
                    chart_radar.plot_area.graphicalProperties = GraphicalProperties(noFill=True)
                    # Texto gris claro para contrastar con fondo oscuro
                    light_cp = CharacterProperties(sz=900, solidFill='D9D9D9')
                    light_txPr = RichText(
                        p=[Paragraph(
                            pPr=ParagraphProperties(defRPr=light_cp),
                            endParaRPr=CharacterProperties()
                        )]
                    )
                    chart_radar.title.txPr = light_txPr
                    # Aplicar color directamente al run del titulo (txPr solo afecta defaults)
                    chart_radar.title.text.rich.paragraphs[0].r[0].rPr = light_cp
                    # Ejes y leyenda tambien en gris claro
                    for ax in [chart_radar.y_axis, chart_radar.x_axis]:
                        ax.txPr = light_txPr
                        ax.delete = False
                        ax.tickLblPos = 'nextTo'
                    chart_radar.legend.txPr = light_txPr
                    gridlines_gp = GraphicalProperties()
                    gridlines_gp.line = LineProperties(solidFill='D9D9D9')
                    chart_radar.y_axis.majorGridlines.spPr = gridlines_gp

                    data_ref = Reference(ws_radar, min_col=2, max_col=len(envs_for_radar) + 1,
                                         min_row=header_row, max_row=last_data_row)
                    cats_ref = Reference(ws_radar, min_col=1, min_row=header_row + 1,
                                         max_row=last_data_row)
                    chart_radar.add_data(data_ref, titles_from_data=True)
                    chart_radar.set_categories(cats_ref)
                    chart_radar.width = 16
                    chart_radar.height = 12

                    chart_col = len(envs_for_radar) + 3
                    chart_cell = f"{get_column_letter(chart_col)}{current_row}"
                    ws_radar.add_chart(chart_radar, chart_cell)

                    current_row = last_data_row + 3

                ws_radar.column_dimensions["A"].width = 16
                for j in range(2, len(envs_for_radar) + 2):
                    ws_radar.column_dimensions[get_column_letter(j)].width = 12

if not archivos_encontrados:
    print("\n⚠️ No se encontraron archivos CSV. Ejecuta primero el script Bash.")
else:
    print("\n" + "=" * 90)
    print("¡ÉXITO! Archivo Excel generado correctamente")
    print(f"   → {OUTPUT_EXCEL.name}")
    print(f"   → Ruta completa: {OUTPUT_EXCEL.resolve()}")
    print("\nHojas de datos  : CLUSTERS | DEPLOYMENTS | SERVICES | CLOUDSQL | CLOUDSQL_DATABASES | INGRESS | CLOUDRUN | PUBSUB")
    print("Hojas gráficos : RESUMEN | POR ENTORNO | POR PROYECTO | MACHINE TYPE | SERVICE TYPE | RADAR PROYECTO")

print("=" * 90)
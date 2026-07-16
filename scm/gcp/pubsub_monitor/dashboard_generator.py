"""
DashboardGenerator - Generador de dashboards y reportes

Módulo responsable de generar dashboards HTML, reportes JSON
y exportación a Excel.

Características:
- Dashboard HTML interactivo
- Reportes JSON estructurados
- Exportación Excel
- Gráficos con Plotly
"""

import json
import logging
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path

from rich.console import Console

console = Console()
logger = logging.getLogger(__name__)


class DashboardGenerator:
    """Genera dashboards y reportes."""

    def __init__(self, results: Dict):
        """
        Inicializa el generador.

        Args:
            results: Resultados del análisis
        """
        self.results = results
        self.timestamp = datetime.now().isoformat()

    def generate_html_dashboard(self, output_path: str) -> str:
        """
        Genera dashboard HTML.

        Args:
            output_path: Ruta de salida

        Returns:
            Ruta del archivo generado
        """
        html = self._build_html_structure()

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)

        console.print(f"[green]✅ Dashboard HTML generado:[/green] {output_file}")
        return str(output_file)

    def _build_html_structure(self) -> str:
        """Construye la estructura HTML."""
        projects = self.results.get("projects", {})

        html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Pub/Sub Monitor Dashboard</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        
        header {{
            background: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        h1 {{
            color: #333;
            margin-bottom: 10px;
        }}
        
        .timestamp {{
            color: #666;
            font-size: 14px;
        }}
        
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .metric-card {{
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            text-align: center;
        }}
        
        .metric-value {{
            font-size: 32px;
            font-weight: bold;
            color: #667eea;
            margin: 10px 0;
        }}
        
        .metric-label {{
            color: #666;
            font-size: 14px;
        }}
        
        .projects-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
        }}
        
        .project-card {{
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        .project-name {{
            font-size: 18px;
            font-weight: bold;
            color: #333;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 2px solid #667eea;
        }}
        
        .project-stat {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid #eee;
        }}
        
        .project-stat-label {{
            color: #666;
        }}
        
        .project-stat-value {{
            font-weight: bold;
            color: #333;
        }}
        
        .status-healthy {{
            color: #10b981;
        }}
        
        .status-warning {{
            color: #f59e0b;
        }}
        
        .status-critical {{
            color: #ef4444;
        }}
        
        footer {{
            text-align: center;
            color: white;
            margin-top: 30px;
            padding: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 Pub/Sub Monitor Dashboard</h1>
            <p class="timestamp">Generado: {self.timestamp}</p>
        </header>
        
        <div class="summary">
            <div class="metric-card">
                <div class="metric-label">Proyectos Monitoreados</div>
                <div class="metric-value">{len(projects)}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Topics Totales</div>
                <div class="metric-value">{sum(len(p.get('topics', [])) for p in projects.values())}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Subscriptions Totales</div>
                <div class="metric-value">{sum(len(p.get('subscriptions', [])) for p in projects.values())}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Alertas Activas</div>
                <div class="metric-value">{sum(len(p.get('alerts', [])) for p in projects.values())}</div>
            </div>
        </div>
        
        <h2 style="color: white; margin-bottom: 20px;">Proyectos</h2>
        <div class="projects-grid">
"""

        for project_id, project_data in projects.items():
            topics_count = len(project_data.get("topics", []))
            subs_count = len(project_data.get("subscriptions", []))
            alerts_count = len(project_data.get("alerts", []))

            html += f"""
            <div class="project-card">
                <div class="project-name">{project_id}</div>
                <div class="project-stat">
                    <span class="project-stat-label">Topics:</span>
                    <span class="project-stat-value">{topics_count}</span>
                </div>
                <div class="project-stat">
                    <span class="project-stat-label">Subscriptions:</span>
                    <span class="project-stat-value">{subs_count}</span>
                </div>
                <div class="project-stat">
                    <span class="project-stat-label">Alertas:</span>
                    <span class="project-stat-value {self._get_alert_class(alerts_count)}">{alerts_count}</span>
                </div>
            </div>
"""

        html += """
        </div>
        
        <footer>
            <p>DevSecOps Toolbox - Pub/Sub Monitor v1.0.0</p>
        </footer>
    </div>
</body>
</html>
"""

        return html

    def generate_json_report(self, output_path: str) -> str:
        """
        Genera reporte JSON.

        Args:
            output_path: Ruta de salida

        Returns:
            Ruta del archivo generado
        """
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)

        console.print(f"[green]✅ Reporte JSON generado:[/green] {output_file}")
        return str(output_file)

    def generate_excel_report(self, output_path: str) -> str:
        """
        Genera reporte Excel.

        Args:
            output_path: Ruta de salida

        Returns:
            Ruta del archivo generado
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            console.print("[yellow]⚠️ openpyxl no instalado. Saltando Excel.[/yellow]")
            return ""

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Hoja de resumen
        ws_summary = wb.create_sheet("Resumen")
        self._add_summary_sheet(ws_summary)

        # Hoja de proyectos
        ws_projects = wb.create_sheet("Proyectos")
        self._add_projects_sheet(ws_projects)

        # Hoja de alertas
        ws_alerts = wb.create_sheet("Alertas")
        self._add_alerts_sheet(ws_alerts)

        wb.save(output_file)
        console.print(f"[green]✅ Reporte Excel generado:[/green] {output_file}")
        return str(output_file)

    def _add_summary_sheet(self, ws) -> None:
        """Agrega hoja de resumen."""
        ws['A1'] = "Pub/Sub Monitor - Resumen"
        ws['A1'].font = Font(bold=True, size=14)

        projects = self.results.get("projects", {})
        ws['A3'] = "Proyectos Monitoreados"
        ws['B3'] = len(projects)

        ws['A4'] = "Topics Totales"
        ws['B4'] = sum(len(p.get('topics', [])) for p in projects.values())

        ws['A5'] = "Subscriptions Totales"
        ws['B5'] = sum(len(p.get('subscriptions', [])) for p in projects.values())

        ws['A6'] = "Alertas Activas"
        ws['B6'] = sum(len(p.get('alerts', [])) for p in projects.values())

    def _add_projects_sheet(self, ws) -> None:
        """Agrega hoja de proyectos."""
        ws['A1'] = "Proyecto"
        ws['B1'] = "Topics"
        ws['C1'] = "Subscriptions"
        ws['D1'] = "Alertas"

        projects = self.results.get("projects", {})
        row = 2

        for project_id, project_data in projects.items():
            ws[f'A{row}'] = project_id
            ws[f'B{row}'] = len(project_data.get("topics", []))
            ws[f'C{row}'] = len(project_data.get("subscriptions", []))
            ws[f'D{row}'] = len(project_data.get("alerts", []))
            row += 1

    def _add_alerts_sheet(self, ws) -> None:
        """Agrega hoja de alertas."""
        ws['A1'] = "Proyecto"
        ws['B1'] = "Severidad"
        ws['C1'] = "Categoría"
        ws['D1'] = "Título"
        ws['E1'] = "Recomendación"

        projects = self.results.get("projects", {})
        row = 2

        for project_id, project_data in projects.items():
            for alert in project_data.get("alerts", []):
                ws[f'A{row}'] = project_id
                ws[f'B{row}'] = alert.get("severity")
                ws[f'C{row}'] = alert.get("category")
                ws[f'D{row}'] = alert.get("title")
                ws[f'E{row}'] = alert.get("recommendation")
                row += 1

    def _get_alert_class(self, count: int) -> str:
        """Obtiene clase CSS para alertas."""
        if count == 0:
            return "status-healthy"
        elif count <= 3:
            return "status-warning"
        else:
            return "status-critical"

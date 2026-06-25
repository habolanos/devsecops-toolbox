#!/usr/bin/env python3
"""
Export Manager - Módulo centralizado para estandarizar exportación de datos

Proporciona funciones consistentes para exportar datos a JSON, CSV y Excel
con estructura y nombres de archivo estandarizados.
"""

import json
import csv
import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Any
from zoneinfo import ZoneInfo

__version__ = "1.0.0"


# --- Directorio de salida centralizado (DEVSECOPS_OUTPUT_DIR) ---
try:
    from utils import get_output_dir
except ImportError:
    import os as _os
    from pathlib import Path as _Path
    def get_output_dir(default="."):
        env = _os.getenv("DEVSECOPS_OUTPUT_DIR")
        if env:
            p = _Path(env)
            p.mkdir(parents=True, exist_ok=True)
            return p
        p = _Path(default)
        p.mkdir(parents=True, exist_ok=True)
        return p
# -------------------------------------------------------------------


class ExportManager:
    """Gestor centralizado de exportación de datos."""
    
    def __init__(self, tool_name: str, tool_version: str = "1.0.0"):
        """
        Inicializa el gestor de exportación.
        
        Args:
            tool_name: Nombre de la herramienta (ej: 'pr_master_checker')
            tool_version: Versión de la herramienta
        """
        self.tool_name = tool_name
        self.tool_version = tool_version
        self.output_dir = get_output_dir("outcome")
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def export_json(
        self,
        data: List[Dict],
        metadata: Optional[Dict] = None,
        summary: Optional[Dict] = None,
        organization: Optional[str] = None,
        project: Optional[str] = None,
        timezone: str = "UTC"
    ) -> str:
        """
        Exporta datos a JSON con estructura estandarizada.
        
        Args:
            data: Lista de diccionarios con los datos
            metadata: Metadata adicional (opcional)
            summary: Resumen de datos (opcional)
            organization: Nombre de la organización (opcional)
            project: Nombre del proyecto (opcional)
            timezone: Zona horaria para timestamp
            
        Returns:
            Ruta del archivo generado
        """
        filepath = self.output_dir / f"{self.tool_name}_{self.timestamp}.json"
        
        # Construir metadata estándar
        std_metadata = {
            "tool": self.tool_name,
            "version": self.tool_version,
            "generated_at": datetime.now(ZoneInfo(timezone)).isoformat(),
        }
        
        if organization:
            std_metadata["organization"] = organization
        if project:
            std_metadata["project"] = project
        
        # Agregar metadata adicional si existe
        if metadata:
            std_metadata.update(metadata)
        
        # Construir resumen estándar
        std_summary = {
            "total": len(data),
            "filtered": len(data),
            "status": "success" if data else "empty"
        }
        
        # Agregar resumen adicional si existe
        if summary:
            std_summary.update(summary)
        
        # Construir payload
        payload = {
            "metadata": std_metadata,
            "summary": std_summary,
            "data": data
        }
        
        # Escribir archivo
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False, default=str)
        
        return str(filepath)
    
    def export_csv(
        self,
        data: List[Dict],
        fieldnames: Optional[List[str]] = None
    ) -> Optional[str]:
        """
        Exporta datos a CSV con estructura estandarizada.
        
        Args:
            data: Lista de diccionarios con los datos
            fieldnames: Nombres de columnas (si None, usa keys del primer item)
            
        Returns:
            Ruta del archivo generado o None si no hay datos
        """
        if not data:
            return None
        
        filepath = self.output_dir / f"{self.tool_name}_{self.timestamp}.csv"
        
        # Determinar fieldnames
        if fieldnames is None:
            fieldnames = list(data[0].keys())
        
        # Escribir archivo
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        
        return str(filepath)
    
    def export_excel(
        self,
        data: List[Dict],
        sheet_name: str = "Data",
        metadata: Optional[Dict] = None,
        summary: Optional[Dict] = None,
        include_metadata_sheet: bool = True,
        include_summary_sheet: bool = True
    ) -> Optional[str]:
        """
        Exporta datos a Excel con estructura estandarizada.
        
        Args:
            data: Lista de diccionarios con los datos
            sheet_name: Nombre de la hoja principal
            metadata: Metadata a incluir en hoja separada (opcional)
            summary: Resumen a incluir en hoja separada (opcional)
            include_metadata_sheet: Incluir hoja de metadata
            include_summary_sheet: Incluir hoja de resumen
            
        Returns:
            Ruta del archivo generado o None si no hay datos o falta pandas
        """
        if not data:
            return None
        
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            filepath = self.output_dir / f"{self.tool_name}_{self.timestamp}.xlsx"
            
            # Crear DataFrame principal
            df_main = pd.DataFrame(data)
            
            # Crear ExcelWriter
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                # Hoja principal
                df_main.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Hoja de metadata
                if include_metadata_sheet and metadata:
                    meta_data = {
                        "Property": list(metadata.keys()),
                        "Value": [str(v) for v in metadata.values()]
                    }
                    df_meta = pd.DataFrame(meta_data)
                    df_meta.to_excel(writer, sheet_name="Metadata", index=False)
                
                # Hoja de resumen
                if include_summary_sheet and summary:
                    summary_data = {
                        "Metric": list(summary.keys()),
                        "Value": [str(v) for v in summary.values()]
                    }
                    df_summary = pd.DataFrame(summary_data)
                    df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            # Aplicar formato básico
            wb = load_workbook(filepath)
            for ws in wb.sheetnames:
                worksheet = wb[ws]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            return str(filepath)
        
        except ImportError:
            return None
    
    def export_all(
        self,
        data: List[Dict],
        formats: List[str] = ["json", "csv", "excel"],
        metadata: Optional[Dict] = None,
        summary: Optional[Dict] = None,
        organization: Optional[str] = None,
        project: Optional[str] = None,
        timezone: str = "UTC",
        fieldnames: Optional[List[str]] = None,
        sheet_name: str = "Data"
    ) -> Dict[str, Optional[str]]:
        """
        Exporta datos a múltiples formatos.
        
        Args:
            data: Lista de diccionarios con los datos
            formats: Formatos a exportar (json, csv, excel)
            metadata: Metadata adicional
            summary: Resumen adicional
            organization: Nombre de la organización
            project: Nombre del proyecto
            timezone: Zona horaria
            fieldnames: Nombres de columnas para CSV
            sheet_name: Nombre de hoja para Excel
            
        Returns:
            Diccionario con rutas de archivos generados
        """
        results = {}
        
        if "json" in formats:
            try:
                results["json"] = self.export_json(
                    data, metadata, summary, organization, project, timezone
                )
            except Exception as e:
                results["json"] = f"Error: {str(e)}"
        
        if "csv" in formats:
            try:
                results["csv"] = self.export_csv(data, fieldnames)
            except Exception as e:
                results["csv"] = f"Error: {str(e)}"
        
        if "excel" in formats:
            try:
                results["excel"] = self.export_excel(
                    data, sheet_name, metadata, summary
                )
            except Exception as e:
                results["excel"] = f"Error: {str(e)}"
        
        return results


def export_json_simple(
    data: List[Dict],
    tool_name: str,
    tool_version: str = "1.0.0",
    organization: Optional[str] = None,
    project: Optional[str] = None,
    timezone: str = "UTC",
    metadata: Optional[Dict] = None,
    summary: Optional[Dict] = None
) -> str:
    """
    Función simplificada para exportar JSON.
    
    Args:
        data: Lista de diccionarios
        tool_name: Nombre de la herramienta
        tool_version: Versión de la herramienta
        organization: Nombre de la organización
        project: Nombre del proyecto
        timezone: Zona horaria
        metadata: Metadata adicional
        summary: Resumen adicional
        
    Returns:
        Ruta del archivo generado
    """
    manager = ExportManager(tool_name, tool_version)
    return manager.export_json(data, metadata, summary, organization, project, timezone)


def export_csv_simple(
    data: List[Dict],
    tool_name: str,
    fieldnames: Optional[List[str]] = None
) -> Optional[str]:
    """
    Función simplificada para exportar CSV.
    
    Args:
        data: Lista de diccionarios
        tool_name: Nombre de la herramienta
        fieldnames: Nombres de columnas
        
    Returns:
        Ruta del archivo generado o None
    """
    manager = ExportManager(tool_name)
    return manager.export_csv(data, fieldnames)


def export_excel_simple(
    data: List[Dict],
    tool_name: str,
    sheet_name: str = "Data",
    metadata: Optional[Dict] = None,
    summary: Optional[Dict] = None
) -> Optional[str]:
    """
    Función simplificada para exportar Excel.
    
    Args:
        data: Lista de diccionarios
        tool_name: Nombre de la herramienta
        sheet_name: Nombre de la hoja
        metadata: Metadata adicional
        summary: Resumen adicional
        
    Returns:
        Ruta del archivo generado o None
    """
    manager = ExportManager(tool_name)
    return manager.export_excel(data, sheet_name, metadata, summary)


if __name__ == "__main__":
    # Ejemplo de uso
    sample_data = [
        {"id": 1, "name": "Item 1", "status": "active"},
        {"id": 2, "name": "Item 2", "status": "inactive"},
        {"id": 3, "name": "Item 3", "status": "active"},
    ]
    
    manager = ExportManager("test_tool", "1.0.0")
    
    print("Exportando a JSON...")
    json_file = manager.export_json(
        sample_data,
        organization="test-org",
        project="test-project",
        summary={"active": 2, "inactive": 1}
    )
    print(f"✅ JSON: {json_file}")
    
    print("Exportando a CSV...")
    csv_file = manager.export_csv(sample_data)
    print(f"✅ CSV: {csv_file}")
    
    print("Exportando a Excel...")
    excel_file = manager.export_excel(
        sample_data,
        sheet_name="Test Data",
        summary={"active": 2, "inactive": 1}
    )
    print(f"✅ Excel: {excel_file}")

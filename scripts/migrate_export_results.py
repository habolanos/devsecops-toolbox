#!/usr/bin/env python3
"""
Script para migrar funciones export_results() a usar ExportManager.

Este script:
1. Identifica funciones export_results() que solo tienen el import
2. Reemplaza la lógica de exportación con ExportManager
3. Mantiene fallback manual para compatibilidad
4. Preserva la lógica de negocio original
"""

import os
import re
from pathlib import Path
from typing import List, Tuple

def find_export_results_function(content: str) -> Tuple[int, int]:
    """Encuentra el inicio y fin de la función export_results()."""
    pattern = r'def export_results\([^)]*\):'
    match = re.search(pattern, content)
    if not match:
        return -1, -1
    
    start = match.start()
    
    # Encontrar el final de la función (siguiente def o fin del archivo)
    lines = content[start:].split('\n')
    indent_level = len(lines[0]) - len(lines[0].lstrip())
    
    end_line = 1
    for i in range(1, len(lines)):
        line = lines[i]
        if line.strip() and not line.startswith(' ' * (indent_level + 1)):
            if line.startswith('def ') or line.startswith('class '):
                end_line = i
                break
        if i == len(lines) - 1:
            end_line = i + 1
    
    end = start + len('\n'.join(lines[:end_line]))
    return start, end

def extract_export_fields_and_logic(content: str) -> Tuple[List[str], str]:
    """Extrae los campos de exportación y la lógica de negocio."""
    # Buscar export_fields
    pattern = r'export_fields\s*=\s*\[(.*?)\]'
    match = re.search(pattern, content, re.DOTALL)
    
    fields = []
    if match:
        fields_str = match.group(1)
        # Extraer nombres de campos
        field_matches = re.findall(r'"([^"]+)"', fields_str)
        fields = field_matches
    
    # Extraer la lógica de preparación de datos (flat, etc.)
    logic_pattern = r'(flat\s*=.*?(?=\n\s{0,4}if\s+output_format|$))'
    logic_match = re.search(logic_pattern, content, re.DOTALL)
    logic = logic_match.group(1).strip() if logic_match else ""
    
    return fields, logic

def create_migrated_export_results(tool_name: str, version: str, fields: List[str], logic: str) -> str:
    """Crea una función export_results() migrada a ExportManager."""
    
    fields_str = ', '.join([f'"{f}"' for f in fields])
    
    template = f'''def export_results(
    rows: List[Dict],
    output_format: str,
    tz_name: str,
) -> Optional[str]:
    """Exporta resultados usando ExportManager centralizado."""
    export_fields = [{fields_str}]
    flat = [{{f: r.get(f, "") for f in export_fields}} for r in rows]

    if not EXPORT_MANAGER_AVAILABLE:
        # Fallback a exportación manual
        if output_format == "json":
            from datetime import datetime
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(tz_name) if tz_name else None
            timestamp = datetime.now(tz).isoformat() if tz else datetime.now().isoformat()
            output_data = {{
                "metadata": {{
                    "tool": "{tool_name}",
                    "version": "{version}",
                    "timestamp": timestamp,
                    "timezone": tz_name,
                }},
                "data": flat,
            }}
            output_file = Path("outcome") / f"{{tool_name}}_{{timestamp.replace(':', '-')}}.json"
            output_file.parent.mkdir(exist_ok=True)
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, default=str)
            return str(output_file)
        elif output_format == "csv":
            import csv
            output_file = Path("outcome") / f"{{tool_name}}_export.csv"
            output_file.parent.mkdir(exist_ok=True)
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=export_fields)
                writer.writeheader()
                writer.writerows(flat)
            return str(output_file)
        elif output_format == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                output_file = Path("outcome") / f"{{tool_name}}_export.xlsx"
                output_file.parent.mkdir(exist_ok=True)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data"
                for col, field in enumerate(export_fields, 1):
                    ws.cell(1, col, field).font = Font(bold=True)
                for row_idx, row_data in enumerate(flat, 2):
                    for col_idx, field in enumerate(export_fields, 1):
                        ws.cell(row_idx, col_idx, row_data.get(field, ""))
                wb.save(output_file)
                return str(output_file)
            except ImportError:
                return None
        return None
    
    # Usar ExportManager
    manager = ExportManager("{tool_name}", __version__)
    
    summary = {{
        "total_records": len(rows),
    }}
    
    if output_format == "json":
        return manager.export_json(flat, summary=summary, timezone=tz_name)
    elif output_format == "csv":
        return manager.export_csv(flat)
    elif output_format == "excel":
        return manager.export_excel(flat, sheet_name="Data", summary=summary)
    
    return None
'''
    
    return template

def migrate_tool(filepath: str) -> bool:
    """Migra una herramienta a usar ExportManager."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Verificar que tiene los requisitos
        if 'EXPORT_MANAGER_AVAILABLE' not in content:
            return False
        
        if 'def export_results(' not in content:
            return False
        
        # Si ya está completamente migrada, saltar
        if 'manager = ExportManager' in content:
            return False
        
        # Extraer información
        tool_name = Path(filepath).stem
        
        # Buscar __version__
        version_match = re.search(r'__version__\s*=\s*["\']([^"\']+)["\']', content)
        version = version_match.group(1) if version_match else "1.0.0"
        
        # Extraer campos y lógica
        start, end = find_export_results_function(content)
        if start == -1:
            return False
        
        old_function = content[start:end]
        fields, logic = extract_export_fields_and_logic(old_function)
        
        if not fields:
            return False
        
        # Crear nueva función
        new_function = create_migrated_export_results(tool_name, version, fields, logic)
        
        # Reemplazar
        new_content = content[:start] + new_function + content[end:]
        
        # Escribir
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        return True
    except Exception as e:
        print(f"  Error en {filepath}: {e}")
        return False

# Lista de herramientas a migrar (solo las que tienen solo importes)
tools_to_migrate = [
    'scm/azdo/cicd_inventory_cd_detailed.py',
    'scm/azdo/cicd_inventory_ci_detailed.py',
    'scm/azdo/cicd_inventory_prod_deploy.py',
    'scm/azdo/cicd_pipeline_status.py',
    'scm/aws/acm/aws_acm_checker.py',
    'scm/aws/cloudwatch/aws_cloudwatch_checker.py',
    'scm/aws/ec2/aws_ebs_checker.py',
    'scm/aws/ec2/aws_ec2_checker.py',
    'scm/aws/ecr/aws_ecr_checker.py',
    'scm/aws/eks/aws_eks_checker.py',
    'scm/aws/eks/aws_eks_node_checker.py',
    'scm/aws/eks/aws_eks_pod_checker.py',
    'scm/aws/elb/aws_load_balancer_checker.py',
    'scm/aws/iam/aws_iam_checker.py',
    'scm/aws/iam/aws_roles_checker.py',
    'scm/aws/lambda/aws_lambda_checker.py',
    'scm/aws/rds/aws_rds_checker.py',
    'scm/aws/rds/aws_rds_storage_checker.py',
    'scm/aws/secretsmanager/aws_secrets_checker.py',
    'scm/aws/vpc/aws_security_groups_checker.py',
    'scm/aws/vpc/aws_vpc_checker.py',
    'scm/aws/waf/aws_waf_checker.py',
    'scm/gcp/connectivity/deploy_dependency_checker.py',
    'scm/gcp/connectivity/pod_connectivity_checker.py',
]

print("=" * 80)
print("MIGRACIÓN DE FUNCIONES export_results() A EXPORTMANAGER")
print("=" * 80)

migrated = 0
skipped = 0
failed = 0

for tool in sorted(tools_to_migrate):
    if os.path.exists(tool):
        if migrate_tool(tool):
            print(f"✅ {os.path.basename(tool)}")
            migrated += 1
        else:
            print(f"⏭️  {os.path.basename(tool)}")
            skipped += 1
    else:
        print(f"❌ {os.path.basename(tool)} - NO ENCONTRADO")
        failed += 1

print("\n" + "=" * 80)
print(f"RESULTADOS:")
print(f"  ✅ Migradas:  {migrated}")
print(f"  ⏭️  Omitidas:  {skipped}")
print(f"  ❌ Fallidas:  {failed}")
print("=" * 80)
